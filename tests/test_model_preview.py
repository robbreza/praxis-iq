"""Pin the IR-Inbox model-attachment fixes:
 1. _fmt_xlsx_cell renders Excel number formats (percent / currency / plain) for the in-app preview,
 2. the rebuilt demo model is a valid, cleanly-encoded .xlsx (no mojibake) with the right numbers.
Both are pure (no DB / no NiceGUI), so they run in the standard suite."""
import io

from core.demo_model import build_model_xlsx
from page_modules_nicegui.investors_page import _fmt_xlsx_cell


def test_fmt_xlsx_cell_formats():
    assert _fmt_xlsx_cell(0.264, "0.0%") == "26.4%"
    assert _fmt_xlsx_cell(0.30, "0%") == "30%"
    assert _fmt_xlsx_cell(42.7, "$#,##0.00") == "$42.70"
    assert _fmt_xlsx_cell(0.13, "$0.00") == "$0.13"
    assert _fmt_xlsx_cell(104.6, "#,##0.0") == "104.6"
    assert _fmt_xlsx_cell(1400, "#,##0") == "1,400"
    assert _fmt_xlsx_cell("BUY", "General") == "BUY"      # text passes through
    assert _fmt_xlsx_cell(None, "0.0%") == ""             # blank stays blank


def test_build_model_xlsx_is_clean_and_valid():
    import openpyxl
    raw = build_model_xlsx()
    assert raw[:2] == b"PK"                                # valid OOXML zip
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    text_cells, values = [], {}
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str):
                text_cells.append(v)
    blob = " ".join(text_cells)
    assert "�" not in blob                            # NO mojibake / replacement chars
    assert "—" in blob                                     # the em-dash survived as real Unicode
    assert "Northlake Payments, Inc.  (NASDAQ: NLKP)" in blob
    assert "Software & Platform" in blob
    # a couple of anchor numbers are present and correct
    flat = [c.value for r in ws.iter_rows() for c in r]
    assert 42.7 in flat and 25.9 in flat and 0.13 in flat
