"""Tag contacts that opened BOTH Lytham 1x1 invites (ProStar + IBI, Jan 2022) as DOUBLY-ENGAGED —
a strong 'takes small-cap meetings' signal. Recovers the cross-campaign signal that was lost when
provenance was overwritten on the second ingest. Matches by opener email; appends a queryable
'DOUBLY-ENGAGED' marker + bumps validation_status/confidence. Idempotent (skips already-marked)."""
import openpyxl
from core.security import load_environment, get_database_url; load_environment()
import psycopg2
def openers(f):
    ws=openpyxl.load_workbook(f, read_only=True, data_only=True)["Opens"]
    hdr=[str(c).lower() if c else "" for c in next(ws.iter_rows(values_only=True))]
    ei=next((i for i,h in enumerate(hdr) if "email" in h), None)
    return {(r[ei] or "").strip().lower() for r in ws.iter_rows(min_row=2, values_only=True) if r and ei is not None and len(r)>ei and r[ei]}
both=sorted(e for e in (openers(r"E:\MAPS Email Campaign Report as of 012522.xlsx")
                        & openers(r"E:\Copy of IBI Group Email Report as of 013122.xlsx")) if "@" in e)
conn=psycopg2.connect(get_database_url()); cur=conn.cursor()
cur.execute("""UPDATE contacts SET provenance = COALESCE(provenance,'') || %s, validation_status='engaged',
               confidence = GREATEST(COALESCE(confidence,0),90), updated_at=now()
               WHERE lower(email)=ANY(%s) AND (provenance IS NULL OR provenance NOT LIKE '%%DOUBLY-ENGAGED%%')""",
            (" | DOUBLY-ENGAGED: opened BOTH ProStar + IBI 1x1 invites (Jan 2022) — strong small-cap meeting-taker signal", both))
print("tagged:", cur.rowcount); conn.commit(); conn.close()
