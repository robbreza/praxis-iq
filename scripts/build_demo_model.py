"""Rebuild the illustrative demo tenant's sell-side analyst model as a clean .xlsx and
write it over the stored document, so the IR Inbox "model" attachment pulls up a legible,
properly-encoded spreadsheet (the previous ad-hoc file had mojibake baked into a few cells).

Idempotent: re-run any time. Finds the demo tenant's `model` document and replaces its bytes
in place (same doc_id, so the inbox queue item keeps resolving). Run from the project root:

    python scripts/build_demo_model.py

The model itself is illustrative — a fictional covering analyst (Cascade Securities / J. Meridian)
on Northlake Payments (NLKP), consistent with the rest of the demo tenant. Nothing real.
"""
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

CID = "demo"
FIRM = "Cascade Securities"
FILENAME = "NLKP_Q2_2026_Model.xlsx"
CTYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PERIODS = ["Q1'26A", "Q2'26E", "Q3'26E", "Q4'26E", "FY2026E", "FY2027E"]


def build_model_xlsx() -> bytes:
    """Construct the clean workbook and return its bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NLKP Model"

    ink = "1E293B"
    accent = "1E40AF"
    grey = "64748B"
    thin = Side(style="thin", color="D3DBE4")
    box = Border(bottom=thin)

    def put(cell, value, *, bold=False, size=11, color=ink, align="left", fill=None,
            italic=False, border=False, num=None):
        c = ws[cell]
        c.value = value
        c.font = Font(bold=bold, size=size, color=color, italic=italic,
                      name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if border:
            c.border = box
        if num:
            c.number_format = num

    # ── Header ──
    put("A1", "Northlake Payments, Inc.  (NASDAQ: NLKP)", bold=True, size=15, color=accent)
    put("A2", "Sell-Side Earnings Model — Q2 2026 Preview", bold=True, size=11, color=ink)
    put("A3", "Cascade Securities · Equity Research   |   J. Meridian, CFA   |   August 3, 2026",
        size=10, color=grey, italic=True)

    put("A5", "Rating", bold=True); put("B5", "BUY", bold=True, color="15803D")
    put("A6", "Price target"); put("B6", 42.70, num='$#,##0.00', align="right")
    put("A7", "Last price"); put("B7", 32.84, num='$#,##0.00', align="right")
    put("A8", "Implied upside"); put("B8", 0.30, num='0%', align="right", color="15803D", bold=True)

    # ── P&L table ──
    hdr = 10
    put(f"A{hdr}", "($ in millions, except per-share)", bold=True, size=10, color=grey, border=True)
    for j, p in enumerate(PERIODS):
        col = chr(ord("B") + j)
        put(f"{col}{hdr}", p, bold=True, size=10, color=ink, align="right", fill="EEF2F7", border=True)

    def row(label, vals, r, *, indent=False, num='#,##0.0', bold=False, color=ink):
        put(f"A{r}", ("   " + label) if indent else label, bold=bold, color=color)
        for j, v in enumerate(vals):
            col = chr(ord("B") + j)
            put(f"{col}{r}", v, align="right", num=(num if isinstance(v, (int, float)) else None),
                bold=bold, color=color)

    row("Revenue", [24.6, 25.9, 26.7, 27.4, 104.6, 117.2], 11, bold=True)
    row("Payment Processing", [18.7, 19.5, 19.9, 20.3, 78.4, 85.9], 12, indent=True, color=grey)
    row("Software & Platform", [5.9, 6.4, 6.8, 7.1, 26.2, 31.3], 13, indent=True, color=grey)
    row("Revenue growth (YoY)", [0.081, 0.089, 0.093, 0.101, 0.092, 0.120], 14, num='0.0%', color=grey)
    row("Gross profit", [6.5, 6.9, 7.2, 7.5, 28.1, 32.4], 15)
    row("Gross margin", [0.264, 0.266, 0.270, 0.274, 0.269, 0.276], 16, num='0.0%', color=grey)
    row("Operating income", [2.1, 2.3, 2.5, 2.7, 9.6, 12.1], 17)
    row("Operating margin", [0.085, 0.089, 0.094, 0.099, 0.092, 0.103], 18, num='0.0%', color=grey)
    row("Net income", [1.6, 1.7, 1.9, 2.0, 7.2, 9.1], 19)
    row("Diluted EPS", [0.11, 0.13, 0.13, 0.16, 0.53, 0.63], 20, num='$0.00', bold=True)
    row("Diluted shares (M)", [14.3, 14.3, 14.4, 14.4, 14.4, 14.5], 21, num='#,##0.0', color=grey)

    # ── Valuation ──
    put("A23", "Valuation", bold=True, size=11, color=accent)
    val = [
        ("Market capitalization", "$473M", "$32.84 × 14.4M shares"),
        ("Less: net cash", "($18M)", "cash less debt, 6/30 est."),
        ("Enterprise value", "$455M", ""),
        ("EV / Gross Profit (FY26E)", "16.2x", "peer median ~13x — premium on mix shift + growth"),
        ("EV / Gross Profit (FY27E)", "14.0x", ""),
        ("Price target basis", "$42.70", "~18x FY27E gross profit, ~52x FY27E EPS"),
    ]
    for i, (k, v, note) in enumerate(val):
        r = 24 + i
        put(f"A{r}", k)
        put(f"B{r}", v, align="right", bold=True)
        if note:
            put(f"C{r}", note, color=grey, italic=True, size=10)

    # ── Thesis ──
    tr = 24 + len(val) + 1
    put(f"A{tr}", "Thesis", bold=True, size=11, color=accent)
    thesis = (
        "NLKP is compounding Software & Platform revenue (25% of mix, +19% YoY) on top of a steady "
        "payment-processing base, lifting blended gross margin ~70bps/yr. We model Q2 revenue of "
        "$25.9M (Street $25.7M) and EPS of $0.12. A guide reiteration plus a software-attach update "
        "should support the multiple; our $42.70 target is ~18x FY27E gross profit."
    )
    put(f"A{tr + 1}", thesis, color=ink, size=10)
    ws[f"A{tr + 1}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{tr + 1}:G{tr + 4}")

    # column widths
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # so .env resolves
    import core.security as security
    security.load_environment()
    import config.client_config as cc
    from core import documents

    cc.reload_registry()
    if CID not in cc.CLIENT_REGISTRY:
        print(f"[demo model] tenant '{CID}' not found — run scripts/seed_illustrative_tenant.py first.")
        return 1
    cc.set_active_client_id(CID)

    doc = documents.get_latest_document(firm=FIRM, doc_type="model", client_id=CID)
    if not doc:
        # fall back to any model doc for the tenant
        docs = documents.list_documents(doc_type="model", client_id=CID)
        doc = docs[0] if docs else None
    if not doc:
        print("[demo model] no existing 'model' document found for the demo tenant; nothing to replace.")
        return 1

    xlsx = build_model_xlsx()
    ok = documents.update_document_bytes(doc["id"], xlsx, filename=FILENAME,
                                         content_type=CTYPE, client_id=CID)
    print(f"[demo model] {'updated' if ok else 'FAILED to update'} doc_id={doc['id']} "
          f"-> {FILENAME} ({len(xlsx):,} bytes), clean Unicode.")

    # Keep the inbox-queue item's displayed filename in sync with the document (the old file had a
    # " 2.xlsx" suffix that showed on the attachment button).
    from core import db
    queue = db.load_json("inbox_queue.json", [], client_id=CID) or []
    changed = 0
    for it in queue:
        if it.get("doc_id") == doc["id"] and it.get("filename") != FILENAME:
            it["filename"] = FILENAME
            changed += 1
    if changed:
        db.save_json("inbox_queue.json", queue, client_id=CID)
        print(f"[demo model] synced {changed} inbox-queue item filename(s) -> {FILENAME}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
