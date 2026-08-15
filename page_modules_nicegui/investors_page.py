"""
page_modules_nicegui/investors_page.py — Investors (buy-side intelligence,
NDR planning, meeting hub, target database), NiceGUI version.

This is the largest, most complex page in the original app (~5,000 lines of
app.py's "Investors" section). Ported here with the same 4-tab structure and
the same information architecture, but with these documented simplifications
(nothing silently dropped — everything below is still reachable in app.py):

- Big Picture synthesis (including the recency cross-check and the
  most-visited-city contrast warning), the pre/post-earnings mode toggle,
  the engagement funnel, institution cards, a per-institution meeting log
  with the 45-day repeat-meeting alert (including the +20 ephemeral score
  boost and warning when a repeat meeting is logged, and the global
  "Repeat Meeting Alerts" banner above the full institution list), the
  full Buy-Side filter set (tier, holder status, min score, Turnover
  Style, Metro, Digital Intent Score, Active/Passive ownership), and the
  dynamic pre-earnings outreach generator are all ported with full
  fidelity. NOT ported: the post-earnings "Update Intelligence Data" panel
  (CSV upload for Q2 Chorus Call listeners / IR visitor surge, plus a
  manual analyst-upgrade log form) — post-earnings scoring still runs, it
  just falls back to the Q1 call-listener flag instead of an uploaded Q2
  list until that panel is ported. Still available in app.py.
- Peer Cross-Targeting is rebuilt as a real Peer Universe manager: a
  persisted peer list (core.db, key "peer_universe.csv" — the same key
  config.client_config.CP() already reads, so a peer added/edited here is
  immediately available in outreach-draft valuation comparisons too) with
  add/remove/quick-add, plus a cross-targeting analysis with a
  peer multi-select AND a working minimum-AUM filter. Note: app.py's
  original min-AUM selectbox was computed but never actually applied to
  the results (a no-op left in the original) — this port applies it for
  real. Placed in the Target Database tab here (app.py nests it inside
  Buy-Side Intelligence) since it's a targeting/prospecting tool, which
  is where a user would look for it alongside the prospect search.
- Meeting Hub is ported with full fidelity for its 3 core sub-tabs (Upcoming
  Meetings, Schedule Meeting, Post-Meeting Notes), including a real Claude
  API call to structure raw meeting notes (same prompt/model as the original
  structure_notes_with_ai, now reading the key via core.security instead of
  a raw environment-variable read), with the same rule-based fallback if the
  API call fails. NOT yet ported: the "Current Full Meeting Schedule" table
  (a sortable dataframe of every scheduled meeting, not just upcoming ones)
  and its "Export schedule as CSV" / "Export meeting notes CSV"
  download buttons, and the Q&A Theme Tracker (a 4th Meeting Hub view that
  aggregates recurring questions/concerns/positives across every structured
  post-meeting note, to surface what to address on the next earnings call).
  All still available in app.py.
- Buy-Side Intelligence has 3 remaining sub-features not yet ported, still
  in app.py: the Chorus Call Listener Log (CSV upload of the actual Q1
  earnings-call listener export — name/firm/duration — with automatic
  "HOT LEAD" flagging of non-covering listeners and a one-click add to the
  New Analyst Onboarding Pipeline); Retail Investor Communication Priority
  (cross-references inbound IR questions against the NOBO shareholder file
  and IR website visitor log to triage who gets a call-back vs. a form
  email); the Closed-Loop Earnings Call Outreach generator (personalized
  pre-earnings emails with call logistics, reviewed/approved before sending
  via IRConnect); and the New Analyst Onboarding Tracker (a 6-stage
  pipeline — Cold through Evangelist — for every new sell-side/buy-side
  analyst, separate from the Engagement Funnel above, which only tracks
  already-known buy-side institutions). Analyst Coverage Network Targeting
  (cross-references institutions against a covering analyst's other covered
  tickers) IS now ported — see Target Database below, where it's grouped
  with the rest of the prospecting pipeline instead of Buy-Side Intelligence
  (same reasoning as Peer Cross-Targeting's placement, below).
- The "Scan IRConnect Inbox for Meeting Confirmations" feature is ported,
  but folded into the existing "Sync IR Inbox" pipeline above (see
  core/email_classifier.py's "meeting_confirmation" category) rather than
  rebuilt as its own ad hoc button — app.py's version prompted for the
  IRConnect mailbox password on every scan instead of using the .env-based
  IMAP credentials core/mail_gateway.py already uses for every other email
  feature; typing a mailbox password into a text field each time would have
  been a real regression in how this app handles secrets, not a faithful
  port. Confirmed meeting emails now go through the same AI
  classify-then-human-confirm flow as every other inbound category.
- NDR Planner: trip creation (name, sponsor, dates, type, city, focus, team,
  objectives) and the Active NDRs list with a plain-text itinerary export are
  ported. The original's auto-suggested, geographically-matched target list
  (a 25-institution curated database matched against the trip's city/type,
  with a two-column checkbox grid for building day-by-day schedules) is NOT
  ported in this pass — trips are created with objectives/team/dates, and
  specific meetings are added to a trip by hand. Prep Cards and Post-NDR
  Debrief sub-tabs are not yet ported; both remain available in app.py.
- Target Database: institution/prospect search+filter, a manual "add
  prospect" form, and the full automated prospecting pipeline are all
  ported — Analyst Coverage Network Targeting (core.analyst_coverage),
  live-13F-backed Automated Prospecting (core.prospecting.
  generate_coverage_prospects — rebuilt on core.sec_filings' real SEC 13F
  fetcher rather than app.py's original hardcoded KNOWN_13F_HOLDERS demo
  dict), NOBO cross-reference (core.prospecting.match_against_nobo), and
  bulk paste-from-website parsing (core.prospecting.parse_pasted_table).
  See each section's on-page caption for specifics — most notably, a ticker
  only contributes real Automated Prospecting results once its 13F data has
  actually been refreshed from the SEC Intelligence tab (13F refresh is
  heavy and stays a manual, explicit action — see core/sec_filings.py).

All per-client state persists via core.db (SQLite), under the keys
meeting_log.csv, peer_universe.csv, ndr_trips.json, scheduled_meetings.json,
post_meeting_notes.json, prospects.json, and buyside_mode.json — same
pattern as today_page.py, calendar_page.py, and reports_page.py. The two
external-system CSV exports this page reads in _render_big_picture()
(q1_2026_call_listener_log.csv from Chorus Call, ir_website_visitor_log.csv
from Google Analytics/Irwin/Q4) are deliberately still real files via
client_data_path() — they're meant to be replaced by literally uploading a
fresh export, not app-internal state. See core/db.py's module docstring for
the reasoning behind what moved to SQLite and what stayed as files.
"""

import asyncio
import csv
import io
import json
import os
import uuid
from datetime import datetime, timedelta
from urllib.parse import quote

import pandas as pd
from nicegui import ui

from config.client_config import CA, CE, CI, CP, CT, client_data_path, get_active_client_id, ndr_defaults
from config.theme_tokens import ACTIVE as COLORS
from core import analyst_coverage, consensus, db, documents, fit_score, inbox_queue, mail_gateway, market_data, nobo_engine, prospecting, risk_scorecard, sec_filings, ui_context
from core.investor_scoring import (
    INTERACTION_SCORE_MAX,
    OUTCOME_POINTS,
    compute_interaction_score as _compute_interaction_score,
    days_since_last_contact as _days_since_last_contact,
    get_fund_meetings as _get_fund_meetings,
    load_meeting_log as _load_meeting_log,
    save_meeting_log as _save_meeting_log,
    score_institutions as _score_institutions,
)
from core.security import get_anthropic_api_key
from core.textfmt import pretty_name
from data.seed.buyside_institutions import get_seed_buyside_institutions
from data.seed.conferences import get_seed_conferences
from data.seed.consensus_estimates import ALL_PERIODS
from data.seed.institution_contacts import get_institution_contacts
from page_modules_nicegui import nav
from page_modules_nicegui.responsive import responsive_table

# Post-Meeting Notes topic scaffolding — shared between the "click a topic to
# add it" chip row (_render_meeting_hub_tab) and _structure_notes_with_ai's
# prompt/fallback, so the two stay in lockstep. A blank raw-notes textarea
# right after a call tends to get skipped or filled with two rushed
# sentences; these chips nudge coverage of the financial topics that matter
# most without forcing rigid per-topic fields, and the inserted headers give
# the AI a much stronger signal to structure from than unlabeled prose.
NOTE_TOPICS = ["Revenue growth", "Segment/product growth", "Margins", "Cash flow",
               "Guidance/outlook", "Balance sheet & capital allocation",
               "Competitive positioning", "Valuation/multiple"]

# ─────────────────────────────────────────────────────────────────────────
# Small persistence helpers — every one of these is client-scoped via
# core.db (SQLite), so client #2 never sees client #1's pipeline data. The
# key strings below are the same names these used to be as filenames before
# the SQLite migration — core.db imports any pre-existing file under that
# name on first read, so nothing from earlier testing is lost.
# ─────────────────────────────────────────────────────────────────────────
def _load_json(name, default):
    return db.load_json(name, default)


def _save_json(name, data):
    db.save_json(name, data)


# _load_meeting_log / _save_meeting_log now live in core/investor_scoring.py
# (imported near the top of this file) — today_page.py's Investor Pipeline
# widget reads/writes the exact same meeting log, so an email logged from
# either page shows up on that fund's record everywhere.


# ─────────────────────────────────────────────────────────────────────────
# NDR Requests — inbound analyst requests to slot a management meeting into
# a city ("Scott is asking to slot 1x1s while management is in NY"). Used
# to be a hardcoded literal list with fixed "received" dates (Jun 28-30,
# 2026) baked into _render_big_picture()'s Metro Priority scoring — those
# dates were already in the past the moment "today" moved past them, which
# meant the "This Week's Priority" recommendation kept citing a request as
# fresh signal long after it had gone stale. Now a real persisted,
# loggable/resolvable list (core.db, key "ndr_requests.json"), same pattern
# as ndr_trips.json — seeded once with 3 example requests whose "received"
# dates are computed relative to whenever the seed is first created (not
# hardcoded absolute dates), so they never render as literally impossible.
# ─────────────────────────────────────────────────────────────────────────

def _load_ndr_requests(client_id=None):
    """This client's inbound NDR requests. EMPTY when none have been logged.

    There used to be a hardcoded fallback here — three invented requests naming real
    sell-side analysts (Scott Buck / H.C. Wainwright, Gary Prestopino / Barrington,
    Barry Sine / Litchfield Hills). Because it fired for ANY tenant with no saved
    requests, every client's pipeline showed USIO's analysts asking for meetings —
    and it PERSISTED them to that tenant's store. Northlake's "Ranked actions for
    management" read "Respond to Barry Sine (Litchfield Hills Research)".

    That is the same trap load_meeting_log documents: a fabricated REQUEST asserts a
    named person asked for something they didn't. An unworked request list is empty.

    Scope is EXPLICIT (client_id resolved here, not left to db's default) so a mis-bound
    active-tenant ContextVar in a callback can't read another client's requests — the same
    class of leak the hardcoded fallback above once caused."""
    return db.load_json("ndr_requests.json", default=[],
                        client_id=client_id or get_active_client_id()) or []


def _save_ndr_requests(records, client_id=None):
    db.save_json("ndr_requests.json", records, client_id=client_id or get_active_client_id())


def _mailto(to, subject, body, label):
    href = f"mailto:{to}?subject={quote(subject)}&body={quote(body)}"
    return ui.link(f"{label}", href).style(f"color:{COLORS['accent_light']};")


def _tel_href(phone):
    """A tel: link target from a display phone string — keeps digits and a
    leading +, drops spaces/dashes/parens so the dialer gets a clean number."""
    return "tel:" + "".join(ch for ch in str(phone) if ch.isdigit() or ch == "+")


def _refresh():
    nav.go_to("Ownership")


# ─────────────────────────────────────────────────────────────────────────
# Peer Universe — persisted via core.db under the "peer_universe.csv" key,
# the same key config.client_config.CP() reads (ticker/name/ev_rev), so
# peers added or edited here immediately show up in outreach-draft
# valuation comparisons too. sector is stored here for display but ignored
# by CP().
# ─────────────────────────────────────────────────────────────────────────
# weight/tier drive core/fit_score.py's Peer conviction + Comparability fit
# components — "core" = closest-size direct comp (weight 2, full Fit
# points, +5 conviction bonus for holding one), "close" = a real but
# less-comparable peer (weight 1, default), "large" = mega-cap/weak-signal
# peer (weight 0.5, weakest Fit tier). GDOT and PRTH are USIO's closest-size
# direct comps; EEFT/FOUR are larger-cap with weaker signal; the rest
# default to "close" until reviewed — same "seed is default, edit persists"
# pattern as everywhere else, editable in Manage Peer Universe below.
# Aligned with the valuation peer set (config.CP()). The Fit-Score tier vocab is
# core (tight comp, weight 2) / close (broader, weight 1) / large (mega-cap
# reference, weight 0.5); CP() normalizes "large" -> valuation "reference" (out
# of the median) and core/close -> "primary". segment / closest_analog feed the
# benchmarking display. Kept in sync so the two systems (which share the DB key
# "peer_universe.csv") never fight — dropping the retired PRTH/EEFT/PAX here is
# what stops _load_peer_universe re-adding them.
DEFAULT_PEER_UNIVERSE = [
    {"ticker": "RPAY", "name": "Repay Holdings", "sector": "Integrated card + ACH + billing/output", "ev_rev": 2.8, "weight": 2.0, "tier": "core", "segment": "Integrated card + ACH + billing/output", "closest_analog": True},
    {"ticker": "CASS", "name": "Cass Information Systems", "sector": "Payment information / billing", "ev_rev": 2.2, "weight": 2.0, "tier": "core", "segment": "Payment information / billing"},
    {"ticker": "CSGS", "name": "CSG Systems", "sector": "Billing / customer comms", "ev_rev": 2.2, "weight": 2.0, "tier": "core", "segment": "Billing / customer comms (output)"},
    {"ticker": "PSFE", "name": "Paysafe", "sector": "Integrated payments", "ev_rev": 1.5, "weight": 2.0, "tier": "core", "segment": "Integrated payments"},
    {"ticker": "PAY", "name": "Paymentus Holdings", "sector": "Bill presentment / EBPP", "ev_rev": 2.5, "weight": 2.0, "tier": "core", "segment": "Bill presentment / EBPP (output)"},
    {"ticker": "FOUR", "name": "Shift4 Payments", "sector": "Card acceptance / PayFac", "ev_rev": 4.2, "weight": 1.0, "tier": "close", "segment": "Card acceptance / PayFac"},
    {"ticker": "GDOT", "name": "Green Dot", "sector": "Prepaid / card issuing", "ev_rev": 1.2, "weight": 1.0, "tier": "close", "segment": "Prepaid / card issuing"},
    {"ticker": "FI", "name": "Fiserv", "sector": "Large-cap processor", "ev_rev": 5.0, "weight": 0.5, "tier": "large", "segment": "Large-cap processor (reference)"},
    {"ticker": "GPN", "name": "Global Payments", "sector": "Large-cap processor", "ev_rev": 4.5, "weight": 0.5, "tier": "large", "segment": "Large-cap processor (reference)"},
    {"ticker": "TOST", "name": "Toast", "sector": "Large-cap fintech", "ev_rev": 3.5, "weight": 0.5, "tier": "large", "segment": "Large-cap fintech (reference)"},
]


def _load_peer_universe():
    records = db.load_json("peer_universe.csv", default=None)
    # Illustrative tenants use their OWN configured comps (e.g. PYRA/CLRT/VNTG) — never the built-in
    # USIO payments default set, and never the missing-default backfill below (which would re-add
    # RPAY/CASS/… and make Peer Cross-Targeting contradict the rest of the demo).
    from core.curated_targets import _is_illustrative
    from config.client_config import get_active_client_id, CP
    if _is_illustrative(get_active_client_id()):
        if records:
            return records
        return [{"ticker": p["ticker"], "name": p.get("name", p["ticker"]),
                 "ev_rev": p.get("ev_rev"), "sector": "Payments / Fintech",
                 "weight": 1.0, "tier": p.get("tier", "close")} for p in CP()]
    if records is not None:
        known_ev = {p["ticker"]: p["ev_rev"] for p in DEFAULT_PEER_UNIVERSE if p["ev_rev"] is not None}
        for r in records:
            if r.get("ev_rev") in (None, ""):
                r["ev_rev"] = known_ev.get(r.get("ticker"))
            if not r.get("sector"):
                r["sector"] = "Payments / Fintech"
            # Backfill for peers saved before core/fit_score.py's weight/tier
            # fields existed — default weight=1.0/tier="close" (neutral,
            # matches core.fit_score._DEFAULT_PEER_ENTRY) so an un-reviewed
            # peer degrades gracefully into the Fit Score instead of KeyError.
            if r.get("weight") in (None, ""):
                r["weight"] = 1.0
            if not r.get("tier"):
                r["tier"] = "close"
        # Backfill: an earlier save of this list only had a subset of the
        # 7 built-in DEFAULT_PEER_UNIVERSE tickers (e.g. GDOT/PRTH/EEFT but
        # not FOUR/CASS/RPAY/PAX) — the loop above only patches fields on
        # records that already exist, it never adds a default ticker that's
        # simply missing entirely. Add back any missing default so nobody's
        # stuck with a stale, incomplete peer set from before all 7 existed.
        # Only ADDS — never removes a peer the user deleted on purpose, and
        # never touches a peer the user added themselves.
        existing_tickers = {r.get("ticker") for r in records}
        missing_defaults = [dict(p) for p in DEFAULT_PEER_UNIVERSE if p["ticker"] not in existing_tickers]
        if missing_defaults:
            records = records + missing_defaults
            _save_peer_universe(records)
        return records
    return [dict(p) for p in DEFAULT_PEER_UNIVERSE]


def _save_peer_universe(peers):
    db.save_json("peer_universe.csv", peers)


def _parse_aum_millions(aum_str):
    """'$2.1B' -> 2100.0, '$380M' -> 380.0, '$8.2T' -> 8_200_000.0. Returns 0
    if it can't be parsed, so an unparseable AUM never satisfies a nonzero
    minimum filter."""
    if not aum_str:
        return 0.0
    s = str(aum_str).strip().upper().replace("$", "").replace(",", "")
    mult = {"T": 1_000_000.0, "B": 1_000.0, "M": 1.0}
    for suffix, m in mult.items():
        if s.endswith(suffix):
            try:
                return float(s[:-1]) * m
            except ValueError:
                return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _client_ev_rev(period="Q2 2026E"):
    """This client's own EV/Revenue multiple, computed the same way
    markets_page.py's PT Justification cards compute it (live price via
    core.market_data x shares outstanding + net debt, over the guidance
    revenue estimate for the given period) — was a hardcoded 0.4x literal
    in the pre-earnings outreach draft below; that number moves with the
    stock and shouldn't be typed in. Returns None if market data or the
    seed financial-position/guidance fields aren't available yet, so
    callers can fall back to a peer-agnostic outreach line instead of
    showing a stale or fabricated multiple."""
    # get_consensus(), not the seed: this reads period_guidance to build an EV/Revenue
    # multiple for outreach copy, and the guidance override is exactly the thing that
    # moves when the CFO revises. Reading the seed here would put a stale multiple in
    # front of an investor — which is the failure this function's docstring says it
    # exists to avoid.
    from core.consensus import get_consensus
    snap = market_data.get_snapshot(CT("ticker"))
    last_price = snap["last_price"] if snap and snap.get("last_price") is not None else CT("last_price", None)
    if last_price is None:
        return None
    seed = get_consensus(get_active_client_id())
    fin = seed.get("financial_position", {})
    shares_out = fin.get("shares_out_m")
    net_debt = fin.get("net_debt_m", 0)
    rev_est = seed.get("period_guidance", {}).get(period, {}).get("Revenue Est ($M)")
    if not shares_out or not rev_est:
        return None
    ev = last_price * shares_out + net_debt
    return round(ev / rev_est, 2)


# OUTCOME_POINTS, INTERACTION_SCORE_MAX, _compute_interaction_score,
# _score_institutions, and _get_fund_meetings now live in
# core/investor_scoring.py (imported near the top of this file) — pulled
# out so today_page.py's Investor Pipeline widget scores institutions the
# exact same way this page does, instead of a second hand-maintained copy.


def _get_repeat_signal(meetings):
    if len(meetings) < 2:
        return None
    try:
        dates = sorted([datetime.strptime(m["Date"], "%Y-%m-%d").date() for m in meetings], reverse=True)
    except Exception:
        return None
    gap = (dates[0] - dates[1]).days
    return gap if gap <= 45 else None


# ─────────────────────────────────────────────────────────────────────────
# AI note structuring — ported directly from app.py's structure_notes_with_ai.
# No Streamlit dependency in the original, so this is a straight port.
# ─────────────────────────────────────────────────────────────────────────
def _structure_notes_with_ai(raw_notes, contact, firm, meeting_type):
    import re
    try:
        import urllib.error
        import urllib.request
        topics_list = ", ".join(NOTE_TOPICS)
        prompt = f"""You are an expert IR analyst. Structure these raw post-meeting notes from a call with {contact} at {firm} ({meeting_type}) into the following JSON format:
{{
  "key_questions": ["list of questions the investor/analyst asked"],
  "concerns_raised": ["list of concerns or risks they mentioned"],
  "positive_signals": ["things they said that indicate interest or conviction"],
  "commitments_made": ["anything IR/management committed to follow up on"],
  "follow_up_actions": ["specific next steps"],
  "financial_kpi_takeaways": ["specific comments tied to financial/operating topics such as {topics_list} — one item per topic actually discussed, phrased as e.g. 'Revenue growth: investor pushed back on Q3 deceleration'; omit topics that weren't discussed"],
  "sentiment": "Positive / Neutral / Negative / Mixed",
  "summary": "2-3 sentence executive summary of the call"
}}

Raw notes:
{raw_notes}

Return ONLY valid JSON, no other text."""
        payload = json.dumps({
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 1000,
            "messages": [{"role": "user", "content": prompt}],
        }).encode()
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={"Content-Type": "application/json", "x-api-key": get_anthropic_api_key(),
                     "anthropic-version": "2023-06-01"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read())
            raw_text = "".join(
                block.get("text", "") for block in result.get("content", []) if block.get("type") == "text"
            ).strip()
            if raw_text.startswith("```"):
                # Claude sometimes wraps the JSON in a ```json ... ``` fence
                # even when told to return raw JSON only. json.loads chokes
                # on the leading backtick with "Expecting value: line 1
                # column 1 (char 0)" — this was the real cause of every
                # real (key-is-valid) call silently falling back to the
                # keyword-matching summary below. Strip the fence first.
                lines = raw_text.split("\n")
                if lines and lines[0].startswith("```"):
                    lines = lines[1:]
                if lines and lines[-1].strip() == "```":
                    lines = lines[:-1]
                raw_text = "\n".join(lines).strip()
            if not raw_text:
                raise ValueError("Claude returned an empty response body")
            parsed = json.loads(raw_text)
            print("[ai_notes] Claude call succeeded — using real AI-structured summary.")
            return parsed
    except urllib.error.HTTPError as e:
        # Anthropic's real error body (e.g. authentication_error, a bad model
        # name, insufficient credit) — this used to be swallowed by a bare
        # `except Exception: pass`, so a bad/expired API key silently fell
        # back to the keyword-matching summary below with zero visible sign
        # anything was wrong. Logging it here (server console, not the UI —
        # this is a diagnostic, not something to alarm the end user with on
        # every call) makes that failure mode debuggable.
        try:
            body = e.read().decode(errors="replace")
        except Exception:
            body = "(no response body)"
        print(f"[ai_notes] Claude API call failed — HTTP {e.code}: {body} — falling back to keyword-based summary.")
    except Exception as e:
        print(f"[ai_notes] Claude API call failed — {type(e).__name__}: {e} — falling back to keyword-based summary.")

    questions = [s.strip() for s in re.split(r'[.!?\n]', raw_notes) if '?' in s and len(s.strip()) > 10]
    positives = [s.strip() for s in re.split(r'[.!\n]', raw_notes)
                 if any(w in s.lower() for w in ["interested", "positive", "like", "impressed", "compelling", "strong"]) and len(s.strip()) > 10]
    concerns = [s.strip() for s in re.split(r'[.!\n]', raw_notes)
                if any(w in s.lower() for w in ["concern", "worry", "risk", "decline", "miss", "issue", "problem"]) and len(s.strip()) > 10]
    actions = [s.strip() for s in re.split(r'[.!\n]', raw_notes)
               if any(w in s.lower() for w in ["follow", "send", "provide", "share", "schedule", "call back", "deck"]) and len(s.strip()) > 10]
    # Pulls "Revenue growth: ..." style lines left by the topic chips (see
    # NOTE_TOPICS) straight through as-is — this is the one part of the
    # fallback that doesn't need real NLP, since the topic label is already
    # right there in the text the chip inserted.
    kpi_takeaways = []
    for line in raw_notes.split("\n"):
        line = line.strip()
        if not line:
            continue
        for topic in NOTE_TOPICS:
            if line.lower().startswith(topic.lower() + ":"):
                detail = line[len(topic) + 1:].strip()
                if detail:
                    kpi_takeaways.append(f"{topic}: {detail}")
                break
    sentiment = "Positive" if len(positives) > len(concerns) else "Mixed" if positives else "Neutral"
    return {
        "key_questions": questions[:4], "concerns_raised": concerns[:3], "positive_signals": positives[:3],
        "commitments_made": [], "follow_up_actions": actions[:3], "financial_kpi_takeaways": kpi_takeaways[:6],
        "sentiment": sentiment,
        "summary": f"Meeting with {contact} at {firm}. {len(questions)} questions raised. Sentiment: {sentiment}.",
    }


def _enrich_peer_holdings_with_live_13f(institutions, peer_tickers):
    """Replaces each institution's seed Peer_Holdings with a freshness-
    aware effective value, mutating institutions in place (same pattern
    core.investor_scoring.score_institutions already uses on this list).

    Every original DEFAULT_PEER_UNIVERSE record before this feature had
    hand-typed Peer_Holdings — never real SEC data, even for the original
    3 tracked tickers (GDOT/PRTH/EEFT). Adding new peer tickers (FOUR,
    CASS, RPAY, PAX) made that visible: selecting a newly-added ticker in
    Peer Cross-Targeting could never surface a match, because nothing in
    the seed data has ever heard of those tickers. core.sec_filings has a
    real, working SEC 13F fetcher (see refresh_13f_holders / the "Refresh
    13F Institutional Holders" button on the SEC Intelligence tab) — this
    function is what actually connects that real data back into Peer
    Cross-Targeting and the "Peers held" display, instead of leaving them
    reading a permanently-static list.

    Per ticker, per institution:
    - If that ticker's 13F data has never been fetched (not_fetched):
      keep whatever the seed said, unchanged — so nothing regresses before
      anyone clicks Refresh, and the original 3 tickers keep behaving
      exactly as they always have until a real refresh happens for them.
    - If that ticker HAS been fetched: use ONLY the real, confirmed
      result — a real 13F filing always outranks a hand-typed seed guess,
      even when they disagree (see core.sec_filings.live_peer_overlap_map's
      docstring on why matching is exact-name, not fuzzy — some seed
      institutions are specific strategies, not 13F filers of record, and
      will legitimately show "no confirmed match" once real data exists).

    Also stamps inst["Peer_Holdings_Source"] = {ticker: "live"|"seed"} so
    the UI can show which tickers are backed by a real filing vs. still on
    the original seed guess, instead of presenting both identically."""
    overlap_map = sec_filings.live_peer_overlap_map([i["Fund"] for i in institutions], peer_tickers)
    for inst in institutions:
        overlap = overlap_map.get(inst["Fund"], {"confirmed": [], "not_fetched": peer_tickers})
        not_fetched = set(overlap["not_fetched"])
        seed_holdings = set(inst.get("Peer_Holdings", []))
        effective = set(overlap["confirmed"]) | {t for t in seed_holdings if t in not_fetched}
        inst["Peer_Holdings"] = sorted(effective)
        inst["Peer_Holdings_Source"] = {
            t: ("live" if t not in not_fetched else "seed") for t in peer_tickers
        }


def _enrich_engagement_signals(institutions, client_id):
    """Populate each institution's IR-website + call-listener signals.

    - Web-flow (ALL tenants): where a tracked fund matches an IR-site visitor, fold in real visit
      count / last-visit / digital-intent. This is what makes the Website Ownership tab and the
      Buy-Side cards agree, and it's honest for real tenants (only fires on a matched visitor).
    - Illustrative tenants ONLY: fill the REMAINING call-listener / visit signals with believable
      illustrative values, so the demo reads as an actively-covered book instead of "no call data /
      no visit data" on every card. Real tenants keep the honest "no data" — no integration exists.
    Deterministic (hash of the fund name) so it's stable across renders and reseeds."""
    from core.curated_targets import _is_illustrative
    from core import web_flow
    illus = _is_illustrative(client_id)
    _wmap = web_flow.intent_map(client_id)
    _dur = ["Full call (42 min)", "Full call (42 min)", "Q&A only (18 min)", "Replay — first 15 min"]
    for inst in institutions:
        fund = inst.get("Fund") or ""
        v = web_flow.intent_for(fund, _map=_wmap) if _wmap else None
        if v:
            if v.get("visits") is not None:
                inst["IR_Visits_30d"] = v["visits"]
            if v.get("last_visit"):
                inst["Last_Visit"] = v["last_visit"]
            if v.get("intent") is not None:
                inst["Digital_Intent_Score"] = v["intent"]
        if not illus:
            continue
        h = sum(ord(c) for c in fund)
        if inst.get("Call_Listener") is None:            # ~60% listened; the rest a measured "no"
            listened = (h % 5) < 3
            inst["Call_Listener"] = listened
            inst["Listen_Duration"] = _dur[h % len(_dur)] if listened else None
        if inst.get("IR_Visits_30d") is None:            # web-flow value wins; else a small count
            n = h % 4
            inst["IR_Visits_30d"] = n
            if n:
                inst["Last_Visit"] = (datetime.now() - timedelta(days=2 + (h % 20))).strftime("%b %d")
        if not inst.get("Digital_Intent_Score"):
            inst["Digital_Intent_Score"] = 25 + (h % 60)


# ─────────────────────────────────────────────────────────────────────────
# Data provenance — every name in the universe carries a Source tag so the UI
# shows where it came from. The seed is "Seed (demo)"; real SEC-sourced names
# (13F holders, 13D/13G filers, funds holding the peer set) are merged in from
# cached EDGAR data, each with its own tag. This is the honest answer to "why
# isn't the database bigger": it grows from authoritative sources, not by hand.
# ─────────────────────────────────────────────────────────────────────────
SOURCE_COLORS = {
    "Seed (demo)": "#64748B", "Seed + SEC-confirmed": "#0F766E",
    "SEC 13F": "#15803D", "SEC 13D/G": "#B45309", "SEC 13F + 13D/G": "#15803D",
    "Peer 13F": "#1E40AF", "Peer 13F (T1)": "#1E3A8A", "Peer 13F (T2)": "#4F46E5",
}

# Comp tiering for peer-overlap prospecting — peers first, then industry.
# Tier 1 = tightest small/mid-cap payment comps (highest-signal holders — a fund
# in Repay or Cass but not USIO is a real conversion target); Tier 2 = broader
# payments names. Anything not listed defaults to T2. The large-cap REFERENCE
# peers (FI / GPN / TOST) are deliberately excluded from prospecting entirely —
# a fund holding Fiserv because it's in the S&P 500 tells you nothing about
# micro-cap-payments interest, so their index/pension holders would be pure noise
# (see the tier=="reference" skip in _sec_universe_records).
_PEER_TIER = {"RPAY": 1, "CASS": 1, "CSGS": 1, "PSFE": 1, "PAY": 1, "FOUR": 2, "GDOT": 2}


def _source_color(src):
    return SOURCE_COLORS.get(src, "#15803D")


def _apply_peer_tier(rec, tier):
    """Stamp a peer-overlap prospect with its comp tier — the tighter the comp
    it holds, the higher-signal the prospect."""
    rec["_peer_tier"] = tier
    rec["Source"] = f"Peer 13F (T{tier})"
    rec["Peer_Score"] = 35 if tier == 1 else 20
    rec["Action"] = ("Holds a close comp — high-signal NDR prospect" if tier == 1
                     else "Holds an industry name — broader prospect")


_NORM_SUFFIXES = {"inc", "incorporated", "llc", "lp", "llp", "corp", "corporation", "co", "company",
                  "ltd", "limited", "plc", "trust", "sa", "ag", "nv", "the", "group", "holdings"}


def _norm_name(n):
    """Normalize a manager name for cross-source dedup — lowercase, drop
    punctuation, and strip common corporate suffixes so 'Dimensional Fund
    Advisors' (seed) and 'DIMENSIONAL FUND ADVISORS LP' (SEC) match."""
    tokens = "".join(c if c.isalnum() else " " for c in str(n).lower()).split()
    return "".join(t for t in tokens if t not in _NORM_SUFFIXES)


# SEC 13F filer cities → the same metro-region labels the seed universe uses,
# so real filers unify into the geographic breakdown ("Where they are") instead
# of a flat "Unknown (SEC)" bucket. Keyed by uppercase city. Unmapped US cities
# fall back to "City, ST"; non-US filers group under "International".
_US_STATES = {"AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI",
              "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI",
              "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC",
              "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT",
              "VT", "VA", "WA", "WV", "WI", "WY", "DC"}
# Roadshow metros — a ~60-mile radius (a day's drive), the way an IR team actually plans an NDR,
# NOT a broad Census region. Cities beyond 60mi of a hub are their OWN stop; a state full of those
# (Wisconsin: Milwaukee ~90mi from Chicago, Madison further) is a scattered, hard-to-cover book.
# The old map lumped Wayzata + Minneapolis + Milwaukee into "Chicago / Midwest" — 350+ miles you
# can't roadshow in one trip — and did the same to Texas (Dallas/Austin/Houston) and Florida.
_SEC_CITY_METRO = {
    # New York metro — incl. lower-CT (Gold Coast) + near-NJ + Westchester, all inside 60mi
    "NEW YORK": "New York, NY", "BROOKLYN": "New York, NY", "NEW YORK CITY": "New York, NY",
    "GREENWICH": "New York, NY", "STAMFORD": "New York, NY", "DARIEN": "New York, NY",
    "WESTPORT|CT": "New York, NY", "NORWALK|CT": "New York, NY", "NEW CANAAN": "New York, NY",
    "RYE|NY": "New York, NY", "WHITE PLAINS": "New York, NY", "PURCHASE": "New York, NY",
    "JERSEY CITY": "New York, NY", "SHORT HILLS": "New York, NY", "SUMMIT|NJ": "New York, NY",
    "GARDEN CITY|NY": "New York, NY", "MELVILLE": "New York, NY",
    # Boston
    "BOSTON": "Boston, MA", "CAMBRIDGE|MA": "Boston, MA", "WELLESLEY": "Boston, MA",
    "WALTHAM": "Boston, MA", "NEWTON|MA": "Boston, MA", "NEEDHAM": "Boston, MA", "PLYMOUTH|MA": "Boston, MA",
    # Chicago (NOT Minneapolis/Milwaukee)
    "CHICAGO": "Chicago, IL", "OAK BROOK": "Chicago, IL", "NAPERVILLE": "Chicago, IL",
    "EVANSTON|IL": "Chicago, IL", "NORTHBROOK": "Chicago, IL", "LAKE FOREST|IL": "Chicago, IL",
    "ROSEMONT": "Chicago, IL", "LISLE": "Chicago, IL",
    # Minneapolis–St. Paul (Twin Cities) — its own stop, ~60mi radius
    "MINNEAPOLIS": "Minneapolis-St. Paul, MN", "ST. PAUL": "Minneapolis-St. Paul, MN",
    "WAYZATA": "Minneapolis-St. Paul, MN", "ROSEMOUNT": "Minneapolis-St. Paul, MN",
    "MINNETONKA": "Minneapolis-St. Paul, MN", "PLYMOUTH|MN": "Minneapolis-St. Paul, MN",
    "EDEN PRAIRIE": "Minneapolis-St. Paul, MN", "EDINA": "Minneapolis-St. Paul, MN",
    "BLOOMINGTON|MN": "Minneapolis-St. Paul, MN", "MENDOTA HEIGHTS": "Minneapolis-St. Paul, MN",
    # Wisconsin — the toughest book: Milwaukee is ~90mi from Chicago (a flight/long haul), Madison
    # further, and funds scatter across the state with no dense hub.
    "MILWAUKEE": "Milwaukee / Wisconsin", "BROOKFIELD": "Milwaukee / Wisconsin",
    "MADISON|WI": "Milwaukee / Wisconsin", "MEQUON": "Milwaukee / Wisconsin",
    # SoCal / Bay / Philly
    "SOUTH PASADENA": "Los Angeles, CA", "PASADENA": "Los Angeles, CA",
    "LOS ANGELES": "Los Angeles, CA", "IRVINE": "Los Angeles, CA", "NEWPORT BEACH": "Los Angeles, CA",
    "SANTA MONICA": "Los Angeles, CA", "EL SEGUNDO": "Los Angeles, CA",
    "SAN FRANCISCO": "San Francisco / Bay Area", "FOLSOM": "San Francisco / Bay Area",
    "SAN MATEO": "San Francisco / Bay Area", "PALO ALTO": "San Francisco / Bay Area",
    "MENLO PARK": "San Francisco / Bay Area", "SAN JOSE": "San Francisco / Bay Area",
    "OAKLAND": "San Francisco / Bay Area", "BERKELEY": "San Francisco / Bay Area",
    "MALVERN": "Philadelphia, PA", "BALA CYNWYD": "Philadelphia, PA",
    "PHILADELPHIA": "Philadelphia, PA", "RADNOR": "Philadelphia, PA", "CONSHOHOCKEN": "Philadelphia, PA",
    "WAYNE|PA": "Philadelphia, PA", "WEST CONSHOHOCKEN": "Philadelphia, PA",
    "BALTIMORE": "Baltimore / DC", "WASHINGTON": "Baltimore / DC", "BETHESDA": "Baltimore / DC",
    "MCLEAN": "Baltimore / DC", "ARLINGTON": "Baltimore / DC",
    # Texas — three separate stops (Dallas–Austin–Houston are 200+ mi apart)
    "DALLAS": "Dallas, TX", "FORT WORTH": "Dallas, TX", "PLANO": "Dallas, TX", "IRVING": "Dallas, TX",
    "AUSTIN": "Austin, TX",
    "HOUSTON": "Houston, TX",
    "DENVER": "Denver, CO", "BOULDER": "Denver, CO", "GREENWOOD VILLAGE": "Denver, CO",
    # Florida — Miami/South-FL vs Tampa are ~280mi apart, separate
    "MIAMI": "Miami / South FL", "PALM BEACH": "Miami / South FL", "BOCA RATON": "Miami / South FL",
    "NAPLES": "Miami / South FL", "FORT LAUDERDALE": "Miami / South FL", "AVENTURA": "Miami / South FL",
    "TAMPA": "Tampa, FL", "ST. PETERSBURG": "Tampa, FL",
    "ATLANTA": "Atlanta, GA", "SEATTLE": "Seattle, WA", "BELLEVUE": "Seattle, WA",
    # St. Louis — all spelling variants normalise to "ST. LOUIS" (see _norm_city) — + inner-ring suburbs
    "ST. LOUIS": "St. Louis, MO", "CLAYTON|MO": "St. Louis, MO", "CHESTERFIELD": "St. Louis, MO",
    "DES PERES": "St. Louis, MO", "CREVE COEUR": "St. Louis, MO", "TOWN AND COUNTRY": "St. Louis, MO",
    "BALLWIN": "St. Louis, MO", "KIRKWOOD": "St. Louis, MO", "MARYLAND HEIGHTS": "St. Louis, MO",
    # Kansas City metro (spans MO/KS)
    "KANSAS CITY": "Kansas City, MO", "OVERLAND PARK": "Kansas City, MO", "LEAWOOD": "Kansas City, MO",
    # ── Fuller CA coverage — state-qualified. Bay Area (Marin/East Bay/Peninsula), LA, San Diego,
    #    Sacramento — NOT all lumped into LA; each city to its true ~60-mile metro. ──
    "MILL VALLEY|CA": "San Francisco / Bay Area", "ORINDA|CA": "San Francisco / Bay Area",
    "WALNUT CREEK|CA": "San Francisco / Bay Area", "GREENBRAE|CA": "San Francisco / Bay Area",
    "NOVATO|CA": "San Francisco / Bay Area", "BURLINGAME|CA": "San Francisco / Bay Area",
    "FOSTER CITY|CA": "San Francisco / Bay Area", "LAFAYETTE|CA": "San Francisco / Bay Area",
    "CONCORD|CA": "San Francisco / Bay Area", "LARKSPUR|CA": "San Francisco / Bay Area",
    "SAN RAFAEL|CA": "San Francisco / Bay Area", "TIBURON|CA": "San Francisco / Bay Area",
    "SAUSALITO|CA": "San Francisco / Bay Area", "EMERYVILLE|CA": "San Francisco / Bay Area",
    "REDWOOD CITY|CA": "San Francisco / Bay Area", "SAN CARLOS|CA": "San Francisco / Bay Area",
    "MOUNTAIN VIEW|CA": "San Francisco / Bay Area",
    "LAGUNA BEACH|CA": "Los Angeles, CA", "VALENCIA|CA": "Los Angeles, CA", "CALABASAS|CA": "Los Angeles, CA",
    "TORRANCE|CA": "Los Angeles, CA", "MANHATTAN BEACH|CA": "Los Angeles, CA", "BEVERLY HILLS|CA": "Los Angeles, CA",
    "WESTLAKE VILLAGE|CA": "Los Angeles, CA", "LONG BEACH|CA": "Los Angeles, CA", "MARINA DEL REY|CA": "Los Angeles, CA",
    "CULVER CITY|CA": "Los Angeles, CA", "GLENDALE|CA": "Los Angeles, CA", "SHERMAN OAKS|CA": "Los Angeles, CA",
    "SAN DIEGO|CA": "San Diego, CA", "CORONADO|CA": "San Diego, CA", "RANCHO SANTA FE|CA": "San Diego, CA",
    "ENCINITAS|CA": "San Diego, CA", "LA JOLLA|CA": "San Diego, CA", "DEL MAR|CA": "San Diego, CA",
    "CARLSBAD|CA": "San Diego, CA", "SOLANA BEACH|CA": "San Diego, CA",
    "SACRAMENTO|CA": "Sacramento, CA", "WEST SACRAMENTO|CA": "Sacramento, CA", "FOLSOM|CA": "Sacramento, CA",
    "ROSEVILLE|CA": "Sacramento, CA", "EL DORADO HILLS|CA": "Sacramento, CA",
    # ── Fuller PA coverage — Philadelphia (Main Line) vs Pittsburgh (both real in the book) ──
    "PITTSBURGH|PA": "Pittsburgh, PA", "WARRENDALE|PA": "Pittsburgh, PA", "SEWICKLEY|PA": "Pittsburgh, PA",
    "CRANBERRY TOWNSHIP|PA": "Pittsburgh, PA", "WEXFORD|PA": "Pittsburgh, PA",
    "BERWYN|PA": "Philadelphia, PA", "NEWTOWN SQUARE|PA": "Philadelphia, PA", "OAKS|PA": "Philadelphia, PA",
    "DEVON|PA": "Philadelphia, PA", "PAOLI|PA": "Philadelphia, PA", "VILLANOVA|PA": "Philadelphia, PA",
    "WEST CHESTER|PA": "Philadelphia, PA", "KING OF PRUSSIA|PA": "Philadelphia, PA",
    # ── Sweep (2026-07-21): fold remaining suburbs into their true ~60-mile metro, state-qualified. ──
    # Detroit / Grand Rapids (MI)
    "HUNTINGTON WOODS|MI": "Detroit, MI", "BINGHAM FARMS|MI": "Detroit, MI", "BLOOMFIELD HILLS|MI": "Detroit, MI",
    "NOVI|MI": "Detroit, MI", "SOUTHFIELD|MI": "Detroit, MI", "BIRMINGHAM|MI": "Detroit, MI", "PLYMOUTH|MI": "Detroit, MI",
    "TROY|MI": "Detroit, MI", "FARMINGTON HILLS|MI": "Detroit, MI", "ANN ARBOR|MI": "Detroit, MI",
    "HOLLAND|MI": "Grand Rapids, MI", "ADA|MI": "Grand Rapids, MI", "GRAND RAPIDS|MI": "Grand Rapids, MI",
    # Richmond (VA) + NoVa / Baltimore into the Baltimore/DC cluster
    "RICHMOND|VA": "Richmond, VA", "GLEN ALLEN|VA": "Richmond, VA",
    "HERNDON|VA": "Baltimore / DC", "MCLEAN|VA": "Baltimore / DC", "RESTON|VA": "Baltimore / DC",
    "ALEXANDRIA|VA": "Baltimore / DC", "ROCKVILLE|MD": "Baltimore / DC", "HUNT VALLEY|MD": "Baltimore / DC",
    "TOWSON|MD": "Baltimore / DC", "COLUMBIA|MD": "Baltimore / DC",
    # Upstate NY + Long Island / CT gold-coast into New York
    "ALBANY|NY": "Albany, NY", "GLENMONT|NY": "Albany, NY",
    "BUFFALO|NY": "Buffalo, NY", "ORCHARD PARK|NY": "Buffalo, NY",
    "ROCHESTER|NY": "Rochester, NY", "FAIRPORT|NY": "Rochester, NY", "PITTSFORD|NY": "Rochester, NY",
    "HAUPPAUGE|NY": "New York, NY", "ROWAYTON|CT": "New York, NY",
    # Salt Lake City
    "SALT LAKE CITY|UT": "Salt Lake City, UT", "PARK CITY|UT": "Salt Lake City, UT", "MIDVALE|UT": "Salt Lake City, UT",
    "MURRAY|UT": "Salt Lake City, UT", "SANDY|UT": "Salt Lake City, UT", "DRAPER|UT": "Salt Lake City, UT",
    "COTTONWOOD HEIGHTS|UT": "Salt Lake City, UT",
    # Denver suburbs
    "ENGLEWOOD|CO": "Denver, CO", "LONE TREE|CO": "Denver, CO", "GOLDEN|CO": "Denver, CO",
    "LITTLETON|CO": "Denver, CO", "CENTENNIAL|CO": "Denver, CO", "AURORA|CO": "Denver, CO",
    # Hartford
    "HARTFORD|CT": "Hartford, CT", "FARMINGTON|CT": "Hartford, CT", "WINDSOR|CT": "Hartford, CT",
    "MERIDEN|CT": "Hartford, CT", "WATERBURY|CT": "Hartford, CT", "WEST HARTFORD|CT": "Hartford, CT",
    # Phoenix
    "PHOENIX|AZ": "Phoenix, AZ", "PEORIA|AZ": "Phoenix, AZ", "SCOTTSDALE|AZ": "Phoenix, AZ",
    "MESA|AZ": "Phoenix, AZ", "TEMPE|AZ": "Phoenix, AZ", "CHANDLER|AZ": "Phoenix, AZ", "GILBERT|AZ": "Phoenix, AZ",
    # Chicago suburbs
    "OAKBROOK TERRACE|IL": "Chicago, IL", "SCHAUMBURG|IL": "Chicago, IL", "WHEATON|IL": "Chicago, IL",
    "HINSDALE|IL": "Chicago, IL", "ITASCA|IL": "Chicago, IL", "DEERFIELD|IL": "Chicago, IL",
    # Nashville / Indianapolis
    "NASHVILLE|TN": "Nashville, TN", "BRENTWOOD|TN": "Nashville, TN", "FRANKLIN|TN": "Nashville, TN",
    "CARMEL|IN": "Indianapolis, IN", "INDIANAPOLIS|IN": "Indianapolis, IN", "FISHERS|IN": "Indianapolis, IN",
    "ZIONSVILLE|IN": "Indianapolis, IN",
    # Atlanta suburbs
    "ALPHARETTA|GA": "Atlanta, GA", "SUWANEE|GA": "Atlanta, GA", "ROSWELL|GA": "Atlanta, GA",
    "MARIETTA|GA": "Atlanta, GA", "SANDY SPRINGS|GA": "Atlanta, GA",
    # Charlotte (+ SC border suburbs)
    "CHARLOTTE|NC": "Charlotte, NC", "ROCK HILL|SC": "Charlotte, NC", "FORT MILL|SC": "Charlotte, NC",
    # Portland OR (state-qualified vs Portland ME) + Seattle
    "PORTLAND|OR": "Portland, OR", "LAKE OSWEGO|OR": "Portland, OR", "TIGARD|OR": "Portland, OR",
    "BEAVERTON|OR": "Portland, OR", "CAMAS|WA": "Portland, OR",
    "PUYALLUP|WA": "Seattle, WA", "TACOMA|WA": "Seattle, WA", "REDMOND|WA": "Seattle, WA", "KIRKLAND|WA": "Seattle, WA",
    "BOTHELL|WA": "Seattle, WA", "ISSAQUAH|WA": "Seattle, WA", "SAMMAMISH|WA": "Seattle, WA",
    "MERCER ISLAND|WA": "Seattle, WA", "RENTON|WA": "Seattle, WA", "EVERETT|WA": "Seattle, WA",
    "WOODINVILLE|WA": "Seattle, WA", "BAINBRIDGE ISLAND|WA": "Seattle, WA", "EDMONDS|WA": "Seattle, WA",
    # Kansas City (KS side)
    "SHAWNEE MISSION|KS": "Kansas City, MO", "LENEXA|KS": "Kansas City, MO",
}


# Major international financial hubs — a foreign fund lands in its city cluster (a real roadshow
# stop: a London or Toronto swing) rather than one undifferentiated "International" blob. Keyed by
# normalised city; province/country suffixes ("TORONTO ONTARIO", "PARIS FRANCE") are stripped first.
# Only reached for non-US state codes, so ambiguous names (Burlington ON) can't collide with US.
_INTL_CITY_METRO = {
    "TORONTO": "Toronto, Canada", "MISSISSAUGA": "Toronto, Canada", "OAKVILLE": "Toronto, Canada",
    "MARKHAM": "Toronto, Canada", "VAUGHAN": "Toronto, Canada", "RICHMOND HILL": "Toronto, Canada",
    "BURLINGTON": "Toronto, Canada", "WATERLOO": "Toronto, Canada",
    "MONTREAL": "Montreal, Canada", "VANCOUVER": "Vancouver, Canada", "CALGARY": "Calgary, Canada",
    "LONDON": "London, UK", "EDINBURGH": "Edinburgh, UK",
    "PARIS": "Paris, France", "FRANKFURT": "Frankfurt, Germany", "FRANKFURT AM MAIN": "Frankfurt, Germany",
    "MUNICH": "Munich, Germany", "BERLIN": "Berlin, Germany",
    # Switzerland — German-Swiss hub (Zurich, incl. the Zug/Pfäffikon hedge-fund belt) vs French-
    # Swiss (Geneva). Zurich↔Geneva is ~175mi, so two stops even at a 60-mile radius.
    "ZURICH": "Zurich, Switzerland", "ZUG": "Zurich, Switzerland", "BASEL": "Zurich, Switzerland", "BERN": "Zurich, Switzerland",
    "PFAFFIKON": "Zurich, Switzerland", "LUCERNE": "Zurich, Switzerland", "WINTERTHUR": "Zurich, Switzerland",
    "SCHINDELLEGI": "Zurich, Switzerland", "BAAR": "Zurich, Switzerland", "WOLLERAU": "Zurich, Switzerland", "FREIENBACH": "Zurich, Switzerland",
    "GENEVA": "Geneva, Switzerland", "GENEVE": "Geneva, Switzerland", "CAROUGE": "Geneva, Switzerland", "LAUSANNE": "Geneva, Switzerland",
    "NYON": "Geneva, Switzerland", "COLOGNY": "Geneva, Switzerland", "VEVEY": "Geneva, Switzerland",
    # Italian-Swiss (Ticino) private-banking centre — its own stop; ~30mi to Milan but a
    # distinct cross-border booking hub, so kept Swiss rather than folded into Milan.
    "LUGANO": "Lugano, Switzerland", "PARADISO": "Lugano, Switzerland", "BELLINZONA": "Lugano, Switzerland",
    "MANNO": "Lugano, Switzerland", "CHIASSO": "Lugano, Switzerland", "MENDRISIO": "Lugano, Switzerland",
    "TOKYO": "Tokyo, Japan", "HONG KONG": "Hong Kong", "CAUSEWAY BAY": "Hong Kong", "CENTRAL": "Hong Kong",
    "SINGAPORE": "Singapore", "SEOUL": "Seoul, South Korea", "DUBAI": "Dubai, UAE", "MILAN": "Milan, Italy",
    "BRUSSELS": "Brussels, Belgium", "LUXEMBOURG": "Luxembourg", "SENNINGERBERG": "Luxembourg",
    "STOCKHOLM": "Stockholm, Sweden", "AMSTERDAM": "Amsterdam, Netherlands", "OSLO": "Oslo, Norway",
    "COPENHAGEN": "Copenhagen, Denmark", "COPENHAGEN V": "Copenhagen, Denmark", "HELSINKI": "Helsinki, Finland",
    "SYDNEY": "Sydney, Australia", "MELBOURNE": "Melbourne, Australia", "SAO PAULO": "Sao Paulo, Brazil",
    "WARSAW": "Warsaw, Poland", "DUBLIN": "Dublin, Ireland",
}


def _norm_city(city):
    """Collapse the St. / Saint / St spelling variants that otherwise fragment a metro into
    separate rows (the data carries 'ST LOUIS', 'ST. LOUIS' and 'SAINT LOUIS' for the same city)."""
    c = (city or "").strip().upper()
    if c.startswith("SAINT "):
        return "ST. " + c[6:]
    if c.startswith("ST ") and not c.startswith("ST. "):
        return "ST. " + c[3:]
    return c


def _metro_from_city(city, state):
    """Map a 13F filer's HQ city/state to a ~60-mile roadshow metro label, so SEC-sourced holders
    and prospects land in the same geographic buckets. Spelling variants are normalised first (see
    _norm_city). Falls back to the normalised 'City, ST' for unmapped US cities and 'International'
    for non-US filers (SEC uses non-US state codes like X0/V8/M0/K3)."""
    c = _norm_city(city)
    st = (state or "").strip().upper()
    # State-qualified key first ("CONCORD|CA" ≠ Concord MA; "LONG BEACH|CA" ≠ Long Beach NY),
    # then the plain city key for unambiguous majors.
    if f"{c}|{st}" in _SEC_CITY_METRO:
        return _SEC_CITY_METRO[f"{c}|{st}"]
    if c in _SEC_CITY_METRO:
        return _SEC_CITY_METRO[c]
    if st and st not in _US_STATES:
        # Cluster the major international hubs (Toronto, London, Zurich...) — real roadshow stops —
        # rather than one undifferentiated "International" bucket. Foreign city strings vary wildly:
        # "TORONTO ONTARIO" / "TORONTO, ONTARIO" (city+province), "LONDON, EC4N 8AF" (city+postcode),
        # "CHIYODA-KU, TOKYO" (ward+city). Try the whole string and each comma-separated token, with
        # province/country suffixes stripped, and take the first that matches a known hub.
        parts = [c] + [p.strip() for p in c.split(",")]
        for cand in parts:
            for _suf in (" ONTARIO", " QUEBEC", " FRANCE", " GERMANY", " ENGLAND", " JAPAN", " CANADA",
                         " SWITZERLAND", " UK"):
                if cand.endswith(_suf):
                    cand = cand[: -len(_suf)].strip()
            if cand in _INTL_CITY_METRO:
                return _INTL_CITY_METRO[cand]
        # Non-hub foreign filer: label by country/province from the EDGAR state code (the verified
        # map in firm_book) instead of one opaque "International" blob — so a UK fund outside London
        # reads "United Kingdom", a Swiss fund outside Zurich/Geneva reads "Switzerland", etc. Only
        # a truly unmapped code falls through to "International".
        from core.firm_book import _state_name
        _cn = _state_name(st)
        return _cn if _cn != st else "International"
    if c and st:
        return f"{c.title()}, {st}"      # normalised so St / St. / Saint variants don't split
    if c:
        return c.title()
    return "Unknown (SEC)"


def _loc_line(r):
    """A prospect card's location, reconciled with the roadshow-metro aggregation. Shows the raw
    'City, ST' and, when the fund clusters into a DIFFERENT metro (e.g. Greenwich, CT → New York),
    appends that metro so the card matches the by-metro table instead of contradicting it."""
    city, state = r.get("city"), r.get("state")
    raw = ", ".join(x for x in [city, state] if x)
    metro = _metro_from_city(city, state)
    if not raw:
        return metro or "—"
    if metro and metro not in ("International", "Unknown (SEC)") and metro != raw:
        return f"{raw} · {metro} metro"
    return raw


def _shortlist_record(c):
    """A peer-prospect candidate captured as a 'shortlisted' NDR target (pipeline entry, no slot
    yet). Kept in trip['shortlist'] — NOT trip['meetings'] — so the slot-based schedule renderer is
    untouched until a target confirms (see NDR-pipeline design, phases 1→3)."""
    from datetime import datetime as _dt
    comps = ", ".join(sorted((c.get("comps") or {}).keys()))
    return {
        "institution": c.get("filer"),
        "city": c.get("city"), "state": c.get("state"),
        "metro": _metro_from_city(c.get("city"), c.get("state")),
        "conviction": c.get("conviction"),
        "peers": comps or ("Curated target" if c.get("kind") == "curated" else ""),
        "bucket": c.get("tier"),
        "status": "shortlisted", "source": "outbound",
        "added_at": _dt.now().strftime("%Y-%m-%d %H:%M"),
    }


def _shortlist_from_inst(inst):
    """Shortlist record from a scored Buy-Side institution (different shape than a peer-prospect
    candidate). Lets the Buy-Side list feed the same NDR pipeline (Phase 4B)."""
    from datetime import datetime as _dt
    peers = inst.get("Peer_Holdings") or []
    return {
        "institution": inst.get("Fund"),
        "city": inst.get("City"), "state": None, "metro": inst.get("Metro"),
        "conviction": inst.get("Engagement_Score"),
        "peers": ", ".join(peers) if peers else "",
        "bucket": "Holder" if inst.get("USIO_Holder") else "Tracked",
        "status": "shortlisted", "source": "outbound",
        "added_at": _dt.now().strftime("%Y-%m-%d %H:%M"),
    }


_DIR_LABEL = {"add": "Adding", "adding": "Adding", "increase": "Adding",
              "trim": "Trimming", "trimming": "Trimming", "decrease": "Trimming",
              "new": "New", "flat": "Steady", "exit": "Exited", "exited": "Exited"}


def _dir_label(d):
    """Holder trajectory tag from a 13F position-history direction (Adding / Trimming / New / …)."""
    return _DIR_LABEL.get((d or "").strip().lower(), "")


def _holder_dollars(v):
    """Compact position-size label for a holder ($1.3M / $940K / $2.1B)."""
    if not isinstance(v, (int, float)) or not v:
        return "—"
    if v >= 1e9: return f"${v / 1e9:.1f}B"
    if v >= 1e6: return f"${v / 1e6:.1f}M"
    if v >= 1e3: return f"${v / 1e3:.0f}K"
    return f"${v:.0f}"


def _holder_position_cells(it):
    """The 'Position' readout for a HOLDER row (replaces the meaningless buy-conviction/engagement
    score): size in $ for the score column, and trajectory + %-of-their-book + how that weight
    compares to how they hold our COMPS (Over/In-line/Under-weight) for the detail column. A holder
    already bought us — what matters is how big a bet we are, the direction, and whether we're
    under our own comp in their book (the upsell case), NOT a likelihood-to-buy number."""
    bp = it.get("Book_Pct")
    parts = []
    dl = _dir_label(it.get("Direction"))
    if dl:
        parts.append(dl)
    if isinstance(bp, (int, float)) and bp:
        parts.append(f"{bp:.1f}% of their book")
    vs, comp, cw = it.get("Weight_Vs_Comp"), it.get("Comp_Best"), it.get("Comp_Weight_Pct")
    if vs and comp:
        parts.append(f"{vs} vs {comp} ({cw}%)")
    return _holder_dollars(it.get("Position_Value")), (" · ".join(parts) or "Existing position")


def _enrich_holders_vs_comps(institutions, cid):
    """For every existing holder, compare OUR weight in their book to how THEY weight our tight
    comps (from the peer 13F books) → Overweight / In-line / Underweight vs the comp they hold
    biggest. Turns a context-free "1.4% of their book" into an actionable read: "underweight vs
    CLRT — we should be at least where our peer is." Only holders with a real comp overlap get the
    tag; the rest just show size + trajectory. comp% = comp_value / their_book_total * 100, and
    since Book_Pct = Position_Value / their_book_total * 100, comp% = comp_value * Book_Pct /
    Position_Value — so no separate book-total lookup is needed."""
    from core import sec_filings, peer_prospects
    try:
        comps = peer_prospects.tight_comps(cid) or set()
    except Exception:
        comps = set()
    if not comps:
        return
    comp_val = {}                                   # holder cik -> {comp_ticker: position_value}
    for tk in comps:
        for h in (sec_filings.get_cached_13f_holders(tk).get("holders") or []):
            k = str(h.get("cik") or "").lstrip("0")
            if k:
                comp_val.setdefault(k, {})[tk] = h.get("value") or 0
    for inst in institutions:
        if not inst.get("USIO_Holder"):
            continue
        bpct, pv = inst.get("Book_Pct"), inst.get("Position_Value")
        cvs = comp_val.get(str(inst.get("cik") or "").lstrip("0"))
        if not (bpct and pv and cvs):
            continue
        best_tk, best_v = max(cvs.items(), key=lambda kv: kv[1])
        if not best_v:
            continue
        comp_pct = best_v * bpct / pv               # their comp weight, same book-total basis
        inst["Comp_Best"] = best_tk
        inst["Comp_Weight_Pct"] = round(comp_pct, 1)
        diff = bpct - comp_pct
        inst["Weight_Vs_Comp"] = ("Overweight" if diff > 0.15 else
                                  "Underweight" if diff < -0.15 else "In-line")


def _is_quant_inst(it):
    """A quant/systematic manager — tracked, but NOT invitable to a 1x1 (they don't take
    management meetings). Read from the institution record's style tag."""
    return bool(it.get("Quant") or it.get("Ownership_Style") == "Quant/Systematic")


def _open_metro_select_dialog(metro, funds, holders=None):
    """Click-a-metro → tick who to see there → shortlist them onto an NDR. ONE list that combines
    existing HOLDERS (defend the relationship) with PEER-OWNERS (convert — own a comp, not us),
    the way an IR team actually builds a roadshow. Quant/systematic shops are shown but NOT
    invitable (no management 1x1s)."""
    holders = list(holders or [])
    funds = sorted(funds, key=lambda c: -(c.get("conviction") or 0))
    items = holders + funds                       # holders first — existing relationships lead
    _key = lambda it: it.get("filer") or pretty_name(it.get("Fund") or "")
    by_key = {_key(it): it for it in items}
    with ui.dialog() as dlg, ui.card().style("min-width:min(880px,95vw);max-width:95vw;"):
        n_h, n_f = len(holders), len(funds)
        with ui.row().classes("w-full justify-between items-center"):
            ui.label(f"{metro} — {n_h} holder{'s' if n_h != 1 else ''} to defend · "
                     f"{n_f} peer-owner{'s' if n_f != 1 else ''} to convert").classes("text-lg font-bold")
            ui.button(icon="close", on_click=dlg.close).props("flat round dense")
        if not items:
            ui.label("Nobody tracked in this metro yet.").style(f"color:{COLORS['text_muted']};")
            dlg.open(); return
        _open = [(i, t) for i, t in enumerate(_load_json("ndr_trips.json", [])) if t.get("status") != "Completed"]
        trip_opts = {str(i): (t.get("name") or f"NDR {i+1}") for i, t in _open}
        trip_opts["__new__"] = "＋ New NDR…"
        # If this metro already has an open NDR, default the selector to it and pre-check the names
        # already on it — so re-opening a planned metro reads as "here's the plan," not a blank slate.
        _metro_pair = next(((i, t) for i, t in _open if (t.get("city") or "") == metro), None)
        _already = {}
        if _metro_pair:
            _mt_name = _metro_pair[1].get("name") or "this NDR"
            for _e in (_metro_pair[1].get("shortlist") or []) + (_metro_pair[1].get("meetings") or []):
                if _e.get("institution"):
                    _already[_e["institution"]] = _mt_name
        with ui.row().classes("w-full items-end gap-2").style(
                f"background:{COLORS['surface_hover_bg']};border-radius:8px;padding:10px 12px;margin-bottom:6px;"):
            ui.label("Tick names, then add them to:").style(
                f"color:{COLORS['text_heading']};font-weight:600;font-size:var(--fs-sm);")
            trip_sel = ui.select(trip_opts, value=(str(_metro_pair[0]) if _metro_pair else "__new__"),
                                 label="NDR").props("dense outlined").style("min-width:190px;")
            new_name = ui.input("New NDR name", value=f"{metro} NDR").props("dense outlined").style("min-width:170px;")
            new_name.bind_visibility_from(trip_sel, "value", backward=lambda v: v == "__new__")
            add_btn = ui.button("Add to NDR", icon="playlist_add").props("dense color=primary")
        d_rows, _preselect = [], []
        for it in items:
            q = _is_quant_inst(it)
            if it.get("filer"):                   # peer-owner candidate (convert)
                conv = it.get("conviction")
                peers = ", ".join(sorted((it.get("comps") or {}).keys()))
                cat = "Peer-owner (convert)"
                score = round(conv) if isinstance(conv, (int, float)) else "—"
                detail = peers or ("Curated target" if it.get("kind") == "curated" else "—")
            else:                                 # existing holder (defend) — or a quant to skip
                cat = "Quant — no 1×1" if q else "Holder (defend)"
                score, detail = _holder_position_cells(it)   # position SIZE + trajectory, not buy-conviction
            _on = _already.get(it.get("filer") or it.get("Fund"))   # already on this metro's NDR?
            if _on:
                detail = f"✓ on {_on}" + (f" · {detail}" if detail and detail != "—" else "")
            _row = {"_filer": _key(it),
                    "Fund": pretty_name(it.get("filer") or it.get("Fund") or "—"),
                    "City": (it.get("city") or it.get("City") or "—"),
                    "Category": cat, "Score": score, "Detail": detail}
            d_rows.append(_row)
            if _on:
                _preselect.append(_row)
        d_cols = [{"name": k, "label": ("Conviction / Position" if k == "Score" else k),
                   "field": k, "align": "right" if k == "Score" else "left"}
                  for k in ("Fund", "City", "Category", "Score", "Detail")]
        tbl = ui.table(columns=d_cols, rows=d_rows, row_key="_filer",
                       selection="multiple").classes("w-full").props("dense flat")
        # Click a fund NAME → open its Account 360 detail — consistent with every other place a name is
        # clickable. @click.stop so opening the profile doesn't also toggle the row's selection checkbox.
        tbl.add_slot("body-cell-Fund", '''
            <q-td :props="props">
              <span class="name-link"
                    @click.stop="() => $parent.$emit('nameclick', props.row._filer)">{{ props.row.Fund }}</span>
            </q-td>
        ''')

        def _open_name(e):
            _k = e.args[0] if isinstance(e.args, (list, tuple)) else e.args
            it = by_key.get(_k)
            if it:
                _open_account_profile(it)
        tbl.on("nameclick", _open_name)
        if _preselect:                            # names already on this metro's NDR start ticked
            tbl.selected = _preselect
        _hint = ("Names already on this metro's NDR are ticked and marked ✓. " if _preselect else "")
        ui.label(_hint + "Peer-owners show conviction to buy (0–100) · Holders show position size + "
                 "trajectory (Adding / Trimming / New) — a holder already owns you, so size and direction "
                 "matter, not a buy score. Quant/systematic shops are listed but can't be invited to a 1×1.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-top:2px;")

        def _do_add():
            sel = tbl.selected
            if not sel:
                ui.notify("Tick at least one name first.", type="warning"); return
            trips2 = _load_json("ndr_trips.json", [])
            if trip_sel.value == "__new__":
                nm = (new_name.value or "").strip()
                if not nm:
                    ui.notify("Name the new NDR.", type="warning"); return
                trips2.append({
                    "name": nm, "sponsor_bank": "", "dates": "TBD", "ndr_type": "in-person",
                    "city": metro, "focus": "", "team": [], "notes": "", "meetings": [], "shortlist": [],
                    "status": "Planning", "debrief": {}, "days": 2, "slots_per_day": 6,
                    "created": datetime.now().strftime("%Y-%m-%d"),
                })
                target, tname = trips2[-1], nm
            else:
                target = trips2[int(trip_sel.value)]
                tname = target.get("name") or "NDR"
            target.setdefault("shortlist", [])
            have = {s.get("institution") for s in target["shortlist"]} | \
                   {m.get("institution") for m in target.get("meetings", [])}
            added = quant_skip = 0
            for row in sel:
                it = by_key.get(row.get("_filer"))
                if not it:
                    continue
                nm2 = it.get("filer") or it.get("Fund")
                if not nm2 or nm2 in have:
                    continue
                if _is_quant_inst(it):
                    quant_skip += 1; continue     # quant/systematic — no management 1x1
                rec = _shortlist_record(it) if it.get("filer") else _shortlist_from_inst(it)
                target["shortlist"].append(rec)
                have.add(nm2); added += 1
            _save_json("ndr_trips.json", trips2)
            dup = len(sel) - added - quant_skip
            msg = f"Added {added} target(s) to '{tname}' (Planning)"
            if quant_skip:
                msg += f" · skipped {quant_skip} quant shop(s) — no 1×1"
            if dup:
                msg += f" · {dup} already on it"
            ui.notify(msg + ". Find it in Outbound → NDR → Active NDRs.", type="positive")
            dlg.close()
        add_btn.on_click(_do_add)
    dlg.open()


def _open_shortlist_outreach(entry, contact, on_invited):
    """NDR pipeline Phase 2 — the 'Contact' draft. Opens a pre-filled outreach email the user
    SENDS themselves (never auto-sent), and a button to mark the target Invited + log it. If no
    email is on file, the draft is still shown so they can reach out via their own channel."""
    fund = entry.get("institution", "")
    tkr, cname = CT("ticker"), CT("name")
    ir = CI()
    where = entry.get("metro") or entry.get("city") or "your area"
    peers = entry.get("peers") or "names in our space"
    greeting = (contact.get("name", "").split()[0] if contact.get("name") else "there")
    subject = f"{tkr} — meeting request during our {where} investor visit"
    body = (f"Hi {greeting},\n\nWe're planning an investor visit around {where} and would value 30 "
            f"minutes with your team. {pretty_name(fund)} holds {peers}, so {tkr}'s story should be "
            f"directly relevant.\n\nWould a meeting work during that window?\n\n"
            f"{ir.get('name', '')}\n{ir.get('title', 'Investor Relations')} · {cname} (NASDAQ: {tkr})")
    with ui.dialog() as dlg, ui.card().style(f"background:{COLORS['surface_bg']};min-width:min(520px,94vw);"):
        ui.label(f"Contact — {pretty_name(fund)}").classes("text-lg font-bold")
        ui.textarea(value=body).classes("w-full").props("rows=10 outlined")
        email = contact.get("email", "")
        if email:
            _mailto(email, subject, body, f"✉ Open email to {contact.get('name', 'the contact')} ({email})")
        else:
            ui.label("No contact email on file — reach out via your own channel, then mark Invited below.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        with ui.row().classes("w-full justify-end gap-2").style("margin-top:6px;"):
            ui.button("Cancel", on_click=dlg.close).props("flat")

            def _mark():
                on_invited()
                dlg.close()
                ui.notify(f"{pretty_name(fund)} marked Invited and logged.", type="positive")
            ui.button("Mark as invited & log", icon="outgoing_mail", on_click=_mark).props("color=primary")
    dlg.open()


def _slot_time(slot_index):
    """Clock time for the Nth slot of a day — 90-min cadence from 9:00 AM (matches the Plan New NDR
    grid). Slot 0 → 9:00 AM, slot 1 → 10:30 AM, …"""
    total = slot_index * 90
    hr, mn = 9 + total // 60, total % 60
    hr12 = hr if hr <= 12 else hr - 12
    return f"{hr12 or 12}:{mn:02d} {'AM' if hr < 12 else 'PM'}"


def _ndr_capacity(trip):
    return int(trip.get("days") or 2), int(trip.get("slots_per_day") or 6)


def _next_open_slot(trip):
    """(day, slot_index) of the next free slot in the trip's day×slots grid, filling days in order.
    Counts existing meetings per day (works for legacy trips with no slot_index). (None, None) = full."""
    days, per = _ndr_capacity(trip)
    count = {}
    for m in trip.get("meetings", []):
        if m.get("type") == "break":
            continue
        d = m.get("day") or 1
        count[d] = count.get(d, 0) + 1
    for d in range(1, days + 1):
        if count.get(d, 0) < per:
            return d, count.get(d, 0)
    return None, None


def _open_schedule_request_dialog(request, on_done):
    """NDR pipeline Phase 4 — an inbound analyst request IS a confirmation, so it goes straight into
    an NDR's schedule at the next open slot (flagged source='inbound'), and the request is resolved.
    Inbound and outbound converge into one Active NDR instead of two tabs."""
    trips = _load_json("ndr_trips.json", [])
    open_trips = [(i, t) for i, t in enumerate(trips) if t.get("status") != "Completed"]

    def _match(t):
        tc = (t.get("city") or "").lower()
        rm, rc = (request.get("metro") or "").lower(), (request.get("city") or "").lower()
        return bool(tc) and (rm and (rm in tc or tc in rm) or rc and (rc in tc or tc in rc))

    opts = {str(i): (("★ " if _match(t) else "") + (t.get("name") or f"NDR {i+1}")) for i, t in open_trips}
    opts["__new__"] = "＋ New NDR…"
    default = next((str(i) for i, t in open_trips if _match(t)),
                   (str(open_trips[0][0]) if open_trips else "__new__"))
    with ui.dialog() as dlg, ui.card().style(f"background:{COLORS['surface_bg']};min-width:min(520px,94vw);"):
        ui.label(f"Schedule request — {request.get('analyst', '')} ({request.get('firm', '')}) → {request.get('city', '')}").classes("text-lg font-bold")
        ui.label("An inbound request is a confirmed meeting — it drops straight into the schedule at the "
                 "next open slot.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        sel = ui.select(opts, value=default, label="Add to NDR").props("dense outlined").classes("w-full")
        new_name = ui.input("New NDR name", value=f"{request.get('city', '')} NDR").props("dense outlined").classes("w-full")
        new_name.bind_visibility_from(sel, "value", backward=lambda v: v == "__new__")
        with ui.row().classes("w-full justify-end gap-2").style("margin-top:6px;"):
            ui.button("Cancel", on_click=dlg.close).props("flat")

            def _go():
                trips2 = _load_json("ndr_trips.json", [])
                if sel.value == "__new__":
                    nm = (new_name.value or "").strip()
                    if not nm:
                        ui.notify("Name the new NDR.", type="warning"); return
                    trips2.append({
                        "name": nm, "sponsor_bank": request.get("firm", ""), "dates": "TBD",
                        "ndr_type": "in-person", "city": request.get("city", ""), "focus": "", "team": [],
                        "notes": "", "meetings": [], "shortlist": [], "status": "Planning", "debrief": {},
                        "days": 2, "slots_per_day": 6, "created": datetime.now().strftime("%Y-%m-%d"),
                    })
                    trip, tname = trips2[-1], nm
                else:
                    trip = trips2[int(sel.value)]
                    tname = trip.get("name") or "NDR"
                day, slot = _next_open_slot(trip)
                if day is None:
                    ui.notify("That NDR's slots are full — raise its capacity or pick another.", type="warning"); return
                trip.setdefault("meetings", []).append({
                    "institution": request.get("firm", ""), "contact": request.get("analyst", ""),
                    "day": day, "slot_index": slot, "time": _slot_time(slot), "type": "1x1",
                    "format": "In-person", "status": "scheduled", "notes": request.get("reason", ""),
                    "non_holder": True, "score": None, "source": "inbound",
                    "inbound_request_id": request.get("id"),
                    "confirmed_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
                _save_json("ndr_trips.json", trips2)
                reqs = _load_ndr_requests()
                for rr in reqs:
                    if rr.get("id") == request.get("id"):
                        rr["resolved"] = True
                        rr["resolved_at"] = datetime.now().strftime("%b %d, %Y")
                        rr["scheduled_into"] = tname
                _save_ndr_requests(reqs)
                try:
                    from core import activity_log
                    activity_log.log_event("ndr_inbound_scheduled", entity=request.get("firm", ""),
                                           launched_from=f"NDR · {tname}")
                except Exception:
                    pass
                dlg.close()
                ui.notify(f"Scheduled {request.get('analyst', '')} → Day {day} · {_slot_time(slot)}. Request resolved.",
                          type="positive")
                on_done()
            ui.button("Schedule into NDR", icon="event_available", on_click=_go).props("color=primary")
    dlg.open()


def _sec_holder_record(name, source, holder, peer_of=None, city=None, state=None, style="Active"):
    """A full institution record for a real SEC-sourced name, with honest
    'unknown' defaults for every enrichment field the scorer and cards read —
    so a 13F/13D-G/peer-overlap holder lives in the same universe as the seed
    without breaking scoring. It simply scores low until it's enriched.
    city/state (from the 13F filing address) drive the Metro breakdown.
    `style` = manager style ("Active" / "Quant/Systematic" / …); quant shops are
    tracked but NOT 1x1-invitable (they don't take management meetings)."""
    action = {
        "SEC 13F": "Confirmed 13F holder — enrich & prioritize",
        "Peer 13F": "Holds a peer — real NDR prospect",
        "SEC 13D/G": "Filed a 5%+ stake — engage directly",
    }.get(source, "SEC-sourced name")
    return {
        "Fund": name, "Type": "Institutional (SEC)", "AUM": "—", "Coverage_Priority": 3,
        "USIO_Holder": holder, "Shares": 0, "QoQ_Change": 0, "Call_Listener": False,
        "Listen_Duration": "—", "Peer_Holdings": ([peer_of] if peer_of else []),
        "IR_Visits_30d": 0, "Last_Visit": "—", "Conviction": "—",
        "Call_Score": 0, "Peer_Score": 0, "Visit_Score": 0,
        "Turnover_Style": "Unknown (SEC)", "Metro": _metro_from_city(city, state), "Ownership_Style": style,
        "Quant": (style == "Quant/Systematic"),   # tracked but not 1x1-invitable
        "Action": action, "Source": source,
    }


def _sec_universe_records(client_id):
    """Build source-tagged institution records from CACHED SEC data (no live
    calls at render): USIO 13F holders, 13D/13G filers, and funds holding the
    peer set (real NDR prospects). Deduped by filer name."""
    recs = {}
    ticker = CT("ticker")
    for h in (sec_filings.get_cached_13f_holders(ticker).get("holders", []) or []):
        name = (h.get("filer") or "").strip()
        if not name:
            continue
        r = recs.setdefault(name, _sec_holder_record(
            name, "SEC 13F", True, city=h.get("city"), state=h.get("state"),
            style=(h.get("style") or "Active")))
        # Real, exact position size from the bulk 13F dataset (size_known).
        if h.get("size_known") and h.get("shares"):
            r["Shares"] = h["shares"]
        # Backfill metro if this record was first created from a 13D/G (no city).
        if r.get("Metro") == "Unknown (SEC)" and h.get("city"):
            r["Metro"] = _metro_from_city(h.get("city"), h.get("state"))
    for f in (sec_filings.get_cached_13d_13g(ticker, refresh_if_stale=False).get("filings", []) or []):
        name = (f.get("filer") or f.get("name") or "").strip()
        if not name:
            continue
        if name in recs:
            recs[name]["Source"] = "SEC 13F + 13D/G"
            recs[name]["Action"] = "Holder with a 5%+ filing — engage directly"
        else:
            recs[name] = _sec_holder_record(name, "SEC 13D/G", True)
    # Peer-overlap prospects are NO LONGER auto-injected here. Dumping every
    # comp's top holders into the universe filled it with index/passive noise and
    # false positives (funds that already own USIO sub-13F-threshold). They now go
    # through a reviewed, conviction-ranked queue (core.peer_prospects, the "Peer
    # Prospects" tab) and only enter the pipeline when the IR lead promotes one —
    # at which point it lands in prospects.json like any other manual prospect.
    for r in recs.values():
        r.pop("_peer_tier", None)  # internal ranking key — not part of the record
    return list(recs.values())


def _promoted_prospect_records(client_id):
    """Promoted non-holders. Delegates to core.targets.promoted_prospects — the report generators
    need the same list, and a second copy here would drift."""
    from core import targets as targets_mod
    return targets_mod.promoted_prospects(client_id)


def _score_val(inst, key="Engagement_Score"):
    """A score for COMPARISON / sorting only.

    Scores are None when nothing was measurable — there is no website-analytics or call-listener
    integration, so Digital_Intent_Score is None for every real 13F institution and
    Engagement_Score is None when no pillar has an input. None can't be compared to an int, and
    that TypeError takes the entire page down (see tests/smoke_render.py).

    Treated as 0 for filtering/tiering, while the UI still DISPLAYS "—" for None — so an absent
    measurement is never rendered as a confident zero.
    """
    v = inst.get(key)
    return 0 if v is None else v


def _merge_sec_universe(seed_institutions, client_id):
    """Append real SEC-sourced names to the seed universe. A name that already
    matches a seed fund (normalized) doesn't duplicate — the seed record wins
    and is marked 'Seed + SEC-confirmed' so the demo shows it's validated by a
    real filing."""
    by_norm = {}
    for i in seed_institutions:
        by_norm.setdefault(_norm_name(i["Fund"]), i)
    for rec in _sec_universe_records(client_id):
        key = _norm_name(rec["Fund"])
        if key in by_norm:
            s = by_norm[key]
            if "SEC" not in s.get("Source", ""):
                s["Source"] = "Seed + SEC-confirmed"
            continue
        seed_institutions.append(rec)
    return seed_institutions


# ─────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────
def render_investors_page(section="ownership"):
    """Investor Targeting, split into three focused rail sections that share this render:
      • ownership — Buyside · NOBO · Website  (with the "who owns you" Big Picture + mode toggle)
      • targeting — Target Database · Peer Prospects · Import
      • roadshow  — Meeting Hub · NDR · CRM
    Each shows only its own ~3 tabs, so no page carries the old nine-tab strip."""
    client_id = get_active_client_id()
    mode_state = _load_json("buyside_mode.json", {})
    earnings_date_str = CE().get("earnings_date", "2026-08-12")
    earnings_date = datetime.strptime(earnings_date_str, "%Y-%m-%d").date()
    days_to_earnings = (earnings_date - datetime.now().date()).days
    auto_post = days_to_earnings < 0
    mode = mode_state.get("mode", "post" if auto_post else "pre")

    q2_listeners = set(mode_state.get("q2_listeners", []))
    meeting_log = _load_meeting_log()
    # REAL 13F holders are the target universe (core/targets.py) — position size, % of the
    # holder's own book, add/trim/new direction, and the contact behind the filing. This replaced
    # data/seed/buyside_institutions.py, whose 62 records carried fabricated conviction and
    # engagement scores; showing a client an invented "Conviction: High" next to a real filing is
    # exactly the credibility risk this platform exists to avoid.
    from core import targets as targets_mod
    raw_institutions = targets_mod.targets_as_institutions(client_id=client_id)
    # Merge real SEC-sourced names (13D/G filers, peer-overlap prospects) from cached EDGAR data —
    # each carries its own Source tag, and this is what the "Refresh from SEC EDGAR" button on the
    # SEC Intelligence tab grows. Reads cache only here (no live network call at page render).
    raw_institutions = _merge_sec_universe(raw_institutions, client_id)
    # Promoted non-holders from the Peer Prospects queue. Without this the universe is holders-only
    # — a target database with no targets to win.
    _existing = {_norm_name(i["Fund"]) for i in raw_institutions}
    raw_institutions += [p for p in _promoted_prospect_records(client_id)
                         if _norm_name(p["Fund"]) not in _existing]
    # Upgrades Peer_Holdings from static seed data to real SEC 13F filings
    # wherever a ticker has actually been refreshed (see
    # _enrich_peer_holdings_with_live_13f's docstring) — must run BEFORE
    # scoring, since post-earnings mode's Catalyst Fit pillar reads
    # Peer_Holdings too, not just Peer Cross-Targeting below.
    # The illustrative demo is the ONE exception: its comp books (PYRA/CLRT/VNTG) are cached to
    # drive Peer Prospects, so the enrichment sees them as "fetched" and — finding our own holders
    # correctly NOT among the comps' filers — wipes the seeded peer_overlap, leaving every holder
    # card reading "Peers held: —". For a self-contained demo the seed IS the intended truth, so we
    # keep it. (Seeding the holders INTO the comp books would confirm the overlap but inflate the
    # comps' holder counts, flipping the peer-average stability signal — see _peer_holder_avg.)
    from core.curated_targets import _is_illustrative
    if not _is_illustrative(client_id):
        _enrich_peer_holdings_with_live_13f(raw_institutions, [p["ticker"] for p in _load_peer_universe()])
    # Populate the IR-website + call-listener engagement signals BEFORE scoring, so the Engagement
    # Score actually reflects the inputs its tooltip advertises (and the demo shows an active book
    # instead of "no call/visit data" on every card).
    _enrich_engagement_signals(raw_institutions, client_id)
    institutions = _score_institutions(raw_institutions, mode, q2_listeners, meeting_log)

    _TITLES = {"ownership": "Ownership — who owns you",
               "targeting": "Targeting — who should own you",
               "roadshow": "Roadshow & CRM"}
    ui.label(f"{CT('name')} — {_TITLES.get(section, 'Investor Targeting')}").classes(
        "text-2xl font-bold").style(f"color:{COLORS['text_heading']};")

    # Mode-dependent sections re-render in place when the Pre/Post toggle changes — no full reload.
    _mode_ctx = {"mode": mode, "institutions": institutions}

    @ui.refreshable
    def _big_picture_section():
        _render_big_picture(_mode_ctx["institutions"], "summary" if section == "ownership" else "geo")

    @ui.refreshable
    def _mode_desc_section():
        _render_mode_description(_mode_ctx["mode"])

    @ui.refreshable
    def _buyside_section():
        _render_buyside_tab(_mode_ctx["institutions"], meeting_log, _mode_ctx["mode"])

    _invalidate_mode_tabs = lambda: None      # real impl assigned once loaded_tabs exists (below)

    def on_mode_change(new_mode):
        _mode_ctx["mode"] = new_mode
        _mode_ctx["institutions"] = _score_institutions(raw_institutions, new_mode, q2_listeners, meeting_log)
        mode_state["mode"] = new_mode
        _save_json("buyside_mode.json", mode_state)
        if section == "ownership":
            _big_picture_section.refresh()
            _buyside_section.refresh()
        _mode_desc_section.refresh()
        _invalidate_mode_tabs()

    # Preamble: the "who owns you" Big Picture leads the Ownership section only. The Pre/Post-earnings
    # mode toggle drives scoring for Ownership + Targeting so it shows on both; Roadshow opens
    # straight into its tools using the saved mode.
    # The "Big Picture — Where Things Stand" synthesis (which includes the geographic "Institutions &
    # Peer Ownership By City/Metro Area" table) leads TARGETING — it's the strategic "where to go"
    # overview, and Targeting is now the first section. Ownership opens straight into its Buy-side/
    # NOBO/Website tabs (whose Buyside Read already carries the who-owns-you summary).
    # Ownership leads with the "who owns you" summary. Targeting's geographic block (WHERE TO GO
    # read + the metro table) no longer sits here — it now renders inside the Target Database tab,
    # directly under the summary cards (see _render_target_db_tab), per the requested layout.
    if section == "ownership":
        _big_picture_section()
        ui.markdown("---")
    # Ownership shows the Pre/Post-earnings mode toggle at the top. Targeting's toggle no longer sits
    # here — it renders inside the Target Database tab, directly BELOW the metro table (per request),
    # via this callback passed into _render_target_db_tab.
    if section == "ownership":
        _render_mode_toggle_control(_mode_ctx["mode"], on_mode_change)
        _mode_desc_section()
        ui.markdown("---")

    def _render_targeting_mode_controls():
        _render_mode_toggle_control(_mode_ctx["mode"], on_mode_change)
        _mode_desc_section()

    # Each section's ~3 tabs: (label, build fn, mode-dependent?). Only the active section is built.
    _SPECS = {
        # ── Ownership — who owns the stock
        "ownership": [("Buyside Ownership", lambda: _buyside_section(), False),
                      ("NOBO Ownership", lambda: _render_nobo_tab(), False),
                      ("Website Ownership", lambda: _render_web_flow_tab(client_id), False)],
        # ── Targeting — who should own it
        "targeting": [("Target Database", lambda: _render_target_db_tab(_mode_ctx["institutions"], client_id, mode_controls=_render_targeting_mode_controls), True),
                      ("Peer Prospects", lambda: _render_peer_prospects_tab(client_id), False),
                      ("Import list", lambda: _render_import_verify_tab(client_id), False)],
        # ── Outbound — reach & track. Tab identities are short (Meeting Hub / NDR / CRM) and match the
        # sidebar's NAV_SUBITEMS["Roadshow"] pointers exactly, so each pointer deep-links to its tab.
        "roadshow":  [("Meeting Hub", lambda: _render_meeting_hub_tab(), False),
                      ("NDR", lambda: _render_ndr_tab(_mode_ctx["institutions"], meeting_log, client_id, _mode_ctx["mode"]), True),
                      ("CRM", lambda: _render_accounts_tab(client_id, _mode_ctx["institutions"]), False)],
    }
    specs = _SPECS.get(section, _SPECS["ownership"])

    # Outbound (roadshow) alone has sidebar sub-items (Meeting Hub · NDR · CRM in NAV_SUBITEMS), which
    # deep-link to these exact tabs via consume_target_tab() below — so its on-page strip is redundant
    # on desktop, same as Markets/Earnings: hide it via .page-tabstrip (kept <1024px, no docked sidebar).
    # Ownership/Targeting have NO sidebar sub-items, so their strips are the only nav — keep them shown.
    if section == "roadshow":
        _tabstrip_cls = "w-full page-tabstrip"          # hidden on desktop — the left-nav sub-items are the nav
    elif section == "ownership":
        _tabstrip_cls = "w-full seg-tabs seg-tabs-own"  # segmented buttons + two-line labels (OWNERSHIP at bottom)
    else:
        _tabstrip_cls = "w-full seg-tabs"               # segmented buttons (single line)
    with ui.tabs().classes(_tabstrip_cls) as tabs:
        _tabobjs = {name: ui.tab(name) for name, _fn, _md in specs}
    # A deep-link (go_to(section, tab)) can open on any of the section's tabs; else the first.
    default_tab = _tabobjs.get(nav.consume_target_tab()) or next(iter(_tabobjs.values()))
    _panels = {}
    with ui.tab_panels(tabs, value=default_tab).classes("w-full"):
        for _name, _tobj in _tabobjs.items():
            with ui.tab_panel(_tobj) as _pnl:
                ui.spinner(size="lg").classes("mx-auto").style("margin-top:32px;")
            _panels[_name] = _pnl

    # Lazy tab loading — a tab's content is built the first time it's opened, then cached, so a
    # section never runs every tab's Neon queries up front.
    lazy_panels = {name: (_panels[name], fn) for name, fn, _md in specs}
    from core import lazy_tab_probe
    lazy_tab_probe.register(section.title(), lazy_panels)   # no-op unless smoke is capturing
    loaded_tabs = set()
    # A mode toggle invalidates the section's mode-dependent tabs so they rebuild with fresh scores.
    _mode_dependent_tabs = tuple(name for name, _fn, md in specs if md)
    _invalidate_mode_tabs = lambda: [loaded_tabs.discard(n) for n in _mode_dependent_tabs]

    async def _load_tab_on_demand(e):
        name = e.value
        # Keep the sidebar's sub-item highlight in sync with in-page tab clicks.
        # Guard it: on a STALE client (e.g. after a server restart while the
        # page stayed open) render_nav() raises "client has been deleted".
        # That must not abort the tab build below and strand the panel on its
        # spinner forever.
        try:
            nav.tab_changed(name)
        except Exception:
            pass
        if name not in lazy_panels or name in loaded_tabs:
            return
        container, build_fn = lazy_panels[name]
        # Yield to the event loop once so the spinner that's already in
        # `container` actually reaches the browser before this function
        # goes on to do the (synchronous) database reads that tab needs —
        # without this await, the spinner and the real content would both
        # get flushed to the browser in the same batch, i.e. no visible
        # "loading" feedback at all, just a delay.
        await asyncio.sleep(0)
        try:
            container.clear()
            with container:
                build_fn()
        except Exception as ex:
            # Never leave the tab stuck on its spinner. Replace it with an
            # actionable error and — critically — do NOT mark the tab loaded,
            # so re-selecting it retries the build.
            try:
                container.clear()
                with container:
                    ui.label("This tab didn't finish loading.").style(
                        f"color:{COLORS['text_heading']};font-weight:600;")
                    ui.label(str(ex)).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    ui.label("If this persists, reload the page (Ctrl-R) — the server may "
                             "have restarted.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    ui.button("Reload page", on_click=lambda: ui.navigate.reload()).props("outline dense")
            except Exception:
                pass  # client is truly gone; only a browser refresh can recover
            return
        # Success — now it's safe to remember it's built.
        loaded_tabs.add(name)

    tabs.on_value_change(_load_tab_on_demand)

    # Eager-build whichever tab we open on (deep-linked or the default t1) — its
    # panel currently shows only a spinner and _load_tab_on_demand fires on
    # change, not initial load.
    _dname = default_tab.props["name"]
    loaded_tabs.add(_dname)
    _dcont, _dbuild = lazy_panels[_dname]
    _dcont.clear()
    with _dcont:
        _dbuild()


# The three rail sections Investor Targeting was split into — each dispatched by the app as
# render_{section}_page() and rendering only its own tabs. render_investors_page stays as the
# Ownership entry point (and back-compat alias for anything that still calls it).
def render_ownership_page():
    render_investors_page("ownership")


def render_targeting_page():
    render_investors_page("targeting")


def render_roadshow_page():
    render_investors_page("roadshow")


def _render_import_verify_tab(client_id):
    """Import & Verify a List — paste an outside list of firms (conference / sell-side / broker /
    AI-generated), verify each is a REAL SEC 13F filer (unconfirmed ones flagged, never trusted),
    reconcile against what we already track, and promote the verified-new into the curated target
    book. See core/list_verify.py."""
    import asyncio

    from core import list_verify

    ui.label("Import & Verify a List").classes("text-lg font-bold").style(f"color:{COLORS['text_heading']};")
    ui.label("Paste a list of firms — from a conference, a sell-side desk, a broker, or an AI tool. Each is "
             "checked against SEC EDGAR as a real 13F filer (anything we can't confirm is flagged, never "
             "trusted), reconciled against what we already track, and the verified-new ones can be promoted "
             "into your curated target book. One firm per line.").style(
        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    box = ui.textarea(placeholder="Whitebox Advisors, LLC\nCorsair Capital Management, L.P.\n1620 Investment Advisors").props(
        "outlined autogrow").classes("w-full").style("font-size:var(--fs-sm);margin-top:6px;")
    results = ui.column().classes("w-full").style("gap:6px;margin-top:6px;")
    state = {"checks": []}

    def _render_results(rows):
        results.clear()
        state["checks"] = []
        with results:
            real = [r for r in rows if r["is_real"]]
            unconf = [r for r in rows if not r["is_real"]]
            new_real = [r for r in real if r["status"] == "new"]
            ui.label(f"{len(rows)} firms · {len(real)} verified real · {len(unconf)} unconfirmed · "
                     f"{len(new_real)} verified & new").classes("font-bold").style(
                f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
            _order = {"new": 0, "contacts": 1, "prospect": 2, "holder": 3}
            for r in sorted(rows, key=lambda x: (not x["is_real"], _order.get(x["status"], 9))):
                clr = "#B91C1C" if not r["is_real"] else ("#B45309" if r["status"] == "new" else "#15803D")
                with ui.card().classes("w-full").style(
                        f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};"
                        f"border-left:4px solid {clr};padding:5px 10px;"):
                    with ui.row().classes("w-full items-center").style("gap:8px;"):
                        if r["is_real"]:
                            cb = ui.checkbox(value=(r["status"] == "new")).props("dense")
                            state["checks"].append((cb, r))
                            ui.icon("check_circle").style("color:#15803D;font-size:var(--fs-lg);")
                        else:
                            ui.icon("help_outline").style("color:#B91C1C;font-size:var(--fs-lg);")
                        ui.label(r["name"]).style(f"color:{COLORS['text_heading']};font-size:var(--fs-sm);font-weight:600;")
                        ui.space()
                        if r["is_real"]:
                            ui.label(f"CIK {r['cik'] or '—'}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        badge = ("#B91C1C", "unconfirmed — review") if not r["is_real"] else (clr, r["status_label"])
                        ui.label(badge[1]).style(
                            f"background:{badge[0]}1a;color:{badge[0]};font-size:var(--fs-2xs);font-weight:700;"
                            "padding:1px 8px;border-radius:9px;white-space:nowrap;")

            if not real:
                return

            def _promote():
                sel = [r for cb, r in state["checks"] if cb.value]
                if not sel:
                    ui.notify("Tick at least one verified firm to promote.", type="warning")
                    return
                n = list_verify.promote(sel, client_id)
                ui.notify(f"Promoted {n} firm(s) to the curated target book (Target Database → Curated)."
                          if n else "Those firms are already in the curated book.", type="positive")
            ui.button("Promote selected to targets", icon="playlist_add", on_click=_promote).props(
                "color=primary dense").style("margin-top:6px;")
            ui.label("Only verified firms can be promoted — unconfirmed ones are never added.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);")

    async def _verify():
        names = [ln.strip() for ln in (box.value or "").splitlines() if ln.strip()]
        if not names:
            ui.notify("Paste at least one firm (one per line).", type="warning")
            return
        if len(names) > 100:
            ui.notify("Keeping it to the first 100 firms.", type="warning")
            names = names[:100]
        results.clear()
        with results:
            ui.spinner(size="lg")
            ui.label(f"Verifying {len(names)} firm(s) against SEC EDGAR…").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        try:
            rows = await asyncio.to_thread(list_verify.verify_and_reconcile, names, client_id)
        except Exception as exc:
            results.clear()
            with results:
                ui.label(f"Verification failed: {exc}").style("color:#B91C1C;font-size:var(--fs-sm);")
            return
        _render_results(rows)

    ui.button("Verify & reconcile", icon="fact_check", on_click=_verify).props("color=primary dense").style(
        "margin-top:4px;")


def _render_web_flow_tab(client_id):
    """Website Engagement (web flow) — who visits the IR site and whether they DOWNLOAD
    material or just READ. Real tenants show a Waiting signal (no web-analytics feed is
    wired); the illustrative demo tenant is seeded so the capability is demonstrable.
    See core/web_flow.py."""
    from core import web_flow
    from page_modules_nicegui.signals import waiting_signal

    ui.label("Website Ownership").classes("text-lg font-bold").style(f"color:{COLORS['text_heading']};")
    ui.label("Who comes to your IR site — and whether they DOWNLOAD material (deck, model, filings) or "
             "just READ. A download is a far stronger buy-side intent signal than a page skim, so the two "
             "are scored separately.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    # Capture read-only ONCE, here, where the page role/page context is bound. A rebuild
    # triggered from an upload or on_click callback runs with the ui_context slot unbound,
    # so re-checking is_read_only() there would wrongly return True and hide every control
    # (same class of bug as the tenant-context callback issue). client_id is captured too.
    _read_only = ui_context.is_read_only()
    _body = ui.column().classes("w-full gap-0")

    def _handle_upload(e):
        try:
            rows = web_flow.parse_visitor_csv(e.content.read())
        except Exception as ex:
            ui.notify(f"Couldn't read that CSV: {ex}", type="negative")
            return
        if not rows:
            ui.notify("No rows recognized — the file needs at least an Organization column.", type="warning")
            return
        web_flow.save_visitor_data(rows, client_id)
        ui.notify(f"Loaded {len(rows)} visitor row(s) — web flow now runs on your data.", type="positive")
        _build()

    def _upload_control():
        ui.upload(label="Upload web-analytics export (.csv)", auto_upload=True,
                  on_upload=_handle_upload).props('accept=.csv flat').classes("w-full").style("max-width:440px;")
        ui.label("Recognized columns (any subset): Organization · Category · Pages_Viewed · "
                 "Time_On_Site_Min · Visits · Date · Downloads (comma-separated assets) · Holder. "
                 "Works with a Google Analytics, Q4 or IRWIN export.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

    def _stat(val, label, sub, color=None):
        with ui.card().classes("flex-1").style(
                f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};"
                "min-width:120px;padding:10px 12px;"):
            ui.label(str(val)).classes("font-bold").style(f"color:{color or COLORS['accent']};font-size:var(--fs-2xl);")
            ui.label(label).style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);font-weight:600;")
            ui.label(sub).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

    def _build():
        _body.clear()
        with _body:
            d = web_flow.compose(client_id)
            if not d.get("available"):
                waiting_signal(
                    "a web-analytics export (Google Analytics · Q4 · IRWIN)",
                    detail="Upload your IR-site visitor export — visits, pages, time-on-site and asset "
                           "downloads — and this fills in automatically. No web-analytics feed is wired "
                           "yet, so nothing is shown rather than invented.",
                    unlocks="download-vs-read intent per visitor, the most-pulled assets, and hot prospects "
                            "(non-holders downloading your deck or model).")
                if not _read_only:
                    ui.label("Have an export? Drop it in:").style(
                        f"color:{COLORS['text_body']};font-size:var(--fs-sm);font-weight:600;margin-top:10px;")
                    _upload_control()
                return

            uploaded = d["source"] == "your uploaded web-analytics export"
            ui.label(f"Source: {d['source']}.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-style:italic;margin:2px 0 6px;")

            # THE WEBSITE READ — plain-English BLUF (same callout as NOBO / Buyside / Big Picture).
            _top_assets = ", ".join(a for a, _ in d["by_asset"][:2]) or "—"
            _wread = [f"{d['download_pct']}% of visitors downloaded material — {d['n_downloaders']} of "
                      f"{d['n_visitors']} orgs; the rest read pages without pulling anything."]
            if d["hot_prospects"]:
                _wread.append(f"{len(d['hot_prospects'])} hot prospect(s) pulled real diligence "
                              f"(deck / model / filings) but don't own you yet — the sharpest targets here.")
            if d["by_asset"]:
                _wread.append(f"Downloads skew to the {_top_assets}.")
            with ui.card().classes("w-full").style(
                    "background:rgba(30,64,175,.06);border:1.5px solid #1E40AF;border-left:6px solid #1E40AF;"
                    "border-radius:8px;margin-top:6px;"):
                ui.label("THE WEBSITE READ — who's on your site and pulling material").style(
                    "color:#1E3A8A;font-size:var(--fs-xs);font-weight:700;letter-spacing:.04em;")
                for _wp in _wread:
                    ui.label("• " + _wp).style(
                        f"color:{COLORS['text_heading']};font-size:var(--fs-base);line-height:1.6;font-weight:500;")

            with ui.row().classes("w-full gap-3").style("margin-top:8px;"):
                _stat(d["n_visitors"], "Visitors", "orgs on the site")
                _stat(d["n_downloaders"], "Downloaders", f"{d['download_pct']}% of visitors", "#15803D")
                _stat(d["n_readers"], "Readers only", "browsed, no download")
                _stat(d["total_downloads"], "Assets pulled", "total downloads")
                _stat(len(d["hot_prospects"]), "Hot prospects", "non-holders pulling assets", "#B45309")

            if d["by_asset"]:
                ui.label("Most-downloaded assets").classes("section-head").style("margin-top:12px;")
                _maxc = d["by_asset"][0][1] or 1
                for name, cnt in d["by_asset"]:
                    with ui.row().classes("w-full items-center").style("gap:10px;"):
                        ui.label(name).style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);min-width:150px;")
                        with ui.element("div").style(
                                f"flex:1;background:{COLORS['surface_hover_bg']};border-radius:6px;height:14px;"):
                            ui.element("div").style(
                                f"width:{round(100*cnt/_maxc)}%;background:{COLORS['accent']};height:14px;border-radius:6px;")
                        ui.label(f"{cnt}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);min-width:20px;")

            if d["hot_prospects"]:
                ui.label("Hot prospects — high intent, not yet holders").classes(
                    "section-head").style("margin-top:14px;")
                ui.label("These orgs pulled real diligence material (deck / model / filings) but don't own you "
                         "yet — the sharpest targets this view surfaces. Push one into the Target Database "
                         "pipeline (the same store the Peer Prospects Promote button feeds).").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                _hot_box = ui.column().classes("w-full gap-0")

                def _render_hot():
                    _hot_box.clear()
                    queued = web_flow.target_queue_keys(client_id)
                    with _hot_box:
                        unqueued = [v for v in d["hot_prospects"] if not web_flow.is_in_queue(v["org"], queued)]
                        if unqueued and not _read_only:
                            def _add_all(unqueued=unqueued):
                                n = sum(1 for v in unqueued if web_flow.add_to_target_queue(v, client_id))
                                ui.notify(f"Added {n} hot prospect(s) to the Target Database pipeline.", type="positive")
                                _render_hot()
                            ui.button(f"Add all {len(unqueued)} to target queue", icon="playlist_add",
                                      on_click=_add_all).props("color=primary dense").style("margin:6px 0;")
                        for v in d["hot_prospects"]:
                            already = web_flow.is_in_queue(v["org"], queued)
                            with ui.card().classes("w-full").style(
                                    f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};"
                                    "border-left:3px solid #B45309;padding:8px 10px;margin-top:4px;"):
                                with ui.row().classes("w-full justify-between items-center").style("gap:8px;"):
                                    ui.label(v["org"]).style(f"color:{COLORS['text_heading']};font-size:var(--fs-base);font-weight:600;")
                                    ui.label(f"intent {v['intent']}").style(
                                        "background:#FEF3C7;color:#B45309;padding:1px 8px;border-radius:6px;"
                                        "font-size:var(--fs-xs);font-weight:700;")
                                ui.label(f"{v['category']}  ·  pulled: {', '.join(v['downloads'])}  ·  "
                                         f"{v['pages']} pages / {v['minutes']} min  ·  last {v['last_visit']}").style(
                                    f"color:{COLORS['text_muted']};font-size:var(--fs-sm);margin-top:2px;")
                                if already:
                                    ui.label("✓ In target pipeline").style(
                                        "color:#15803D;font-size:var(--fs-xs);font-weight:600;margin-top:4px;")
                                elif not _read_only:
                                    def _add_one(v=v):
                                        if web_flow.add_to_target_queue(v, client_id):
                                            ui.notify(f"'{v['org']}' added to the Target Database pipeline.",
                                                      type="positive")
                                        _render_hot()
                                    ui.button("Add to target queue", icon="add", on_click=_add_one).props(
                                        "flat dense color=primary").style("margin-top:4px;")
                _render_hot()

            ui.label("All visitors — ranked by intent").classes("section-head").style("margin-top:14px;")
            rows = [{
                "org": v["org"],
                "cat": v["category"] + ("  · holder" if v["is_holder"] else ""),
                "reads": f"{v['pages']} pages · {v['minutes']} min" + (f" · {v['visits']} visits" if v["visits"] > 1 else ""),
                "downloads": ", ".join(v["downloads"]) if v["downloads"] else "—",
                "intent": f"{v['intent']} · {v['intent_label']}",
            } for v in d["visitors"]]
            responsive_table(
                [{"name": "org", "label": "Visitor", "field": "org", "align": "left"},
                 {"name": "cat", "label": "Category", "field": "cat", "align": "left"},
                 {"name": "reads", "label": "Read", "field": "reads", "align": "left"},
                 {"name": "downloads", "label": "Downloaded", "field": "downloads", "align": "left"},
                 {"name": "intent", "label": "Intent", "field": "intent", "align": "left"}],
                rows, table_classes="w-full dense-table", table_props="flat dense wrap-cells", primary="org")
            ui.label("Intent = downloads (weighted by asset — a model/deck pull counts most) + capped page/time "
                     "reading. Upload a fresh export any time to replace this with live numbers.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-top:6px;")

            if not _read_only:
                with ui.expansion("Replace or update this data (upload CSV)", icon="upload_file").classes(
                        "w-full").style("margin-top:12px;"):
                    _upload_control()
                    if uploaded:
                        def _clear():
                            web_flow.clear_visitor_data(client_id)
                            ui.notify("Cleared the uploaded feed.", type="info")
                            _build()
                        ui.button("Clear uploaded data", icon="delete", on_click=_clear).props(
                            "flat dense").style(f"color:{COLORS['danger']};margin-top:6px;")

    _build()


def _render_nobo_tab():
    """NOBO Ownership tab — relocated here from Market Intelligence, where it sat
    among market-data views; it belongs with the investor base. The renderer
    (CEO's-read BLUF, composition/concentration, two-pull flow, 13D/13G threshold
    watch, pipeline cross-reference, and the Broadridge NOBO-file upload) stays in
    markets_page as the single home for the nobo_engine surface; imported lazily
    to keep module load order simple."""
    from page_modules_nicegui.markets_page import _render_nobo
    _render_nobo()


def _render_curated_targets(client_id):
    """Hand-curated NDR targets — accounts you KNOW are a fit but that the 13F
    peer-holder crawl can't surface because they don't hold a comp today (a
    Geneva/Lugano private bank, a relationship carried from a prior seat). Two
    scopes: this client's own book, and a shared global book that accretes as we
    onboard more issuers. Shown in the metro view above tagged 'Curated'."""
    from core import curated_targets, peer_prospects

    ui.label("Curated targets — known accounts the crawl can't find").classes(
        "text-md font-bold").style(f"color:{COLORS['text_heading']};margin-top:6px;")
    ui.label("A peer-holder crawl only finds funds that already own a comp. Accounts you know are a fit but "
             "that don't hold a peer today — a Geneva/Lugano private bank, a relationship from a prior seat — "
             "go here by hand and appear in the metro view above, tagged 'Curated' (never 'holds a comp'). "
             "Client scope is this issuer's own book; Global is the house book, shared across every client and "
             "growing as we scale.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

    _cur_filter = {"scope": "total"}

    def _set_scope(s):
        # Clicking the active tile again clears back to the full "Total in view".
        _cur_filter["scope"] = "total" if _cur_filter["scope"] == s else s
        _panel.refresh()

    @ui.refreshable
    def _panel():
        cc = curated_targets.counts(client_id)
        _scope = _cur_filter["scope"]
        # Same clickable-filter tiles as the Target Database summary (_hub_metric): click one to
        # filter the list below by scope; the active tile highlights with the accent top edge.
        with ui.row().classes("w-full gap-3").style("margin-top:6px;"):
            _hub_metric("This client", cc["client"], "This issuer's own book",
                        _scope == "client", lambda: _set_scope("client"))
            _hub_metric("Global (house book)", cc["global"], "House book — all clients",
                        _scope == "global", lambda: _set_scope("global"))
            _hub_metric("Total in view", cc["total"], "Client + global",
                        _scope == "total", lambda: _set_scope("total"))

        # ── Add form ──────────────────────────────────────────────────────────
        with ui.card().classes("w-full").style(
                f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};margin-top:6px;"):
            ui.label("Add a curated target").classes("font-bold").style(
                f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
            with ui.row().classes("w-full gap-2 items-end no-wrap").style("flex-wrap:wrap;"):
                f_name = ui.input("Firm name").props("dense outlined").style("min-width:220px;flex:2;")
                f_city = ui.input("City").props("dense outlined").style("min-width:120px;flex:1;")
                f_state = ui.input("State / country", placeholder="NY · Switzerland").props(
                    "dense outlined").style("min-width:130px;flex:1;")
            f_why = ui.input("Why (rationale)").props("dense outlined").classes("w-full").style("margin-top:4px;")
            with ui.row().classes("items-center gap-3").style("margin-top:4px;"):
                f_scope = ui.toggle({"client": "This client", "global": "Global (house book)"},
                                    value="client").props("dense no-caps")
                # Live metro preview, so you see where it will cluster before saving.
                _preview = ui.label("").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

                def _refresh_preview():
                    c, s = (f_city.value or "").strip(), (f_state.value or "").strip()
                    _preview.set_text(f"→ clusters into: {_metro_from_city(c, s)}" if c else "")
                f_city.on("blur", lambda *_: _refresh_preview())
                f_state.on("blur", lambda *_: _refresh_preview())

                def _add():
                    if not (f_name.value or "").strip():
                        ui.notify("Firm name is required.", type="warning")
                        return
                    curated_targets.add(f_name.value, f_city.value, f_state.value,
                                        f_why.value, scope=f_scope.value, cid=client_id)
                    ui.notify(f"Added {f_name.value.strip()} to the "
                              f"{'global house book' if f_scope.value == 'global' else 'client'} list.",
                              type="positive")
                    _panel.refresh()
                ui.button("Add target", icon="add", on_click=_add).props("dense color=primary")

        # ── Current list (filtered by the active scope tile above) ──────────────
        entries = curated_targets.merged(client_id)
        if _scope == "client":
            entries = [e for e in entries if e.get("scope") == "client"]
        elif _scope == "global":
            entries = [e for e in entries if e.get("scope") != "client"]
        if not entries:
            ui.label("No curated targets in this view — add one above, or click Total in view."
                     if _scope != "total" else "No curated targets yet — add one above.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);margin-top:6px;")
            return
        for r in sorted(entries, key=lambda x: (x.get("scope") != "client", x.get("filer", "").lower())):
            is_client = r.get("scope") == "client"
            badge_clr = COLORS["accent"] if is_client else "#6D28D9"
            metro = _metro_from_city(r.get("city"), r.get("state"))
            with ui.card().classes("w-full").style(
                    f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};"
                    f"border-left:4px solid {badge_clr};margin-top:4px;"):
                with ui.column().classes("gap-0").style("min-width:0;"):
                    with ui.row().classes("items-center gap-2"):
                        ui.label(r.get("filer")).classes("font-bold").style(
                            f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
                        ui.label("This client" if is_client else "Global").style(
                            f"background:{'rgba(30,64,175,.10)' if is_client else 'rgba(109,40,217,.10)'};"
                            f"color:{badge_clr};border-radius:6px;padding:1px 7px;font-size:var(--fs-2xs);font-weight:700;")
                        if r.get("seed"):
                            ui.label("default").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);border:1px solid {COLORS['border']};"
                                "border-radius:6px;padding:0 6px;")
                    loc = _loc_line(r)
                    ui.label(f"{loc}  ·  metro: {metro}").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                    if r.get("rationale"):
                        ui.label(r["rationale"]).style(f"color:{COLORS['text_secondary']};font-size:var(--fs-xs);")
                # Actions in a bottom row set off by a thin divider — consistent with the
                # peer-prospect, institution, and Today dashboard cards.
                with ui.row().classes("w-full gap-1 items-center").style(
                        f"margin-top:4px;padding-top:4px;border-top:1px solid {COLORS['border']};"):
                    ui.button("Add to pipeline", on_click=lambda r=r: (
                        peer_prospects.promote(r, cid=client_id),
                        ui.notify(f"Added {r['filer']} to the pipeline (Target Database).", type="positive"),
                    )).props("dense size=sm color=primary")
                    if not r.get("seed"):
                        ui.button(icon="delete", on_click=lambda r=r: (
                            curated_targets.remove(r["key"], scope=r.get("scope", "client"), cid=client_id),
                            _panel.refresh(),
                        )).props("flat dense size=sm color=negative")

    _panel()


def _render_peer_prospects_tab(client_id):
    """Reviewed, conviction-ranked peer-overlap candidates (core.peer_prospects).
    Funds that hold a close comp but not the client, filtered for noise and false
    positives, each with its evidence and a promote/dismiss gate — nothing enters
    the pipeline unvetted."""
    from core import peer_prospects

    # Ticker and comp names come from THIS client. They were hardcoded to USIO's, so
    # every other tenant read "…but not USIO (Repay / Cass / CSG…)" on their own page.
    _tk = CT("ticker")
    _tight = sorted(peer_prospects.tight_comps(client_id))
    _comps = " / ".join(_tight[:5]) if _tight else "its closest comps"
    ui.label("Peer Prospects — comp overlap, to qualify").classes("text-lg font-bold")
    ui.label(f"Funds that hold a close comp ({_comps}) but not {_tk}, "
             "ranked by conviction — position weight in their own book, focus on the tightest comps, active "
             "vs. index breadth, and size fit. Anyone already yours (13F / NOBO / tracked / 13D-G), plus "
             "passive/index and quasi-index books, is filtered out. Nothing hits the pipeline until you promote it.").style(
        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
    _size_calibration_note(client_id)

    # The roadshow-metro view (with click-to-select) is now the single "Where they are" table in
    # Buy-Side Intelligence — no duplicate here. The conviction-ranked prospects below feed the
    # same NDR pipeline.
    ui.label("📍 Roadshow metro view — with click-a-metro → select → Add to NDR — now lives in "
             "Buyside Ownership → “Where they are”. The conviction-ranked prospects below feed the "
             "same pipeline.").style(
        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);background:{COLORS['surface_hover_bg']};"
        "border-radius:8px;padding:8px 10px;")
    ui.markdown("---")
    _render_curated_targets(client_id)
    ui.markdown("---")

    # Sort toggle — conviction (the smart default) vs raw 13F position size (what
    # a plain 13F screen shows) vs concentration. Lives outside the refreshable so
    # flipping it just re-ranks the list.
    _state = {"sort": "conviction", "show_all": False}
    with ui.row().classes("items-center gap-2").style("margin-top:6px;"):
        ui.label("Rank by").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        _sort_toggle = ui.toggle(
            {"conviction": "Conviction", "size": "Position size", "concentration": "Concentration"},
            value="conviction").props("dense no-caps")

    @ui.refreshable
    def _list():
        # Show ALL qualified peer-owners by default is heavy (222 cards), so start at 40 with a
        # one-click "show all" — but never a silent cap: the count card says how many exist and the
        # control lists every one (same principle as the RIA/diversified buckets below).
        _lim = None if _state.get("show_all") else 40
        cands = peer_prospects.build_candidates(client_id, limit=_lim, sort=_state["sort"])
        c = peer_prospects.counts(client_id)
        with ui.row().classes("w-full gap-3").style("margin-top:6px;"):
            for lbl, val, clr in [("To qualify", c["candidates"], COLORS["accent"]),
                                  ("RIA / wealth", c.get("rias", 0), "#B45309"),
                                  ("Diversified", c.get("diversified", 0), "#6D28D9"),
                                  ("Promoted", c["promoted"], "#15803D"),
                                  ("Dismissed", c["dismissed"], COLORS["text_muted"])]:
                with ui.card().classes("flex-1").style(
                        f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};min-width:110px;"):
                    ui.label(str(val)).classes("font-bold").style(f"color:{clr};font-size:var(--fs-2xl);")
                    ui.label(lbl).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

        if not cands:
            ui.label("No candidates to qualify. Run the SEC 13F refresh from Settings → Data Sources if you "
                     "haven't yet, or you've reviewed them all.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);margin-top:8px;")
            return

        from core import contacts as _cmod
        _rcounts = _cmod.firm_roster_counts()   # firm_cik -> named investment contacts we hold
        for r in cands:
            promoted = r.get("promoted")
            border = "#15803D" if promoted else COLORS["border"]
            with ui.card().classes("w-full").style(
                    f"background:{COLORS['surface_bg']};border:1px solid {border};"
                    f"border-left:4px solid {'#15803D' if promoted else COLORS['accent']};margin-top:4px;"):
                with ui.row().classes("w-full items-start justify-between no-wrap"):
                    with ui.column().classes("gap-0").style("flex:1;min-width:0;"):
                        _nm = pretty_name(r["filer"]) + ("  ✓ promoted" if promoted else "")
                        ui.label(_nm).classes("font-bold").style(f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
                        loc = _loc_line(r)
                        conc = f"{r['concentration']*100:.1f}% of book" if r.get("concentration") is not None else "book n/a"
                        bp = f"{r['book_positions']} positions" if r.get("book_positions") else "breadth n/a"
                        pv = f"${r['peer_value']/1e6:.1f}M position" if r.get("peer_value") else ""
                        ui.label(f"{loc} · {conc} in payment comps{(' · ' + pv) if pv else ''} · {bp} · "
                                 f"filed {(r.get('file_date') or '—')}").style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        with ui.row().classes("gap-1 flex-wrap").style("margin-top:3px;"):
                            for ct, cv in sorted(r["comps"].items(), key=lambda kv: -kv[1]["value"]):
                                tight = cv["tight"]
                                ui.label(ct + (" ◆" if tight else "")).style(
                                    f"background:{'rgba(30,64,175,.10)' if tight else COLORS['surface_hover_bg']};"
                                    f"color:{COLORS['accent_strong'] if tight else COLORS['text_secondary']};"
                                    "border-radius:6px;padding:1px 7px;font-size:var(--fs-xs);font-weight:600;")
                    with ui.column().classes("gap-0 items-end").style("flex-shrink:0;"):
                        ui.label(f"{r['conviction']:.0f}").classes("font-bold").style(
                            f"color:{COLORS['accent']};font-size:var(--fs-2xl);line-height:1;")
                        ui.label("conviction").style(f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);")
                # Actions in a bottom row set off by a thin divider — same treatment as the
                # Today dashboard and institution cards.
                with ui.row().classes("w-full gap-1 items-center").style(
                        f"margin-top:4px;padding-top:4px;border-top:1px solid {COLORS['border']};"):
                    if promoted:
                        ui.button("Undo", on_click=lambda r=r: (peer_prospects.reset(r["key"]), _list.refresh())).props(
                            "flat dense size=sm")
                    else:
                        def _promote(r=r):
                            peer_prospects.promote(r)
                            ui.notify(f"Promoted {r['filer']} to the pipeline (Target Database).", type="positive")
                            _list.refresh()

                        def _dismiss(r=r):
                            peer_prospects.dismiss(r["key"])
                            _list.refresh()

                        ui.button("Promote", on_click=_promote).props("dense size=sm color=primary")
                        ui.button("Dismiss", on_click=_dismiss).props("flat dense size=sm")
                    _tn = _rcounts.get(_cmod.norm_cik(r.get("cik"))) if r.get("cik") else None
                    if _tn:
                        ui.button(f"Team ({_tn})", icon="groups",
                                  on_click=lambda r=r: _open_account_profile(r)) \
                            .props("flat dense size=sm").style(f"color:{COLORS['accent']};font-weight:600;")

        # Show-all control — every qualified peer-owner is listable, not just the top 40.
        _total = c["candidates"]
        if _total > len(cands) and not _state.get("show_all"):
            def _showall():
                _state["show_all"] = True
                _list.refresh()
            ui.button(f"Show all {_total} funds that own a peer but not {CT('ticker')}",
                      icon="expand_more", on_click=_showall).props("flat dense").style("margin-top:8px;")
        elif _state.get("show_all"):
            ui.label(f"Showing all {len(cands)} qualified peer-owners (funds holding a comp, not USIO).").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-top:8px;")

            def _showtop():
                _state["show_all"] = False
                _list.refresh()
            ui.button("Show top 40 only", icon="expand_less", on_click=_showtop).props("flat dense")

        # ── RIA / wealth bucket ─────────────────────────────────────────────
        # These genuinely own the comps but have no PM to pitch, so they're a
        # video-call tier rather than an NDR target. Previously discarded
        # outright; now surfaced separately, ranked by position size.
        # No limit: the header count must match the "RIA / wealth" card above, and
        # a capped list would silently hide names the card says exist.
        _rias = peer_prospects.build_candidates(client_id, limit=None, kind="ria")
        if _rias:
            with ui.expansion(f"RIA / wealth managers holding your comps ({len(_rias)})",
                              icon="account_balance_wallet", value=False).classes("w-full").style(
                    f"border:1px solid {COLORS['border']};border-radius:8px;margin-top:12px;"):
                ui.label("Real positions in your peer set, but held through client accounts — no portfolio "
                         "manager to pitch. Video-call tier: an assistant can set these up, no management "
                         "travel. Ranked by position size; a large one can still be worth a call.").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                for r in _rias:
                    with _list_item_card():
                      with ui.row().classes("w-full items-center justify-between no-wrap"):
                        with ui.column().classes("gap-0").style("flex:1;min-width:0;"):
                            ui.label(pretty_name(r["filer"])).classes("name-link").on(
                                "click", lambda r=r: _open_account_profile(r)).style(
                                f"color:{COLORS['text_heading']};font-size:var(--fs-lg);font-weight:700;")
                            loc = _loc_line(r)
                            comps = ", ".join(sorted(r["comps"].keys()))
                            ui.label(f"{loc} · ${r['peer_value']/1e6:.1f}M across {comps}").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                        with ui.row().classes("gap-1").style("flex-shrink:0;"):
                            if r.get("promoted"):
                                ui.button("Undo", on_click=lambda r=r: (
                                    peer_prospects.reset(r["key"]), _list.refresh())).props("flat dense size=sm")
                            else:
                                def _promote_ria(r=r):
                                    peer_prospects.promote(r)
                                    ui.notify(f"Promoted {r['filer']} — video-call tier.", type="positive")
                                    _list.refresh()

                                def _dismiss_ria(r=r):
                                    peer_prospects.dismiss(r["key"])
                                    _list.refresh()

                                ui.button("Promote", on_click=_promote_ria).props("flat dense size=sm color=primary")
                                ui.button("Dismiss", on_click=_dismiss_ria).props("flat dense size=sm")

        # ── Large diversified / index-family managers ───────────────────────
        # These were DROPPED outright on the assumption that a big book "never takes a micro-cap
        # NDR". That's a bad assumption: they may not chase a meeting, but if you're already in
        # their city and they own you or a comp they'll usually take one — and their analysts often
        # WANT to meet a competitor they DON'T own, for candid industry colour. So they're a review
        # bucket now, ranked by how much of the peer set they actually own.
        _div = peer_prospects.build_candidates(client_id, limit=None, kind="diversified")
        if _div:
            with ui.expansion(f"Large diversified & index-family managers ({len(_div)})",
                              icon="corporate_fare", value=False).classes("w-full").style(
                    f"border:1px solid {COLORS['border']};border-radius:8px;margin-top:12px;"):
                ui.label("Index families, bank asset-management arms, and wide-book active managers "
                         "(Capital Group, Fidelity, Clearbridge…). They run real fundamental strategies. "
                         "Worth a meeting when you're already in their city, when they own a comp, or "
                         "when they want industry colour on a name they don't own. Ranked by peer position.").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                for r in _div:
                    with _list_item_card():
                      with ui.row().classes("w-full items-center justify-between no-wrap"):
                        with ui.column().classes("gap-0").style("flex:1;min-width:0;"):
                            ui.label(pretty_name(r["filer"])).classes("name-link").on(
                                "click", lambda r=r: _open_account_profile(r)).style(
                                f"color:{COLORS['text_heading']};font-size:var(--fs-lg);font-weight:700;")
                            loc = _loc_line(r)
                            comps = ", ".join(sorted(r["comps"].keys()))
                            why = "wide book" if r.get("broad_book") else "index / bank AM"
                            ui.label(f"{loc} · ${r['peer_value']/1e6:.1f}M across {comps} · {why}").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                        with ui.row().classes("gap-1").style("flex-shrink:0;"):
                            if r.get("promoted"):
                                ui.button("Undo", on_click=lambda r=r: (
                                    peer_prospects.reset(r["key"]), _list.refresh())).props("flat dense size=sm")
                            else:
                                def _promote_div(r=r):
                                    peer_prospects.promote(r)
                                    ui.notify(f"Promoted {r['filer']} for review.", type="positive")
                                    _list.refresh()

                                def _dismiss_div(r=r):
                                    peer_prospects.dismiss(r["key"])
                                    _list.refresh()

                                ui.button("Promote", on_click=_promote_div).props("flat dense size=sm color=primary")
                                ui.button("Dismiss", on_click=_dismiss_div).props("flat dense size=sm")

        # ── Market makers / HFT / ETF mechanics ─────────────────────────────
        # Genuinely no fundamental PM to meet — they hold the comps as inventory or index
        # mechanics. Shown for completeness/transparency (so a holder isn't silently missing from
        # the count), but read-only: there's no promote, because there's nobody to put on an NDR.
        _mm = peer_prospects.build_candidates(client_id, limit=None, kind="market_maker")
        if _mm:
            with ui.expansion(f"Market makers / HFT / ETF mechanics ({len(_mm)}) — no PM to meet",
                              icon="bolt", value=False).classes("w-full").style(
                    f"border:1px solid {COLORS['border']};border-radius:8px;margin-top:12px;"):
                ui.label("These hold your comps as trading inventory or index mechanics, not a "
                         "fundamental view — there is no portfolio manager to pitch. Listed for a "
                         "complete picture of who owns the peer set; not an NDR queue.").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                for r in sorted(_mm, key=lambda x: -(x.get("peer_value") or 0)):
                    loc = _loc_line(r)
                    comps = ", ".join(sorted(r["comps"].keys()))
                    with _list_item_card():
                        ui.label(f"{pretty_name(r['filer'])} · {loc} · ${(r.get('peer_value') or 0)/1e6:.1f}M across {comps}").classes("name-link").on(
                            "click", lambda r=r: _open_account_profile(r)).style(
                            f"color:{COLORS['text_body']};font-size:var(--fs-sm);")

    _sort_toggle.on_value_change(lambda e: (_state.update(sort=e.value), _list.refresh()))
    _list()


def _render_mode_toggle_control(mode, on_change_cb):
    """Just the Pre/Post toggle buttons. Stays put when the mode changes (it's
    not inside the refreshed region), so it keeps focus and doesn't flicker;
    the change is applied in place by on_change_cb rather than a page reload."""
    with ui.row().classes("w-full items-center gap-4"):
        toggle = ui.toggle({"pre": "Pre-Earnings — Engagement Mode", "post": "Post-Earnings — Prospecting Mode"}, value=mode)
        toggle.on_value_change(lambda: on_change_cb(toggle.value))


def _render_mode_description(mode):
    # Was: solid fixed-dark background (from the old dark theme) with the
    # descriptive text left uncolored — it inherited the page's default text
    # color, which under the active CREAM (light) theme resolves dark/near-
    # black, making it nearly invisible against the leftover dark green/navy
    # fill (the washed-out text in the screenshot). Same fix as _tier_card /
    # _render_repeat_alert_banner: light card, colored left border, explicit
    # theme-token text color throughout.
    if mode == "pre":
        accent = COLORS["accent"]
        ui.html(
            f"<div style='background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};border-left:4px solid {accent};"
            f"border-radius:8px;padding:8px 14px;font-size:var(--fs-sm);color:{COLORS['text_body']};'>"
            f"<b style='color:{accent};'>Pre-Earnings Mode</b> · Scoring weights: Call listener 40% · Peer holder 35% · "
            f"IR visits 25% · Goal: who already knows the story → send call invitation → close the loop</div>"
        )
    else:
        accent = COLORS["positive"]
        ui.html(
            f"<div style='background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};border-left:4px solid {accent};"
            f"border-radius:8px;padding:8px 14px;font-size:var(--fs-sm);color:{COLORS['text_body']};'>"
            f"<b style='color:{accent};'>Post-Earnings Prospecting Mode</b> · Scoring weights: Q2 call listener 40% · "
            f"Catalyst fit 35% · New IR activity 25% · Goal: who heard the beat → fits the upgrade thesis → call in first 5 days</div>"
        )


# ─────────────────────────────────────────────────────────────────────────
# Big Picture synthesis
# ─────────────────────────────────────────────────────────────────────────
def _header_tooltips(table, tips):
    """Hover-explain column HEADINGS: adds a per-cell header slot (leaves any selection
    checkbox column intact, unlike a full header slot) with a q-tooltip and a help
    cursor. `tips` maps column name -> tooltip text. Reusable across any ui.table."""
    for name, text in tips.items():
        safe = (text or "").replace("<", "&lt;").replace(">", "&gt;")
        table.add_slot(f"header-cell-{name}", (
            '<q-th :props="props" class="cursor-help">'
            '{{ props.col.label }}'
            '<q-tooltip>' + safe + '</q-tooltip>'
            '</q-th>'))
    return table


def _card_rec(inst):
    """Adapt a Buy-Side institution dict into the shape the 360 profile reads (its comps
    come from Peer_Holdings; carry city/cik through)."""
    peers = inst.get("Peer_Holdings") or []
    if isinstance(peers, str):
        peers = [peers]
    return {**inst, "comps": {t: {} for t in peers if isinstance(t, str)},
            "city": inst.get("City"), "cik": inst.get("cik")}


def _quick_log_call(inst):
    """One-click 'I just talked to them' from a card — appends a call event (always safe)."""
    from core import account_api
    account_api.log_interaction(name=inst["Fund"], cik=inst.get("cik"), type="call",
                                summary="Logged from Buy-Side card")
    ui.notify(f"Logged a call for {pretty_name(inst['Fund'])}.", type="positive")


def _open_person_notes(person_name, firm):
    """Per-person notes timeline for one roster member — this client's notes only (contact_notes
    defaults to the active client). Keyed by (name, firm) with no CIK, the same convention as
    add_contact_note, so what's written here and elsewhere lines up on one timeline."""
    from core import contacts as _cmod
    contact_id = _cmod.contact_id_for(None, person_name, firm)
    with ui.dialog() as d, ui.card().style("min-width:min(94vw,480px);max-width:560px;"):
        ui.label(f"Notes — {pretty_name(person_name)}").classes("text-lg font-bold").style(
            f"color:{COLORS['text_heading']};")
        ui.label(firm).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        ui.separator().style("margin:6px 0;")
        _box = ui.column().classes("w-full").style("gap:5px;")

        def _paint():
            _box.clear()
            with _box:
                items = _cmod.contact_notes(contact_id)
                if not items:
                    ui.label("No notes yet — add the first one below.").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                for it in items:
                    with ui.element("div").style(
                            f"border-left:3px solid {COLORS['accent']};padding:1px 0 1px 9px;"):
                        ui.label(" · ".join(x for x in (it.get("ts"), it.get("source"), it.get("by")) if x)).style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);")
                        ui.label(it.get("note", "")).style(
                            f"color:{COLORS['text_body']};font-size:var(--fs-sm);white-space:pre-wrap;")

        _in = ui.textarea("Add a note").props("autogrow dense outlined").classes("w-full").style("margin-top:4px;")

        def _add():
            txt = (_in.value or "").strip()
            if not txt:
                ui.notify("Nothing to save yet.", type="warning"); return
            _cmod.add_contact_note(person_name, firm, txt, source="Account 360",
                                   by=(CI().get("name") or "IR Team"))
            _in.value = ""
            _paint()
            ui.notify(f"Note added to {pretty_name(person_name)}.", type="positive")
        with ui.row().classes("w-full justify-end gap-2"):
            ui.button("Close", on_click=d.close).props("flat")
            ui.button("Add note", icon="add", on_click=_add).props("color=primary")
        _paint()
    d.open()


def _open_account_profile(rec):
    """Account 360 — one card fusing the AUTO-derived book (holdings, peers, fund
    lineup, NDR pipeline) with the HUMAN relationship layer (quality, last contact,
    notes). `rec` is the original candidate/institution dict from wherever it was
    clicked. Prototype: read-rich, with an editable relationship strip that persists
    to the global house book (core.relationship_notes)."""
    from core import fund_lineup, account_api

    name = rec.get("filer") or rec.get("Fund") or "—"
    disp = pretty_name(name)
    is_cand = "filer" in rec
    comps = rec.get("comps") or {}
    peers = ", ".join(sorted(comps.keys()))
    metro = rec.get("Metro") or _metro_from_city(rec.get("city"), rec.get("state")) if is_cand else rec.get("Metro")
    loc = (rec.get("city") or rec.get("City") or metro or "—")
    category = rec.get("tier") or ("Current holder" if rec.get("USIO_Holder") else "Non-holder")
    score = rec.get("conviction") if is_cand else rec.get("Engagement_Score")

    def _section(title):
        ui.label(title).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-weight:600;"
                              "letter-spacing:.04em;text-transform:uppercase;margin-top:10px;")

    def _kv(label, value):
        with ui.row().classes("w-full items-baseline gap-2").style("padding:1px 0;"):
            ui.label(label).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);min-width:130px;")
            ui.label(str(value)).style(f"color:{COLORS['text_secondary']};font-size:var(--fs-sm);")

    acct_cik = rec.get("cik")
    acct = account_api.get_account(name=name, cik=acct_cik)   # one read: identity + both layers
    note = acct["relationship"]

    with ui.dialog() as dlg, ui.card().style("min-width:min(680px,95vw);max-width:95vw;"):
        with ui.row().classes("w-full justify-between items-start"):
            with ui.column().classes("gap-0"):
                ui.label(disp).classes("text-xl font-bold").style(f"color:{COLORS['text_heading']};")
                ui.label(f"{category} · {loc}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            with ui.row().classes("items-center gap-2"):
                q = note.get("quality")
                if q:
                    ui.label(account_api.QUALITY.get(q, q)).style(
                        f"background:{COLORS['positive'] if q in ('good','responsive') else COLORS['surface_hover_bg']};"
                        f"color:{'#fff' if q in ('good','responsive') else COLORS['text_secondary']};"
                        "font-size:var(--fs-xs);font-weight:600;padding:2px 8px;border-radius:6px;")
                ui.button(icon="close", on_click=dlg.close).props("flat round dense")

        # ── Ownership & filings (auto) ─────────────────────────────
        _section("Ownership & filings")
        if is_cand:
            _kv("Owns (your comps)", peers or "—")
            _kv("Peer position value", f"${(rec.get('peer_value') or 0):,.0f}")
            _kv("Conviction", round(score) if isinstance(score, (int, float)) else "—")
            _kv("Book breadth", f"{rec.get('book_positions') or '—'} positions")
        else:
            _kv("Relationship to you", "Holder" if rec.get("USIO_Holder") else "Non-holder")
            if rec.get("Position_Value") is not None:
                _kv("Position value", f"${(rec.get('Position_Value') or 0):,.0f}")
            _kv("Engagement score", round(score) if isinstance(score, (int, float)) else "—")
            if peers:
                _kv("Owns (your comps)", peers)
            if rec.get("Action"):
                _kv("Signal", rec.get("Action"))

        # ── Addressable size (auto) — ONLY for a peer-owning prospect (a holder already
        # owns you, so it's meaningless there). The prize: if they took a position the size
        # of the capital they already hold in your comps, how big a USIO holder would that
        # make them (% of the company). Feasibility (days of ADV to build) is shown only
        # when the stock is too thin to accumulate quickly. Display-only, not a prediction.
        _hold360 = (not is_cand) and rec.get("USIO_Holder")
        _addr_val = rec.get("peer_value") or (None if _hold360 else rec.get("Position_Value"))
        if not _hold360 and _addr_val:
            from config.client_config import CF as _CF
            _shares = (_CF().get("shares_out_m") or 0) * 1e6
            _ar = market_data.addressable_readout(
                _addr_val, market_data.get_snapshot(CT("ticker"), refresh_if_stale=False),
                shares_out=_shares or None)
            if _ar:
                _section("Addressable size")
                _pct = f" ≈ {_ar['pct_of_company']*100:.1f}% of USIO" if _ar.get("pct_of_company") else ""
                _kv("Comp capital", f"${_ar['value']/1e6:,.2f}M in your comps{_pct}")
                _dd = _ar.get("days_to_build")
                if _dd and _dd >= 5:
                    _kv("Feasibility", f"~{_dd:,.0f} days of USIO ADV to build (thin — accumulate patiently)")

        # ── Fund lineup (auto) ─────────────────────────────────────
        lu = fund_lineup.lineup_for_manager(name)
        if lu and lu.get("funds"):
            _section(f"Fund lineup — {lu.get('registrant')}")
            names = [f["name"] for f in lu["funds"]]
            ui.label(" · ".join(names[:8]) + (f"  (+{len(names) - 8} more)" if len(names) > 8 else "")).style(
                f"color:{COLORS['text_secondary']};font-size:var(--fs-sm);")

        # ── Investment team — the named people to actually call ────
        # THE value-add: a peer-owner ("Firm X holds your peers") resolves to the real PMs/
        # analysts we hold on the house roster, decision-makers first, each with a one-click
        # pre-filled outreach draft. Turns "who owns this" into "who do I email".
        from core import contacts as _cmod
        _team = _cmod.roster_for_firm(cik=acct_cik, firm=name)
        if _team:
            _rlbl = {"buy_side_pm": "PM", "buy_side_sector_pm": "Sector PM", "cio": "CIO",
                     "director_of_research": "Dir. Research", "associate_pm": "Assoc. PM",
                     "bs_analyst": "Analyst", "bs_sector_analyst": "Sector Analyst",
                     "bs_generalist": "Generalist", "principal_csuite": "Principal", "trader": "Trader"}
            _rtkr, _rcname, _rir = CT("ticker"), CT("name"), CI()

            def _draft_to(to, fname):
                subj = f"{_rtkr} — meeting request"
                body = (f"Hi {fname or 'there'},\n\nWe'd value 30 minutes with your team. {disp} holds "
                        f"{peers or 'names in our space'}, so {_rtkr}'s story should be directly relevant.\n\n"
                        f"Would a short meeting work?\n\n{_rir.get('name', '')}\n"
                        f"{_rir.get('title', 'Investor Relations')} · {_rcname} (NASDAQ: {_rtkr})")
                ui.run_javascript("window.location.href = " + json.dumps(
                    "mailto:" + to + "?subject=" + quote(subj) + "&body=" + quote(body)))

            _withemail = sum(1 for p in _team if p.get("email"))
            _section(f"Investment team — {len(_team)} on file · {_withemail} reachable")
            ui.label("The named PMs / analysts we hold at this firm — decision-makers first. Click Notes "
                     "for a per-person timeline.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
            _pnotes = db.load_json("contact_notes.json", default={}) or {}   # one read for the counts
            with ui.column().classes("w-full gap-0").style("max-height:236px;overflow-y:auto;margin-top:2px;"):
                for p in _team:
                    with ui.row().classes("w-full items-center gap-2").style(
                            f"padding:3px 0;border-bottom:1px solid {COLORS['border']};"):
                        ui.label(pretty_name(p["name"])).style(
                            f"color:{COLORS['text_heading']};font-size:var(--fs-sm);font-weight:600;min-width:150px;")
                        ui.label(_rlbl.get(p.get("primary_role"), "—")).style(
                            f"color:{COLORS['accent_light']};font-size:var(--fs-xs);font-weight:600;min-width:82px;")
                        ui.label((p.get("title") or "")[:38]).style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);flex:1;min-width:0;")
                        if p.get("phone"):
                            ui.label(p["phone"]).style(f"color:{COLORS['text_secondary']};font-size:var(--fs-xs);")
                        if p.get("email"):
                            ui.button(icon="mail", on_click=lambda _=None, e=p["email"],
                                      f=pretty_name(p["name"]).split()[0]: _draft_to(e, f)) \
                                .props("flat dense round size=sm").tooltip(f"Draft to {p['email']}")
                        else:
                            ui.label("no email").style(f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);")
                        _pcnt = len(_pnotes.get(_cmod.contact_id_for(None, p["name"], name), []))
                        ui.button(f"Notes{f' ({_pcnt})' if _pcnt else ''}", icon="sticky_note_2",
                                  on_click=lambda p=p: _open_person_notes(p["name"], name)) \
                            .props("flat dense size=sm").tooltip("Per-person notes timeline")

        # ── Contact & outreach ─────────────────────────────────────
        # The metro → fund flow ended at "who is this" with no way to reach them. This closes
        # it: the contact on file (if any), an editable email, and a one-click pre-filled draft
        # the user SENDS themselves (never auto-sent).
        _section("Contact & outreach")
        _contact = get_institution_contacts().get(name, {}) if name else {}
        if _contact.get("name") or _contact.get("title"):
            _kv("Contact", " · ".join(x for x in (_contact.get("name"), _contact.get("title")) if x))
        _tkr, _cname, _ir = CT("ticker"), CT("name"), CI()
        with ui.row().classes("w-full items-end gap-2"):
            _c_email = ui.input("Contact email", value=_contact.get("email", "")).props(
                "dense outlined clearable").style("min-width:280px;flex:1;")

            def _draft_outreach():
                to = (_c_email.value or "").strip()
                if not to:
                    ui.notify("Add a contact email to draft to.", type="warning"); return
                _p = peers or "names in our space"
                subj = f"{_tkr} — meeting request"
                body = (f"Hi {(_contact.get('name') or '').split()[0] or 'there'},\n\n"
                        f"We'd value 30 minutes with your team. {disp} holds {_p}, so {_tkr}'s story "
                        f"should be directly relevant.\n\nWould a short meeting work?\n\n"
                        f"{_ir.get('name', '')}\n{_ir.get('title', 'Investor Relations')} · {_cname} (NASDAQ: {_tkr})")
                href = "mailto:" + to + "?subject=" + quote(subj) + "&body=" + quote(body)
                ui.run_javascript(f"window.location.href = {json.dumps(href)}")
            ui.button("Draft outreach email", icon="mail", on_click=_draft_outreach).props("dense color=primary")
        if not _contact.get("email"):
            ui.label("No contact on file — enter an email above to draft, or reach out via your own channel.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

        # ── Contact notes (the per-person timeline, client-scoped) ─────────────────
        # The same notes surfaced on House Contacts, here on the client's own account view — this
        # client's notes only (contact_notes defaults to the active client, so one client never sees
        # another's notes on a shared person). Keyed to the primary contact on file.
        if _contact.get("name"):
            _section(f"Contact notes — {pretty_name(_contact['name'])}")
            _note_cid = _cmod.contact_id_for(None, _contact["name"], name)
            _notes_box = ui.column().classes("w-full").style("gap:5px;margin-top:2px;")

            def _paint_contact_notes():
                _notes_box.clear()
                with _notes_box:
                    items = _cmod.contact_notes(_note_cid)
                    if not items:
                        ui.label("No notes yet — add the first one below.").style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    for it in items:
                        with ui.element("div").style(
                                f"border-left:3px solid {COLORS['accent']};padding:1px 0 1px 9px;"):
                            ui.label(" · ".join(x for x in (it.get("ts"), it.get("source"), it.get("by")) if x)).style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);")
                            ui.label(it.get("note", "")).style(
                                f"color:{COLORS['text_body']};font-size:var(--fs-sm);white-space:pre-wrap;")

            _cn_in = ui.textarea("Add a note").props("autogrow dense outlined").classes("w-full").style("margin-top:4px;")

            def _add_contact_note():
                txt = (_cn_in.value or "").strip()
                if not txt:
                    ui.notify("Nothing to save yet.", type="warning"); return
                _cmod.add_contact_note(_contact["name"], name, txt, source="Account 360",
                                       by=(CI().get("name") or "IR Team"),
                                       email=(_contact.get("email") or None))
                _cn_in.value = ""
                _paint_contact_notes()
                ui.notify(f"Note added to {pretty_name(_contact['name'])}.", type="positive")
            ui.button("Add note", icon="add", on_click=_add_contact_note).props("dense color=primary").style("margin-top:4px;")
            _paint_contact_notes()

        # ── Interactions (derived — touches & last contact are computed, not typed) ──
        _section("Interactions")

        @ui.refreshable
        def _interactions():
            summ = account_api.get_account(name=name, cik=acct_cik)["interactions"]
            with ui.row().classes("items-baseline gap-4"):
                ui.label(f"{summ['touches']} touch{'es' if summ['touches'] != 1 else ''}").style(
                    f"color:{COLORS['text_heading']};font-weight:700;font-size:var(--fs-base);")
                ui.label(f"Last contact: {summ['last_contact'] or '—'}").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            for e in summ["events"][:6]:
                src = " · NDR" if e.get("source") == "ndr" else ""
                with ui.row().classes("w-full items-baseline gap-2").style("padding:1px 0;"):
                    ui.label(e.get("date") or "—").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);min-width:88px;")
                    ui.label(f"{account_api.TYPES.get(e.get('type'), e.get('type'))} — "
                             f"{e.get('summary') or ''}{src}").style(
                        f"color:{COLORS['text_secondary']};font-size:var(--fs-sm);")
            if not summ["events"]:
                ui.label("No interactions logged yet.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        _interactions()
        # log an event → appends (never overwrites); touches/last-contact recompute
        with ui.row().classes("w-full items-end gap-2").style("margin-top:6px;"):
            it_type = ui.select({k: v for k, v in account_api.TYPES.items()}, value="meeting",
                                label="Type").props("dense outlined").style("min-width:120px;")
            it_date = ui.input("Date", value=datetime.now().strftime("%Y-%m-%d")).props("dense outlined").style("min-width:130px;")
            it_sum = ui.input("What happened").props("dense outlined").style("flex:1;min-width:150px;")

            def _log_it():
                account_api.log_interaction(name=name, cik=acct_cik, type=it_type.value,
                                            date=(it_date.value or "").strip() or None,
                                            summary=(it_sum.value or "").strip() or None)
                it_sum.value = ""
                ui.notify(f"Logged {account_api.TYPES.get(it_type.value, 'event').lower()} for {disp}.",
                          type="positive")
                _interactions.refresh()
            ui.button("Log", icon="add", on_click=_log_it).props("dense color=primary")

        # ── Relationship (human opinion — quality + free note; global to the firm) ──
        _section("Relationship — your notes")
        q_opts = {"": "— quality —", **account_api.QUALITY}
        q_sel = ui.select(q_opts, value=note.get("quality") or "", label="Quality").props(
            "dense outlined").style("min-width:200px;")
        note_box = ui.textarea(value=note.get("note") or "", label="Note").classes("w-full").props("rows=2 outlined")

        def _save_note():
            account_api.save_relationship(name=name, cik=acct_cik, quality=(q_sel.value or None),
                                          note=(note_box.value or "").strip() or None)
            ui.notify(f"Saved relationship note for {disp}.", type="positive")
            dlg.close()
        with ui.row().classes("w-full justify-end gap-2").style("margin-top:6px;"):
            ui.button("Save note", icon="save", on_click=_save_note).props("dense color=primary")
    dlg.open()


def _fund_lineup_label(manager_name, head=4):
    """Short inline lineup for a 13F manager: the individual fund names EDGAR shows
    for that family ('Value Fund · Value Plus · Mid Cap Value'), or "" if the manager
    doesn't resolve to a registered fund family. Cache-backed (see core.fund_lineup);
    unmapped managers return "" after a cheap no-network check, so it's safe to call
    per row. Never guesses — a name we can't confidently map shows nothing."""
    try:
        from core import fund_lineup
        if not fund_lineup.has_lineup(manager_name):
            return ""
        lu = fund_lineup.lineup_for_manager(manager_name)
        if not lu or not lu.get("funds"):
            return ""
        # Always DISPLAY the fund names we have — a reader who knows the family spots a
        # gap, and showing something beats a blank. State the true total from the
        # authoritative ticker-file count, and mark "partial" only when we genuinely
        # couldn't capture the whole roster (rare now that we aggregate across filings).
        names = [f["name"] for f in lu["funds"]]
        total = lu.get("expected_series") or len(names)
        total = max(total, len(names))
        label = " · ".join(names[:head])
        extra = total - min(len(names), head)
        if extra > 0:
            label += f"  (+{extra} more)"
        if not lu.get("complete"):
            label += " · partial"
        return label
    except Exception:
        return ""


def _open_lineup_crosswalk_dialog():
    """Human-in-the-loop review of the manager→fund-family crosswalk — the links that
    let the drill-down show a holder's individual funds. Confirm the ambiguous matches,
    reject any wrong auto-match; decisions survive a re-scan (they're never clobbered)."""
    from core import fund_lineup as fl
    dlg = ui.dialog()

    def _row(manager, norm):
        return pretty_name(manager or norm)

    @ui.refreshable
    def _body():
        entries = fl.crosswalk_entries()
        pend = [e for e in entries if not e.get("confirmed") and not e.get("rejected")]
        review = [e for e in pend if e.get("confidence") == "review"]
        high = [e for e in pend if e.get("confidence") == "high"]
        confirmed = [e for e in entries if e.get("confirmed")]
        rejected = [e for e in entries if e.get("rejected")]

        ui.label(f"Needs your confirmation ({len(review)})").classes("text-md font-bold").style(
            f"color:{COLORS['text_heading']};margin-top:4px;")
        if not review:
            ui.label("Nothing waiting — every ambiguous match is resolved.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        for e in review:
            with _list_item_card():
              with ui.row().classes("w-full items-center gap-2 no-wrap"):
                ui.label(_row(e.get("manager"), e["norm"])).style("min-width:210px;font-weight:600;")
                if e.get("note"):
                    ui.label(e["note"]).style(f"color:{COLORS['warning']};font-size:var(--fs-xs);max-width:200px;")
                cands = fl.candidate_names(e.get("ciks", []))
                opts = {str(c["cik"]): f"{pretty_name(c['name'])} (CIK {c['cik']})" for c in cands}
                opts["__none__"] = "No registered fund family"
                sel = ui.select(opts, value=next(iter(opts)), label="Correct fund family").props(
                    "dense outlined").style("min-width:300px;flex:1;")

                def _do(ev=None, norm=e["norm"], sel=sel):
                    if sel.value == "__none__":
                        fl.reject_entry(norm); ui.notify("Marked: no fund family", type="info")
                    else:
                        fl.confirm_entry(norm, int(sel.value)); ui.notify("Confirmed", type="positive")
                    _body.refresh()
                ui.button("Confirm", on_click=_do).props("dense color=primary")

        ui.label(f"Auto-matched — high confidence ({len(high)})").classes("text-md font-bold").style(
            f"color:{COLORS['text_heading']};margin-top:14px;")
        ui.label("Validated against the family's current SEC name. Skim for anything off.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
        for e in high:
            reg = e.get("registrant") or fl._registrant_name(e.get("cik"))
            roster = fl.series_roster(e.get("cik"))
            cnt = (roster or {}).get("expected_series")
            with _list_item_card():
              with ui.row().classes("w-full items-center gap-2 no-wrap"):
                ui.label(_row(e.get("manager"), e["norm"])).style("min-width:210px;font-weight:600;")
                ui.label(f"→ {pretty_name(reg or '—')}" + (f"  ·  {cnt} funds" if cnt else "")).style(
                    f"color:{COLORS['text_secondary']};font-size:var(--fs-sm);flex:1;")
                ui.button("Looks right", on_click=lambda ev=None, n=e["norm"]: (
                    fl.confirm_entry(n), ui.notify("Locked in", type="positive"), _body.refresh())).props(
                    "dense flat color=positive")
                ui.button("Wrong", on_click=lambda ev=None, n=e["norm"]: (
                    fl.reject_entry(n), ui.notify("Removed", type="warning"), _body.refresh())).props(
                    "dense flat color=negative")

        if confirmed:
            with ui.expansion(f"Confirmed ({len(confirmed)})", value=False).classes("w-full").style("margin-top:8px;"):
                for e in confirmed:
                    with ui.row().classes("w-full items-center gap-2"):
                        ui.label(_row(e.get("manager"), e["norm"])).style("min-width:210px;")
                        ui.label(f"→ {pretty_name(e.get('registrant') or '—')}").style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);flex:1;")
                        ui.button("Undo", on_click=lambda ev=None, n=e["norm"]: (
                            fl.restore_entry(n), _body.refresh())).props("dense flat")
        if rejected:
            with ui.expansion(f"Rejected ({len(rejected)})", value=False).classes("w-full"):
                for e in rejected:
                    with ui.row().classes("w-full items-center gap-2"):
                        ui.label(_row(e.get("manager"), e["norm"])).style(
                            f"min-width:210px;color:{COLORS['text_muted']};")
                        ui.button("Restore", on_click=lambda ev=None, n=e["norm"]: (
                            fl.restore_entry(n), _body.refresh())).props("dense flat")

    def _rescan():
        res = fl.bootstrap_from_book(get_active_client_id())
        ui.notify(f"Re-scanned book — {len(res['high'])} matched, {len(res['review'])} to confirm",
                  type="info")
        _body.refresh()

    with dlg, ui.card().style("min-width:min(920px,96vw);max-width:96vw;max-height:90vh;overflow:auto;"):
        with ui.row().classes("w-full justify-between items-center"):
            ui.label("Fund-lineup matches — manager → registered fund family").classes("text-lg font-bold")
            with ui.row().classes("items-center gap-1"):
                ui.button("Re-scan book", icon="refresh", on_click=_rescan).props("dense flat")
                ui.button(icon="close", on_click=dlg.close).props("flat round dense")
        ui.label("These links are what let the drill-down name a holder's individual funds. Confirm the "
                 "ambiguous ones, reject any wrong auto-match — your decisions stick across re-scans.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        _body()
    dlg.open()


def _render_big_picture(institutions, part="all"):
    client_id = get_active_client_id()
    # Tag each holder with how our weight in their book compares to how they weight our comps
    # (Over/In-line/Under) — read by _holder_position_cells in the metro select + drill dialogs.
    _enrich_holders_vs_comps(institutions, client_id)

    metro_lookup = {i["Fund"]: i["Metro"] for i in institutions}
    metro_lookup.update({
        "H.C. Wainwright": "New York Metro", "Ladenburg Thalmann": "New York Metro",
        "Maxim Group": "New York Metro", "Litchfield Hills Research": "New York Metro",
        "Barrington Research": "Chicago / Midwest", "Northland Securities": "Chicago / Midwest",
        "Royce Investment Partners": "New York Metro", "Kennedy Capital Management": "Chicago / Midwest",
        "Conestoga Capital Advisors": "Philadelphia / Baltimore", "Robotti & Company": "New York Metro",
    })

    trip_rows, total_meetings = [], 0
    trips = _load_json("ndr_trips.json", [])
    for t in trips:
        if t.get("city") and t.get("city") != "Virtual":
            n = sum(len(d.get("meetings", [])) for d in t.get("schedule", []))
            total_meetings += n
            trip_rows.append((t["city"], t.get("dates", "—"), ", ".join(t.get("team", [])) or "—",
                               t.get("sponsor_bank", "—") or "—", n))

    ndr_requests = [r for r in _load_ndr_requests() if not r.get("resolved")]
    conferences = get_seed_conferences(client_id)
    wainwright_conf = next((c for c in conferences if "Wainwright" in c.get("Event", "")), None)

    call_by_metro = {}
    try:
        cl_path = client_data_path("q1_2026_call_listener_log.csv")
        if os.path.exists(cl_path):
            cl = pd.read_csv(cl_path)
            hot = cl[cl["Flag"].astype(str).str.contains("HOT LEAD|Newly engaged", case=False, na=False)]
            for _, r in hot.iterrows():
                m = metro_lookup.get(r["Firm"], "Unmapped")
                call_by_metro[m] = call_by_metro.get(m, 0) + 1
    except Exception:
        pass
    call_signal = sum(call_by_metro.values())

    visitor_signal = 0
    try:
        vis_path = client_data_path("ir_website_visitor_log.csv")
        if os.path.exists(vis_path):
            vl = pd.read_csv(vis_path)
            visitor_signal = int((vl["Category"] == "New - Unidentified").sum())
    except Exception:
        pass
    prospects = _load_json("prospects.json", [])
    prospect_q_count = len(prospects)

    holders_by_metro = {}
    for i in institutions:
        if i["USIO_Holder"]:
            holders_by_metro[i["Metro"]] = holders_by_metro.get(i["Metro"], 0) + 1
    holder_count = sum(holders_by_metro.values())
    tracked_total = len(institutions)

    city_to_metro = {
        "Boston": "Boston / New England", "New York": "New York Metro",
        "San Francisco": "West Coast (SF/LA)", "Chicago": "Chicago / Midwest",
        "Minneapolis / Midwest": "Chicago / Midwest", "Texas / Denver": "Texas / Denver",
        "Philadelphia / Baltimore": "Philadelphia / Baltimore",
    }
    visited_metros = {city_to_metro.get(c, c) for c, *_ in trip_rows}

    new_total = call_signal + visitor_signal + prospect_q_count
    bp_denom = max(new_total + holder_count, 1)
    new_pct = round(new_total / bp_denom * 100)
    if new_pct >= 70:
        mix_label, mix_color = "Strong — well above the 50/50 baseline management would be happy with", "#15803D"
    elif new_pct >= 45:
        mix_label, mix_color = "Balanced — roughly the 50/50 mix management would be content with", "#B45309"
    else:
        mix_label, mix_color = "Existing-heavy — below the mix that keeps management happy; needs more new-investor prospecting", "#B91C1C"

    new_from_recent_trip = sum(n for m, n in call_by_metro.items() if m in visited_metros)
    recency_note = ""
    if call_signal > 0 and visited_metros:
        recent_pct = round(new_from_recent_trip / call_signal * 100)
        if recent_pct >= 50:
            recency_note = (f"{recent_pct}% of the call-engagement signal ties back to a city you just visited "
                             f"({', '.join(visited_metros & set(call_by_metro.keys()))}) — worth watching whether this holds up "
                             f"once that trip's momentum fades, not just banking it as durable growth.")
        else:
            recency_note = (f"Checked against your recent trip(s): only {recent_pct}% of the call-engagement signal "
                             f"ties back to a recently-visited city — this isn't just NDR afterglow, it's broader than that.")

    visits_by_metro = {}
    for city, *_r in trip_rows:
        m = city_to_metro.get(city, city)
        visits_by_metro[m] = visits_by_metro.get(m, 0) + 1

    # Cluster raw "City, ST" holder metros into the same ~60-mile roadshow metros the unified
    # prospect view uses (Twin Cities as one stop, not Wayzata/Minneapolis split). Holders don't
    # carry raw city/state, so parse the "City, ST" label back through _metro_from_city.
    def _cluster(label):
        if not label or label in ("Unknown (SEC)", "International") or ", " not in label:
            return label or "Unknown (SEC)"
        city, st = label.rsplit(", ", 1)
        return _metro_from_city(city, st)

    metro_summary = {}
    for i in institutions:
        m = _cluster(i["Metro"])
        d = metro_summary.setdefault(m, {"count": 0, "tier1_nonholder": 0, "holders": 0, "top": None, "top_score": -1,
                                         "insts": [], "holder_list": [], "t1_list": []})
        d["count"] += 1
        d["insts"].append(i)
        if _score_val(i) >= 80 and not i["USIO_Holder"]:
            d["tier1_nonholder"] += 1
            d["t1_list"].append(i)                  # exact list behind the "Tier-1 ready" count, for the cell drill-down
        if i["USIO_Holder"]:
            d["holders"] += 1
            d["holder_list"].append(i)              # exact list behind the "Holders" count
        if _score_val(i) > (d["top_score"] or 0):
            d["top_score"], d["top"], d["top_inst"] = i["Engagement_Score"], i["Fund"], i

    # Fold the qualified peer-prospect universe into the metro rollup as NON-HOLDERS — the funds
    # that own a peer/comp but not us, i.e. the actual NDR conversion targets. Without this the table
    # counts only current holders, so a prospect-rich metro (New York = 30 real prospects) reads
    # "0 non-holders" — the biggest roadshow opportunity, invisible. See peer_prospects.build_candidates.
    try:
        from core import peer_prospects
        _prospects = peer_prospects.build_candidates(get_active_client_id(), limit=None) or []
    except Exception:
        _prospects = []
    for c in _prospects:
        city, state = (c.get("city") or "").strip(), (c.get("state") or "").strip()
        m = _metro_from_city(city, state)     # same ~60-mile roadshow clustering as the holders above
        d = metro_summary.setdefault(m, {"count": 0, "tier1_nonholder": 0, "holders": 0,
                                         "top": None, "top_score": -1, "insts": [],
                                         "holder_list": [], "t1_list": []})
        d["count"] += 1
        _pinst = {"Fund": c.get("filer"), "Metro": m, "USIO_Holder": False,
                  "City": city.title(), "Position_Value": c.get("peer_value"),
                  "Conviction": None, "Engagement_Score": c.get("conviction"), "AUM": None,
                  "comps": c.get("comps"),          # carry the specific comps so the drill-down
                  "Action": "Prospect — owns a peer/comp, not you"}  # can show "owns CASS, FOUR"
        if (c.get("conviction") or 0) >= 70:        # strong, meeting-worthy prospect (Tier-1 proxy)
            d["tier1_nonholder"] += 1
            d["t1_list"].append(_pinst)             # prospects also count toward Tier-1 ready
        if d["top"] is None:                        # prospect-only metro: name its top prospect
            d["top"] = c.get("filer")
        d["insts"].append(_pinst)
    tracked_total = len(institutions) + len(_prospects)

    requests_by_metro = {}
    for req in ndr_requests:
        requests_by_metro[req["metro"]] = requests_by_metro.get(req["metro"], 0) + 1

    metro_priority = []
    for m in set(list(metro_summary.keys()) + list(requests_by_metro.keys())):
        d = metro_summary.get(m, {"count": 0, "tier1_nonholder": 0, "holders": 0, "top": None, "top_score": -1})
        visits = visits_by_metro.get(m, 0)
        req_count = requests_by_metro.get(m, 0)
        priority = (d["tier1_nonholder"] * 3 + d["holders"] * 1.5 + d["count"] + req_count * 2.5) / (visits + 1)
        metro_priority.append((m, priority, d, visits, req_count))
    metro_priority.sort(key=lambda x: -x[1])
    if not metro_priority:
        # A genuinely new client — no 13F holders on file and no inbound NDR requests
        # logged yet — has no metros to rank. This crashed until the fabricated seed
        # requests were removed; the seed had been masking it by guaranteeing three.
        ui.label("No geography to prioritise yet.").classes("section-head")
        ui.label("This view ranks where to spend roadshow time from your 13F holder book and any inbound "
                 "NDR requests. Neither is on file for this client yet — run a 13F refresh on the SEC "
                 "Intelligence tab, and log requests as they come in.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        return
    # Peer-owner universe by metro (all buckets), computed ONCE here so both "where to go" reads (the
    # Ownership IR READ and the Targeting WHERE TO GO), the "Where to start" tile, and the metro table
    # all rank from the same live data every render — and therefore always agree and update automatically.
    try:
        from core import peer_prospects as _pp
        _all_cands = _pp.all_candidates(get_active_client_id())
    except Exception:
        _all_cands = []
    _tier_key = {"Institutional": "inst", "RIA / wealth": "ria", "Diversified": "div",
                 "Market maker": "mm", "Curated": "curated"}
    peer_by_metro = {}
    peer_funds_by_metro = {}          # metro -> [candidate dicts], for the click-to-select dialog
    for _c in _all_cands:
        _m = _metro_from_city(_c.get("city"), _c.get("state"))
        peer_funds_by_metro.setdefault(_m, []).append(_c)
        _pm = peer_by_metro.setdefault(_m, {"funds": 0, "inst": 0, "ria": 0, "div": 0, "mm": 0, "curated": 0})
        _pm["funds"] += 1
        _k = _tier_key.get(_c.get("tier"))
        if _k:
            _pm[_k] += 1

    # THE top destination = the metro table's #1 row: ranked by ORGANIC peer-owners (Inst+RIA+Divsfd+MM,
    # curated EXCLUDED) with the same holders/Tier-1 tiebreak the table uses. Every "where to go" surface
    # reads this, so they name the same city. (The old visit/request-weighted priority formula survives
    # only for the ranked "Priority cities" detail list on the Where-to-start tile.)
    def _organic(_mm):
        _p = peer_by_metro.get(_mm, {})
        return _p.get("funds", 0) - _p.get("curated", 0)
    _geo_rank = sorted(
        (set(metro_summary) | set(peer_by_metro)),
        key=lambda _mm: (-_organic(_mm),
                         -metro_summary.get(_mm, {}).get("holders", 0),
                         -metro_summary.get(_mm, {}).get("tier1_nonholder", 0)))
    top_metro = _geo_rank[0] if _geo_rank else metro_priority[0][0]
    top_d = metro_summary.get(top_metro, {})
    top_visits = visits_by_metro.get(top_metro, 0)
    top_metro_request = next((r for r in ndr_requests if r["metro"] == top_metro), None)
    top_metro_conf = None
    if top_metro_request and wainwright_conf and "Wainwright" in top_metro_request["firm"]:
        top_metro_conf = wainwright_conf
    _wtg_rq = (f" — plus {top_metro_request['analyst']} ({top_metro_request['firm']}) asked for it directly"
               if top_metro_request else "")

    # weakest_request / quiet_period_days removed with the "respond before the quiet
    # period begins in N days" line. That countdown collapsed "not configured" and
    # "already started" to the same 0, so BOTH tenants rendered "in 0 days". Reports'
    # quiet-period banner is the correct pattern if a countdown is wanted again: it
    # renders only when a quiet_start exists AND the count is positive.

    if part in ("all", "summary"):
        ui.label("Big Picture — Where Things Stand").classes("text-xl font-bold").style(f"color:{COLORS['text_heading']};margin-top:12px;")

        # THE IR READ — a plain-English BLUF up top (same pattern as NOBO's CEO read), built from the
        # numbers already on this page so the "so what" lands before the tiles.
        _new_pos = [i for i in institutions if (i.get("Direction") or "").lower() == "new"]
        _dirl = lambda i: (i.get("Direction") or "").lower()
        _bp_adding = [i for i in institutions if i.get("USIO_Holder") and _dirl(i) in ("adding", "new")]
        _bp_trimming = [i for i in institutions if i.get("USIO_Holder") and _dirl(i) in ("trimming", "exited")]
        try:                                          # who should own you = the conviction target universe
            from core import peer_prospects as _bpp
            _bp_targets = _bpp.build_candidates(client_id, limit=None) or []
        except Exception:
            _bp_targets = []
        _bp_tier1 = [c for c in _bp_targets if (c.get("conviction") or 0) >= 70]
        # Tier-1-ready count from the ONE canonical definition, shared with the Target Database (see
        # _tier1_ready_names) so the card and the click-through can never disagree.
        _bp_tier1_ready = len(_tier1_ready_names(institutions, client_id))
        _bp_top_targets = sorted(_bp_targets, key=lambda c: -(c.get("conviction") or 0))
        _bp_targets_count = max(tracked_total - holder_count, 0)   # ALL non-holder targets (tracked + peer-owners) — matches the split list
        _ir_read = []
        _ir_read.append(
            f"Start in {top_metro}: {top_d.get('tier1_nonholder', 0)} Tier-1 non-holder(s) ready to convert and "
            f"{top_d.get('holders', 0)} holder(s) to defend{_wtg_rq}.")
        if ndr_requests:
            _ir_read.append(
                f"{len(ndr_requests)} inbound NDR request(s) on the desk across "
                f"{len({r['metro'] for r in ndr_requests})} cities — slot them while you're already routing.")
        if _new_pos:
            _nm = ", ".join(pretty_name(i['Fund']) for i in _new_pos[:3])
            _ir_read.append(
                f"{len(_new_pos)} new 13F position(s) entered the book this cycle ({_nm}"
                f"{' …' if len(_new_pos) > 3 else ''}) — new money to engage early.")
        _ir_read.append(
            f"{holder_count} holder(s) own you today; {_bp_targets_count} non-holder target(s) are being worked, "
            f"{_bp_tier1_ready} at Tier-1.")
        with ui.card().classes("w-full").style(
                "background:rgba(30,64,175,.06);border:1.5px solid #1E40AF;border-left:6px solid #1E40AF;"
                "border-radius:8px;margin-top:8px;"):
            ui.label("THE IR READ — where your ownership stands").style(
                "color:#1E3A8A;font-size:var(--fs-sm);font-weight:800;letter-spacing:.04em;text-transform:uppercase;")
            # A divider under the header, then bullets with a hanging indent — the bullet sits in its own
            # column so wrapped lines align under the TEXT, not under the bullet.
            ui.element("div").style("border-bottom:1.5px solid #1E40AF55;margin:6px 0 9px;width:100%;")
            for _p in _ir_read:
                with ui.row().classes("w-full no-wrap items-start").style("gap:8px;margin-top:1px;"):
                    ui.label("•").style("color:#1E40AF;font-size:var(--fs-base);line-height:1.6;font-weight:800;")
                    ui.label(_p).style(
                        f"color:{COLORS['text_heading']};font-size:var(--fs-base);line-height:1.6;font-weight:500;")

        # The four questions (matches the Buyside Ownership tab and NOBO's composition — one shape across
        # every ownership surface): who owns you · who should · who's moving · where to start. Each card now
        # names itself via its eyebrow header, so no redundant subheader row above the grid.
        with ui.row().classes("w-full gap-3 items-stretch").style("margin-top:10px;"):
            _bp_metric("Current Institutional Investors", str(holder_count),
                       [f"{holder_count} tracked holder(s) · {tracked_total - holder_count} non-holder prospect(s)",
                        ("By city: " + ", ".join(f"{m} ({n})" for m, n in sorted(holders_by_metro.items(), key=lambda x: -x[1])))
                        if holders_by_metro else "No holders tracked yet."],
                       sub=_peer_holder_sub(holder_count, client_id),
                       on_click=lambda: nav.go_to("Targeting", "Target Database", db_filter="holders"))
            _bp_metric("Target Investors", f"{_bp_tier1_ready} / {_bp_targets_count}",
                       [f"{pretty_name(c.get('filer', ''))} — conviction {round(c.get('conviction') or 0)}"
                        for c in _bp_top_targets[:5]] or ["No conviction targets yet."],
                       on_click=lambda: nav.go_to("Targeting", "Target Database", db_filter="tier1"))
            _bp_metric("Adding / Trimming", f"{len(_bp_adding)} ↑ / {len(_bp_trimming)} ↓",
                       ([f"Adding/new: {', '.join(pretty_name(i['Fund']) for i in _bp_adding[:4])}"] if _bp_adding else ["None adding this cycle."])
                       + ([f"Trimming/exited: {', '.join(pretty_name(i['Fund']) for i in _bp_trimming[:4])}"] if _bp_trimming else []),
                       on_click=lambda: nav.go_to("Ownership", "Buyside Ownership"))
            # "Where to start" is the roadshow-geography answer: the #1 city, then a ranked summary of
            # the other cities in play, and — the part IR actually acts on — who asked for each one.
            _wts_detail = ["Priority cities (convert · defend):"]
            for _m in _geo_rank[:5]:
                _d = metro_summary.get(_m, {})
                _rc = requests_by_metro.get(_m, 0)
                _seg = f"  {_m}: {_d.get('tier1_nonholder', 0)} to convert · {_d.get('holders', 0)} to defend"
                if _rc:
                    _seg += f" · {_rc} request(s)"
                _wts_detail.append(_seg)
            if ndr_requests:
                _wts_detail.append("Who's asking:")
                for _r in ndr_requests:
                    _wts_detail.append(f"  {_r['metro']} — {_r.get('analyst', '—')} ({_r.get('firm', '—')})")
            # Card face shows ONLY the #1 city (per request); the ranked city list stays in the detail dialog.
            _wts_sub = f"{len(ndr_requests)} inbound request(s) on the desk" if ndr_requests else None
            _bp_metric("Where to start", top_metro, _wts_detail, sub=_wts_sub,
                       on_click=lambda: nav.go_to("Roadshow", "NDR"))

        # Holder moves this cycle — the approved colored-header card pattern (green adds / red trims,
        # count badge, status pills). Only shown when the 13F diff actually surfaced movement.
        if _bp_adding or _bp_trimming:
            ui.label("Institutional Activity").classes("section-head").style("margin-top:16px;")

            def _move_pill(it):
                d = (it.get("Direction") or "").lower()
                return {"new": ("rpill-good", "new"), "adding": ("rpill-good", "adding"),
                        "trimming": ("rpill-warn", "trimming"), "exited": ("rpill-def", "exited")}.get(d, ("rpill-def", d or "—"))

            def _moves_card(title, items, tone, val_color):
                def _move_row(it):
                    _pc, _pt = _move_pill(it)
                    with ui.row().classes("rrow"):
                        ui.label(pretty_name(it.get("Fund", "") or "—")).style("font-weight:600;")
                        ui.label(_pt).classes(f"rpill {_pc}")
                        _dv = _holder_dollars(it.get("Position_Value"))
                        if _dv and _dv != "—":
                            ui.label(_dv).classes("rval").style(f"color:{val_color};")

                with ui.card().classes("flex-1 report-card"):
                    with ui.row().classes(f"rhead rhead-{tone}"):
                        ui.label(title)
                        ui.label(str(len(items))).classes("count")
                    if not items:
                        with ui.row().classes("rrow"):
                            ui.label("None this cycle").classes("rmeta")
                    for it in items[:6]:
                        _move_row(it)
                    # Expand to show ALL of them (the card can hold every mover, not just the first 6).
                    if len(items) > 6:
                        with ui.expansion(f"Show all {len(items)}").classes("w-full").props("dense"):
                            for it in items[6:]:
                                _move_row(it)

            with ui.row().classes("w-full gap-4 items-start").style("margin-top:8px;"):
                _moves_card("New money & adds", _bp_adding, "good", "#127A4A")
                _moves_card("Trimming & exits", _bp_trimming, "bad", "#B1352D")

        # (Removed the opaque "New vs. existing" mix bar — its 52/48 came from a call/visitor/prospect
        # signal blend that never reconciled with the visible holder/target counts; the four cards above
        # carry the composition now.)
        # Data-source footnote — deliberately NOT a headline; provenance/cadence are set in onboarding.
        ui.label("Sourced from SEC 13F filings + peer 13F books · refreshed each 13F cycle."
                 + (f" · {recency_note}" if recency_note else "")).style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);margin-top:6px;font-style:italic;")

        most_visited = max(visits_by_metro.items(), key=lambda x: x[1]) if visits_by_metro else None
        if most_visited and most_visited[1] >= 2 and most_visited[0] != top_metro:
            ui.html(
                f"<div style='font-size:var(--fs-base);color:#B45309;margin-top:10px;'>"
                f"By contrast, <b>{most_visited[0]}</b> has had {most_visited[1]} trips already for only "
                f"{metro_summary.get(most_visited[0], {}).get('count', '?')} tracked institution(s) — before scheduling a third trip "
                f"there, weigh it against the untapped opportunity in {top_metro}.</div>"
            )

        # ("Top opportunity — where to start" card removed from here per request — the four-question
        # "Where to start" summary card above already answers it, and the metro table below carries the
        # geographic detail. Can be relocated elsewhere if wanted.)

    _geo_table_exp = None                     # the metro-table collapsible, returned so a card click can collapse it
    if part in ("all", "geo"):
        # Geographic breakdown — answers the obvious question the counts above raise:
        # "these institutions... where ARE they?" One row per metro with its holders,
        # ready-to-convert Tier-1 non-holders, NDR trips so far, and its top name.
        ui.markdown("---")
        # peer_by_metro / peer_funds_by_metro and the top destination (top_metro / top_d / _wtg_rq) are
        # computed once in the shared section above, so this WHERE TO GO read names the same city as the
        # Ownership IR READ and the metro table below.
        with ui.column().classes("w-full").style(
                "background:rgba(30,64,175,.06);border:1px solid #1E40AF;border-left:5px solid #1E40AF;"
                "border-radius:8px;padding:9px 14px;gap:5px;margin-top:6px;"):
            with ui.row().classes("w-full items-baseline no-wrap").style("gap:10px;flex-wrap:wrap;"):
                ui.label("WHERE TO GO").style(
                    "color:#1E3A8A;font-size:var(--fs-2xs);font-weight:800;letter-spacing:.05em;white-space:nowrap;")
                ui.label(f"Start in {top_metro}: {top_d.get('tier1_nonholder', 0)} Tier-1 non-holder(s) ready to "
                         f"convert and {top_d.get('holders', 0)} holder(s) to defend{_wtg_rq}.").style(
                    f"color:{COLORS['text_heading']};font-size:var(--fs-base);font-weight:600;line-height:1.5;")
            # The "N current holders and M peer-owners across K metros…" instruction, FOLDED INTO this box
            # (was a separate line under the title). Populated once the peer/metro totals are computed.
            _geo_note = ui.label("").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);line-height:1.45;")
        # The metro table lives in a collapsible so a summary-card click (see set_db_filter) can collapse
        # it to focus on the filtered results; the header re-opens it any time. The WHERE TO GO read above
        # stays put — only the table collapses.
        _geo_table_exp = ui.expansion("Institutions & Peer Ownership By City/Metro Area", icon="map",
                                      value=True).classes("w-full").style("margin-top:10px;")

        # (peer_by_metro / peer_funds_by_metro were computed above, before the WHERE TO GO read, so the
        # read can point to the table's #1 destination.)

        # NDR status per metro, from the trip book — replaces the old "full-day / half-day"
        # viability read, which said nothing you'd act on (you plan a full day regardless;
        # the slot count gets worked out later). This shows where a roadshow actually STANDS:
        # Planning (trip opened) → Scheduled (dates/meetings set) → Completed. A metro can
        # carry several trips; keep the most-active.
        _st_rank = {"Completed": 0, "Planning": 1, "Scheduled": 2}
        ndr_status_by_metro = {}
        for _t in trips:
            _c = _t.get("city")
            if not _c or _c == "Virtual":
                continue
            _mk = _metro_from_city(*_c.rsplit(", ", 1)) if ", " in _c else _c
            if (_t.get("status") or "") == "Completed":
                _disp = "Completed"
            elif _t.get("meetings") or (_t.get("dates") and _t.get("dates") not in ("TBD", "")):
                _disp = "Scheduled"
            else:
                _disp = "Planning"
            if _mk not in ndr_status_by_metro or _st_rank[_disp] > _st_rank[ndr_status_by_metro[_mk]]:
                ndr_status_by_metro[_mk] = _disp

        # Union all metro keys (a metro may have holders but no peer-owners, or vice-versa).
        _all_metros = set(metro_summary) | set(peer_by_metro)
        _blank_pm = {"funds": 0, "inst": 0, "ria": 0, "div": 0, "mm": 0, "curated": 0}
        geo_rows = sorted(
            [{"metro": m, "holders": (d := metro_summary.get(m, {})).get("holders", 0),
              "t1": d.get("tier1_nonholder", 0), "trips": visits_by_metro.get(m, 0),
              "funds": (pm := peer_by_metro.get(m, _blank_pm))["funds"], "inst": pm["inst"],
              "ria": pm["ria"], "div": pm["div"], "mm": pm["mm"], "curated": pm["curated"],
              "read": ndr_status_by_metro.get(m, "—"), "top": pretty_name(d.get("top")) if d.get("top") else "—"}
             for m in _all_metros],
            # Rank destinations by ORGANIC peer-owners (Inst+RIA+Divsfd+MM), EXCLUDING hand-entered
            # Curated targets: curated are known-fit accounts you added by hand, not evidence of where
            # peer-ownership actually clusters, so they must not push a metro up the destination order.
            # Curated still shows in its own column; it just no longer drives the ranking.
            key=lambda r: (-(r["funds"] - r["curated"]), -r["holders"], -r["t1"]))
        _tk = CT("ticker")
        geo_cols = [
            {"name": "metro", "label": "Metro / region", "field": "metro", "align": "left", "sortable": True,
             "tooltip": "Fund HQ clustered into ~60-mile roadshow metros — a day's drive, the way an NDR is planned"},
            {"name": "holders", "label": "Holders", "field": "holders", "align": "right", "sortable": True,
             "tooltip": f"Funds here that own {_tk} today (13F)"},
            {"name": "t1", "label": "Tier-1\nready", "field": "t1", "align": "right", "sortable": True,          # two-row header
             "tooltip": "Tier-1 ready — high-conviction NON-holders ready to convert (score ≥80 / conviction ≥70)"},
            {"name": "funds", "label": "Peer-\nowners", "field": "funds", "align": "right", "sortable": True,    # two-row header
             "tooltip": f"Funds that own a peer/comp but not {_tk} — your prospecting universe (= Inst+RIA+Divsfd+MM+Curated)"},
            {"name": "inst", "label": "Inst", "field": "inst", "align": "right", "sortable": True,
             "tooltip": "Institutional peer-owners — the primary NDR-target bucket"},
            {"name": "ria", "label": "RIA", "field": "ria", "align": "right", "sortable": True,
             "tooltip": "RIA / wealth peer-owners — video-call tier, usually no PM to pitch"},
            {"name": "div", "label": "Divsfd", "field": "div", "align": "right", "sortable": True,
             "tooltip": "Diversified — large multi-strategy / bank-AM houses"},
            {"name": "mm", "label": "MM", "field": "mm", "align": "right", "sortable": True,
             "tooltip": "Market makers — no fundamental PM to pitch"},
            {"name": "curated", "label": "Curated", "field": "curated", "align": "right", "sortable": True,
             "tooltip": "Hand-entered targets that don't hold a comp yet (a private bank, a prior-seat relationship)"},
            {"name": "trips", "label": "NDRs", "field": "trips", "align": "right", "sortable": True,
             "tooltip": "Trips logged for this metro — any status (planning, scheduled, completed)"},
            {"name": "read", "label": "NDR status", "field": "read", "align": "left", "sortable": True,
             "tooltip": "State of the most-active NDR here: Planning → Scheduled → Completed (— none yet)"},
            {"name": "top", "label": "Top holder", "field": "top", "align": "left", "sortable": True,
             "tooltip": "Highest-scoring tracked name in this metro",
             "style": "white-space:normal;min-width:220px;", "headerStyle": "white-space:normal;"},
        ]
        # Click a metro row to see exactly WHO is there — the counts above always
        # raised "which 5 institutions?"; this opens the named list on demand
        # Click a metro row → tick the peer-owners there → shortlist onto an NDR (one selectable list;
        # replaces the old tracked-only readout and the separate Peer Prospects metro table).
        # ── Cell drill-down: click any number to see the exact names behind it ────────────
        # Every count in this table is backed by a real list, so a number is a doorway, not a
        # dead end. Holders/Tier-1 come from the lists captured while counting (can't disagree
        # with the cell); the bucket counts filter the same peer-owner universe the row uses.
        _CLICKABLE = {"holders": "Holders", "t1": "Tier-1 ready", "funds": "Peer-owners",
                      "inst": "Institutional", "ria": "RIA / wealth", "div": "Diversified",
                      "mm": "Market maker", "curated": "Curated targets"}

        def _drill_list(metro, col):
            d = metro_summary.get(metro, {})
            if col == "holders":
                return d.get("holder_list", [])
            if col == "t1":
                return d.get("t1_list", [])
            funds = peer_funds_by_metro.get(metro, [])
            if col == "funds":
                return funds
            return [c for c in funds if _tier_key.get(c.get("tier")) == col]

        def _drill_row(x):
            # Two shapes reach here: peer-owner candidates ("filer"/"tier"/"conviction"/"comps")
            # and tracked-institution dicts ("Fund"/"USIO_Holder"/"Engagement_Score"). Normalise both.
            raw_name = x.get("filer") or x.get("Fund") or ""
            if "filer" in x:
                conv = x.get("conviction")
                peers = ", ".join(sorted((x.get("comps") or {}).keys()))
                return {"_filer": x.get("filer"),          # row key for selection (not a column)
                        "Fund": pretty_name(x.get("filer") or "—"),
                        "City": (x.get("city") or "—").title(),
                        "Category": x.get("tier") or "—",
                        "Score": round(conv) if isinstance(conv, (int, float)) else "—",
                        "Detail": (f"Owns {peers}" if peers else ("Curated target" if x.get("kind") == "curated" else "—")),
                        "Funds": _fund_lineup_label(raw_name)}
            sc = x.get("Engagement_Score")
            # A peer-prospect carried its comps in — show which comp it owns ("owns CASS, FOUR"),
            # matching the Peer-owners drill. Real tracked institutions have no comps → fall back.
            peers = ", ".join(sorted((x.get("comps") or {}).keys()))
            _q = _is_quant_inst(x)
            if x.get("USIO_Holder"):                 # existing holder → position size + trajectory
                _cat = "Quant — no 1×1" if _q else "Holder"
                _score, detail = _holder_position_cells(x)
            else:                                    # tracked non-holder / promoted prospect → conviction
                _cat = "Non-holder"
                _score = round(sc) if isinstance(sc, (int, float)) else "—"
                detail = f"Owns {peers}" if peers else (x.get("Conviction") or x.get("Action") or "—")
            return {"_filer": pretty_name(x.get("Fund") or "—"),   # row key for selection
                    "Fund": pretty_name(x.get("Fund") or "—"),
                    "City": x.get("City") or (x.get("Metro") or "—"),
                    "Category": _cat,
                    "Score": _score,
                    "Detail": detail,
                    "Funds": _fund_lineup_label(raw_name)}

        _drill_dialog = ui.dialog()

        # Every bucket drill-down is selectable → Add to NDR (same shortlist flow as the metro
        # row-click). Holders/Tier-1 (tracked institutions) are invitable too — you defend existing
        # relationships on a roadshow — EXCEPT quant/systematic shops, excluded at add time.
        _selectable_cols = {"funds", "inst", "ria", "div", "mm", "curated", "holders", "t1"}

        def _open_cell_drill(e):
            metro, col = e.args.get("metro"), e.args.get("col")
            items = _drill_list(metro, col)
            d_rows = [_drill_row(x) for x in items]
            is_peer = col in _selectable_cols
            _drill_dialog.clear()
            with _drill_dialog, ui.card().style("min-width:min(900px,95vw);max-width:95vw;"):
                with ui.row().classes("w-full justify-between items-center"):
                    ui.label(f"{metro} — {_CLICKABLE.get(col, col)} · {len(d_rows)} "
                             f"name{'s' if len(d_rows) != 1 else ''}").classes("text-lg font-bold")
                    ui.button(icon="close", on_click=_drill_dialog.close).props("flat round dense")
                if not d_rows:
                    ui.label("No names behind this number.").style(f"color:{COLORS['text_muted']};")
                else:
                    _n_lineups = sum(1 for r in d_rows if r.get("Funds"))
                    # Add-to-NDR bar (restored) — only for the peer-owner buckets, which are the
                    # NDR-prospecting universe. Tick names → shortlist onto a new or existing trip.
                    trip_sel = new_name = tbl = None
                    item_by_key = {(it.get("filer") or pretty_name(it.get("Fund") or "")): it
                                   for it in items} if is_peer else {}
                    if is_peer:
                        _open = [(i, t) for i, t in enumerate(_load_json("ndr_trips.json", [])) if t.get("status") != "Completed"]
                        trip_opts = {str(i): (t.get("name") or f"NDR {i+1}") for i, t in _open}
                        trip_opts["__new__"] = "＋ New NDR…"
                        with ui.row().classes("w-full items-end gap-2").style(
                                f"background:{COLORS['surface_hover_bg']};border-radius:8px;padding:8px 10px;margin-bottom:6px;"):
                            ui.label("Tick names, then add them to an NDR as shortlisted targets.").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);flex:1;")
                            trip_sel = ui.select(trip_opts, value="__new__", label="NDR").props("dense outlined").style("min-width:170px;")
                            new_name = ui.input("New NDR name", value=f"{metro} NDR").props("dense outlined").style("min-width:150px;")
                            new_name.bind_visibility_from(trip_sel, "value", backward=lambda v: v == "__new__")
                            add_btn = ui.button("Add selected", icon="playlist_add").props("dense color=primary")
                    _dcol_label = {"Funds": "Fund lineup (SEC)", "Category": "Category",
                                   "Score": "Conviction / Position"}
                    dcols = [{"name": k, "label": _dcol_label.get(k, k),
                              "field": k, "sortable": True,
                              "align": "right" if k == "Score" else "left",
                              **({"style": "white-space:normal;min-width:230px;",
                                  "headerStyle": "white-space:normal;"} if k == "Funds" else {})}
                             for k in ("Fund", "City", "Category", "Score", "Detail", "Funds")]
                    _extra = {"selection": "multiple"} if is_peer else {}
                    tbl = ui.table(columns=dcols, rows=d_rows, row_key=("_filer" if is_peer else "Fund"),
                                   pagination=25, **_extra).classes("w-full").props("dense flat")   # 25 per page
                    _header_tooltips(tbl, {
                        "Fund": "The fund / manager — click the name for its Account 360",
                        "City": "Fund HQ city",
                        "Category": "Ownership bucket / holder status",
                        "Score": "Conviction (peer-owners) or engagement score (holders), 0-100",
                        "Detail": f"Which of {CT('ticker')}'s comps they own, or their signal",
                        "Funds": "The manager's individual '40-Act funds, from its latest SEC N-CEN/485 filing",
                    })
                    # Fund name → Account 360 profile. @click.stop so it doesn't toggle the row checkbox.
                    _orig_by_fund = {pretty_name(it.get("filer") or it.get("Fund") or ""): it for it in items}
                    tbl.add_slot("body-cell-Fund", (
                        '<q-td :props="props">'
                        '<span class="cursor-pointer" '
                        f'style="color:{COLORS["accent"]};text-decoration:underline dotted;text-underline-offset:2px;" '
                        '@click.stop="() => $parent.$emit(\'openProfile\', props.row.Fund)">'
                        '{{ props.value }}</span></q-td>'))
                    tbl.on("openProfile", lambda e: _open_account_profile(
                        _orig_by_fund.get(e.args) or {"Fund": e.args}))

                    if is_peer:
                        def _do_add():
                            sel = tbl.selected
                            if not sel:
                                ui.notify("Tick at least one name first.", type="warning"); return
                            trips2 = _load_json("ndr_trips.json", [])
                            if trip_sel.value == "__new__":
                                nm = (new_name.value or "").strip()
                                if not nm:
                                    ui.notify("Name the new NDR.", type="warning"); return
                                trips2.append({
                                    "name": nm, "sponsor_bank": "", "dates": "TBD", "ndr_type": "in-person",
                                    "city": metro, "focus": "", "team": [], "notes": "", "meetings": [], "shortlist": [],
                                    "status": "Planning", "debrief": {}, "days": 2, "slots_per_day": 6,
                                    "created": datetime.now().strftime("%Y-%m-%d"),
                                })
                                target, tname = trips2[-1], nm
                            else:
                                target = trips2[int(trip_sel.value)]
                                tname = target.get("name") or "NDR"
                            target.setdefault("shortlist", [])
                            have = {s.get("institution") for s in target["shortlist"]} | \
                                   {m.get("institution") for m in target.get("meetings", [])}
                            added = quant_skip = 0
                            for row in sel:
                                it = item_by_key.get(row.get("_filer"))
                                if not it:
                                    continue
                                nm2 = it.get("filer") or it.get("Fund")
                                if not nm2 or nm2 in have:
                                    continue
                                if _is_quant_inst(it):
                                    quant_skip += 1; continue     # quant/systematic — no management 1x1
                                rec = _shortlist_record(it) if it.get("filer") else _shortlist_from_inst(it)
                                target["shortlist"].append(rec)
                                have.add(nm2); added += 1
                            _save_json("ndr_trips.json", trips2)
                            dup = len(sel) - added - quant_skip
                            _m = f"Added {added} target(s) to '{tname}' (Planning)"
                            if quant_skip:
                                _m += f" · skipped {quant_skip} quant shop(s) — no 1×1"
                            if dup:
                                _m += f" · {dup} already on it"
                            ui.notify(_m + ". Find it in Outbound → NDR → Active NDRs.", type="positive")
                            _drill_dialog.close()
                        add_btn.on_click(_do_add)
                    if _n_lineups:
                        ui.label(f"“Fund lineup” names the individual '40-Act funds the manager runs (from its "
                                 f"latest SEC N-CEN/485 filing) — the strategy sleeve a 13F hides. Shown for "
                                 f"{_n_lineups} of {len(d_rows)} here; blank where the holder has no registered "
                                 "fund family (hedge fund / SMA) or isn't yet confirmed in the crosswalk.").style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-top:6px;")
            _drill_dialog.open()

        # pagination=25 → the metro list itself pages 25 at a time; sortable headers sort the full set.
        geo_table = ui.table(columns=geo_cols, rows=geo_rows, row_key="metro", pagination=25).classes(
            "w-full cursor-pointer").props("dense flat")
        # Custom header so a "\n" in a label renders as a two-row heading (pre-line) — lets the numeric
        # columns stay narrow and hands the freed width to Top name so it isn't truncated. QTh with
        # :props keeps the sort click/arrow, so the two-row header stays sortable.
        geo_table.add_slot("header", r'''
            <q-tr :props="props">
              <q-th v-for="col in props.cols" :key="col.name" :props="props"
                    style="white-space:pre-line;vertical-align:bottom;line-height:1.15;"
                    :class="col.tooltip ? 'cursor-help' : ''">
                {{ col.label }}
                <q-tooltip v-if="col.tooltip">{{ col.tooltip }}</q-tooltip>
              </q-th>
            </q-tr>
        ''')
        # Make each count a clickable, underlined link that opens the drill-down. @click.stop keeps the
        # row-click ("Add to NDR" select dialog) from also firing. Zero renders as a muted, inert number.
        for _cc in _CLICKABLE:
            geo_table.add_slot(f"body-cell-{_cc}", (
                '<q-td :props="props" class="text-right">'
                '<span v-if="props.value" class="cursor-pointer" '
                f'style="color:{COLORS["accent"]};text-decoration:underline dotted;text-underline-offset:2px;" '
                '@click.stop="() => $parent.$emit(\'cellClick\', {metro: props.row.metro, col: \'%s\'})">'
                '{{ props.value }}</span>'
                '<span v-else style="opacity:.45;">{{ props.value }}</span>'
                '</q-td>'
            ) % _cc)
        # NDR status as a colored badge: Scheduled (green) / Planning (blue) / Completed (grey) / — none.
        # When an NDR EXISTS for the metro, the badge is a shortcut STRAIGHT TO that NDR (Active NDRs) —
        # @click.stop so it doesn't also fire the row-click's "add names to an NDR" dialog. A metro with
        # no NDR yet (—) still opens that dialog via the row click, to start one.
        geo_table.add_slot("body-cell-read", (
            '<q-td :props="props">'
            '<q-badge v-if="props.value===\'Scheduled\'" color="green" label="Scheduled" '
            'class="cursor-pointer" @click.stop="() => $parent.$emit(\'openndr\', props.row.metro)"/>'
            '<q-badge v-else-if="props.value===\'Planning\'" color="blue" label="Planning" '
            'class="cursor-pointer" @click.stop="() => $parent.$emit(\'openndr\', props.row.metro)"/>'
            '<q-badge v-else-if="props.value===\'Completed\'" outline color="grey" label="Completed" '
            'class="cursor-pointer" @click.stop="() => $parent.$emit(\'openndr\', props.row.metro)"/>'
            '<span v-else style="opacity:.4;">—</span></q-td>'))
        geo_table.on("cellClick", _open_cell_drill)
        geo_table.on("openndr", lambda e: nav.go_to("Roadshow", "NDR"))
        geo_table.on("rowClick", lambda e: _open_metro_select_dialog(
            e.args[1]["metro"], peer_funds_by_metro.get(e.args[1]["metro"], []),
            holders=metro_summary.get(e.args[1]["metro"], {}).get("holder_list", [])))
        _peer_total = sum(v["funds"] for v in peer_by_metro.values())
        # Fill the instruction moved up under the title (see _geo_note) now that the totals exist.
        _geo_note.set_text(
            f"{holder_count} current holders and {_peer_total} peer-owners (own a comp, not you) across "
            f"{len(_all_metros)} metros. Peer-owners break into Inst / RIA / Diversified / MM / Curated "
            "(they sum to Peer-owners).")
        _manage_btn = ui.button("Manage fund-lineup matches", icon="account_tree",
                                on_click=_open_lineup_crosswalk_dialog).props("flat dense").style("font-size:var(--fs-xs);margin-top:2px;")
        # Tuck the metro table + its Manage button inside the collapsible header created above (they
        # render at top level because the table build sits between title and table; move() preserves
        # all their slots and row/cell event handlers).
        geo_table.move(_geo_table_exp)
        _manage_btn.move(_geo_table_exp)

    return _geo_table_exp


def _peer_holder_avg(cid):
    """Average number of institutional 13F holders across the client's tight comps — turns a bare
    holder count into a judgement (better- or worse-owned than peers). None when no comp 13F is
    cached (real clients before a peer 13F refresh), so the card just omits the comparison."""
    try:
        from core import peer_prospects as _pp, sec_filings as _sf
        counts = [len(_sf.get_cached_13f_holders(tk).get("holders", []) or [])
                  for tk in (_pp.tight_comps(cid) or set())]
        counts = [c for c in counts if c > 0]
        return round(sum(counts) / len(counts)) if counts else None
    except Exception:
        return None


def _tier1_ready_names(institutions, client_id):
    """Canonical 'Tier-1 ready' non-holder set — ONE definition so the count matches on every surface
    (the Ownership summary card, the geography table, and the Target Database). A non-holder is Tier-1
    if its engagement score >= 80 OR it clears the conviction model at >= 70 — RAW conviction, not the
    rounded fit-score (a 69.7 that displays as '70' must NOT flip it into Tier-1; that rounding was why
    the card read 13 while the Target DB read 14). Returns a set of UPPER-cased fund names."""
    def _n(s):
        return (s or "").strip().upper()
    names = {_n(i.get("Fund")) for i in institutions
             if not i.get("USIO_Holder") and (i.get("Engagement_Score") or 0) >= 80}
    try:
        from core import peer_prospects
        names |= {_n(c.get("filer")) for c in (peer_prospects.build_candidates(client_id, limit=None) or [])
                  if (c.get("conviction") or 0) >= 70}
    except Exception:
        pass
    names.discard("")
    return names


def _peer_holder_sub(holder_count, cid):
    """Subtitle for the 'Who owns you' card — 'vs peer avg N · above/below/in line', or honest context
    when the comparison would mislead. '' if no comp data."""
    pa = _peer_holder_avg(cid)
    if not pa:
        return ""
    # A raw holder-COUNT comparison is only meaningful among size-comparable companies. When the client
    # sits well below the peer average — a micro-cap benchmarked against much larger comps (USIO's 28
    # vs a 187 peer avg dominated by names 10-80x its size) — "below peers" is trivially true and
    # misleading, so state the honest metric instead of a false judgement.
    if holder_count < 0.6 * pa:
        return ""
    rel = "above peers" if holder_count > pa else ("below peers" if holder_count < pa else "in line with peers")
    return f"vs peer avg {pa} · {rel}"


def _bp_metric(label, value, detail_lines, sub=None, on_click=None):
    # A clickable box in the same style as the colored-header cards below: a defined header BAND, then
    # the number and the analysis. When `on_click` is given the card deep-links (e.g. straight to the
    # Target Database filtered to this bucket — the exact list layout). Otherwise it pulls up its
    # detail in a dialog styled to match the Target Database drill-down (bold title + close X).
    handler = on_click
    if handler is None:
        _dlg = ui.dialog()
        with _dlg, ui.card().style("min-width:min(560px,94vw);max-width:640px;"):
            with ui.row().classes("w-full justify-between items-center"):
                ui.label(label).classes("text-lg font-bold")
                ui.button(icon="close", on_click=_dlg.close).props("flat round dense")
            for line in (detail_lines or ["No further detail available."]):
                ui.label(str(line)).style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);line-height:1.5;")
        handler = _dlg.open
    card = ui.card().classes("flex-1 report-card cursor-pointer")
    with card:
        with ui.row().classes("rhead rhead-accent"):
            ui.label(label)                          # the defined, colored header band
        with ui.column().classes("w-full").style("padding:12px 15px;gap:2px;"):
            ui.label(value).classes("rkpi-n")        # the number
            if sub:
                ui.label(sub).classes("rkpi-sub")    # the analysis
    card.on("click", handler)


# ─────────────────────────────────────────────────────────────────────────
# Accounts (CRM) tab — the relationship book, the Account 360 destination
# ─────────────────────────────────────────────────────────────────────────
def _acct_cadence(last_contact):
    """Days-since-last-contact bucket, from the DERIVED last_contact date."""
    if not last_contact:
        return "No contact"
    try:
        n = (datetime.now().date() - datetime.strptime(last_contact[:10], "%Y-%m-%d").date()).days
    except Exception:
        return "No contact"
    return "Active" if n <= 30 else "Cooling" if n <= 90 else "Gone quiet"


def _size_calibration_note(client_id):
    """Onboarding-critical: show the market-cap size tier the whole peer/CRM book is
    calibrated to — and warn loudly if no market cap is on record (which silently
    defaults to the strictest micro-cap targeting, wrong for a mid/large issuer)."""
    from core import peer_prospects
    p = peer_prospects.size_profile(client_id)
    mc = p.get("market_cap_m") or 0
    cap_str = (f"${mc/1000:.1f}B" if mc >= 1000 else f"${mc:,.0f}M") if mc else "—"
    tier = p["tier"]
    div = ("large diversified / bank-AM houses are PRIMARY targets"
           if p["diversified_are_targets"] else
           "large diversified / index houses are a review bucket, not primary targets")
    if not p.get("has_market_cap"):
        ui.label(f"⚠ No market cap on record for {CT('name')} — peer targeting is defaulting to the strictest "
                 "micro-cap profile. Set the market cap at onboarding so the book is calibrated to the issuer's size.").style(
            f"background:#FBEFE1;color:{COLORS['warning']};border:1px solid {COLORS['warning']};"
            "border-radius:8px;padding:6px 10px;font-size:var(--fs-sm);font-weight:600;margin:6px 0;")
        return
    ui.label(f"Sized as a {tier}-cap issuer ({cap_str}) — peer targeting is calibrated to this: {div}. "
             "Set at onboarding; a different size band (e.g. a mid-cap) reshapes the whole book.").style(
        f"background:{COLORS['surface_hover_bg']};color:{COLORS['text_secondary']};border-radius:8px;"
        "padding:6px 10px;font-size:var(--fs-sm);margin:6px 0;")


def _render_accounts_tab(client_id, institutions=None):
    from core import account_api, accounts, interactions
    ui.label("Accounts — your relationship book").classes("text-lg font-bold").style(
        f"color:{COLORS['text_heading']};")
    ui.label(f"Your {CT('ticker')} holders and every firm you've noted or logged an interaction with. Quality and "
             "notes are yours; touches and last contact are computed from logged + NDR activity. Click a name for the full 360.").style(
        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
    _size_calibration_note(client_id)

    # Start from the CRM book (noted/logged/seeded), then fold in the current holder book
    # and the peer-owner universe so your whole investable relationship set is in the CRM
    # out of the box. Both are joined LIVE off the 13F (re-scored each visit) so
    # holder/peer status never goes stale. One bulk interaction index avoids per-account
    # scans across the ~hundreds of peer-owners.
    _ixidx = interactions.active_index(client_id)

    def _fold(nm, cik, flag, tier=None):
        if not nm:
            return
        key = accounts.resolve(nm, cik=cik, register=False)
        if not key:
            return
        a = by_key.get(key)
        if a:
            a[flag] = True
            a["open_name"] = a.get("open_name") or nm
            if tier and not a.get("tier"):
                a["tier"] = tier
            if not a.get("cik") and cik:
                a["cik"] = accounts._cik_int(cik)
        else:
            s = _ixidx.get(key, {})
            by_key[key] = {"key": key, "name": pretty_name(nm), "open_name": nm,
                           "cik": accounts._cik_int(cik), "quality": None, "note": None,
                           "touches": s.get("touches", 0), "last_contact": s.get("last_contact"),
                           flag: True, "tier": tier}

    by_key = {a["key"]: a for a in account_api.list_accounts(client_id)}
    for inst in (institutions or []):
        if inst.get("USIO_Holder"):
            _fold(inst["Fund"], inst.get("cik"), "holder")
    try:
        from core import peer_prospects
        for c in peer_prospects.all_candidates(client_id):
            _fold(c.get("filer"), c.get("cik"), "peer", c.get("tier"))
    except Exception:
        pass

    def _rank(a):
        return 2 if a.get("holder") else 1 if (a["quality"] or a["touches"]) else 0
    all_accts = sorted(by_key.values(),
                       key=lambda a: (_rank(a), a.get("last_contact") or "", a["touches"]),
                       reverse=True)

    n_holders = sum(1 for a in all_accts if a.get("holder"))
    n_peers = sum(1 for a in all_accts if a.get("peer") and not a.get("holder"))
    good = sum(1 for a in all_accts if a["quality"] in ("good", "responsive"))
    with ui.row().classes("gap-6").style("margin:8px 0;"):
        for lbl, val in [("Accounts", len(all_accts)), (f"{CT('ticker')} holders", n_holders),
                         ("Peer-owners", n_peers), ("Good to deal with", good)]:
            with ui.column().classes("gap-0"):
                ui.label(str(val)).classes("text-lg font-bold").style(f"color:{COLORS['text_heading']};")
                ui.label(lbl).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

    with ui.row().classes("w-full items-end gap-2").style("margin-top:4px;"):
        search = ui.input("Search name").props("dense outlined clearable").style("min-width:190px;")
        typ = ui.select({"": "All accounts", "holder": "Holders", "peer": "Peer-owners",
                         "rel": "Has notes / activity"},
                        value="").props("dense outlined").style("min-width:160px;")
        tier = ui.select({"": "All tiers", "Institutional": "Institutional", "RIA / wealth": "RIA / wealth",
                          "Diversified": "Diversified", "Market maker": "Market maker", "Curated": "Curated"},
                         value="").props("dense outlined").style("min-width:150px;")
        qual = ui.select({"": "All quality", **account_api.QUALITY, "__unset__": "— unset —"},
                         value="").props("dense outlined").style("min-width:160px;")
        cad = ui.select({"": "All cadence", "Active": "Active", "Cooling": "Cooling",
                         "Gone quiet": "Gone quiet", "No contact": "No contact"},
                        value="").props("dense outlined").style("min-width:150px;")

    @ui.refreshable
    def _list():
        rows, by_name = [], {}
        term = (search.value or "").strip().lower()
        for a in all_accts:
            if term and term not in (a["name"] or "").lower():
                continue
            if typ.value == "holder" and not a.get("holder"):
                continue
            if typ.value == "peer" and not (a.get("peer") and not a.get("holder")):
                continue
            if typ.value == "rel" and not (a["quality"] or a["touches"]):
                continue
            if tier.value and a.get("tier") != tier.value:   # tier filter → peer-owners of that bucket
                continue
            if qual.value == "__unset__" and a["quality"]:
                continue
            if qual.value and qual.value != "__unset__" and a["quality"] != qual.value:
                continue
            cd = _acct_cadence(a["last_contact"])
            if cad.value and cd != cad.value:
                continue
            seg = "Holder" if a.get("holder") else "Peer-owner" if a.get("peer") else ""
            rows.append({"Name": a["name"], "Segment": seg, "Tier": a.get("tier") or "—",
                         "Quality": account_api.QUALITY.get(a["quality"], "—") if a["quality"] else "—",
                         "Touches": a["touches"], "Last contact": a["last_contact"] or "—", "Cadence": cd})
            by_name[a["name"]] = a
        if not rows:
            ui.label("No accounts match — your book holds your holders, the peer-owner universe, and "
                     "anything you note or log.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            return
        cols = [{"name": k, "label": k, "field": k, "sortable": True,
                 "align": "right" if k == "Touches" else "left"}
                for k in ("Name", "Segment", "Tier", "Quality", "Touches", "Last contact", "Cadence")]
        tbl = ui.table(columns=cols, rows=rows, row_key="Name", pagination=25).classes(
            "w-full cursor-pointer").props("dense flat")
        _header_tooltips(tbl, {
            "Name": "The firm — click for its Account 360",
            "Segment": "Holder (owns you) · Peer-owner (owns a comp, not you) · — (your own note)",
            "Tier": "Peer-owner bucket (Institutional / RIA / Diversified / Market maker / Curated)",
            "Quality": "Your relationship read — good to deal with, responsive, neutral, hard to reach, low-touch",
            "Touches": "Interactions logged (calls, meetings, NDR touches) — computed, not typed",
            "Last contact": "Most recent interaction date — computed from the event log",
            "Cadence": "Active ≤30d · Cooling ≤90d · Gone quiet >90d · No contact",
        })
        tbl.add_slot("body-cell-Name", (
            '<q-td :props="props"><span class="cursor-pointer" '
            f'style="color:{COLORS["accent"]};text-decoration:underline dotted;text-underline-offset:2px;" '
            '@click.stop="() => $parent.$emit(\'openAcct\', props.row.Name)">{{ props.value }}</span></q-td>'))
        tbl.add_slot("body-cell-Segment", (
            '<q-td :props="props"><q-badge v-if="props.value===\'Holder\'" color="blue-grey" label="Holder"/>'
            '<q-badge v-else-if="props.value===\'Peer-owner\'" outline color="teal" label="Peer-owner"/>'
            '<span v-else style="opacity:.4;">—</span></q-td>'))
        tbl.add_slot("body-cell-Cadence", (
            '<q-td :props="props"><q-badge outline '
            ':color="props.value===\'Gone quiet\'?\'red\':props.value===\'Cooling\'?\'orange\':'
            'props.value===\'Active\'?\'green\':\'grey\'" :label="props.value"/></q-td>'))

        def _open(e):
            a = by_name.get(e.args)
            if a:
                _open_account_profile({"Fund": a.get("open_name") or a["name"], "cik": a.get("cik"),
                                       "USIO_Holder": a.get("holder", False)})
        tbl.on("openAcct", _open)

    for w in (search, typ, tier, qual, cad):
        w.on_value_change(lambda *_: _list.refresh())
    _list()


# ─────────────────────────────────────────────────────────────────────────
# Buy-Side Intelligence tab
# ─────────────────────────────────────────────────────────────────────────
def _render_buyside_tab(institutions, meeting_log, mode):
    contacts = get_institution_contacts()

    if mode == "pre":
        ui.label("Buyside Ownership — Pre-Earnings Engagement").classes("text-lg font-bold")
        ui.label("Existing holders to defend · non-holder targets to convert · ranked by conviction "
                 "(likelihood to buy) and position (size, trajectory).").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
    else:
        ui.label("Buyside Ownership — Post-Earnings Prospecting").classes("text-lg font-bold")
        ui.label("Post-earnings: who heard the beat and fits the upgrade thesis · holders to defend, "
                 "targets to convert, ranked by conviction and position.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    # Split by the core buyside question — existing HOLDERS to defend vs NON-HOLDER targets to
    # convert — and rank on the SAME model as the metro selector: Position for holders, Conviction
    # for targets. Replaces the opaque "Engagement Funnel" (Tier 1/2/3 by an unexplained 0-100
    # score that collided with Coverage Priority and the conviction model — the "3 confusing cards").
    _enrich_holders_vs_comps(institutions, get_active_client_id())
    _holders = [i for i in institutions if i.get("USIO_Holder")]
    _dirl = lambda i: (i.get("Direction") or "").lower()
    _adding = [i for i in _holders if _dirl(i) in ("adding", "new")]
    _trimming = [i for i in _holders if _dirl(i) in ("trimming", "exited")]
    _top_holders = sorted(_holders, key=lambda x: -(x.get("Position_Value") or 0))
    # "Who should own you" = the conviction-ranked peer-owner universe (the SAME model the metro
    # selector / NDR planner use), NOT the buyside engagement score — so the conviction numbers and
    # the Tier-1 count match everywhere on the platform (the whole point of the unify-the-score fix).
    try:
        from core import peer_prospects as _pp
        _targets = _pp.build_candidates(get_active_client_id(), limit=None) or []
    except Exception:
        _targets = []
    _tier1 = [c for c in _targets if (c.get("conviction") or 0) >= 70]   # Tier-1 conviction (metro-table def)
    _top_targets = sorted(_targets, key=lambda c: -(c.get("conviction") or 0))
    # Unified non-holder count = tracked non-holders + peer-owners (deduped) — matches the split
    # Non-Holders list below, so the card, read, and list all show the same number.
    _tracked_nonh = [i for i in institutions if not i.get("USIO_Holder")]
    _seen_t = {_norm_name(c.get("filer", "")) for c in _targets}
    _targets_count = len(_targets) + len([i for i in _tracked_nonh if _norm_name(i["Fund"]) not in _seen_t])

    # THE BUYSIDE READ — plain-English BLUF (same pattern as NOBO / the Big Picture read).
    _read = [f"{len(_holders)} holder(s) to defend and {_targets_count} non-holder target(s) to convert — "
             f"{len(_tier1)} at Tier-1 conviction, ready for a 1×1 this week."]
    if _adding or _trimming:
        _read.append(f"Holder momentum this 13F cycle: {len(_adding)} adding/new vs {len(_trimming)} trimming/exiting.")
    if _top_targets[:3]:
        _read.append("Call first: " + ", ".join(pretty_name(c.get("filer", "")) for c in _top_targets[:3]) + ".")
    with ui.card().classes("w-full").style(
            "background:rgba(30,64,175,.06);border:1.5px solid #1E40AF;border-left:6px solid #1E40AF;"
            "border-radius:8px;margin-top:8px;"):
        ui.label("THE BUYSIDE READ — who owns you, who should").style(
            "color:#1E3A8A;font-size:var(--fs-xs);font-weight:700;letter-spacing:.04em;")
        for _p in _read:
            ui.label("• " + _p).style(
                f"color:{COLORS['text_heading']};font-size:var(--fs-base);line-height:1.6;font-weight:500;")

    # Composition — the four questions (replaces the Tier 1/2/3 funnel), same tile style as NOBO/Big
    # Picture. Each card self-identifies via its eyebrow header, so no redundant subheader row.
    with ui.row().classes("w-full gap-3").style("margin-top:10px;"):
        _bp_metric("Current Investors", str(len(_holders)),
                   [f"{pretty_name(i['Fund'])} — {_holder_dollars(i.get('Position_Value'))}"
                    + (f" · {_dir_label(i.get('Direction'))}" if _dir_label(i.get('Direction')) else "")
                    for i in _top_holders[:6]] or ["No tracked holders yet."],
                   sub=_peer_holder_sub(len(_holders), get_active_client_id()))
        _bp_metric("Who should own you", f"{len(_tier1)} / {_targets_count}",
                   [f"{pretty_name(c.get('filer', ''))} — conviction {round(c.get('conviction') or 0)}"
                    for c in _top_targets[:6]] or ["No conviction targets yet."])
        _bp_metric("Adding / New (holders)", str(len(_adding)),
                   [pretty_name(i['Fund']) for i in _adding[:6]] or ["None adding this cycle."])
        _bp_metric("Trimming / Exited (holders)", str(len(_trimming)),
                   [pretty_name(i['Fund']) for i in _trimming[:6]] or ["None trimming this cycle."])

    ui.markdown("---")

    # One left-aligned summary line — "Q1 call listeners" reads as a
    # continuation of the tracked count rather than floating off to the far
    # right (justify-between) disconnected from anything.
    with ui.row().classes("items-baseline gap-2"):
        ui.label(f"{len(institutions)} institutions tracked").classes("font-bold")
        ui.label(f"· {sum(1 for i in institutions if i['Call_Listener'])} were Q1 call listeners").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    with ui.expansion("Filter Criteria — Coverage Priority, Holder Status, Turnover, Metro, Intent, Ownership",
                      value=False, icon="tune").classes("w-full").style(
            f"border:1px solid {COLORS['border']};border-radius:8px;background:{COLORS['surface_bg']};"
            f"color:{COLORS['text_heading']};font-weight:600;"):
        ui.label("These narrow the institution list below — they don't affect the Engagement Funnel or metrics "
                 "above, which always show the complete tracked set. All defaults below are unfiltered (everything "
                 "shown), so the list you see with no filters applied matches the same ranked order Today's "
                 "Investor Pipeline widget pulls from — narrow it from here if you want a subset.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
        with ui.row().classes("w-full gap-4"):
            # Coverage Priority is a manually-assigned seed classification
            # (data/seed/buyside_institutions.py's "Coverage_Priority" field,
            # renamed from "Tier") — a different concept from the live
            # Engagement_Score-derived Tier 1/2/3 buckets in the Engagement
            # Funnel above, which recompute from real signals every render.
            # Was confusable with those (both called "Tier") and previously
            # defaulted to [1, 2] here, silently hiding every Coverage
            # Priority 3 fund regardless of its live Engagement_Score —
            # defaulting to all three so nothing is hidden until a user
            # deliberately narrows it.
            tier_filter = ui.select({1: "Priority 1", 2: "Priority 2", 3: "Priority 3"}, multiple=True, value=[1, 2, 3]).classes("min-w-[140px]").props("outlined dense label='Coverage priority'")
            holder_filter = ui.select(["All", "Current holders", "Non-holders only"], value="All").classes("min-w-[160px]").props("outlined dense label='Holder status'")
            score_filter = ui.number(label="Min score", value=0, min=0, max=100).props("outlined dense")
            # Coerce to str: a record with Metro=None makes sorted() raise
            # TypeError('<' not supported between NoneType and str) and takes the whole
            # page down. Defensive here as well as at the source, because this list is fed
            # by several record shapes (13F holders, promoted prospects, SEC universe).
            metro_options = ["All Regions"] + sorted({(i.get("Metro") or "Unknown (SEC)")
                                                      for i in institutions})
            metro_filter = ui.select(metro_options, value="All Regions").classes("min-w-[160px]").props("outlined dense label='Metro'")
        turnover_options = ["Low (Long-Term Value)", "Medium (Growth/GARP)", "High (Hedge/Trading)"]
        # Include any other turnover value actually present in the universe
        # (e.g. "Unknown (SEC)" on real SEC-sourced names) so the default
        # all-selected filter doesn't silently hide them — same silent-hiding
        # issue as the tier filter above.
        turnover_options = turnover_options + sorted(
            {i["Turnover_Style"] for i in institutions} - set(turnover_options))
        turnover_filter = ui.select(turnover_options, multiple=True,
                                     value=list(turnover_options)).classes("w-full").props("outlined dense label='Fund Turnover Style'")
        # Previously defaulted to 40, hiding any fund with Digital_Intent_Score
        # below that even if its Engagement_Score ranked it top-5 — that score
        # is an intentionally separate, unrelated metric (see
        # core/investor_scoring.py), not something that should gate visibility
        # by default.
        intent_filter = ui.number(label="Minimum Digital Intent Score (Web Traffic & Call Replays)", value=0, min=0, max=100, step=10).classes("w-full").props("outlined dense")
        ownership_filter = ui.toggle(["All", "Active only", "Passive only"], value="All")
        ui.label("Ownership Type is tagged per manager (not parsed from the 13F itself — filings don't carry an "
                 "active/passive flag). Excluding passive/index holders surfaces who can actually take a 1x1 meeting.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
        filter_btn = ui.button("Apply filters")

    ui.markdown("---")
    banner_container = ui.column().classes("w-full")
    ui.label("Full Institution List").classes("font-bold")

    # ── Add-to-NDR bar (NDR pipeline Phase 4B) — the Buy-Side list feeds the same
    # pipeline as Peer Prospects. Filter to the set you want (metro, holder, score),
    # tick institutions, and shortlist them onto an NDR.
    bs_checks = {}
    _bs_open = [(i, t) for i, t in enumerate(_load_json("ndr_trips.json", [])) if t.get("status") != "Completed"]
    _bs_opts = {str(i): (t.get("name") or f"NDR {i+1}") for i, t in _bs_open}
    _bs_opts["__new__"] = "＋ New NDR…"
    with ui.row().classes("w-full items-end gap-2").style(
            f"background:{COLORS['surface_hover_bg']};border-radius:8px;padding:8px 10px;margin:6px 0;"):
        ui.label("Tick institutions below, then add them to an NDR as shortlisted targets.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);flex:1;")
        _bs_sel = ui.select(_bs_opts, value="__new__", label="NDR").props("dense outlined").style("min-width:170px;")
        _bs_name = ui.input("New NDR name", value="Buy-Side NDR").props("dense outlined").style("min-width:150px;")
        _bs_name.bind_visibility_from(_bs_sel, "value", backward=lambda v: v == "__new__")

        def _bs_add():
            picked = [(f, inst) for f, (cb, inst) in bs_checks.items() if cb.value]
            if not picked:
                ui.notify("Tick at least one institution first.", type="warning"); return
            trips2 = _load_json("ndr_trips.json", [])
            if _bs_sel.value == "__new__":
                nm = (_bs_name.value or "").strip()
                if not nm:
                    ui.notify("Name the new NDR.", type="warning"); return
                trips2.append({
                    "name": nm, "sponsor_bank": "", "dates": "TBD", "ndr_type": "in-person",
                    "city": "Multiple", "focus": "", "team": [], "notes": "", "meetings": [],
                    "shortlist": [], "status": "Planning", "debrief": {}, "days": 2, "slots_per_day": 6,
                    "created": datetime.now().strftime("%Y-%m-%d"),
                })
                target, tname = trips2[-1], nm
            else:
                target = trips2[int(_bs_sel.value)]
                tname = target.get("name") or "NDR"
            target.setdefault("shortlist", [])
            have = {s.get("institution") for s in target["shortlist"]} | \
                   {m.get("institution") for m in target.get("meetings", [])}
            added = 0
            for f, inst in picked:
                if f in have:
                    continue
                target["shortlist"].append(
                    _shortlist_record(inst) if inst.get("filer") else _shortlist_from_inst(inst))
                have.add(f); added += 1
            _save_json("ndr_trips.json", trips2)
            skipped = len(picked) - added
            ui.notify(f"Shortlisted {added} to '{tname}'"
                      + (f" · {skipped} already on it" if skipped else "")
                      + ". See Outbound → NDR → Active NDRs.", type="positive")
        ui.button("Add selected", icon="playlist_add", on_click=_bs_add).props("dense color=primary")

    list_container = ui.column().classes("w-full gap-3")

    # Per-group "sort by city" toggle (the little arrow in each group header). False = default order
    # (holders by position, non-holders by conviction); True = grouped alphabetically by metro.
    _sortcity = {"h": False, "n": False}

    def apply_and_render():
        banner_container.clear()
        list_container.clear()
        bs_checks.clear()

        def _toggle_city(group):
            _sortcity[group] = not _sortcity[group]
            apply_and_render()
        sel_tiers = tier_filter.value or [1, 2, 3]
        sel_turnover = turnover_filter.value or turnover_options

        def _pass(i):
            return (i["Coverage_Priority"] in sel_tiers
                    and (holder_filter.value == "All"
                         or (holder_filter.value == "Current holders" and i["USIO_Holder"])
                         or (holder_filter.value == "Non-holders only" and not i["USIO_Holder"]))
                    and _score_val(i) >= (score_filter.value or 0)
                    and (metro_filter.value == "All Regions" or i["Metro"] == metro_filter.value)
                    and i["Turnover_Style"] in sel_turnover
                    and _score_val(i, "Digital_Intent_Score") >= (intent_filter.value or 0)
                    and (ownership_filter.value == "All"
                         or (ownership_filter.value == "Active only" and i["Ownership_Style"] == "Active")
                         or (ownership_filter.value == "Passive only" and i["Ownership_Style"] == "Passive")))

        filtered = [i for i in institutions if _pass(i)]
        holders_f = sorted([i for i in filtered if i["USIO_Holder"]],
                           key=lambda x: -(x.get("Position_Value") or 0))
        nonh = [i for i in filtered if not i["USIO_Holder"]]

        # Fold the conviction targets (peer-owners) into the NON-HOLDER group — the SAME universe as
        # the metro table / the "who should own you" card — so the metro-identified names appear HERE
        # (ranked by conviction), not only on Peer Prospects. This is what made the two disagree.
        _seen = {_norm_name(i["Fund"]) for i in nonh}
        if holder_filter.value != "Current holders":
            try:
                from core import peer_prospects as _pp
                _cands = _pp.build_candidates(get_active_client_id(), limit=None) or []
            except Exception:
                _cands = []
            for c in _cands:
                _m = _metro_from_city(c.get("city"), c.get("state"))
                if metro_filter.value != "All Regions" and _m != metro_filter.value:
                    continue
                if _norm_name(c.get("filer", "")) in _seen:
                    continue
                nonh.append(c); _seen.add(_norm_name(c.get("filer", "")))
        _cscore = lambda x: (x.get("conviction") if x.get("filer") else _score_val(x)) or 0
        nonh.sort(key=lambda x: -_cscore(x))

        # "Sort by city" toggle overrides the default order: group by metro, then keep the default
        # ranking within each city. Non-holders can be tracked institutions OR peer-owner candidates.
        def _row_metro(x):
            if x.get("filer"):
                return _metro_from_city(x.get("city"), x.get("state")) or "~"
            return x.get("Metro") or "~"
        if _sortcity["h"]:
            holders_f.sort(key=lambda x: ((x.get("Metro") or "~"), -(x.get("Position_Value") or 0)))
        if _sortcity["n"]:
            nonh.sort(key=lambda x: (_row_metro(x), -_cscore(x)))

        with banner_container:
            _render_repeat_alert_banner(holders_f, meeting_log)

        with list_container:
            # Holders and Non-Holders sit SIDE BY SIDE (two columns) on a wide screen and stack on a
            # narrow one — halves this section's height vs stacking, and each has its own inner scroll
            # so neither pushes the page into an endless scroll.
            with ui.row().classes("w-full gap-3 items-start").style("flex-wrap:wrap;"):
                # ── Current Holders — who owns you (rich cards, by position size) ──
                with ui.column().style("flex:1 1 400px;min-width:0;"):
                    with ui.expansion(f"Current Holders — who owns you ({len(holders_f)})",
                                      value=True, icon="account_balance").classes("w-full").style(
                            f"border:1px solid {COLORS['border']};border-radius:8px;"
                            f"color:{COLORS['text_heading']};font-weight:600;"):
                        with ui.row().classes("w-full items-center justify-end").style("margin:-4px 0 2px;"):
                            ui.button("City", icon="sort_by_alpha", on_click=lambda: _toggle_city("h")).props(
                                "flat dense size=sm").style(
                                f"color:{COLORS['accent'] if _sortcity['h'] else COLORS['text_muted']};").tooltip(
                                "Sort by city (click again to restore position order)")
                        if not holders_f:
                            ui.label("No current holders match these filters.").style(f"color:{COLORS['text_muted']};")
                        with ui.column().classes("w-full gap-2").style(
                                "max-height:560px;overflow-y:auto;padding-right:6px;"):
                            for inst in holders_f:
                                with ui.row().classes("w-full items-start no-wrap gap-2"):
                                    bs_checks[inst["Fund"]] = (
                                        ui.checkbox().props("dense").style("margin-top:12px;flex:0 0 auto;"), inst)
                                    with ui.column().classes("flex-1").style("min-width:0;"):
                                        _institution_card(inst, meeting_log, contacts)

                # ── Non-Holders — who should own you (conviction targets, incl. the metro-identified peer-owners) ──
                with ui.column().style("flex:1 1 400px;min-width:0;"):
                    with ui.expansion(f"Non-Holders — who should own you ({len(nonh)})",
                                      value=True, icon="person_search").classes("w-full").style(
                            f"border:1px solid {COLORS['border']};border-radius:8px;"
                            f"color:{COLORS['text_heading']};font-weight:600;"):
                        with ui.row().classes("w-full items-center justify-end").style("margin:-4px 0 2px;"):
                            ui.button("City", icon="sort_by_alpha", on_click=lambda: _toggle_city("n")).props(
                                "flat dense size=sm").style(
                                f"color:{COLORS['accent'] if _sortcity['n'] else COLORS['text_muted']};").tooltip(
                                "Sort by city (click again to restore conviction order)")
                        if not nonh:
                            ui.label("No non-holder targets match these filters.").style(f"color:{COLORS['text_muted']};")
                        # Same rich card as the holders — a peer-owner candidate is adapted into the
                        # institution shape first (_candidate_to_inst); a tracked non-holder is already
                        # one. The checkbox stores the ORIGINAL record so _bs_add routes it correctly.
                        with ui.column().classes("w-full gap-2").style(
                                "max-height:560px;overflow-y:auto;padding-right:6px;"):
                            for x in nonh:
                                inst = _candidate_to_inst(x) if x.get("filer") else x
                                with ui.row().classes("w-full items-start no-wrap gap-2"):
                                    bs_checks[inst["Fund"]] = (
                                        ui.checkbox().props("dense").style("margin-top:12px;flex:0 0 auto;"), x)
                                    with ui.column().classes("flex-1").style("min-width:0;"):
                                        _institution_card(inst, meeting_log, contacts)

    filter_btn.on_click(apply_and_render)
    apply_and_render()


def _render_repeat_alert_banner(filtered, meeting_log):
    """Global 'Repeat Meeting Alerts' banner above the institution list —
    surfaces every institution in the current filter set with a repeat
    meeting inside the 45-day window, classified by QoQ direction and
    outcome tone, before the per-institution cards render."""
    alerts = []
    for inst in filtered:
        fund_meetings = _get_fund_meetings(meeting_log, inst["Fund"])
        gap = _get_repeat_signal(fund_meetings)
        if gap is not None:
            alerts.append({"inst": inst, "gap": gap, "meetings": fund_meetings})
    if not alerts:
        return

    ui.label("Repeat Meeting Alerts — Investigate Signal").classes("font-bold").style(f"color:{COLORS['text_heading']};")
    for alert in alerts:
        inst, gap, meetings = alert["inst"], alert["gap"], alert["meetings"]
        last_m, prev_m = meetings[0], meetings[1]
        outcome = last_m.get("Outcome", "")
        qoq = inst.get("QoQ_Change", 0)
        # Was: solid fixed-dark background per status (e.g. "#173321") plus
        # hardcoded light body text ("#E2E8F0"/"#94A3B8") — same leftover-
        # dark-theme pattern as _tier_card above, illegible/mismatched under
        # the active CREAM (light) theme. Now a light card with a colored
        # left border for the status, and theme-token text throughout.
        if qoq > 0:
            sig_icon, sig_txt, sig_clr = "", "Likely diligence — shares increasing", COLORS["positive"]
        elif "positive" in outcome.lower() or "interested" in outcome.lower():
            sig_icon, sig_txt, sig_clr = "", "Monitor — positive tone but verify intent", COLORS["warning"]
        elif qoq < 0:
            sig_icon, sig_txt, sig_clr = "", "Flag — shares declining, possible exit diligence", COLORS["negative"]
        else:
            sig_icon, sig_txt, sig_clr = "", "Unknown intent — escalate to CFO for next call", COLORS["warning"]

        ui.html(
            f"<div style='border:1px solid {COLORS['border']};border-left:4px solid {sig_clr};border-radius:8px;"
            f"padding:10px 14px;background:{COLORS['surface_bg']};margin-bottom:8px;'>"
            f"<div style='display:flex;justify-content:space-between;align-items:center;'>"
            f"<div style='color:{COLORS['text_heading']};'><b>{inst['Fund']}</b> requested a second meeting in <b>{gap} days</b> (within 45-day window)</div>"
            f"<span style='font-weight:600;color:{sig_clr};'>{sig_icon} {sig_txt}</span></div>"
            f"<div style='font-size:var(--fs-base);color:{COLORS['text_muted']};margin-top:4px;'>"
            f"First: {prev_m['Date']} ({prev_m['Type']}) · Last: {last_m['Date']} ({last_m['Type']}) · "
            f"Last outcome: {last_m.get('Outcome','—')}</div></div>"
        )
    ui.markdown("---")


def _tier_card(institutions, label, action, accent_color):
    # Was: solid fixed-dark background (e.g. "#173321") + a fixed light
    # text_color passed in per call site — designed for the old dark theme.
    # Under the now-active CREAM (light) theme, COLORS['text_heading'] /
    # COLORS['text_muted'] resolve to near-black, which is illegible against
    # those leftover dark literals — same root cause as the earlier Earnings
    # stage-card contrast bug. Fixed the same way _institution_card already
    # does it below: theme-aware light card background + a colored left
    # border for the tier accent, so every text color in here can just use
    # the normal theme tokens and stay correct under either theme.
    with ui.card().classes("flex-1 column").style(
        f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};border-left:4px solid {accent_color};"):
        with ui.row().classes("w-full justify-between items-start"):
            with ui.column().classes("gap-0"):
                ui.label(label).style(f"color:{accent_color};font-weight:bold;")
                ui.label(action).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            ui.label(str(len(institutions))).classes("text-2xl font-bold").style(f"color:{COLORS['text_heading']};")
        ui.space()
        # Always render the disclosure — even when empty — so all three tier
        # cards share the same structure and height and each gets an arrow
        # (Tier 1 with 0 institutions used to have none, so it sat shorter).
        with ui.expansion(f"Show {len(institutions)} institution(s)").classes("w-full"):
            if institutions:
                for inst in institutions:
                    ui.label(f"{pretty_name(inst['Fund'])} — {inst['Engagement_Score']}/100").style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
            else:
                ui.label("No institutions currently in this tier.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")


def _fmt_aum(v):
    """Compact dollar label for a book total ($1.1B / $840M / $12,500)."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "—"
    if v >= 1e9:
        return f"${v / 1e9:.1f}B"
    if v >= 1e6:
        return f"${v / 1e6:.0f}M"
    return f"${v:,.0f}"


def _candidate_to_inst(c):
    """Adapt a peer_prospects.build_candidates record (a peer-owner we don't hold yet) into the
    institution shape _institution_card renders, so the Non-Holders group looks IDENTICAL to the
    Current Holders group instead of a separate light row. The card's score slot carries CONVICTION
    here (our one non-holder score); the holdings-of-us fields (Shares / QoQ) are blank because a
    non-holder holds none of us yet, and 'Peers held' carries the comps that make them a prospect.
    Signal fields we never measured (call-listener, IR visits) stay None so the card reads
    'no data', never a fabricated zero. Keeps the raw candidate identity (filer/comps/city/state)
    so the Add-to-NDR checkbox still routes through _shortlist_record."""
    nm = c.get("filer") or c.get("Fund") or "—"
    comps = sorted((c.get("comps") or {}).keys())
    out = {
        "Fund": nm,
        "Engagement_Score": round(c.get("conviction") or 0),
        "USIO_Holder": False,
        "QoQ_Change": 0, "Shares": 0,
        "Type": "Institution", "AUM": _fmt_aum(c.get("book_total")),
        "Metro": _metro_from_city(c.get("city"), c.get("state")) or "—",
        "Turnover_Style": "—",
        "Action": "Prospect — owns your comp, not you yet",
        "Ownership_Style": "Active", "Source": "Peer 13F",
        "cik": c.get("cik"),
        "Peer_Holdings": comps,
        "Peer_Holdings_Source": {t: "live" for t in comps},
        "Call_Listener": None, "Listen_Duration": None, "IR_Visits_30d": None, "Last_Visit": None,
        "filer": c.get("filer"), "comps": c.get("comps"),
        "city": c.get("city"), "state": c.get("state"), "conviction": c.get("conviction"),
    }
    # Same engagement fill as the tracked universe so non-holder cards don't read "no call/visit data".
    _enrich_engagement_signals([out], get_active_client_id())
    return out


def _list_item_card(rail_color=None):
    """The app's ONE standard shell for an actionable list item — a bounded grey tile
    (surface_hover_bg fill + the global `.q-card` border), matching `_institution_card`
    and the Today cards. Use this for EVERY list row (prospects, holders, suggestions,
    requests) instead of hand-rolling `border-bottom` line separators, so lists read the
    same everywhere. The border is what keeps it legible even inside a grey expander
    (where a borderless same-grey tile would blend). Optional `rail_color` adds a left
    attention rail (same cue Today uses for repeat meetings). Returns the card CM."""
    rail = f"border-left:3px solid {rail_color};" if rail_color else ""
    return ui.card().classes("w-full list-tile").style(rail)


def _institution_card(inst, meeting_log, contacts):
    fund_meetings = _get_fund_meetings(meeting_log, inst["Fund"])
    repeat_gap = _get_repeat_signal(fund_meetings)
    _tk = CT("ticker")
    score = inst["Engagement_Score"]
    holder_badge = "Current Holder" if inst["USIO_Holder"] else "Non-Holder"
    qoq_str = f"{inst['QoQ_Change']:+,}" if inst["QoQ_Change"] else "—"
    shares_str = f"{inst['Shares']:,}" if inst["Shares"] else "—"

    # Match the Today page cards (the app card standard): softer surface_hover_bg fill, no hard
    # border; a repeat-meeting alert shows as a left rail, the same "attention" cue Today uses.
    _card_rail = "border-left:3px solid #F9A825;" if repeat_gap else ""
    # Today's Investor Pipeline widget (top_engagement_targets) drops any
    # fund contacted within the last 7 days from its top-5, so it can look
    # "out of order" compared to this full list unless it's clear why a
    # given fund isn't showing up there — this badge makes that visible
    # here instead of the two pages silently disagreeing.
    days_contacted = _days_since_last_contact(inst["Fund"], meeting_log)
    from core import account_api
    _acct = account_api.get_account(name=inst["Fund"], cik=inst.get("cik")) or {}
    _rel, _ix = _acct.get("relationship", {}), _acct.get("interactions", {})
    with ui.card().classes("w-full").style(
            f"background:{COLORS['surface_hover_bg']};{_card_rail}padding:10px 14px;gap:3px;"):
        with ui.row().classes("w-full justify-between items-start"):
            with ui.column().classes("gap-0"):
                ownership_badge = "Passive" if inst.get("Ownership_Style") == "Passive" else "Active"
                # Join only the parts we actually have — real SEC-sourced holders often carry no
                # Type / Turnover_Style, and rendering the literal "None" between the dots ("· None ·")
                # looked broken. Drop any empty / None / placeholder segment instead.
                def _dots(*vals):
                    out = []
                    for v in vals:
                        s = "" if v is None else str(v).strip()
                        if s and s.lower() != "none" and s != "—":
                            out.append(s)
                    return "  ·  ".join(out)
                _aum = inst.get("AUM")
                _aum_seg = f"AUM {_aum}" if (_aum and str(_aum).strip().lower() not in ("none", "—", "")) else None
                _title_txt = _dots(pretty_name(inst["Fund"]), inst.get("Type"), _aum_seg)
                _meta_txt = _dots(inst.get("Metro"), inst.get("Turnover_Style"), ownership_badge, holder_badge)
                with ui.row().classes("items-center gap-2"):
                    ui.label(_title_txt).classes("font-bold name-link").on(
                        "click", lambda inst=inst: _open_account_profile(_card_rec(inst))).style(
                        f"color:{COLORS['text_heading']};font-size:var(--fs-lg);")
                    _src = inst.get("Source", "Seed (demo)")
                    ui.label(_src).style(
                        f"background:{_source_color(_src)};color:#fff;border-radius:6px;padding:1px 6px;"
                        "font-size:var(--fs-2xs);font-weight:700;letter-spacing:.02em;white-space:nowrap;")
                    _q = _rel.get("quality")
                    if _q:
                        ui.label(account_api.QUALITY.get(_q, _q)).style(
                            f"background:{COLORS['positive'] if _q in ('good', 'responsive') else COLORS['surface_hover_bg']};"
                            f"color:{'#fff' if _q in ('good', 'responsive') else COLORS['text_secondary']};"
                            "border-radius:6px;padding:1px 7px;font-size:var(--fs-2xs);font-weight:700;white-space:nowrap;")
                ui.label(_meta_txt).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                if _ix.get("touches"):
                    ui.label(f"{_ix['touches']} touch{'es' if _ix['touches'] != 1 else ''} · last {_ix.get('last_contact') or '—'}").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                if repeat_gap is not None:
                    ui.label(f"Repeat meeting in {repeat_gap}d").style(f"color:#B45309;font-size:var(--fs-sm);font-weight:bold;")
                if days_contacted is not None and days_contacted <= 7:
                    ui.label(f"Contacted {days_contacted}d ago — off Today's Investor Pipeline until day 8").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-style:italic;")
            with ui.column().classes("items-center gap-0"):
                ui.label(str(score)).classes("text-2xl font-bold").style(f"color:{COLORS['accent_light']};")
                with ui.row().classes("items-center gap-1 no-wrap"):
                    ui.label("/100").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                    ui.icon("info").classes("cursor-help").style(
                        f"font-size:var(--fs-base);color:{COLORS['text_muted']};").tooltip(
                        "Engagement score, 0-100 — blends who already knows the story (call-listener), "
                        "peer-holding overlap, and recent IR visits; the weights shift with Pre-/Post-earnings "
                        "mode. Higher = readier to engage.")

        with ui.row().classes("w-full gap-6").style("margin-top:6px;"):
            ui.label(f"Shares: {shares_str}").classes("cursor-help").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);").tooltip(f"Shares of {_tk} held (latest 13F)")
            ui.label(f"QoQ: {qoq_str}").classes("cursor-help").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);").tooltip(
                "Quarter-over-quarter change in shares held (+ adding · − trimming)")
            # None means UNKNOWN, not "no". There is no call-listener or website-analytics
            # integration, so rendering "Did not listen" / "None · None" for a real 13F holder
            # asserts a negative we never measured. Say we don't know.
            if inst.get("Call_Listener") is None:
                _call_str = "no call data"
            elif inst["Call_Listener"]:
                _call_str = inst.get("Listen_Duration") or "listened"
            else:
                _call_str = "Did not listen"
            _visits_str = ("no visit data" if inst.get("IR_Visits_30d") is None
                           else f"{inst['IR_Visits_30d']} · {inst.get('Last_Visit') or '—'}")
            ui.label(f"Q1 call: {_call_str}").classes("cursor-help").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);").tooltip(
                "Whether they listened to the last earnings call (call-listener signal). "
                "'no call data' = we didn't measure it, not a 'no'")
            ui.label(f"IR visits (30d): {_visits_str}").classes("cursor-help").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);").tooltip(
                "IR-website visits in the last 30 days · last visit date. 'no visit data' = not measured")
        # marks a peer holding confirmed by a real SEC 13F filing (see
        # _enrich_peer_holdings_with_live_13f); an unmarked ticker is still
        # the original hand-typed seed guess — that ticker hasn't been
        # 13F-refreshed yet (SEC Intelligence tab's "Refresh 13F
        # Institutional Holders" button).
        # Defensive: several record shapes feed this list; Peer_Holdings_Source is a
        # {ticker: "live"} dict for enriched 13F holders but can be a bare string (or
        # missing) on SEC/peer-prospect records, and Peer_Holdings can be a str — either
        # of which used to crash the whole Buy-Side tab here.
        _peer_src = inst.get("Peer_Holdings_Source") or {}
        if not isinstance(_peer_src, dict):
            _peer_src = {}
        _peers = inst.get("Peer_Holdings") or []
        if isinstance(_peers, str):
            _peers = [_peers]
        _peer_labels = [t + (" " if _peer_src.get(t) == "live" else "") for t in _peers]
        ui.label(f"Peers held: {', '.join(_peer_labels) or '—'}").classes("cursor-help").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);").tooltip(
            f"Which of {_tk}'s comps this fund owns. A checkmark = confirmed by a live SEC 13F; "
            "unmarked = seed guess, not yet 13F-refreshed")
        ui.label(f"Action: {inst['Action']}").classes("cursor-help").style(
            f"color:{COLORS['accent_light']};font-size:var(--fs-sm);font-weight:bold;").tooltip(
            "The suggested next move for this holder, read from its position change and signals")

        contact = contacts.get(inst["Fund"], {})
        # Actions sit in a bottom row set off by a thin top divider — same treatment
        # as the Today dashboard cards, so the whole app reads consistently.
        with ui.row().classes("w-full gap-3 items-center").style(
                f"margin-top:4px;padding-top:4px;border-top:1px solid {COLORS['border']};"):
            if contact.get("email"):
                _mailto(contact["email"], f"{CT('ticker')} — Following up, {inst['Fund']}", "Hi,\n\n", f"Email {contact.get('name','Contact')}")
            ui.button("Account 360", icon="account_circle",
                      on_click=lambda inst=inst: _open_account_profile(_card_rec(inst))).props("flat dense")
            ui.button("Log call", icon="add_call",
                      on_click=lambda inst=inst: _quick_log_call(inst)).props("flat dense")
            ui.button("Draft Pre-Earnings Outreach", on_click=lambda inst=inst, contact=contact: _open_outreach_dialog(inst, contact)).props("flat dense")
            ui.button(f"Meeting Log ({len(fund_meetings)})", on_click=lambda inst=inst, fm=fund_meetings, rg=repeat_gap: _open_meeting_log_dialog(inst, fm, rg)).props("flat dense")


def _open_outreach_dialog(inst, contact):
    peer_lookup = {p["ticker"]: p for p in CP()}
    usio_ev_rev = _client_ev_rev()
    peer_lines = []
    for tkr in inst["Peer_Holdings"]:
        p = peer_lookup.get(tkr)
        if p and p.get("ev_rev") and usio_ev_rev:
            disc = round(p["ev_rev"] / usio_ev_rev, 1)
            peer_lines.append(f"Since {inst['Fund']} holds {tkr} ({p['name']}), you might want to look at our latest "
                               f"numbers — {CT('ticker')} trades at {usio_ev_rev}x EV/Revenue vs {tkr} at {p['ev_rev']}x, "
                               f"roughly a {disc}x discount for comparable economics.")
        else:
            peer_lines.append(f"Since {inst['Fund']} holds {tkr}, our upcoming print is worth a look given the sector re-rating setup.")
    if not peer_lines:
        peer_lines = [f"{CT('ticker')}'s upcoming print is shaping up as a re-rating catalyst — EPS inflection, "
                       "multiple expansion, and analyst coverage resumption all in view."]
    # Guard every signal — a real tenant with no call/web feed carries None here, and interpolating
    # it printed a literal "None IR site visit(s)" mid-sentence.
    _call_clause = (f"attention to the call replay ({inst['Listen_Duration']})"
                    if inst.get("Call_Listener") and inst.get("Listen_Duration")
                    else "recent interest in our materials")
    _visits = inst.get("IR_Visits_30d")
    _visit_clause = (f" and {_visits} IR site visit(s) in the last 30 days"
                     if _visits else "")
    engagement_line = (f"Given {inst['Fund']}'s {_call_clause}{_visit_clause}, this feels like "
                       "the right moment to reconnect ahead of the print.")
    greeting = contact.get("name", "[Contact Name]").split()[0] if contact.get("name") else "[Contact Name]"
    subject = f"{CT('ticker')} Pre-Earnings Note — {inst['Fund']}"
    body = (f"Hi {greeting},\n\n{peer_lines[0]}\n" + "\n".join(peer_lines[1:]) + f"\n\n{engagement_line}\n\n"
            f"{CI().get('name','')}\n{CI().get('title','')} · {CT('name')} (NASDAQ: {CT('ticker')})")

    with ui.dialog() as dialog, ui.card().style(f"background:{COLORS['surface_bg']};min-width:480px;"):
        ui.label(f"Draft Outreach — {inst['Fund']}").classes("text-lg font-bold")
        ui.textarea(value=body).classes("w-full").props("rows=12 outlined")
        to_email = contact.get("email", "")
        if to_email:
            _mailto(to_email, subject, body, f"Email {contact.get('name','Contact')}")
        else:
            ui.label("No email on file for this contact yet.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        ui.button("Close", on_click=dialog.close).props("flat")
    dialog.open()


def _open_meeting_log_dialog(inst, fund_meetings, repeat_gap):
    with ui.dialog() as dialog, ui.card().style(f"background:{COLORS['surface_bg']};min-width:480px;"):
        ui.label(f"Meeting Log — {inst['Fund']}").classes("text-lg font-bold")
        if repeat_gap is not None:
            qoq = inst.get("QoQ_Change", 0)
            if qoq > 0:
                ui.label(f"Likely diligence — shares +{qoq:,} and repeat meeting in {repeat_gap}d. Escalate to CFO.").style("color:#15803D;font-size:var(--fs-sm);")
            elif qoq < 0:
                ui.label(f"Flag — shares {qoq:,} and repeat meeting in {repeat_gap}d. Possible pre-exit diligence.").style("color:#B91C1C;font-size:var(--fs-sm);")
            else:
                ui.label(f"Intent unclear — repeat meeting in {repeat_gap}d, no share change. Ask directly.").style("color:#B45309;font-size:var(--fs-sm);")

        if fund_meetings:
            for mtg in fund_meetings:
                with ui.card().classes("w-full").style(f"background:{COLORS['surface_hover_bg']};"):
                    ui.label(f"{mtg['Date']} · {mtg['Type']} — {mtg.get('Outcome','—')}").classes("font-bold").style("font-size:var(--fs-base);")
                    ui.label(f"Attendees: {mtg.get('Attendees','—')}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    ui.label(f"Notes: {mtg.get('Notes','—')}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    if mtg.get("Logged By"):
                        ui.label(f"Logged by: {mtg['Logged By']}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-style:italic;")
        else:
            ui.label("No meetings logged yet.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        # Pre-fill attendees from the known contact on file for this
        # institution, if there is one — previously always blank, which
        # meant re-typing a name that's already stored in
        # institution_contacts.py on every single meeting.
        known_contact = get_institution_contacts().get(inst["Fund"], {})
        default_attendees = known_contact.get("name", "")

        # "Logged by" is auto-filled from this client's own IR contact
        # (config.client_config.CI()) rather than a free-text field — there's
        # no multi-user login in this app yet, so the configured IR contact
        # is the closest available notion of "who is using this session."
        logged_by = CI().get("name") or "IR Team"

        ui.markdown("**Log a new meeting**")
        m_date = ui.input("Date (YYYY-MM-DD)", value=datetime.now().strftime("%Y-%m-%d")).classes("w-full").props("outlined dense")
        m_type = ui.select(["1x1 — Investor Conference", "Intro call", "Follow-up call", "NDR meeting",
                             "Earnings call Q&A", "Other"], label="Meeting Type", value="Follow-up call").classes("w-full").props("outlined dense")
        m_attend = ui.input("Attendees", value=default_attendees).classes("w-full").props("outlined dense")
        m_notes = ui.textarea("Notes").classes("w-full").props("outlined autogrow")
        m_outcome = ui.select(["Positive — follow up", "Neutral — maintain", "Warm — send materials",
                                "Flag — possible exit", "CFO follow-up required", "No clear signal"],
                               label="Meeting Outcome (scored)", value="Positive — follow up").classes("w-full").props("outlined dense")
        ui.label(f"Outcome feeds this fund's Interaction Score (0–{INTERACTION_SCORE_MAX} pts of the 100-pt "
                 "Engagement Score) — a positive/CFO-escalation outcome raises it, a Flag lowers it.") \
            .style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
        ui.label(f"Logging as: {logged_by}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

        def log_meeting():
            log = _load_meeting_log()
            log.append({"Fund": inst["Fund"], "Date": m_date.value, "Type": m_type.value,
                        "Attendees": m_attend.value, "Notes": m_notes.value, "Outcome": m_outcome.value,
                        "Logged By": logged_by, "Source": "Manual"})
            _save_meeting_log(log)
            # Interaction Score is derived fresh from meeting_log on every
            # render (see _compute_interaction_score / _score_institutions),
            # so this recomputes what it will actually be on next render —
            # a real persisted number, not the old ephemeral same-render-only
            # "+20 repeat meeting" flash that never got saved anywhere.
            updated = _get_fund_meetings(log, inst["Fund"])
            new_interaction = _compute_interaction_score(updated)
            old_interaction = inst.get("Interaction_Score", 0)
            new_engagement = min(100, inst["Engagement_Score"] - old_interaction + new_interaction)
            new_gap = _get_repeat_signal(updated)
            delta = new_interaction - old_interaction
            delta_str = f"+{delta}" if delta >= 0 else str(delta)
            if new_gap is not None:
                ui.notify(f"Repeat meeting alert — {inst['Fund']} met {len(updated)}x in {new_gap}d. "
                          f"Interaction Score {delta_str} → {new_interaction}/{INTERACTION_SCORE_MAX} · "
                          f"Engagement Score → {new_engagement}/100", type="warning")
            else:
                ui.notify(f"Meeting logged for {inst['Fund']} · Interaction Score {delta_str} → "
                          f"{new_interaction}/{INTERACTION_SCORE_MAX} · Engagement Score → {new_engagement}/100")
            dialog.close()
            _refresh()

        ui.button("Log Meeting", on_click=log_meeting).props("color=primary")
        ui.button("Close", on_click=dialog.close).props("flat")
    dialog.open()


# ─────────────────────────────────────────────────────────────────────────
# NDR Planner tab
# ─────────────────────────────────────────────────────────────────────────
def _parse_time_min(s):
    """A meeting time string ('2:00 PM', '14:00', '9 AM') → minutes since
    midnight, or None if unscheduled/unparseable."""
    s = (s or "").strip().upper().replace(".", "")
    if not s or s == "—":
        return None
    for fmt in ("%I:%M %p", "%I %p", "%H:%M"):
        try:
            t = datetime.strptime(s, fmt)
            return t.hour * 60 + t.minute
        except Exception:
            continue
    return None


def _meeting_street(m):
    """The best STREET address for a meeting: its own address override, else the
    fund's office address from the address book (SEC-sourced or manual). Returns
    '' when only a metro/city is known — i.e. no street-level location."""
    own = (m.get("address") or "").strip()
    if own:
        return own
    try:
        from core import fund_addresses
        return (fund_addresses.address_for(m.get("institution", "")) or "").strip()
    except Exception:
        return ""


def _meeting_loc(m, trip):
    """Best location string for the travel calc: the meeting's own address, then
    the fund's office address, then its stored metro, then the trip city."""
    return (_meeting_street(m)
            or (m.get("metro") or "").strip()
            or (trip.get("city") or "").strip())


def _sorted_day_meetings(meetings):
    """Meetings ordered by parsed time; unscheduled ones sink to the end keeping
    their original order (stable), so the itinerary reads chronologically."""
    keyed = [(_parse_time_min(m.get("time")), i, m) for i, m in enumerate(meetings)]
    scheduled = sorted((k for k in keyed if k[0] is not None), key=lambda x: (x[0], x[1]))
    unscheduled = [k for k in keyed if k[0] is None]
    return [m for _, _, m in scheduled] + [m for _, _, m in unscheduled]


def _travel_leg_between(prev_m, m, trip):
    """Travel estimate between two consecutive in-person stops, plus a
    feasibility read from their scheduled times. Prefers REAL routed driving
    distance/time (core.routing / OpenRouteService) and falls back to the
    offline great-circle estimate (core.geo) when there's no API key or an
    address can't be resolved. Returns (leg_dict, tight_note_or_None) or
    (None, None) when a leg can't be estimated (virtual stop or unknown city)."""
    if (prev_m.get("format", "In-person") != "In-person"
            or m.get("format", "In-person") != "In-person"):
        return None, None
    from_loc, to_loc = _meeting_loc(prev_m, trip), _meeting_loc(m, trip)
    # Real routing is only meaningful when BOTH stops have a real street address.
    # If either falls back to a metro/city label, both endpoints collapse to the
    # same city point and routing would report a misleading "0 mi · 0 min" — so
    # for those we use the offline "same city — allow ~20 min" allowance instead.
    both_have_addr = bool(_meeting_street(prev_m)) and bool(_meeting_street(m))
    lg = None
    try:
        from core import routing
        if routing.is_configured() and both_have_addr:
            # Bias geocoding toward the trip's business district so ambiguous
            # street addresses (e.g. "745 5th Ave") resolve in the right metro.
            focus = routing.focus_for(trip.get("city")) or routing.focus_for(from_loc)
            lg = routing.leg(from_loc, to_loc, focus=focus)
            if lg and lg.get("miles", 0) < 0.1:
                lg = None  # addresses collapsed to one point — not a real leg
    except Exception:
        lg = None
    if not lg:
        from core import geo
        lg = geo.leg(from_loc, to_loc)
    if not lg:
        return None, None
    tight = None
    t0, t1 = _parse_time_min(prev_m.get("time")), _parse_time_min(m.get("time"))
    if t0 is not None and t1 is not None and lg.get("drive_min") is not None:
        gap = t1 - t0
        # Gap between starts must cover the prior meeting itself (1h, or 1h15 for a catered lunch)
        # plus the drive to the next stop.
        needed = _meeting_duration(prev_m) + lg["drive_min"]
        if 0 <= gap < needed:
            tight = f"TIGHT — {gap} min between starts vs ~{needed} min needed (meeting + travel)"
    return lg, tight


def _hotel_departure(trip, first_m, buffer_min=10):
    """Where management is staying → the day's first meeting. Returns (pickup_time_str, leg_dict):
    the leg is the routed (or city-level) drive from the hotel, and pickup is the first meeting's
    start minus that drive minus a short buffer — the "what time do they get picked up" answer.
    (None, None) if there's no hotel on the trip or the first stop can't be routed."""
    hotel = (trip.get("hotel") or "").strip()
    if not hotel or first_m.get("format", "In-person") != "In-person":
        return None, None
    lg, _ = _travel_leg_between({"address": hotel, "format": "In-person"}, first_m, trip)
    if not lg:
        return None, None
    t1 = _parse_time_min(first_m.get("time"))
    if t1 is None:
        return None, lg
    return _fmt_min_12h(max(t1 - (lg.get("drive_min") or 0) - buffer_min, 0)), lg


def _hotel_return(trip, last_m):
    """The day's last meeting → back to where management is staying (the trip's ending). Returns
    (arrive_time_str, leg_dict): the leg is the routed drive back to the hotel, and arrive is the
    last meeting's start + its duration + that drive — 'back at the hotel by …'. (None, None) if
    there's no hotel or the last stop can't be routed."""
    hotel = (trip.get("hotel") or "").strip()
    if not hotel or last_m.get("format", "In-person") != "In-person":
        return None, None
    lg, _ = _travel_leg_between(last_m, {"address": hotel, "format": "In-person"}, trip)
    if not lg:
        return None, None
    t0 = _parse_time_min(last_m.get("time"))
    if t0 is None:
        return None, lg
    return _fmt_min_12h(t0 + _meeting_duration(last_m) + (lg.get("drive_min") or 0)), lg


def _travel_total_line(total_miles, leg_count, routed_count, routing_on=False):
    """The trip-level travel summary, honestly labelled by whether the legs were
    really routed (driving) or fell back to the offline straight-line estimate."""
    if not leg_count:
        return None
    miles = f"~{total_miles:,.0f} mi"
    if routed_count == leg_count:
        return (f"Est. driving distance: {miles} across {leg_count} leg(s) — "
                "real driving miles/time, routed via OpenRouteService.")
    if routed_count:
        return (f"Est. travel: {miles} across {leg_count} leg(s) — {routed_count} routed "
                "(driving), the rest straight-line city-level (stops without a street "
                "address). Confirm with your car service.")
    # No leg routed. If the key IS set, the blocker is missing addresses, not the key.
    nudge = ("add a street address to each meeting to get routed driving miles/time"
             if routing_on else "add a maps API key in Settings for routed driving miles")
    return (f"Est. in-person travel: {miles} across {leg_count} leg(s) — straight-line, "
            f"city-level (not driving miles). To upgrade, {nudge}.")


def _build_ndr_itinerary(trip, ticker):
    lines = [f"{ticker} — NON-DEAL ROADSHOW ITINERARY", trip["name"],
              f"{trip['dates']}  ·  {trip['city']}  ·  {'Virtual' if trip['ndr_type']=='virtual' else 'In-Person'}"]
    if trip.get("sponsor_bank"):
        lines.append(f"Sponsoring Bank: {trip['sponsor_bank']}")
    if trip.get("hotel"):
        lines.append(f"Lodging: {trip['hotel']}")
    lines.append(f"Attendees: {', '.join(trip.get('team', [])) or '—'}")
    lines.append(f"Focus: {trip.get('focus','—')}")
    if trip.get("notes"):
        lines.append(f"Trip Objectives: {trip['notes']}")
    lines.append("=" * 64)
    # Grouped by day (falls back to a single unlabeled block for meetings
    # added the old way, before the "day" field existed — see
    # _render_ndr_tab's schema-upgrade note below).
    by_day = {}
    for m in trip.get("meetings", []):
        by_day.setdefault(m.get("day", 1), []).append(m)
    try:
        from core import routing as _routing
        _routing_on = _routing.is_configured()
    except Exception:
        _routing_on = False
    total_miles, leg_count, routed_count = 0.0, 0, 0
    for day_num in sorted(by_day.keys()):
        if len(by_day) > 1:
            lines.append(f"DAY {day_num}")
            lines.append("-" * 64)
        day_ms = _sorted_day_meetings(by_day[day_num])
        # Day opens from the hotel — where management is staying and the pickup time to make the
        # first meeting.
        if trip.get("hotel") and day_ms:
            _pk, _plg = _hotel_departure(trip, day_ms[0])
            if _plg:
                lines.append(f"  Staying at: {trip['hotel']}")
                lines.append(f"  {('Pickup ' + _pk + ' — ') if _pk else ''}{_plg['label']} to "
                             f"{day_ms[0].get('institution', '')}")
        prev = None
        for m in day_ms:
            if prev is not None:
                lg, tight = _travel_leg_between(prev, m, trip)
                if lg:
                    total_miles += lg["miles"]
                    leg_count += 1
                    routed_count += 1 if lg.get("basis") == "routed" else 0
                    lines.append(f"        ↳ travel: {lg['label']}"
                                 + (f"   [!] {tight}" if tight else ""))
            tag = "Non-Holder — Priority Target" if m.get("non_holder", True) else "Existing Holder"
            time_lbl = _meeting_time_range(m)
            lines.append(f"  {time_lbl:>18}   {m.get('institution','')}")
            if m.get("score") is not None:
                lines.append(f"              {m.get('format','In-person')} · {tag} · Engagement Score {m['score']}/100")
            else:
                lines.append(f"              {m.get('format','In-person')} · {tag}")
            known_c = get_institution_contacts().get(m.get("institution", ""), {})
            who = m.get("contact") or ", ".join(p for p in (known_c.get("name"), known_c.get("title")) if p)
            if who:
                lines.append(f"              Meeting with: {who}")
            if m.get("address"):
                lines.append(f"              Location: {m['address']}")
            _virtual = m.get("format", "In-person") != "In-person"
            if _virtual:
                _link = (m.get("meeting_link") or "").strip()
                lines.append(f"              Join ({m.get('format')}): {_link or 'link TBD — add on the Active NDRs tab'}")
            if m.get("notes"):
                lines.append(f"              Note: {m['notes']}")
            if m.get("lunch"):
                lines.append("              Catered working lunch (~1h15) · Dietary: "
                             + ((m.get("dietary") or "").strip() or "— (confirm with caterer)"))
            prev = m
        # Day ends back at where management is staying (the trip ending).
        if trip.get("hotel") and day_ms:
            _ar, _rl = _hotel_return(trip, day_ms[-1])
            if _rl:
                total_miles += _rl["miles"]; leg_count += 1
                routed_count += 1 if _rl.get("basis") == "routed" else 0
                lines.append(f"        ↳ return: {_rl['label']} to hotel"
                             + (f"   (back ~{_ar})" if _ar else ""))
                lines.append(f"              Return to {trip['hotel']}")
    lines.append("=" * 64)
    _tl = _travel_total_line(total_miles, leg_count, routed_count, _routing_on)
    if _tl:
        lines.append(_tl)
    lines.append(f"Prepared {datetime.now().strftime('%b %d, %Y')}")
    return "\n".join(lines)


def _build_ndr_itinerary_html(trip, ticker):
    """Print-formatted itinerary — a clean, self-contained HTML document opened
    in a new window for the browser's print dialog, so the app chrome/nav isn't
    printed. Same content as the .txt export, laid out as a schedule table."""
    import html as _html

    def esc(s):
        return _html.escape(str(s))

    rows = []
    by_day = {}
    for m in trip.get("meetings", []):
        by_day.setdefault(m.get("day", 1), []).append(m)
    try:
        from core import routing as _routing
        _routing_on = _routing.is_configured()
    except Exception:
        _routing_on = False
    total_miles, leg_count, routed_count = 0.0, 0, 0
    for day_num in sorted(by_day.keys()):
        if len(by_day) > 1:
            rows.append(f'<tr><td class="day" colspan="2">Day {day_num}</td></tr>')
        prev = None
        for m in _sorted_day_meetings(by_day[day_num]):
            if prev is not None:
                lg, tight = _travel_leg_between(prev, m, trip)
                if lg:
                    total_miles += lg["miles"]
                    leg_count += 1
                    routed_count += 1 if lg.get("basis") == "routed" else 0
                    warn = (f' <span class="tight">&#9888; {esc(tight)}</span>') if tight else ""
                    rows.append(
                        f'<tr><td class="time travel">&#8627;</td>'
                        f'<td class="travel">{esc(lg["label"])}{warn}</td></tr>')
            known_c = get_institution_contacts().get(m.get("institution", ""), {})
            who = m.get("contact") or ", ".join(p for p in (known_c.get("name"), known_c.get("title")) if p)
            tag = "Non-Holder — Priority Target" if m.get("non_holder", True) else "Existing Holder"
            meta = " · ".join(x for x in [m.get("format", "In-person"), tag,
                    (f"Score {m['score']}/100" if m.get("score") is not None else "")] if x)
            _virtual = m.get("format", "In-person") != "In-person"
            _link = (m.get("meeting_link") or "").strip()
            _join = ""
            if _virtual:
                _join = (f'<b>Join ({esc(m.get("format"))}):</b> <a href="{esc(_link)}">{esc(_link)}</a>'
                         if _link else f'<b>Join ({esc(m.get("format"))}):</b> link TBD')
            det = "<br>".join(filter(None, [
                f"<b>Meeting with:</b> {esc(who)}" if who else "",
                f"<b>Location:</b> {esc(m['address'])}" if m.get("address") else "",
                _join,
                f"<b>Note:</b> {esc(m['notes'])}" if m.get("notes") else "",
            ]))
            rows.append(
                f'<tr><td class="time">{esc(_meeting_time_range(m))}</td>'
                f'<td class="firm">{esc(m.get("institution", ""))}'
                f'<div class="meta">{esc(meta)}</div>'
                f'{("<div class=det>" + det + "</div>") if det else ""}</td></tr>')
            prev = m
    _tl = _travel_total_line(total_miles, leg_count, routed_count, _routing_on)
    if _tl:
        rows.append(f'<tr><td class="time travel"></td>'
                    f'<td class="travel"><b>{esc(_tl)}</b></td></tr>')

    # Route map (best-effort, cached) — a street map of the day pinned in schedule
    # order, so the printed itinerary carries its own map to hand the team.
    map_html = ""
    try:
        _png = _build_ndr_route_map_png(trip, ticker)
        if _png:
            import base64 as _b64
            _uri = "data:image/png;base64," + _b64.b64encode(_png).decode()
            map_html = (f"<div class='routemap'><img src='{_uri}' alt='NDR route map'/>"
                        "<div class='mapcap'>Stops numbered in schedule order; H = hotel (start / end). "
                        "Map data © OpenStreetMap contributors.</div></div>")
    except Exception:
        map_html = ""

    meta_line = (f"{esc(trip.get('dates', ''))} &nbsp;·&nbsp; {esc(trip.get('city', ''))} &nbsp;·&nbsp; "
                 f"{'Virtual' if trip.get('ndr_type') == 'virtual' else 'In-Person'}")
    blocks = []
    if trip.get("sponsor_bank"):
        blocks.append(f"<div><b>Sponsoring Bank:</b> {esc(trip['sponsor_bank'])}</div>")
    blocks.append(f"<div><b>Company Attendees:</b> {esc(', '.join(trip.get('team', [])) or '—')}</div>")
    blocks.append(f"<div><b>Focus:</b> {esc(trip.get('focus', '—'))}</div>")
    if trip.get("notes"):
        blocks.append(f"<div><b>Objectives:</b> {esc(trip['notes'])}</div>")

    return (
        "<!doctype html><html><head><meta charset='utf-8'>"
        f"<title>{esc(ticker)} NDR — {esc(trip.get('name', ''))}</title><style>"
        "body{font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;color:#0F172A;margin:32px;}"
        "h1{font-size:var(--fs-2xl);margin:0 0 2px;}.sub{color:#475569;font-size:var(--fs-base);margin-bottom:12px;}"
        ".meta-block div{font-size:var(--fs-base);margin:2px 0;color:#334155;}"
        "table{width:100%;border-collapse:collapse;margin-top:14px;}"
        "td{border-bottom:1px solid #E2E8F0;padding:8px 6px;vertical-align:top;font-size:var(--fs-base);}"
        "td.time{white-space:nowrap;width:132px;color:#475569;font-weight:600;}"
        "td.firm{font-weight:600;}.meta{font-weight:400;color:#64748B;font-size:var(--fs-xs);margin-top:2px;}"
        ".det{font-weight:400;color:#334155;font-size:var(--fs-sm);margin-top:4px;}"
        "td.travel{border-bottom:none;color:#64748B;font-size:var(--fs-xs);font-style:italic;padding:2px 6px;}"
        "td.time.travel{text-align:right;color:#94A3B8;font-weight:400;}"
        ".tight{color:#B45309;font-style:normal;font-weight:600;}"
        "td.day{background:#F1F5F9;font-weight:700;font-size:var(--fs-sm);letter-spacing:.03em;}"
        ".foot{margin-top:18px;color:#94A3B8;font-size:var(--fs-xs);}"
        ".routemap{margin:14px 0 4px;}"
        ".routemap img{width:100%;max-width:900px;border:1px solid #E2E8F0;border-radius:8px;}"
        ".mapcap{color:#94A3B8;font-size:var(--fs-xs);margin-top:4px;}"
        "@media print{body{margin:12px;}.routemap img{max-width:100%;}.routemap{page-break-inside:avoid;}}"
        "</style></head><body>"
        f"<h1>{esc(ticker)} — Non-Deal Roadshow Itinerary</h1>"
        f"<div class='sub'>{esc(trip.get('name', ''))} &nbsp;|&nbsp; {meta_line}</div>"
        f"<div class='meta-block'>{''.join(blocks)}</div>"
        f"{map_html}"
        f"<table>{''.join(rows)}</table>"
        f"<div class='foot'>Prepared {datetime.now().strftime('%b %d, %Y')} · {esc(ticker)} Investor Relations</div>"
        "</body></html>"
    )


def _ndr_map_stops(trip):
    """Ordered (hotel, stops) for the route map: the hotel as start/end plus every
    meeting in itinerary order (day, then time), each numbered and carrying its best
    street address. Mirrors exactly what the printed schedule lists."""
    hotel_addr = (trip.get("hotel") or "").strip()
    hotel = {"name": hotel_addr.split(",")[0].strip() or "Hotel", "address": hotel_addr} if hotel_addr else None
    ordered = sorted(
        trip.get("meetings", []),
        key=lambda m: (m.get("day", 1),
                       _parse_time_min(m.get("time")) if _parse_time_min(m.get("time")) is not None else 9999))
    stops, n = [], 0
    for m in ordered:
        addr = _meeting_street(m) or _meeting_loc(m, trip)
        if not addr:
            continue
        n += 1
        stops.append({"label": str(n), "name": m.get("institution", ""),
                      "time": (m.get("time") or "").strip(), "address": addr})
    return hotel, stops


def _build_ndr_route_map_png(trip, ticker):
    """PNG bytes of the trip's route map (numbered stops in schedule order, real
    driving route, hotel, legend, scale), or None when the routing key is missing or
    no stop can be located. Cached in db per trip content so re-exports are instant."""
    import base64
    import hashlib

    from core import db, ndr_map, routing
    if not routing.is_configured():
        return None
    hotel, stops = _ndr_map_stops(trip)
    if not stops:
        return None
    sig = hashlib.md5(json.dumps(
        [ticker, trip.get("name"), trip.get("dates"), (hotel or {}).get("address"),
         [(s["label"], s["address"], s["time"]) for s in stops]], sort_keys=True).encode()).hexdigest()
    cache = db.load_json("ndr_route_maps.json", {}) or {}
    if sig in cache:
        try:
            return base64.b64decode(cache[sig])
        except Exception:
            pass
    try:
        focus = routing.focus_for(trip.get("city") or "")
    except Exception:
        focus = None
    png = ndr_map.build_route_map(
        hotel, stops, focus=focus,
        title=trip.get("name") or f"{ticker} NDR",
        subtitle=" · ".join(x for x in [trip.get("dates", ""), ticker, "route & schedule"] if x))
    if not png:
        return None
    cache[sig] = base64.b64encode(png).decode()
    db.save_json("ndr_route_maps.json", cache)
    return png


# ─────────────────────────────────────────────────────────────────────────
# NDR target auto-suggestion — real institutions instead of a hardcoded
# 25-row USIO-specific database (app.py's INST_DB). Reuses the SAME
# per-client institutions list (core.investor_scoring.score_institutions
# output, computed once in render_investors_page() and threaded down here)
# that Buy-Side Intelligence and Target Database already read, so "who's a
# good NDR target in Boston" is always the same live answer as "who's a
# good prospect in Boston" everywhere else on this page — no second,
# drifting copy of the tracked-institution universe to maintain.
# ─────────────────────────────────────────────────────────────────────────
def _ndr_location_options(institutions):
    """Metro-region labels straight from the live institutions list, not a
    hardcoded city list — stays correct for whatever a given client
    actually has tracked, matching the same Metro values already used by
    the NDR Requests tab's metro dropdown and the Big Picture panel's
    Metro Priority scoring."""
    # Same None-safety as the region filter: one record with Metro=None would raise
    # TypeError here and break the NDR planner.
    return sorted({(i.get("Metro") or "Unknown (SEC)") for i in institutions})


def _ndr_target_candidates(institutions, location, ndr_type):
    """Port of app.py's get_institutions_for_location, onto real per-client
    data. In-person: exact match on Metro (the institutions list already
    stores metro-REGION labels like 'Boston / New England', not raw
    cities, so no fuzzy city->region translation table is needed the way
    app.py's INST_DB required). Virtual: every non-holder scoring >=40
    anywhere — same threshold app.py used, since a virtual trip isn't
    geography-bound. Non-holders are sorted first (conversion is the
    point of an NDR), holders after (relationship maintenance)."""
    if ndr_type == "virtual":
        candidates = [i for i in institutions if not i["USIO_Holder"] and _score_val(i) >= 40]
    else:
        candidates = [i for i in institutions if i["Metro"] == location]
    # _score_val (not x["Engagement_Score"]) — a candidate can have a missing or None score
    # (SEC-sourced names, unscored prospects), and `-None` / a missing key raises, which used to
    # abort the whole Active NDRs panel mid-render so a Planning NDR's card never drew.
    non_holders = sorted([i for i in candidates if not i["USIO_Holder"]], key=lambda x: -_score_val(x))
    holders = sorted([i for i in candidates if i["USIO_Holder"]], key=lambda x: -_score_val(x))
    return non_holders + holders


def _last_city_visit(trips, city):
    past = [t for t in trips if str(t.get("city", "")).lower() == str(city).lower()]
    if not past:
        return None
    past.sort(key=lambda t: t.get("created", ""), reverse=True)
    return past[0]


def _ndr_talking_points(inst):
    """Grounded in THIS fund's own real tracked fields (Action, Conviction,
    Peer_Holdings, Call_Listener, Engagement_Score) rather than app.py's
    hardcoded per-peer financial claims (specific EV/Revenue multiples,
    margin comparisons keyed to literal ticker strings like 'GDOT'/'PRTH'
    baked into if/elif branches). Those were real, true facts for one
    specific client at one specific moment — porting them verbatim would
    mean every future client (or even the same client next quarter) sees
    stale, misattributed numbers presented as current talking points.
    Same 'honest, disclosed approximation over fabricated precision'
    philosophy as core/fit_score.py's turnover/contactability heuristics."""
    points = []
    if inst.get("Action"):
        points.append(inst["Action"])
    if inst.get("Peer_Holdings"):
        points.append(f"Owns {', '.join(inst['Peer_Holdings'])} — draw the direct comparison to those peers' "
                       f"positioning rather than starting from a blank slate.")
    if inst.get("Call_Listener") and inst.get("Listen_Duration"):
        points.append(f"Listened to the last earnings call ({inst['Listen_Duration']}) — reference the specific "
                       f"moments they engaged with instead of re-covering the whole deck.")
    if inst.get("Engagement_Score", 0) >= 80:
        points.append(f"Top-tier engagement score ({inst['Engagement_Score']}/100) — this is a "
                       f"{'defend' if inst.get('USIO_Holder') else 'high-priority conversion'} meeting, not a cold intro.")
    if inst.get("Conviction"):
        points.append(f"Conviction on file: {inst['Conviction']}.")
    if not points:
        points.append("No prior signal on file yet — treat as a discovery meeting: confirm mandate fit before "
                       "going deep on the thesis.")
    return points


def _fmt_time_12h(t):
    """Convert a picker's 24-hour 'HH:MM' into the '2:00 PM' style the rest of
    the itinerary displays and _build_ndr_itinerary expects. Leaves anything it
    can't parse (incl. a hand-typed '2:00 PM') untouched, so the field stays
    manually editable."""
    try:
        hh, mm = t.split(":")[:2]
        h, m = int(hh), int(mm)
        ap = "AM" if h < 12 else "PM"
        return f"{h % 12 or 12}:{m:02d} {ap}"
    except Exception:
        return t


def _fmt_min_12h(mins):
    """Minutes-since-midnight → '2:00 PM' style."""
    h, m = divmod(int(mins), 60)
    ap = "AM" if h < 12 else "PM"
    return f"{h % 12 or 12}:{m:02d} {ap}"


def _meeting_duration(m):
    """Minutes a meeting occupies on the calendar. A regular 1x1 runs ~45 min; a catered lunch runs
    ~1h15 (the Street norm — the meeting IS the lunch)."""
    return 75 if m.get("lunch") else 45


def _meeting_time_range(m):
    """'8:00 AM – 8:45 AM' — the meeting's start through its end (start + duration), so the
    itinerary shows how long each slot runs. Falls back to the raw start (or '—') when the
    time can't be parsed, leaving hand-typed / unscheduled entries untouched."""
    start = _parse_time_min(m.get("time"))
    if start is None:
        return (m.get("time") or "—")
    return f"{_fmt_min_12h(start)} – {_fmt_min_12h(start + _meeting_duration(m))}"


def _available_slots(meetings, day, win_start="8:00 AM", win_end="5:00 PM",
                     new_meeting_min=45, step_min=15):
    """Open meeting START times on `day`, inside the management window [win_start, win_end], that fit
    a `new_meeting_min`-long meeting without colliding with an already-scheduled one (each blocked
    for its own duration — a lunch takes 1h15). Returns a chronological list of 12-hour time strings
    — the REAL availability once existing meetings carve up the day (before the first, between, and
    after the last). Empty when the day is full or the window is invalid."""
    ws, we = _parse_time_min(win_start), _parse_time_min(win_end)
    if ws is None or we is None or we <= ws:
        return []
    occ = []
    for m in meetings:
        if m.get("day", 1) != day:
            continue
        t = _parse_time_min(m.get("time"))
        if t is not None:
            occ.append((t, t + _meeting_duration(m)))
    out, t = [], ws
    while t + new_meeting_min <= we:
        if not any(t < b and t + new_meeting_min > a for a, b in occ):   # overlap test
            out.append(_fmt_min_12h(t))
        t += step_min
    return out


def _time_picker_input(label="Time", placeholder="e.g. 2:00 PM", value=""):
    """A text time field with an attached clock (ui.time in a popup menu). The
    field stays hand-editable — the clock just fills it — so a planner can nudge
    a time freely to absorb travel, which is the whole point of an NDR schedule.
    Returns the ui.input element (read .value)."""
    ti = ui.input(label, placeholder=placeholder, value=value).classes("flex-1").props("outlined dense")
    with ti:
        with ui.menu().props("no-parent-event") as _tmenu:
            ui.time(on_change=lambda e: ti.set_value(_fmt_time_12h(e.value)))
        with ti.add_slot("append"):
            ui.icon("schedule").on("click", _tmenu.open).classes("cursor-pointer")
    return ti


def _open_add_to_trip_dialog(idx, fund, contact, non_holder, score, default_metro, on_added,
                             default_address=""):
    """Ask for a time (with a clock) before dropping an investor onto a trip,
    instead of the old silent time='—' quick-add. Captures day / time / format /
    optional address, and stashes the metro so the itinerary's travel calc can
    resolve a city even when no street address is entered yet."""
    # RIAs / wealth managers are a low-touch tier: no PM to pitch, so they get a
    # video call an assistant can set up — never management travel time. Default
    # the format to Virtual so they can't silently land on the in-person route
    # (and so they never enter the itinerary's driving-distance calc).
    try:
        from core import peer_prospects as _pp
        _is_ria_fund = _pp.is_ria(fund)
    except Exception:
        _is_ria_fund = False

    dialog = ui.dialog()
    with dialog, ui.card().style("min-width:380px;"):
        ui.label(f"Schedule {fund}").style(
            f"color:{COLORS['text_heading']};font-weight:700;font-size:var(--fs-md);")
        ui.label("Pick a meeting time — flexible, adjust later for travel.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        if _is_ria_fund:
            with ui.row().classes("items-center gap-1").style("margin-top:2px;"):
                ui.icon("videocam").style("color:#B45309;font-size:var(--fs-md);")
                ui.label("RIA / wealth manager — video call tier. An assistant can set this up; "
                         "it doesn't need management travel.").style("color:#B45309;font-size:var(--fs-sm);")
        # Load the trip's existing schedule + management window so we offer ONLY open slots — the day
        # already has meetings, and management is available a bounded number of hours.
        try:
            _trip_now = _load_json("ndr_trips.json", [])[idx]
        except Exception:
            _trip_now = {}
        _ms = _trip_now.get("meetings", [])
        _ws = _trip_now.get("day_start") or "8:00 AM"
        _we = _trip_now.get("day_end") or "5:00 PM"
        _ndays = max(int(_trip_now.get("days", 1) or 1), 1)

        with ui.row().classes("w-full gap-3 items-end").style("margin-top:6px;"):
            day_in = ui.select({d: f"Day {d}" for d in range(1, _ndays + 1)}, value=1,
                               label="Day").props("dense outlined").classes("w-28")
            slot_in = ui.select([], label="Open slot").props("dense outlined").classes("flex-1")
            format_in = ui.select(["In-person", "Virtual", "Zoom", "Teams", "Phone", "Jitsi"],
                                  value=("Zoom" if _is_ria_fund else "In-person"),
                                  label="Format").props("dense outlined").classes("w-32")
        _slot_note = ui.label("").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

        # Catered lunch meeting — the Street norm of building lunch INTO a meeting. Runs ~1h15 (so it
        # blocks a bigger chunk of the day) and opens a dietary-requests field for the caterer.
        lunch_in = ui.checkbox("Catered lunch meeting (~1h15)").props("dense")
        diet_in = ui.input("Dietary requests", placeholder="e.g. 2 vegetarian, 1 gluten-free · no nuts").classes("w-full").props("outlined dense")
        diet_in.bind_visibility_from(lunch_in, "value")

        def _refresh_slots():
            _mm = 75 if lunch_in.value else 45
            _av = _available_slots(_ms, int(day_in.value or 1), _ws, _we, new_meeting_min=_mm)
            slot_in.set_options(_av, value=(_av[0] if _av else None))
            _kind = "lunch (1h15)" if lunch_in.value else "meeting (45 min)"
            if _av:
                _slot_note.set_text(f"Management available {_ws}–{_we} · {len(_av)} open slot(s) on day "
                                    f"{day_in.value} that fit a {_kind} (existing meetings blocked out).")
            else:
                _slot_note.set_text(f"No open slots on day {day_in.value} that fit a {_kind} within "
                                    f"{_ws}–{_we} — try another day, or widen the management window.")
        day_in.on_value_change(lambda: _refresh_slots())
        lunch_in.on_value_change(lambda: _refresh_slots())
        _refresh_slots()
        address_in = ui.input("Address / location (optional — powers the travel calc)",
                              value=default_address or "",
                              placeholder=f"street, {default_metro or 'city'}").classes("w-full").props("outlined dense")
        # Virtual meetings carry a join link that flows onto the itinerary. Shown
        # only when the format isn't in-person.
        link_in = ui.input("Meeting link", placeholder="Zoom / Teams / Meet URL").classes("w-full").props("outlined dense")
        link_in.bind_visibility_from(format_in, "value", backward=lambda v: v != "In-person")

        def commit():
            trips_ = _load_json("ndr_trips.json", [])
            trips_[idx].setdefault("meetings", []).append({
                "institution": fund, "contact": contact,
                "address": (address_in.value or "").strip(), "notes": "",
                "day": int(day_in.value or 1),
                "time": (slot_in.value or "").strip() or "—",
                "type": "1x1", "format": format_in.value, "status": "scheduled",
                "meeting_link": (link_in.value or "").strip(),
                "non_holder": non_holder, "score": score, "metro": default_metro or "",
                "lunch": bool(lunch_in.value),
                "dietary": (diet_in.value or "").strip() if lunch_in.value else "",
            })
            _save_json("ndr_trips.json", trips_)
            _t = (slot_in.value or "").strip()
            ui.notify(f"{fund} added to trip{(' at ' + _t) if _t else ' — time TBD'}"
                      + (" · catered lunch" if lunch_in.value else "") + ".")
            dialog.close()
            on_added()

        with ui.row().classes("gap-2").style("margin-top:8px;"):
            ui.button("Add to trip", on_click=commit).props("color=primary")
            ui.button("Cancel", on_click=dialog.close).props("flat")
    dialog.open()


def _render_ndr_tab(institutions, meeting_log, client_id, mode="pre"):
    # Count open (non-Completed) NDRs up front so the TAB ITSELF advertises them. A Planning NDR
    # built from the metro table lands here; without a visible count it's easy to be sitting on
    # "Plan New" while your just-created NDR is one tab over (the "where is my NDR?" report).
    _open_ndrs = [t for t in _load_json("ndr_trips.json", []) if t.get("status") != "Completed"]
    _active_lbl = f"Active NDRs ({len(_open_ndrs)})" if _open_ndrs else "Active NDRs"
    with ui.tabs().classes("w-full") as ndr_tabs:
        nt1 = ui.tab("Plan New NDR")
        nt2 = ui.tab("Active NDRs", label=_active_lbl)
        nt4 = ui.tab("Prep Cards")
        nt5 = ui.tab("Post-NDR Debrief")
    # Land on Active NDRs when a roadshow is in flight, so a Planning NDR you just built isn't hidden.
    with ui.tab_panels(ndr_tabs, value=(nt2 if _open_ndrs else nt1)).classes("w-full"):
        with ui.tab_panel(nt1):
            ui.label("Plan a New Non-Deal Roadshow").classes("font-bold")
            # Start from real demand: pick an inbound request (sponsoring bank · city) and it prefills
            # the name, sponsoring bank and city — instead of opening on a blank name field. You can
            # still start blank and type everything by hand.
            _plan_open_reqs = [r for r in _load_ndr_requests() if not r.get("resolved")]
            _req_pick = None
            if _plan_open_reqs:
                _rp_opts = {"": "— start blank —"}
                for _r in _plan_open_reqs:
                    _rp_opts[_r["id"]] = f"{_r.get('firm', '')} · {_r.get('city', '')} — {_r.get('analyst', '')}"
                _req_pick = ui.select(
                    _rp_opts, value="",
                    label="Start from an inbound request (sponsoring bank · city)").classes("w-full").props("outlined")
            with ui.row().classes("w-full gap-4"):
                with ui.column().classes("flex-1"):
                    name_in = ui.input("NDR name / label", placeholder="Post-Q2 Boston NDR").classes("w-full").props("outlined dense")
                    # Suggest, don't pre-fill: covering analysts are OFFERED via autocomplete,
                    # but the field starts EMPTY. Auto-filling the first covering bank tagged
                    # every NDR as its sponsor unless someone remembered to change it.
                    bank_options = [f"{a['firm']} — {a['name']}" for a in CA()] + ["Other / Non-Covering Bank"]
                    sponsor_in = ui.input("Sponsoring Bank",
                                          placeholder="Covering bank, or self-sponsored — pick or type",
                                          autocomplete=bank_options).classes("w-full").props("outlined dense")
                    # Calendar-backed date field — click the field (or the
                    # calendar icon) to pick a start/end range; the input shows
                    # a friendly "Aug 20–21, 2026" and is what save_trip stores.
                    dates_in = ui.input("Dates", placeholder="Click to pick trip dates").classes("w-full").props("outlined dense")

                    def _fmt_ndr_dates():
                        v = _dates_picker.value
                        if isinstance(v, dict) and v.get("from"):
                            a = datetime.strptime(v["from"], "%Y-%m-%d")
                            b = datetime.strptime(v["to"], "%Y-%m-%d") if v.get("to") else a
                            if a == b:
                                dates_in.value = f"{a.strftime('%b')} {a.day}, {a.year}"
                            elif a.month == b.month and a.year == b.year:
                                dates_in.value = f"{a.strftime('%b')} {a.day}–{b.day}, {a.year}"
                            else:
                                dates_in.value = f"{a.strftime('%b')} {a.day} – {b.strftime('%b')} {b.day}, {b.year}"
                        elif isinstance(v, str) and v:
                            d = datetime.strptime(v, "%Y-%m-%d")
                            dates_in.value = f"{d.strftime('%b')} {d.day}, {d.year}"

                    with dates_in:
                        with ui.menu().props("no-parent-event") as _dates_menu:
                            _dates_picker = ui.date().props("range")
                            _dates_picker.on_value_change(lambda: _fmt_ndr_dates())
                        with dates_in.add_slot("append"):
                            ui.icon("edit_calendar").on("click", _dates_menu.open).classes("cursor-pointer")
                    # Per-client default cadence (layered: client override > Praxis Point global).
                    _nd = ndr_defaults(client_id)
                    type_in = ui.toggle({"in_person": "In-Person", "virtual": "Virtual"},
                                        value=("virtual" if _nd["type"] == "virtual" else "in_person"))
                    # Free-text city with autocomplete suggestions from the live
                    # Metro labels (same ones the NDR Requests tab and Metro
                    # Priority scoring use). A plain input guarantees manual entry
                    # always works — a QSelect's new-value-add did not reliably
                    # commit a typed city. Typing a metro that isn't tracked yet
                    # just yields no auto-suggested targets.
                    city_in = ui.input("City / Region", placeholder="Type a city, or pick a tracked metro",
                                       autocomplete=_ndr_location_options(institutions)).classes("w-full").props("outlined dense")
                    # Wire the request picker (created above) now that the fields it fills exist.
                    if _req_pick is not None:
                        def _prefill_from_req():
                            _r = next((x for x in _plan_open_reqs if x["id"] == _req_pick.value), None)
                            if not _r:
                                return
                            sponsor_in.value = _r.get("firm", "")
                            city_in.value = _r.get("metro") or _r.get("city", "")
                            name_in.value = (f"{_r.get('city', '')} NDR — {_r.get('firm', '')}").strip(" —")
                            ui.notify(f"Prefilled from {_r.get('analyst', '')}'s request — adjust as needed.")
                        _req_pick.on_value_change(lambda: _prefill_from_req())
                with ui.column().classes("flex-1"):
                    # Contextual default: follow the active engagement mode (Pre-/Post-earnings)
                    # instead of a fixed "Post-earnings" that fought the current context.
                    _focus_default = ("Pre-earnings — build anticipation" if mode == "pre"
                                      else "Post-earnings — deliver results story")
                    focus_in = ui.select([
                        "Post-earnings — deliver results story", "Pre-earnings — build anticipation",
                        "New institution discovery", "Existing holder relationship maintenance",
                        "Analyst initiation support", "Other",
                    ], value=_focus_default).classes("w-full").props("outlined dense label='NDR focus'")
                    team_in = ui.select([f"{v['name']} ({k})" for k, v in CT("executives", {}).items()] + [CI().get("name", "")],
                                         multiple=True).classes("w-full").props("outlined dense label='Attendees'")
                    notes_in = ui.textarea("Trip objectives", placeholder="e.g. Re-introduce to Putnam. Get Fidelity Small Cap on record post-print.").classes("w-full").props("outlined autogrow")
                    with ui.row().classes("w-full gap-3"):
                        days_in = ui.number("Days", value=_nd["days"], min=1, max=3).classes("flex-1").props("outlined dense")
                        slots_in = ui.number("Meetings per day", value=_nd["slots_per_day"], min=3, max=10).classes("flex-1").props("outlined dense")
                    # Management availability window — the hours the exec team can take meetings.
                    # Set upfront: it bounds the smart slot picker when funds are added to the trip.
                    with ui.row().classes("w-full gap-3"):
                        daystart_in = _time_picker_input("Management available from", value="8:00 AM")
                        dayend_in = _time_picker_input("… until", value="5:00 PM")

            # ── Auto-suggested targets — real tracked institutions instead of
            # app.py's hardcoded 25-row database (see _ndr_target_candidates).
            # Rebuilds whenever city/type/days/slots change, since all four
            # affect either WHICH institutions qualify or WHICH ones default-
            # check; a stale grid left over from a prior city selection would
            # silently mis-target the whole trip. ──────────────────────────
            ui.markdown("---")
            ui.label("Recommended targets").classes("font-bold")
            targets_caption = ui.label(
                "Geographically matched, non-holders prioritized. Pick a city/region above (or Virtual) to see candidates."
            ).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            target_checks = {}  # Fund name -> (checkbox, institution dict) — read by save_trip()

            def rebuild_targets(e=None):
                targets_area.clear()
                target_checks.clear()
                ndr_type = type_in.value
                location = (city_in.value or "").strip() if ndr_type == "in_person" else "Virtual"
                with targets_area:
                    if ndr_type == "in_person" and not location:
                        return
                    if ndr_type == "in_person":
                        prior = _last_city_visit(_load_json("ndr_trips.json", []), location)
                        if prior:
                            ui.label(f"Last time in {location}: {prior.get('dates','—')} — \"{prior.get('name','')}\" "
                                      f"({len(prior.get('meetings', []))} meeting(s) scheduled).").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);margin-bottom:4px;")
                        else:
                            ui.label(f"No prior NDR trip on file to {location} — this would be a first visit.").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);margin-bottom:4px;")
                    candidates = _ndr_target_candidates(institutions, location, ndr_type)
                    if not candidates:
                        ui.label(f"No tracked institutions {'score >=40 for virtual outreach' if ndr_type == 'virtual' else f'in {location}'} yet — "
                                  "add prospects via Target Database or Buy-Side Intelligence first, or add meetings to this trip "
                                  "by hand from the Active NDRs tab.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                        return
                    n_non = sum(1 for i in candidates if not i["USIO_Holder"])
                    max_auto = int((days_in.value or 1) * (slots_in.value or 5))
                    ui.label(f"{len(candidates)} tracked institution(s) — {n_non} non-holder(s) prioritized, "
                              f"{len(candidates) - n_non} existing holder(s). Top {min(max_auto, n_non)} non-holder(s) pre-checked "
                              f"for {int(days_in.value or 1)} day(s) × {int(slots_in.value or 5)} meeting(s)/day.").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);margin-bottom:6px;")
                    with ui.row().classes("w-full gap-3"):
                        col_a = ui.column().classes("flex-1 gap-2")
                        col_b = ui.column().classes("flex-1 gap-2")
                    cols = [col_a, col_b]
                    for idx, inst in enumerate(candidates):
                        default_check = idx < max_auto and not inst["USIO_Holder"]
                        days_since = _days_since_last_contact(inst["Fund"], meeting_log)
                        border_clr = "#15803D" if _score_val(inst) >= 80 else "#B45309" if _score_val(inst) >= 50 else "#94A3B8"
                        with cols[idx % 2]:
                            with ui.row().classes("w-full items-start gap-2").style(
                                    f"background:{COLORS['surface_hover_bg']};border-left:3px solid {border_clr};"
                                    f"border-radius:8px;padding:8px 10px;"):
                                cb = ui.checkbox(value=default_check)
                                with ui.column().classes("gap-0"):
                                    with ui.row().classes("items-center gap-2"):
                                        ui.label(pretty_name(inst["Fund"])).classes("font-bold").style(f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
                                        ui.label(str(inst["Engagement_Score"])).style(
                                            f"background:{border_clr}22;color:{border_clr};font-size:var(--fs-xs);font-weight:700;"
                                            f"padding:1px 6px;border-radius:6px;")
                                    ui.label(f"{inst['Metro']} · {'Holder' if inst['USIO_Holder'] else 'Non-holder'} · {inst.get('Type','—')}").style(
                                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                                    ui.label(f"Contacted {days_since}d ago" if days_since is not None else "No prior contact logged").style(
                                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                            target_checks[inst["Fund"]] = (cb, inst)

            targets_area = ui.column().classes("w-full")
            city_in.on_value_change(rebuild_targets)
            type_in.on_value_change(rebuild_targets)
            days_in.on_value_change(rebuild_targets)
            slots_in.on_value_change(rebuild_targets)
            rebuild_targets()

            def save_trip():
                if not name_in.value:
                    ui.notify("NDR name is required.", type="warning")
                    return
                selected = [(name, inst) for name, (cb, inst) in target_checks.items() if cb.value]
                days = int(days_in.value or 1)
                slots = max(1, int(slots_in.value or 5))
                # Lay the initial meetings out on a 90-min cadence starting at the management window's
                # opening time (not a hardcoded 9:00), so the seeded schedule already respects the hours.
                _day_open = _parse_time_min(daystart_in.value) or (9 * 60)
                meetings = []
                for idx, (fund_name, inst) in enumerate(selected):
                    day = idx // slots + 1
                    slot_in_day = idx % slots
                    time_label = _fmt_min_12h(_day_open + slot_in_day * 90)
                    meetings.append({
                        "institution": fund_name, "day": day, "time": time_label,
                        "type": "virtual" if type_in.value == "virtual" else "1x1",
                        "format": "Zoom" if type_in.value == "virtual" else "In-person",
                        "status": "scheduled", "notes": inst.get("Action", ""),
                        "non_holder": not inst["USIO_Holder"], "score": inst["Engagement_Score"], "contact": "",
                    })
                trips = _load_json("ndr_trips.json", [])
                trips.append({
                    "name": name_in.value, "sponsor_bank": sponsor_in.value, "dates": dates_in.value or "TBD",
                    "ndr_type": type_in.value, "city": city_in.value or ("Virtual" if type_in.value == "virtual" else "TBD"),
                    "focus": focus_in.value, "team": team_in.value or [], "notes": notes_in.value,
                    "day_start": (daystart_in.value or "8:00 AM").strip(),
                    "day_end": (dayend_in.value or "5:00 PM").strip(),
                    "meetings": meetings, "status": "Planning", "debrief": {},
                    "created": datetime.now().strftime("%Y-%m-%d"),
                })
                _save_json("ndr_trips.json", trips)
                n_non = sum(1 for _, i in selected if not i["USIO_Holder"])
                ui.notify(f"NDR '{name_in.value}' created — {len(selected)} meeting(s) "
                          f"({n_non} non-holders, {len(selected) - n_non} holders) across {days} day(s). "
                          "Now showing under Active NDRs.")
                # Refresh just the Active NDRs list and switch to it, instead of
                # a full page reload that bounced the user back to Buy-Side
                # Intelligence (the "I created it but where did it go?" problem).
                _refresh_ndr()
                ndr_tabs.set_value(nt2)

            ui.button("Create NDR Trip", on_click=save_trip).props("color=primary").style("margin-top:8px;")

        # Remember which trips' "NDR pipeline" expander is open, so a status change (Contact → mark
        # invited, Decline, Confirm…) that repaints this panel keeps the pipeline OPEN instead of
        # collapsing it shut and scrolling away. Defined OUTSIDE the refreshable so it survives rebuilds.
        _pipe_open = {}

        def _refresh_ndr():
            # Defer the refresh out of the current click/dialog handler (once=True timer) so the
            # refreshable reliably repaints — a direct in-handler .refresh() from a dialog that lives
            # INSIDE this panel sometimes didn't repaint the schedule ("saved but no change on screen").
            ui.timer(0.01, lambda: getattr(_active_ndrs_panel, "refresh")(), once=True)

        @ui.refreshable
        def _active_ndrs_panel():
            trips = _load_json("ndr_trips.json", [])
            # Active NDRs = Planning / In Progress only. Completed roadshows live under Post-NDR
            # Debrief; showing them here made a finished trip read as unfinished/past-due (the
            # redundancy flagged in review). The per-card status dropdown graduates a trip to Completed.
            _active_idxs = {i for i, t in enumerate(trips) if t.get("status") != "Completed"}
            if not _active_idxs:
                ui.label("No active NDRs. Plan one from the metro table (Buyside Ownership → click a "
                         "metro), or review finished roadshows under Post-NDR Debrief.").style(
                    f"color:{COLORS['text_muted']};")
            status_options = ["scheduled", "completed", "cancelled", "no-show"]

            # ── Shortlist status transitions (NDR pipeline Phase 2) ──────────────
            def _sl_set(ti, si, new_status):
                trips_ = _load_json("ndr_trips.json", [])
                try:
                    entry = trips_[ti]["shortlist"][si]
                except (IndexError, KeyError):
                    _refresh_ndr(); return
                entry["status"] = new_status
                stamp = {"invited": "contacted_at", "confirmed": "confirmed_at",
                         "declined": "declined_at"}.get(new_status)
                if stamp:
                    entry[stamp] = datetime.now().strftime("%Y-%m-%d %H:%M")
                _save_json("ndr_trips.json", trips_)
                try:
                    from core import activity_log
                    ev = {"invited": "ndr_outreach", "confirmed": "ndr_confirmed",
                          "declined": "ndr_declined"}.get(new_status, "ndr_update")
                    activity_log.log_event(ev, entity=entry.get("institution", ""),
                                           launched_from=f"NDR · {trips_[ti].get('name', '')}")
                except Exception:
                    pass
                _refresh_ndr()

            def _sl_contact(ti, si, entry):
                contact = get_institution_contacts().get(entry.get("institution", ""), {})
                _open_shortlist_outreach(entry, contact, lambda: _sl_set(ti, si, "invited"))

            # Confirm → move the target out of the pipeline and into the schedule at the next open
            # slot (NDR pipeline Phase 3). This is the "slot on confirmation" rule: only confirmed
            # targets take calendar capacity; declines never do.
            def _sl_confirm_and_slot(ti, si):
                trips_ = _load_json("ndr_trips.json", [])
                try:
                    trip_ = trips_[ti]
                    entry = trip_["shortlist"][si]
                except (IndexError, KeyError):
                    _refresh_ndr(); return
                day, slot = _next_open_slot(trip_)
                if day is None:
                    ui.notify("This NDR's slots are full — raise its days or meetings/day to schedule more.",
                              type="warning")
                    return
                trip_.setdefault("meetings", []).append({
                    "institution": entry.get("institution"), "day": day, "slot_index": slot,
                    "time": _slot_time(slot), "type": "1x1", "format": "In-person", "status": "scheduled",
                    "notes": entry.get("peers", ""), "non_holder": True, "score": entry.get("conviction"),
                    "contact": "", "confirmed_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "source": entry.get("source", "outbound"),
                })
                del trip_["shortlist"][si]
                _save_json("ndr_trips.json", trips_)
                try:
                    from core import activity_log
                    activity_log.log_event("ndr_confirmed", entity=entry.get("institution", ""),
                                           launched_from=f"NDR · {trip_.get('name', '')}")
                except Exception:
                    pass
                ui.notify(f"{pretty_name(entry.get('institution', ''))} confirmed → Day {day} · {_slot_time(slot)}.",
                          type="positive")
                _refresh_ndr()

            def _set_capacity(ti, field, value):
                # Persist silently (no refresh) so the number input keeps focus while you type.
                trips_ = _load_json("ndr_trips.json", [])
                try:
                    trips_[ti][field] = max(1, int(value or 1))
                except (IndexError, ValueError, TypeError):
                    return
                _save_json("ndr_trips.json", trips_)

            def _set_window(ti, field, value):
                # Management window (day_start / day_end) + lodging. Saved on blur — with a confirming
                # toast, since there's no explicit Save button on these inline fields.
                trips_ = _load_json("ndr_trips.json", [])
                try:
                    trips_[ti][field] = (value or "").strip()
                except (IndexError, TypeError):
                    return
                _save_json("ndr_trips.json", trips_)
                ui.notify(f"{ {'hotel': 'Lodging', 'day_start': 'Management start', 'day_end': 'Management end'}.get(field, field)} saved.",
                          type="positive")
                if field == "hotel":
                    _refresh_ndr()   # the morning pickup + end-of-day return legs recompute with the new address

            for idx, trip in enumerate(trips):
                if trip.get("status") == "Completed":
                    continue                      # Completed → Post-NDR Debrief tab, not here
                all_meetings = trip.get("meetings", [])
                # A malformed trip (e.g. legacy seed data where meetings is a COUNT, not a list)
                # must not crash the whole NDR Planner tab — treat a non-list as no meetings.
                if not isinstance(all_meetings, list):
                    all_meetings = []
                real_meetings = [m for m in all_meetings if isinstance(m, dict) and m.get("type") != "break"]
                done_m = sum(1 for m in real_meetings if m.get("status") == "completed")
                non_h = sum(1 for m in real_meetings if m.get("non_holder", True))
                trip_status = trip.get("status", "Planning")
                if trip_status not in ("Planning", "In Progress", "Completed"):   # legacy/seed values
                    trip_status = {"complete": "Completed", "completed": "Completed",
                                   "in progress": "In Progress"}.get(str(trip_status).lower(), "Planning")
                status_icon = {"Planning": "", "In Progress": "", "Completed": ""}.get(trip_status, "")
                with ui.card().classes("w-full").style(f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};"):
                    with ui.row().classes("w-full justify-between items-start"):
                        with ui.column().classes("gap-0"):
                            _tname = trip.get("name") or trip.get("city") or "NDR trip"
                            _tdates = trip.get("dates") or trip.get("date") or "—"
                            _ttype = "Virtual" if trip.get("ndr_type") == "virtual" else "In-Person"
                            _tsponsor = trip.get("sponsor_bank") or trip.get("sponsor") or "—"
                            ui.label(f"{status_icon} {_tname}").classes("font-bold").style(f"color:{COLORS['text_heading']};")
                            ui.label(f"{_tdates} · {trip.get('city','—')} · {_ttype} · Sponsor: {_tsponsor}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                            ui.label(f"Focus: {trip.get('focus','—')} · Attendees: {', '.join(trip.get('team',[])) or '—'}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                            if trip.get("notes"):
                                ui.label(f"Objectives: {trip['notes']}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                        with ui.column().classes("items-end gap-1"):
                            ui.label(f"{done_m}/{len(real_meetings)} completed · {non_h} non-holder(s)").style(f"color:{COLORS['accent_light']};font-size:var(--fs-sm);")

                            def set_trip_status(e, idx=idx):
                                trips_ = _load_json("ndr_trips.json", [])
                                trips_[idx]["status"] = e.value
                                _save_json("ndr_trips.json", trips_)
                                ui.notify(f"Trip marked {e.value}.")
                                _refresh_ndr()

                            ui.select(["Planning", "In Progress", "Completed"], value=trip_status,
                                      on_change=set_trip_status).props("dense outlined").classes("min-w-[130px]")

                    # Capacity grid (NDR pipeline Phase 3) — days × meetings/day. Confirmed targets
                    # fill it in order. Editable inline; persists silently to keep input focus.
                    _cap_d, _cap_p = _ndr_capacity(trip)
                    _filled = len([m for m in all_meetings if m.get("type") != "break"])
                    with ui.row().classes("w-full items-center gap-2").style("margin-top:4px;"):
                        ui.label("Capacity").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        ui.number(value=_cap_d, min=1, max=10,
                                  on_change=lambda e, ti=idx: _set_capacity(ti, "days", e.value)).props("dense outlined").style("width:66px;").tooltip("Days")
                        ui.label("days ×").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        ui.number(value=_cap_p, min=1, max=12,
                                  on_change=lambda e, ti=idx: _set_capacity(ti, "slots_per_day", e.value)).props("dense outlined").style("width:66px;").tooltip("Meetings per day")
                        ui.label(f"/day  ·  {_filled} of {_cap_d * _cap_p} slots filled").style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

                    # Management availability window — the exec team's meeting hours. Bounds the smart
                    # slot picker when a fund is added. Defaults to 8:00 AM–5:00 PM until set.
                    with ui.row().classes("w-full items-center gap-2").style("margin-top:2px;"):
                        ui.icon("schedule").style(f"color:{COLORS['text_muted']};font-size:var(--fs-base);")
                        ui.label("Management hours").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        _ws_in = ui.input(value=trip.get("day_start") or "8:00 AM").props(
                            "dense outlined").style("width:104px;").tooltip("Earliest meeting start")
                        ui.label("to").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        _we_in = ui.input(value=trip.get("day_end") or "5:00 PM").props(
                            "dense outlined").style("width:104px;").tooltip("Latest meeting end")
                        _ws_in.on("blur", lambda e, ti=idx, el=_ws_in: _set_window(ti, "day_start", el.value))
                        _we_in.on("blur", lambda e, ti=idx, el=_we_in: _set_window(ti, "day_end", el.value))

                    # Where management is staying — the itinerary opens from here (pickup time to make
                    # the first meeting is computed off the routed drive from this address).
                    with ui.row().classes("w-full items-center gap-2").style("margin-top:2px;"):
                        ui.icon("hotel").style(f"color:{COLORS['text_muted']};font-size:var(--fs-base);")
                        ui.label("Lodging").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        _hotel_in = ui.input(value=trip.get("hotel") or "",
                                             placeholder="Hotel name & street address — powers the pickup time").props(
                            "dense outlined").classes("flex-1").tooltip("Where management is staying")
                        _hotel_in.on("blur", lambda e, ti=idx, el=_hotel_in: _set_window(ti, "hotel", el.value))

                    # NDR pipeline (Phase 2) — shortlisted → invited → confirmed → declined.
                    # Held in trip['shortlist'], separate from the slot schedule below (Phase 3
                    # moves a confirmed target into meetings[] with a slot).
                    shortlist = trip.get("shortlist", [])
                    if shortlist:
                        _counts = {}
                        for s in shortlist:
                            k = s.get("status", "shortlisted")
                            _counts[k] = _counts.get(k, 0) + 1
                        _order = ["shortlisted", "invited", "confirmed", "declined"]
                        _summ = " · ".join(f"{_counts[k]} {k}" for k in _order if _counts.get(k))
                        _badge = {"shortlisted": ("#64748B", "Shortlisted"), "invited": ("#B45309", "Invited"),
                                  "confirmed": ("#15803D", "Confirmed"), "declined": ("#94A3B8", "Declined")}
                        _pipe_exp = ui.expansion(f"NDR pipeline — {len(shortlist)} target(s) · {_summ}",
                                                 icon="filter_alt", value=_pipe_open.get(idx, False)).classes(
                                                 "w-full").style("margin-top:6px;")
                        _pipe_exp.on_value_change(lambda e, ti=idx: _pipe_open.update({ti: e.value}))
                        with _pipe_exp:
                            for si, s in enumerate(shortlist):
                                st = s.get("status", "shortlisted")
                                clr, lbl = _badge.get(st, ("#64748B", st.title()))
                                with ui.row().classes("w-full items-center gap-2").style(
                                        f"background:{COLORS['surface_hover_bg']};border-radius:6px;padding:4px 8px;"
                                        f"margin:2px 0;{'opacity:.55;' if st == 'declined' else ''}"):
                                    ui.label(lbl).style(f"background:{clr}22;color:{clr};border-radius:6px;"
                                                        "padding:1px 8px;font-size:var(--fs-2xs);font-weight:700;white-space:nowrap;")
                                    ui.label(pretty_name(s.get("institution", ""))).classes("flex-1 name-link").on(
                                        "click", lambda s=s: _open_account_profile(
                                            {"Fund": s.get("institution"), "cik": s.get("cik"),
                                             "city": s.get("city"), "state": s.get("state"),
                                             "conviction": s.get("conviction")})).style(
                                        f"color:{COLORS['text_body']};font-size:var(--fs-base);")
                                    _sl_loc = ", ".join(x for x in [s.get("city"), s.get("state")] if x) or s.get("metro", "—")
                                    ui.label(_sl_loc).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                                    if s.get("conviction") is not None:
                                        ui.label(f"{s['conviction']}/100").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                                    # Per-status actions
                                    if st == "shortlisted":
                                        ui.button("Contact", icon="mail",
                                                  on_click=lambda ti=idx, si=si, e=s: _sl_contact(ti, si, e)).props("flat dense size=sm color=primary")
                                        ui.button("Decline", on_click=lambda ti=idx, si=si: _sl_set(ti, si, "declined")).props("flat dense size=sm")
                                    elif st == "invited":
                                        ui.button("Confirm & slot", icon="event_available",
                                                  on_click=lambda ti=idx, si=si: _sl_confirm_and_slot(ti, si)).props("flat dense size=sm color=positive")
                                        ui.button("Re-contact", on_click=lambda ti=idx, si=si, e=s: _sl_contact(ti, si, e)).props("flat dense size=sm")
                                        ui.button("Decline", on_click=lambda ti=idx, si=si: _sl_set(ti, si, "declined")).props("flat dense size=sm")
                                    elif st == "confirmed":
                                        # Legacy Phase-2 confirmed (still in the pipeline) — slot it now.
                                        ui.button("Slot now", icon="event_available",
                                                  on_click=lambda ti=idx, si=si: _sl_confirm_and_slot(ti, si)).props("flat dense size=sm color=positive")
                                        ui.button("Undo", on_click=lambda ti=idx, si=si: _sl_set(ti, si, "invited")).props("flat dense size=sm")
                                    elif st == "declined":
                                        ui.button("Restore", on_click=lambda ti=idx, si=si: _sl_set(ti, si, "shortlisted")).props("flat dense size=sm")
                            ui.label("Contact opens a draft you send yourself and marks the target Invited (logged). "
                                     "Confirm → automatic slot assignment arrives in Phase 3.").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);margin-top:4px;")

                    # Sort by DAY then TIME (not insertion order) so a hand-added meeting lands in its
                    # chronological slot — and so the travel legs / hotel pickup below are computed
                    # between the right chronological neighbours. Untimed meetings sort last; the
                    # original index (x[0]) is the stable tiebreaker and is still the delete/edit key.
                    meetings_with_idx = sorted(
                        enumerate(all_meetings),
                        key=lambda x: (x[1].get("day", 1),
                                       _parse_time_min(x[1].get("time")) if _parse_time_min(x[1].get("time")) is not None else 9999,
                                       x[0]))
                    current_day = None
                    _prev_m = None
                    _leg_miles, _leg_count, _routed_count = 0.0, 0, 0
                    try:
                        from core import routing as _routing_mod
                        _routing_on = _routing_mod.is_configured()
                    except Exception:
                        _routing_on = False
                    for flat_idx, m in meetings_with_idx:
                        d = m.get("day", 1)
                        if d != current_day:
                            current_day = d
                            _prev_m = None   # the travel chain restarts each day
                            ui.label(f"Day {d}" if len(meetings_with_idx) and any(mm.get("day", 1) != 1 for _, mm in meetings_with_idx) else "Meetings").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-weight:700;margin-top:8px;")
                        # Day opens from the hotel: where management is staying and the pickup time to
                        # make the first meeting (first-meeting start − routed drive − buffer).
                        if _prev_m is None and (trip.get("hotel") or "").strip():
                            _pk, _plg = _hotel_departure(trip, m)
                            if _plg:
                                with ui.row().classes("items-start gap-1").style("margin:0 0 2px 20px;"):
                                    ui.icon("hotel").style(f"color:{COLORS['accent']};font-size:var(--fs-sm);margin-top:2px;")
                                    with ui.column().classes("gap-0"):
                                        ui.label(f"Staying at {trip['hotel']}").style(
                                            f"color:{COLORS['text_body']};font-size:var(--fs-xs);font-weight:600;")
                                        ui.label((f"Pickup {_pk} — " if _pk else "") + f"{_plg['label']} to "
                                                 f"{pretty_name(m.get('institution', ''))}").style(
                                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        # Inline travel leg from the previous in-person stop — REAL routed driving
                        # miles/time when both stops have a street address (OpenRouteService), else a
                        # city-level estimate. This is the "allow time in the schedule" read, shown
                        # right where you plan the day instead of only in the downloaded itinerary.
                        if _prev_m is not None:
                            _lg, _tight = _travel_leg_between(_prev_m, m, trip)
                            if _lg:
                                _leg_miles += _lg["miles"]; _leg_count += 1
                                _routed_count += 1 if _lg.get("basis") == "routed" else 0
                                with ui.row().classes("items-center gap-1").style("margin:0 0 2px 20px;"):
                                    ui.icon("directions_car").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                                    ui.label(_lg["label"]).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                                    if _tight:
                                        ui.label("· " + _tight).style(f"color:#B45309;font-size:var(--fs-xs);font-weight:600;")
                        nh_badge = "" if m.get("non_holder", True) else ""
                        fmt_badge = "" if m.get("format") in ("Zoom", "Teams") else ""
                        with ui.row().classes("w-full items-center gap-2").style(
                                f"background:{COLORS['surface_hover_bg']};border-radius:6px;padding:4px 8px;margin:2px 0;"):
                            ui.label(f"{fmt_badge} {_meeting_time_range(m)}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);width:150px;")
                            with ui.row().classes("flex-1 items-center gap-1").style("min-width:0;"):
                                ui.label(f"{nh_badge} {pretty_name(m.get('institution',''))}").style(f"color:{COLORS['text_body']};font-size:var(--fs-base);")
                                if m.get("source") == "inbound":
                                    ui.label("inbound").style("background:#B4530922;color:#B45309;border-radius:6px;"
                                                              "padding:0 6px;font-size:var(--fs-micro);font-weight:700;")
                            if m.get("score") is not None:
                                ui.label(f"{m['score']}/100").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

                            def set_meeting_status(e, idx=idx, flat_idx=flat_idx):
                                trips_ = _load_json("ndr_trips.json", [])
                                trips_[idx]["meetings"][flat_idx]["status"] = e.value
                                _save_json("ndr_trips.json", trips_)
                                _refresh_ndr()

                            ui.select(status_options, value=m.get("status", "scheduled"),
                                      on_change=set_meeting_status).props("dense outlined").classes("min-w-[110px]")

                            # Call / email the contact at this institution —
                            # to reschedule or send materials. Uses the known
                            # institution contact; each icon only appears when
                            # we actually have that number/email on file (a
                            # hand-added firm with no contact shows neither).
                            _mc = get_institution_contacts().get(m.get("institution", ""), {})
                            if _mc.get("phone"):
                                with ui.link(target=_tel_href(_mc["phone"])).tooltip(
                                        f"Call {_mc.get('name', '')} · {_mc['phone']}"):
                                    ui.icon("call").style(f"color:{COLORS['accent_light']};font-size:var(--fs-xl);")
                            if _mc.get("email"):
                                _first = _mc.get("name", "").split()[0] if _mc.get("name") else ""
                                _slot = f" for our {m.get('time')} slot" if m.get("time") and m.get("time") != "—" else ""
                                _subj = f"{CT('ticker')} — {m.get('institution', '')} meeting during {trip.get('city', '')} NDR"
                                _bodym = (f"Hi {_first},\n\nLooking forward to our meeting during the "
                                          f"{trip.get('city', '')} NDR ({trip.get('dates', '')}). I wanted to confirm "
                                          f"logistics{_slot} and share some materials ahead of time.\n\n")
                                _href = f"mailto:{_mc['email']}?subject={quote(_subj)}&body={quote(_bodym)}"
                                with ui.link(target=_href).tooltip(f"Email {_mc.get('name', '')} · {_mc['email']}"):
                                    ui.icon("mail").style(f"color:{COLORS['accent_light']};font-size:var(--fs-xl);")

                            # Per-meeting remove — confirmed, since it's
                            # destructive and there's no undo. flat_idx is the
                            # index into the trip's own meetings list, so the
                            # delete lands on the right row even though the
                            # display is re-sorted by day.
                            _m_name = m.get("institution") or "this meeting"
                            with ui.dialog() as _del_dialog, ui.card():
                                ui.label(f"Remove {_m_name} from this trip?").classes("font-bold")
                                ui.label("This can't be undone.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                                with ui.row().classes("justify-end w-full"):
                                    ui.button("Cancel", on_click=_del_dialog.close).props("flat dense")

                                    def do_remove(idx=idx, flat_idx=flat_idx, name=_m_name, dlg=_del_dialog):
                                        trips_ = _load_json("ndr_trips.json", [])
                                        try:
                                            del trips_[idx]["meetings"][flat_idx]
                                        except (IndexError, KeyError):
                                            dlg.close()
                                            ui.notify("List changed — reloading.", type="warning")
                                            _refresh_ndr()
                                            return
                                        _save_json("ndr_trips.json", trips_)
                                        dlg.close()
                                        ui.notify(f"Removed {name} from trip.")
                                        _refresh_ndr()

                                    ui.button("Remove", on_click=do_remove).props("color=negative dense")

                            # Edit meeting — set/adjust time, format, address, and
                            # (for a virtual meeting) the join link that flows to
                            # the itinerary. The key path for making a stop virtual
                            # and pushing its link to the schedule.
                            with ui.dialog() as _edit_dialog, ui.card().classes("w-96"):
                                ui.label(f"Edit — {_m_name}").classes("font-bold").style(f"color:{COLORS['text_heading']};")
                                with ui.row().classes("w-full gap-3 items-end"):
                                    _e_time = ui.input("Time", value=m.get("time", "")).classes("flex-1").props("outlined dense")
                                    with _e_time:
                                        with ui.menu().props("no-parent-event") as _etmenu:
                                            ui.time(on_change=lambda e, ti=_e_time: ti.set_value(_fmt_time_12h(e.value)))
                                        with _e_time.add_slot("append"):
                                            ui.icon("schedule").on("click", _etmenu.open).classes("cursor-pointer")
                                    _e_fmt = ui.select(["In-person", "Virtual", "Zoom", "Teams", "Phone", "Jitsi"],
                                                       value=m.get("format", "In-person"),
                                                       label="Format").props("dense outlined").classes("w-32")
                                _e_addr = ui.input("Address / location", value=m.get("address", "")).classes("w-full").props("outlined dense")
                                _e_lunch = ui.checkbox("Catered lunch meeting (~1h15)", value=bool(m.get("lunch"))).props("dense")
                                _e_diet = ui.input("Dietary requests", value=m.get("dietary", ""),
                                                   placeholder="e.g. 2 vegetarian, 1 gluten-free · no nuts").classes("w-full").props("outlined dense")
                                _e_diet.bind_visibility_from(_e_lunch, "value")
                                _e_link = ui.input("Meeting link", value=m.get("meeting_link", ""),
                                                   placeholder="Zoom / Teams / Meet URL").classes("w-full").props("outlined dense")
                                _e_link.bind_visibility_from(_e_fmt, "value", backward=lambda v: v != "In-person")

                                async def create_zoom(idx=idx, flat_idx=flat_idx, trip=trip, m=m,
                                                      e_time=_e_time, e_link=_e_link, dlg=_edit_dialog):
                                    import asyncio
                                    from zoneinfo import ZoneInfo
                                    from core import ndr_calendar, zoom_meetings
                                    if not zoom_meetings.is_configured():
                                        ui.notify("Add Zoom credentials in Settings → Data Sources first.", type="warning")
                                        return
                                    m_now = dict(m); m_now["time"] = (e_time.value or "").strip()
                                    local = ndr_calendar.meeting_datetime(trip, m_now)
                                    start_utc = local.astimezone(ZoneInfo("UTC")) if local else None
                                    tzname = ndr_calendar.tz_name_for(trip.get("city"))
                                    topic = f"{CT('ticker')} NDR — {m.get('institution', '')}"
                                    ui.notify("Creating Zoom meeting…")
                                    try:
                                        res = await asyncio.to_thread(
                                            zoom_meetings.create_meeting, topic, start_utc, 45, tzname)
                                    except Exception as ex:
                                        ui.notify(f"Zoom create failed: {ex}", type="negative")
                                        return
                                    link = res.get("join_url") or ""
                                    e_link.value = link
                                    # Persist immediately so a created Zoom meeting is never orphaned.
                                    trips_ = _load_json("ndr_trips.json", [])
                                    try:
                                        mm = trips_[idx]["meetings"][flat_idx]
                                    except (IndexError, KeyError):
                                        ui.notify("List changed — reloading.", type="warning")
                                        dlg.close(); _refresh_ndr(); return
                                    mm["format"] = "Zoom"; mm["meeting_link"] = link
                                    mm["zoom_meeting_id"] = str(res.get("id") or "")
                                    _save_json("ndr_trips.json", trips_)
                                    ui.notify("Zoom meeting created — join link added to the schedule.", type="positive")

                                _zoom_btn = ui.button("Create Zoom meeting", icon="videocam",
                                                      on_click=create_zoom).props("outline dense color=primary")
                                _zoom_btn.bind_visibility_from(_e_fmt, "value", backward=lambda v: v == "Zoom")

                                def create_jitsi(idx=idx, flat_idx=flat_idx, m=m, e_link=_e_link, dlg=_edit_dialog):
                                    from core import jitsi_meetings
                                    link = jitsi_meetings.create_link(f"{CT('ticker')} NDR {m.get('institution', '')}")
                                    e_link.value = link
                                    trips_ = _load_json("ndr_trips.json", [])
                                    try:
                                        mm = trips_[idx]["meetings"][flat_idx]
                                    except (IndexError, KeyError):
                                        ui.notify("List changed — reloading.", type="warning")
                                        dlg.close(); _refresh_ndr(); return
                                    mm["format"] = "Jitsi"; mm["meeting_link"] = link
                                    _save_json("ndr_trips.json", trips_)
                                    ui.notify("Jitsi link generated — added to the schedule.", type="positive")

                                _jitsi_btn = ui.button("Generate Jitsi link", icon="link",
                                                       on_click=create_jitsi).props("outline dense color=primary")
                                _jitsi_btn.bind_visibility_from(_e_fmt, "value", backward=lambda v: v == "Jitsi")
                                with ui.row().classes("justify-end w-full"):
                                    ui.button("Cancel", on_click=_edit_dialog.close).props("flat dense")

                                    def do_edit(idx=idx, flat_idx=flat_idx, dlg=_edit_dialog,
                                                e_time=_e_time, e_fmt=_e_fmt, e_addr=_e_addr, e_link=_e_link,
                                                e_lunch=_e_lunch, e_diet=_e_diet):
                                        trips_ = _load_json("ndr_trips.json", [])
                                        try:
                                            mm = trips_[idx]["meetings"][flat_idx]
                                        except (IndexError, KeyError):
                                            dlg.close(); ui.notify("List changed — reloading.", type="warning")
                                            _refresh_ndr(); return
                                        mm["time"] = (e_time.value or "").strip() or "—"
                                        mm["format"] = e_fmt.value
                                        mm["address"] = (e_addr.value or "").strip()
                                        mm["lunch"] = bool(e_lunch.value)
                                        mm["dietary"] = (e_diet.value or "").strip() if e_lunch.value else ""
                                        mm["meeting_link"] = (e_link.value or "").strip() if e_fmt.value != "In-person" else ""
                                        _save_json("ndr_trips.json", trips_)
                                        dlg.close()
                                        ui.notify("Meeting updated.")
                                        _refresh_ndr()

                                    ui.button("Save", on_click=do_edit).props("color=primary dense")

                            ui.button(icon="edit", on_click=_edit_dialog.open).props(
                                "flat dense round size=sm color=grey-7").tooltip("Edit meeting")
                            ui.button(icon="close", on_click=_del_dialog.open).props(
                                "flat dense round size=sm color=grey-7").tooltip("Remove meeting")
                        # Who they're meeting at the firm — the meeting's own
                        # contact if set, else fall back to the known
                        # institution contact so seeded meetings still show a
                        # name instead of a blank.
                        _known_c = get_institution_contacts().get(m.get("institution", ""), {})
                        _who = m.get("contact") or (
                            ", ".join(p for p in (_known_c.get("name"), _known_c.get("title")) if p))
                        if _who:
                            ui.label(f"   Meeting with: {_who}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-left:6px;")
                        if m.get("address"):
                            ui.label(f"   Location: {m['address']}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-left:6px;")
                        if m.get("format", "In-person") != "In-person":
                            _lk = (m.get("meeting_link") or "").strip()
                            if _lk:
                                with ui.row().classes("items-center gap-1").style("margin-left:6px;"):
                                    ui.label(f"   Join ({m.get('format')}):").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                                    with ui.link(target=_lk, new_tab=True):
                                        ui.label(_lk[:48] + ("…" if len(_lk) > 48 else "")).style(f"color:{COLORS['accent_light']};font-size:var(--fs-xs);")
                            else:
                                ui.label(f"   {m.get('format')} · link TBD — click ✎ to add").style(
                                    f"color:#B45309;font-size:var(--fs-xs);margin-left:6px;")
                        if m.get("notes"):
                            ui.label(f"   {m['notes']}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-left:6px;")
                        if m.get("lunch"):
                            with ui.row().classes("items-center gap-1").style("margin-left:6px;"):
                                ui.icon("restaurant").style(f"color:#B45309;font-size:var(--fs-sm);")
                                _dtxt = "Catered working lunch (~1h15)"
                                if (m.get("dietary") or "").strip():
                                    _dtxt += f" · Dietary: {m['dietary'].strip()}"
                                else:
                                    _dtxt += " · Dietary: — (confirm with caterer)"
                                ui.label(_dtxt).style(f"color:#B45309;font-size:var(--fs-xs);font-weight:600;")
                        _prev_m = m

                    # Trip ending: the last meeting → back to where management is staying. Mirrors the
                    # morning pickup and its drive counts toward the day's total.
                    if _prev_m is not None and (trip.get("hotel") or "").strip():
                        _arr, _rlg = _hotel_return(trip, _prev_m)
                        if _rlg:
                            _leg_miles += _rlg["miles"]; _leg_count += 1
                            _routed_count += 1 if _rlg.get("basis") == "routed" else 0
                            with ui.row().classes("items-start gap-1").style("margin:2px 0 2px 20px;"):
                                ui.icon("hotel").style(f"color:{COLORS['accent']};font-size:var(--fs-sm);margin-top:2px;")
                                with ui.column().classes("gap-0"):
                                    ui.label(f"Return to {trip['hotel']}").style(
                                        f"color:{COLORS['text_body']};font-size:var(--fs-xs);font-weight:600;")
                                    ui.label((f"Back by {_arr} — " if _arr else "") + f"{_rlg['label']} from "
                                             f"{pretty_name(_prev_m.get('institution', ''))}").style(
                                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

                    # Trip-level travel total — honestly labelled as routed (driving) vs city-level
                    # straight-line, per _travel_total_line.
                    _tl_line = _travel_total_line(_leg_miles, _leg_count, _routed_count, _routing_on)
                    if _tl_line:
                        with ui.row().classes("items-center gap-1").style("margin-top:6px;"):
                            ui.icon("route").style(f"color:{COLORS['text_secondary']};font-size:var(--fs-base);")
                            ui.label(_tl_line).style(f"color:{COLORS['text_secondary']};font-size:var(--fs-xs);font-weight:600;")

                    # Filler-meeting finder — a light day (like 3 meetings)
                    # is the IR person's cue to fill open slots. Rather than
                    # hunting around, surface tracked funds in this trip's metro
                    # that aren't already scheduled, ranked by engagement score
                    # (non-holders first — the conversion targets), each with
                    # one-click Call / Email / Add-to-trip. Same targeting
                    # engine as Plan New NDR.
                    _scheduled = {mm.get("institution") for mm in all_meetings}
                    _all_cands = [c for c in _ndr_target_candidates(
                                      institutions, trip.get("city"), trip.get("ndr_type", "in-person"))
                                  if c["Fund"] not in _scheduled] if trip_status != "Completed" else []
                    _fillers = _all_cands[:6]
                    if _fillers:
                        with ui.expansion(
                                f"Fill open slots — {len(_all_cands)} available targets in {trip.get('city', 'this metro')}",
                                value=len(real_meetings) < 4).classes("w-full").style("margin-top:6px;"):
                            _n_non = sum(1 for c in _all_cands if not c.get("USIO_Holder", False))
                            _more = f" Showing the top {len(_fillers)}." if len(_all_cands) > len(_fillers) else ""
                            ui.label(f"Tracked funds in this metro you're not already meeting, ranked by engagement "
                                     f"score — non-holders first ({_n_non} non-holder{'s' if _n_non != 1 else ''}, "
                                     f"{len(_all_cands) - _n_non} holder{'s' if (len(_all_cands) - _n_non) != 1 else ''}). "
                                     f"Call or email to invite, then add them to the trip.{_more}").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                            # Street addresses (where available) so you can see WHERE each candidate is
                            # before adding it — and so an added filler carries the address into the
                            # itinerary's routed drive-time calc.
                            from core import fund_addresses
                            for c in _fillers:
                                cc = get_institution_contacts().get(c["Fund"], {})
                                # Canonical accessor — normalizes both stored shapes (structured
                                # dict for SEC-sourced tenants, bare string for the demo) to a str.
                                _c_addr = fund_addresses.address_for(c["Fund"]) or ""
                                with ui.row().classes("w-full items-center gap-2").style(
                                        f"background:{COLORS['surface_hover_bg']};border-radius:6px;padding:4px 8px;margin:2px 0;"):
                                    _cs = _score_val(c)
                                    ui.label(str(round(_cs)) if _cs else "—").style(
                                        f"color:{COLORS['accent_light']};font-size:var(--fs-sm);font-weight:700;width:34px;")
                                    with ui.column().classes("gap-0 flex-1"):
                                        ui.label(c["Fund"]).style(f"color:{COLORS['text_body']};font-size:var(--fs-base);font-weight:600;")
                                        _bits = [_c_addr or c.get("Metro"), c.get("Type")]
                                        if cc.get("name"):
                                            _bits.append(f"{cc['name']} ({cc.get('title', '')})")
                                        _bits.append(cc.get("phone") or "no contact on file")
                                        ui.label(" · ".join(x for x in _bits if x)).style(
                                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                                    if cc.get("phone"):
                                        ui.link("Call", _tel_href(cc["phone"])).style(
                                            f"color:{COLORS['accent_light']};font-size:var(--fs-sm);")
                                    if cc.get("email"):
                                        _first = cc.get("name", "").split()[0] if cc.get("name") else ""
                                        _mailto(cc["email"],
                                                f"{CT('ticker')} — meeting during {trip.get('city', '')} NDR?",
                                                f"Hi {_first},\n\nWe'll be in {trip.get('city', '')} for a non-deal "
                                                f"roadshow ({trip.get('dates', '')}). Would you have time to meet with "
                                                f"{CT('name')} management? Happy to work around your schedule.\n\n",
                                                "Email")

                                    def add_filler(c=c, cc=cc, idx=idx, _addr=_c_addr):
                                        _open_add_to_trip_dialog(
                                            idx, c["Fund"],
                                            ", ".join(p for p in (cc.get("name"), cc.get("title")) if p),
                                            not c.get("USIO_Holder", False),
                                            c.get("Engagement_Score"), c.get("Metro"),
                                            _refresh_ndr, default_address=_addr)

                                    ui.button("Add to trip", on_click=add_filler).props("flat dense color=primary")

                    with ui.expansion("Add a meeting to this trip by hand").classes("w-full").style("margin-top:6px;"):
                        with ui.row().classes("w-full gap-4"):
                            inst_in = ui.input("Institution *").classes("flex-1").props("outlined dense")
                            contact_in = ui.input(
                                "Who you're meeting (name / title)",
                                placeholder="e.g. Jane Smith, PM; John Doe, Analyst").classes("flex-1").props("outlined dense")
                        with ui.row().classes("w-full gap-4"):
                            address_in = ui.input(
                                "Address / location",
                                placeholder="e.g. 55 E 52nd St, 12th Fl, New York, NY").classes("flex-1").props("outlined dense")
                        with ui.row().classes("w-full gap-4 items-end"):
                            day_in = ui.number("Day", value=1, min=1, step=1).classes("w-20").props("outlined dense")
                            time_in = ui.input("Time", placeholder="e.g. 2:00 PM").classes("flex-1").props("outlined dense")
                            with time_in:
                                with ui.menu().props("no-parent-event") as _tmenu:
                                    ui.time(on_change=lambda e, ti=time_in: ti.set_value(_fmt_time_12h(e.value)))
                                with time_in.add_slot("append"):
                                    ui.icon("schedule").on("click", _tmenu.open).classes("cursor-pointer")
                            type_in = ui.select(["1x1", "Group", "Fireside", "Call"], value="1x1",
                                                label="Type").props("dense outlined").classes("w-32")
                            format_in = ui.select(["In-person", "Virtual", "Zoom", "Teams", "Phone", "Jitsi"], value="In-person",
                                                  label="Format").props("dense outlined").classes("w-32")
                        with ui.row().classes("w-full gap-4 items-end"):
                            holder_in = ui.select(["Non-holder", "Holder"], value="Non-holder",
                                                  label="Holder status").props("dense outlined").classes("w-40")
                            notes_m = ui.input("Notes").classes("flex-1").props("outlined dense")
                        link_m = ui.input("Meeting link (virtual)",
                                          placeholder="Zoom / Teams / Meet URL").classes("w-full").props("outlined dense")
                        link_m.bind_visibility_from(format_in, "value", backward=lambda v: v != "In-person")

                        def add_meeting(idx=idx, inst_in=inst_in, contact_in=contact_in, address_in=address_in,
                                        day_in=day_in, time_in=time_in, type_in=type_in, format_in=format_in,
                                        holder_in=holder_in, notes_m=notes_m, link_m=link_m):
                            if not (inst_in.value or "").strip():
                                ui.notify("Institution is required.", type="warning")
                                return
                            trips_ = _load_json("ndr_trips.json", [])
                            trips_[idx].setdefault("meetings", []).append({
                                "institution": inst_in.value.strip(),
                                "contact": (contact_in.value or "").strip(),
                                "address": (address_in.value or "").strip(),
                                "notes": notes_m.value,
                                "day": int(day_in.value or 1),
                                "time": (time_in.value or "").strip() or "—",
                                "type": type_in.value, "format": format_in.value, "status": "scheduled",
                                "meeting_link": (link_m.value or "").strip(),
                                "non_holder": holder_in.value == "Non-holder", "score": None,
                                "metro": trip.get("city", ""),
                            })
                            _save_json("ndr_trips.json", trips_)
                            ui.notify(f"{inst_in.value.strip()} added to trip.")
                            _refresh_ndr()

                        ui.button("Add", on_click=add_meeting).props("dense")

                    def download_itinerary(trip=trip):
                        ui.download(_build_ndr_itinerary(trip, CT("ticker")).encode(), filename=f"{trip['name'].replace(' ','_')}_itinerary.txt")

                    def download_calendar(trip=trip):
                        from core import fund_addresses, ndr_calendar
                        ics = ndr_calendar.build_ics(
                            trip, CT("ticker"),
                            contacts=get_institution_contacts(),
                            address_for=fund_addresses.address_for)
                        ui.download(ics.encode(),
                                    filename=f"{trip['name'].replace(' ','_')}.ics")
                        ui.notify("Calendar file downloaded — open it to add every meeting "
                                  "(with join links) to your calendar.")

                    def print_itinerary(trip=trip):
                        # Open the print-formatted itinerary in a new window and
                        # trigger the browser print dialog, so app nav isn't
                        # printed. Pop-up-blocked case falls back to an alert.
                        doc = _build_ndr_itinerary_html(trip, CT("ticker"))
                        ui.run_javascript(
                            "const w=window.open('','_blank');"
                            "if(w){"
                            f"w.document.write({json.dumps(doc)});"
                            "w.document.close();w.focus();setTimeout(()=>w.print(),350);"
                            "}else{alert('Please allow pop-ups to print the itinerary.');}"
                        )

                    def email_itinerary(trip=trip):
                        # Opens the user's own mail client (mailto) with the
                        # itinerary pre-filled — no recipient, they choose who.
                        href = (f"mailto:?subject={quote(CT('ticker') + ' NDR Itinerary — ' + trip['name'])}"
                                f"&body={quote(_build_ndr_itinerary(trip, CT('ticker')))}")
                        ui.run_javascript(f"window.location.href={json.dumps(href)};")

                    async def download_map(trip=trip):
                        # A street map of the day (numbered stops, driving route, hotel) to
                        # attach to the itinerary. Built off-thread so the first (uncached)
                        # geocode/route/tile fetch doesn't freeze the UI.
                        from nicegui import run
                        ui.notify("Building route map…")
                        try:
                            png = await run.io_bound(_build_ndr_route_map_png, trip, CT("ticker"))
                        except Exception:
                            png = None
                        if not png:
                            ui.notify("Route map needs the routing key (Settings) and at least one "
                                      "located stop.", type="warning")
                            return
                        ui.download(png, filename=f"{trip['name'].replace(' ', '_')}_route_map.png")

                    with ui.row().classes("gap-2 items-center").style("margin-top:6px;"):
                        ui.button("Export itinerary (.txt)", on_click=download_itinerary).props("flat dense")
                        ui.button("Add to calendar (.ics)", icon="event", on_click=download_calendar).props("flat dense")
                        ui.button("Route map (.png)", icon="map", on_click=download_map).props("flat dense")
                        ui.button("Print", icon="print", on_click=print_itinerary).props("flat dense")
                        ui.button("Email", icon="mail", on_click=email_itinerary).props("flat dense")

        with ui.tab_panel(nt2):
            _active_ndrs_panel()

        with ui.tab_panel(nt4):
            _render_ndr_prep_cards_tab(institutions, meeting_log)

        with ui.tab_panel(nt5):
            _render_ndr_debrief_tab()


async def _open_ndr_reply_dialog(req, on_done):
    """Reply to an inbound NDR request FROM the app: a reviewed, human-sent email (Reg FD — never
    auto-sent). On send it's logged onto the request's correspondence trail and the item flips to
    'replied', and a copy is filed to the Zoho Sent folder — so 'did we get back to them?' is
    answerable in-app instead of by digging through a mailbox."""
    import asyncio
    import json

    from config.client_config import CT
    from core import ndr_correspondence, zoho_mail
    _zoho = zoho_mail.is_configured()
    analyst = (req.get("analyst") or "there").strip()
    first = analyst.split()[0] if analyst else "there"
    company = CT("name") or "our"
    city = req.get("city") or ""
    subject = f"Re: your meeting request{f' — {city}' if city else ''}"
    body = (f"Hi {first},\n\n"
            f"Thanks for the request to meet with {company} management"
            f"{f' in {city}' if city else ''}. We'd be glad to coordinate — could you share a "
            f"couple of dates and times that work on your end, and whether you'd prefer a call or "
            f"an in-person meeting?\n\nBest regards,\nInvestor Relations\n{company}")

    with ui.dialog() as dlg, ui.card().style("min-width:480px;max-width:620px;"):
        ui.label(f"Reply to {analyst}" + (f" ({req.get('firm')})" if req.get("firm") else "")).classes(
            "font-bold").style(f"color:{COLORS['text_heading']};font-size:var(--fs-md);")
        ui.label(f"Their request: {req.get('reason') or '—'}").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
        _to = ui.input("To", value=req.get("analyst_email") or req.get("sender_email") or "").props(
            "outlined dense").classes("w-full").style("font-size:var(--fs-sm);")
        _subj = ui.input("Subject", value=subject).props("outlined dense").classes("w-full").style("font-size:var(--fs-sm);")
        _body = ui.textarea("Message", value=body).props("outlined autogrow dense").classes(
            "w-full").style("font-size:var(--fs-sm);")
        ui.label("Reviewed and sent by you — IRconnect never sends shareholder mail automatically "
                 "(Reg FD). A copy is filed to your Sent folder and logged on this request.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);")

        def _log(via, message_id=None):
            ndr_correspondence.record_reply(req["id"], _to.value, _subj.value or "", _body.value or "",
                                            via=via, message_id=message_id)

        with ui.row().classes("justify-end w-full gap-2 items-center").style("margin-top:4px;"):
            ui.button("Cancel", on_click=dlg.close).props("flat dense")

            def _copy():
                ui.run_javascript("navigator.clipboard.writeText(" + json.dumps(
                    f"To: {_to.value}\nSubject: {_subj.value or ''}\n\n{_body.value or ''}") + ")")
                ui.notify("Draft copied.", type="positive")
            ui.button("Copy", icon="content_copy", on_click=_copy).props("flat dense")

            def _openmail():
                if not (_to.value or "").strip():
                    ui.notify("Add the analyst's email first.", type="warning")
                    return
                href = f"mailto:{_to.value}?subject={quote(_subj.value or '')}&body={quote(_body.value or '')}"
                ui.run_javascript(f"window.location.href = {json.dumps(href)}")
                _log("email client")
                ui.notify("Logged on the request (Replied).", type="positive")
                dlg.close()
                on_done()
            ui.button("Open in email", icon="mail", on_click=_openmail).props(
                "flat dense" if _zoho else "color=primary dense")

            if _zoho:
                async def _send():
                    if not (_to.value or "").strip():
                        ui.notify("Add the analyst's email first.", type="warning")
                        return
                    ui.notify("Sending via Zoho…", type="info")
                    import email.utils
                    msgid = email.utils.make_msgid()
                    ok, err = await asyncio.to_thread(
                        zoho_mail.send_email, _to.value, _subj.value or "", _body.value or "",
                        None, True, msgid)
                    if ok:
                        _log("Zoho", message_id=msgid)
                        ui.notify(f"Sent to {_to.value} — filed to Sent and logged on this request.",
                                  type="positive")
                        dlg.close()
                        on_done()
                    else:
                        ui.notify(f"Zoho send failed: {err}", type="negative")
                ui.button("Send via Zoho", icon="send", on_click=_send).props("color=primary dense")
    dlg.open()


def _render_ndr_requests_tab(refresh_fn=None):
    # Self-contained refresh so this renders wherever it's mounted (it now lives on the IR Inbox, not
    # the NDR tab). The caller passes a refresh that repaints just its own section in place; falls back
    # to the page-level _refresh (which navigates) only if none is given.
    _do_refresh = refresh_fn or _refresh
    with ui.row().classes("w-full items-center").style("gap:10px;"):
        ui.label("Inbound NDR / Meeting Requests").classes("font-bold")
        ui.space()

        async def _check_replies():
            import asyncio

            from core import ndr_inbound, zoho_mail
            if not zoho_mail.is_configured():
                ui.notify("Connect Zoho first to pull replies from the mailbox.", type="warning")
                return
            ui.notify("Checking the mailbox for analyst replies…", type="info")
            try:
                r = await asyncio.to_thread(ndr_inbound.poll_replies, [get_active_client_id()])
            except Exception as e:
                ui.notify(f"Reply check failed: {e}", type="negative")
                return
            if not r.get("ok"):
                ui.notify(f"Couldn't check replies: {r.get('reason')}", type="warning")
                return
            ui.notify(f"Checked {r['checked']} message(s) — {r['matched']} threaded onto requests.",
                      type="positive")
            if r["matched"]:
                _do_refresh()
        ui.button("Check for replies", icon="mark_email_unread", on_click=_check_replies).props(
            "flat dense").style(f"color:{COLORS['text_muted']};")
    ui.label(
        "Analyst requests to slot a management meeting into a city — they feed the WHERE TO GO read and "
        "the 'Who's asking' detail on Targeting / Ownership. Resolve a request once it's been scheduled "
        "(or declined) so it stops counting as open demand."
    ).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    with ui.expansion("Log a new request", value=False).classes("w-full"):
        with ui.row().classes("w-full gap-4"):
            r_analyst = ui.input("Analyst name *").classes("flex-1").props("outlined dense")
            # Suggest covering firms via autocomplete, but start empty — a request can
            # come from any firm; auto-filling the first covering bank mislabeled it.
            bank_options = [a["firm"] for a in CA()] + ["Other / Non-Covering Bank"]
            r_firm = ui.input("Firm", placeholder="Analyst's firm — pick or type",
                              autocomplete=bank_options).classes("flex-1").props("outlined dense")
        with ui.row().classes("w-full gap-4"):
            r_city = ui.input("City *").classes("flex-1").props("outlined dense")
            r_metro = ui.input("Metro region",
                               autocomplete=sorted({i["Metro"] for i in get_seed_buyside_institutions(get_active_client_id())})).classes("flex-1").props("outlined dense")
        r_email = ui.input("Analyst email", placeholder="so Reply pre-fills the recipient").classes("w-full").props("outlined dense")
        r_reason = ui.textarea("Reason / context", placeholder="What did they ask for, and why?").classes("w-full").props("outlined autogrow")

        def log_request():
            if not (r_analyst.value and r_city.value):
                ui.notify("Analyst name and city are required.", type="warning")
                return
            reqs = _load_ndr_requests()
            reqs.append({
                "id": datetime.now().strftime("%Y%m%d%H%M%S"), "analyst": r_analyst.value, "firm": r_firm.value,
                "city": r_city.value, "metro": r_metro.value or r_city.value, "reason": r_reason.value or "—",
                "analyst_email": (r_email.value or "").strip() or None,
                "received": datetime.now().strftime("%b %d, %Y"), "resolved": False, "seeded": False,
            })
            _save_ndr_requests(reqs)
            ui.notify(f"Request from {r_analyst.value} logged.")
            _do_refresh()

        ui.button("Log Request", on_click=log_request).props("color=primary")

    ui.markdown("---")
    reqs = _load_ndr_requests()
    open_reqs = [r for r in reqs if not r.get("resolved")]
    resolved_reqs = [r for r in reqs if r.get("resolved")]

    # Rank by the request date — the ones that came in earliest (aging longest on the desk) sort to
    # the top, so the most overdue demand is what you see first. Undated requests fall to the bottom.
    def _req_date(r):
        try:
            return datetime.strptime(r.get("received", ""), "%b %d, %Y")
        except Exception:
            return datetime.max
    open_reqs.sort(key=_req_date)

    # Cross-reference open NDR trips: a request whose city is already being worked reads "IN PROGRESS"
    # (with the trip name) rather than counting as untouched open demand.
    _open_trips = [t for t in _load_json("ndr_trips.json", []) if t.get("status") != "Completed"]

    def _matching_trip(r):
        _keys = {(r.get("metro") or "").strip().lower(), (r.get("city") or "").strip().lower()} - {""}
        for t in _open_trips:
            _tk = {(t.get("metro") or "").strip().lower(), (t.get("city") or "").strip().lower()} - {""}
            if _keys & _tk:
                return t
            for a in _keys:
                for b in _tk:
                    if a and b and (a in b or b in a):
                        return t
        return None

    def _req_email(r):
        if r.get("analyst_email"):
            return r["analyst_email"]
        _nm = (r.get("analyst") or "").split()
        _slug = "".join(ch for ch in (r.get("firm") or "").lower() if ch.isalnum()) or "firm"
        if len(_nm) >= 2:
            return f"{_nm[0].lower()}.{_nm[-1].lower()}@{_slug}.com"
        return f"contact@{_slug}.com"

    ui.label(f"{len(open_reqs)} open request(s)").classes("font-bold")
    if not open_reqs:
        ui.label("No open requests.").style(f"color:{COLORS['text_muted']};")
    for r in open_reqs:
        from core import ndr_correspondence
        seed_tag = " · example" if r.get("seeded") else ""
        _replied = ndr_correspondence.status(r) == "replied"
        _mt = _matching_trip(r)
        _cap = f"Received {r['received']}" + (f"  ·  In progress — {_mt.get('name', 'NDR')}" if _mt else "")
        with ui.expansion(f"{r['analyst']} ({r['firm']}) → {r['city']}{seed_tag}",
                          caption=_cap).classes("w-full").style(
                f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};border-radius:8px;"):
            with ui.row().classes("items-center").style("gap:8px;"):
                # An NDR already exists for this city → this is being worked, not untouched demand.
                if _mt:
                    ui.label(f"IN PROGRESS · {_mt.get('name', 'NDR')}").style(
                        "background:rgba(30,64,175,.14);color:#1E40AF;font-size:var(--fs-micro);font-weight:800;"
                        "padding:2px 8px;border-radius:9px;")
                if _replied:
                    ui.label(f"REPLIED · {r.get('replied_at','')}").style(
                        "background:rgba(21,128,61,.14);color:#15803D;font-size:var(--fs-micro);font-weight:800;"
                        "padding:2px 8px;border-radius:9px;")
                else:
                    ui.label("AWAITING REPLY").style(
                        "background:rgba(180,83,9,.14);color:#B45309;font-size:var(--fs-micro);font-weight:800;"
                        "padding:2px 8px;border-radius:9px;")

            # The original message, as it landed in the IR Inbox — so you can see the actual ask, not
            # just a one-line summary.
            with ui.card().classes("w-full").style(
                    f"background:{COLORS['surface_hover_bg']};border-left:3px solid {COLORS['accent']};"
                    "border-radius:6px;padding:8px 12px;gap:2px;margin-top:4px;"):
                ui.label("FROM THE IR INBOX · NDR request").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);font-weight:700;letter-spacing:.04em;")
                ui.label(f"From: {r['analyst']} <{_req_email(r)}>").style(
                    f"color:{COLORS['text_body']};font-size:var(--fs-xs);")
                ui.label(f"Subject: {CT('ticker')} — management meeting request, {r['city']}").style(
                    f"color:{COLORS['text_body']};font-size:var(--fs-xs);")
                ui.label(f"Received: {r['received']}").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                ui.label(r["reason"]).style(
                    f"color:{COLORS['text_body']};font-size:var(--fs-sm);margin-top:4px;line-height:1.5;")
                ui.button("Open in IR Inbox", icon="inbox",
                          on_click=lambda: nav.go_to("Inbox")).props("flat dense size=sm").style(
                    f"color:{COLORS['accent']};align-self:flex-start;margin-top:2px;")

            # Correspondence trail — the sent (and any received) messages on this request.
            trail = ndr_correspondence.trail(r)
            if trail:
                ui.label("Correspondence").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);font-weight:700;margin-top:4px;")
                for c in trail:
                    out = c.get("direction") == "out"
                    who = f"To {c.get('to')}" if out else f"From {c.get('from')}"
                    with ui.row().classes("items-center").style("gap:6px;"):
                        ui.icon("call_made" if out else "call_received").style(
                            f"color:{'#15803D' if out else COLORS['accent']};font-size:var(--fs-base);")
                        ui.label(f"{c.get('ts')} · {who} · {c.get('subject') or ''}"
                                 + (f"  ({c.get('via')})" if c.get('via') else "")).style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

            def mark_resolved(rid=r["id"]):
                current = _load_ndr_requests()
                for rr in current:
                    if rr["id"] == rid:
                        rr["resolved"] = True
                        rr["resolved_at"] = datetime.now().strftime("%b %d, %Y")
                _save_ndr_requests(current)
                ui.notify("Marked resolved.")
                _do_refresh()

            with ui.row().classes("gap-2").style("margin-top:6px;"):
                ui.button("Reply", icon="reply",
                          on_click=lambda r=r: _open_ndr_reply_dialog(r, _do_refresh)).props("flat dense color=primary")
                ui.button("Schedule into NDR", icon="event_available",
                          on_click=lambda r=r: _open_schedule_request_dialog(r, _do_refresh)).props("flat dense")
                ui.button("Resolve without scheduling", on_click=mark_resolved).props("flat dense")

    if resolved_reqs:
        with ui.expansion(f"{len(resolved_reqs)} resolved").classes("w-full").style("margin-top:8px;"):
            for r in resolved_reqs:
                ui.label(f"{r['analyst']} ({r['firm']}) → {r['city']} · resolved {r.get('resolved_at','—')}").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")


# ─────────────────────────────────────────────────────────────────────────
# NDR — Prep Cards tab
# ─────────────────────────────────────────────────────────────────────────
_PREP_HDR_STYLE = ("color:{c};font-size:var(--fs-xs);font-weight:700;letter-spacing:.04em;"
                   "margin-top:8px;text-transform:uppercase;")


def _last_meeting_brief(firm):
    """The most recent post-meeting debrief for a firm, with its structured fields parsed —
    or None. This is the richest prep input the app holds: what they asked, what worried them,
    and what we committed to send. Captured on the Post-NDR Debrief / meeting-notes flow, the
    same store the CRM reads — so the prep card and the account history never diverge."""
    import ast
    recs = _load_json("post_meeting_notes.json", []) or []
    mine = [r for r in recs if _norm_name(r.get("Firm", "")) == _norm_name(firm)]
    if not mine:
        return None
    mine.sort(key=lambda r: r.get("Date", ""), reverse=True)
    r = mine[0]
    s, st = r.get("Structured"), {}
    if isinstance(s, dict):
        st = s
    elif isinstance(s, str) and s.strip():
        for parse in (json.loads, ast.literal_eval):   # stored as JSON or a py-dict repr
            try:
                st = parse(s)
                break
            except Exception:
                continue
    return {"date": r.get("Date"), "type": r.get("Type"), "contact": r.get("Contact"),
            "raw": r.get("Raw"), **(st if isinstance(st, dict) else {})}


def _ndr_prep_read(inst, prior=None):
    """The one-line READ for a meeting — holder-vs-prospect framing off the Fit score, plus the
    strongest warm signal on file. Mirrors the ownership-summary READ used elsewhere on the page
    so a prep card leads with the same judgement, not a different one."""
    if not inst:
        base = ("No tracked signal on file — treat as a discovery meeting: confirm mandate fit and "
                "time horizon before going deep on the thesis.")
    else:
        fit = inst.get("Fit_Score")
        if inst.get("USIO_Holder"):
            base = "Existing holder — a defend-and-deepen meeting: protect the position and grow it."
        elif isinstance(fit, (int, float)) and fit >= 80:
            base = f"High-fit prospect (Fit {int(fit)}/100) — a priority conversion, not a cold intro."
        elif isinstance(fit, (int, float)) and fit >= 55:
            base = f"Good-fit prospect (Fit {int(fit)}/100) — worth a full pitch; qualify the path to a position."
        else:
            base = "Prospect — discovery meeting: confirm mandate fit before going deep."
        warm = []
        if inst.get("Peer_Holdings"):
            warm.append(f"owns {len(inst['Peer_Holdings'])} of your peers")
        if inst.get("IR_Visits_30d"):
            warm.append(f"on your IR site {inst['IR_Visits_30d']}× in 30d")
        di = inst.get("Digital_Intent_Score")
        if isinstance(di, (int, float)) and di >= 50 and not inst.get("IR_Visits_30d"):
            warm.append(f"digital-intent {int(di)}/100")
        if warm:
            base += " They're warm: " + ", ".join(warm) + "."
    if prior and prior.get("sentiment"):
        base += f" Last conversation read {str(prior['sentiment']).lower()}."
    return base


def _prep_bring(inst, prior):
    """Concrete things to walk in with: what we committed to send + open follow-ups from the last
    meeting (deduped by overlap), then the peer comparison if they hold the comps and not us."""
    out = []
    if prior:
        for x in (prior.get("commitments_made") or []) + (prior.get("follow_up_actions") or []):
            xs = str(x).strip()
            if xs and not any(xs.lower() in o.lower() or o.lower() in xs.lower() for o in out):
                out.append(xs)
    if inst and inst.get("Peer_Holdings"):
        out.append(f"The side-by-side vs {', '.join(inst['Peer_Holdings'])} — they hold the peers, not you yet.")
    return out


def _prep_signals(inst):
    """Why this name matters now, pulled from the SAME tracked fields the scoring and ownership
    surfaces use: real 13F peer overlap, IR-site visits, digital intent, call-listening, and the
    curated note — instead of the two generic lines the card showed before."""
    if not inst:
        return []
    out = []
    ph = inst.get("Peer_Holdings")
    if ph:
        src = inst.get("Peer_Holdings_Source") or ""
        out.append(f"Owns {', '.join(ph)}{f' ({src})' if src else ''} — the direct peer comparison writes itself.")
    if inst.get("IR_Visits_30d"):
        last = inst.get("Last_Visit")
        out.append(f"{inst['IR_Visits_30d']} visit(s) to your IR site in the last 30 days"
                   + (f", most recently {last}." if last else "."))
    di = inst.get("Digital_Intent_Score")
    if isinstance(di, (int, float)) and di >= 50:
        out.append(f"Digital-intent score {int(di)}/100 — actively researching the name.")
    if inst.get("Call_Listener") and inst.get("Listen_Duration"):
        out.append(f"Listened to the last earnings call ({inst['Listen_Duration']}).")
    note = (inst.get("Notes") or "").strip()
    if note:
        out.append(note)
    return out


def _render_ndr_prep_cards_tab(institutions, meeting_log):
    """Per-meeting prep cards for an NDR trip. Ported from app.py's Prep
    Cards tab, but the talking points are no longer hardcoded per-ticker
    (app.py branched on literal strings like `"GDOT" in inst_data["peer"]`,
    which only ever worked for USIO) — they're generated live from each
    institution's own tracked fields via _ndr_talking_points, so this works
    unchanged for any client's roster. A meeting whose institution isn't in
    the tracked list (added by hand, or a name typo) still gets a card, just
    without the institution-specific detail — it degrades gracefully instead
    of crashing or silently disappearing."""
    trips = _load_json("ndr_trips.json", [])
    if not trips:
        ui.label("No NDR trips yet — create one in Plan New NDR.").style(f"color:{COLORS['text_muted']};")
        return

    ui.label("Meeting Prep Cards").classes("font-bold")
    ui.label("Pick a trip to generate a prep card per meeting — pulled live from tracked institution data.").style(
        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    trip_names = [t["name"] for t in trips]
    trip_sel = ui.select(trip_names, value=trip_names[-1], label="Trip").classes("w-full max-w-md").props("outlined dense")
    cards_area = ui.column().classes("w-full")

    inst_by_name = {i["Fund"]: i for i in institutions}

    def rebuild_cards(e=None):
        cards_area.clear()
        trip = next((t for t in trips if t["name"] == trip_sel.value), None)
        with cards_area:
            if not trip:
                return
            meetings = [m for m in trip.get("meetings", []) if m.get("type") != "break"]
            if not meetings:
                ui.label("No meetings on this trip yet.").style(f"color:{COLORS['text_muted']};")
                return
            # Chronological — group by day, then order each day by parsed time (a hand-added
            # 8:00 must lead the day, not sink to insertion-order bottom — the same fix the
            # Active NDRs schedule uses).
            by_day = {}
            for m in meetings:
                by_day.setdefault(m.get("day", 1), []).append(m)
            hdr = _PREP_HDR_STYLE.format(c=COLORS['text_muted'])
            current_day = None
            for d in sorted(by_day.keys()):
                if d != current_day:
                    current_day = d
                    ui.label(f"Day {d}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);font-weight:700;margin-top:10px;")
                for m in _sorted_day_meetings(by_day[d]):
                    firm = m.get("institution", "")
                    inst = inst_by_name.get(firm)
                    contact = get_institution_contacts().get(firm, {}) or {}
                    who = (m.get("contact") or "").strip() or ", ".join(
                        p for p in (contact.get("name"), contact.get("title")) if p)
                    prior = _last_meeting_brief(firm)
                    # Caption: who we're meeting + the same Fit/Engagement read the rest of the page shows.
                    if inst:
                        cap_bits = [who or None, inst.get("Metro"), inst.get("Type"),
                                    "Existing holder" if inst.get("USIO_Holder") else "Prospect"]
                        if inst.get("Fit_Score") is not None:
                            cap_bits.append(f"Fit {inst['Fit_Score']}")
                        if inst.get("Engagement_Score") is not None:
                            cap_bits.append(f"Engagement {inst['Engagement_Score']}")
                        caption = " · ".join(str(x) for x in cap_bits if x)
                    else:
                        caption = (f"{who} · " if who else "") + "not in tracked list — added by hand"
                    with ui.expansion(f"{_meeting_time_range(m)} — {pretty_name(firm)}", caption=caption).classes("w-full").style(
                            f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};border-radius:8px;"):
                        # Who + one-tap reach
                        if who or contact.get("phone") or contact.get("email"):
                            with ui.row().classes("items-center gap-3 flex-wrap").style("margin:2px 0 6px;"):
                                if who:
                                    ui.label(f"Meeting with: {who}").style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
                                if contact.get("phone"):
                                    ui.link(f"Call {contact['phone']}", _tel_href(contact['phone'])).style(
                                        f"color:{COLORS['accent_light']};font-size:var(--fs-sm);")
                                if contact.get("email"):
                                    _mailto(contact["email"], f"{CT('ticker')} — following up on our meeting", "", "Email")
                        # THE READ
                        ui.label(_ndr_prep_read(inst, prior)).style(
                            f"background:{COLORS['surface_hover_bg']};border-left:3px solid {COLORS['accent']};"
                            f"border-radius:6px;padding:8px 10px;color:{COLORS['text_body']};font-size:var(--fs-sm);"
                            "line-height:1.45;margin-bottom:2px;")
                        # WHERE YOU LEFT OFF — the last debrief's structured detail
                        if prior:
                            _t = (prior.get("type") or "meeting").lower()
                            ui.label(f"Where you left off — last {_t} {prior.get('date','')}".strip()).style(hdr)
                            if prior.get("summary"):
                                ui.label(prior["summary"]).style(
                                    f"color:{COLORS['text_body']};font-size:var(--fs-sm);line-height:1.4;")
                            for _lbl, _items, _col in (("They asked", prior.get("key_questions"), COLORS['text_secondary']),
                                                       ("Their concern", prior.get("concerns_raised"), COLORS['warning'])):
                                _items = [x for x in (_items or []) if x]
                                if _items:
                                    ui.label(_lbl).style(f"color:{_col};font-size:var(--fs-xs);font-weight:600;margin-top:4px;")
                                    for _it in _items:
                                        ui.label(f"• {_it}").style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
                        # BRING TO THIS MEETING
                        bring = _prep_bring(inst, prior)
                        if bring:
                            ui.label("Bring to this meeting").style(hdr)
                            for b in bring:
                                ui.label(f"• {b}").style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
                        # WHY THEY MATTER NOW — real tracked signals
                        sigs = _prep_signals(inst)
                        if sigs:
                            ui.label("Why they matter now").style(hdr)
                            for s in sigs:
                                ui.label(f"• {s}").style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
                        if not (prior or bring or sigs or inst):
                            ui.label("• No tracked signal on file — treat as a discovery meeting.").style(
                                f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
                        # Strategy note last (the human overlay)
                        if m.get("notes"):
                            ui.label(f"Strategy note: {m['notes']}").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);font-style:italic;margin-top:6px;")

    trip_sel.on_value_change(rebuild_cards)
    rebuild_cards()


# ─────────────────────────────────────────────────────────────────────────
# NDR — Post-NDR Debrief tab
# ─────────────────────────────────────────────────────────────────────────
def _render_ndr_debrief_tab():
    """Debrief form for a completed NDR trip. Ported from app.py, with one
    real enhancement: app.py's caption said "Key objections will feed into
    next Q&A prep" but never actually wired that anywhere — here, saving a
    debrief with a key objection or narrative gap calls
    risk_scorecard.log_ndr_objection(), which is what turns the Markets IR
    Risk Dashboard's "KPI Understanding" / "Investor Objection Trend" tiles
    from permanently-GRAY (no data source) to a real, disclosed YELLOW
    signal. Only trips marked Completed (Active NDRs tab) show up here —
    debriefing a trip that hasn't happened yet doesn't make sense."""
    trips = _load_json("ndr_trips.json", [])
    completed = [t for t in trips if t.get("status") == "Completed"]

    ui.label("Post-NDR Debrief").classes("font-bold")
    if not completed:
        ui.label("No completed NDR trips yet — mark a trip \"Completed\" in Active NDRs once it wraps, "
                  "then debrief it here.").style(f"color:{COLORS['text_muted']};")
        return

    trip_names = [t["name"] for t in completed]
    trip_sel = ui.select(trip_names, value=trip_names[-1], label="Trip").classes("w-full max-w-md").props("outlined dense")
    form_area = ui.column().classes("w-full")

    def rebuild_form(e=None):
        form_area.clear()
        trip = next((t for t in completed if t["name"] == trip_sel.value), None)
        with form_area:
            if not trip:
                return
            meetings = [m for m in trip.get("meetings", []) if m.get("type") != "break"]
            debrief = trip.get("debrief", {})
            n_held = sum(1 for m in meetings if m.get("status") == "completed")

            with ui.row().classes("w-full gap-4"):
                with ui.column().classes("flex-1"):
                    held_in = ui.number("Meetings actually held", value=debrief.get("meetings_held", n_held), min=0).classes("w-full").props("outlined dense")
                    eff_in = ui.number("Effectiveness score (0-100)", value=debrief.get("effectiveness", 70), min=0, max=100).classes("w-full").props("outlined dense")
                    meeting_names = [m.get("institution", "") for m in meetings] or ["—"]
                    best_in = ui.select(meeting_names, value=debrief.get("best_meeting", meeting_names[0]),
                                         label="Best meeting").classes("w-full").props("outlined dense")
                with ui.column().classes("flex-1"):
                    follow_in = ui.textarea("Follow-ups needed", value=debrief.get("follow_ups", "")).classes("w-full").props("outlined autogrow")
                    new_pos_in = ui.textarea("New positions initiated (if known)", value=debrief.get("new_positions", "")).classes("w-full").props("outlined autogrow")

            ui.markdown("---")
            ui.label("Feeds the IR Risk Dashboard (Markets page) once saved:").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
            with ui.row().classes("w-full gap-4"):
                objection_in = ui.textarea("Key objection heard", value=debrief.get("key_objection", ""),
                                            placeholder="e.g. \"Concerned about customer concentration in top 3 accounts.\"").classes("flex-1").props("outlined autogrow")
                gap_in = ui.textarea("Narrative gap (question the current deck/script doesn't answer)",
                                      value=debrief.get("narrative_gap", "")).classes("flex-1").props("outlined autogrow")
            next_in = ui.textarea("Next targets for this metro", value=debrief.get("next_targets", "")).classes("w-full").props("outlined autogrow")

            def save_debrief(trip_name=trip["name"]):
                trips_ = _load_json("ndr_trips.json", [])
                for t in trips_:
                    if t["name"] == trip_name:
                        t["debrief"] = {
                            "meetings_held": int(held_in.value or 0), "effectiveness": int(eff_in.value or 0),
                            "best_meeting": best_in.value, "follow_ups": follow_in.value,
                            "new_positions": new_pos_in.value, "key_objection": objection_in.value,
                            "narrative_gap": gap_in.value, "next_targets": next_in.value,
                            "saved_at": datetime.now().isoformat(),
                        }
                _save_json("ndr_trips.json", trips_)
                risk_scorecard.log_ndr_objection(trip_name=trip_name, objection=objection_in.value, narrative_gap=gap_in.value)
                ui.notify(f"Debrief saved for '{trip_name}'.")
                _refresh()

            ui.button("Save Debrief", on_click=save_debrief).props("color=primary").style("margin-top:8px;")

            if debrief:
                ui.markdown("---")
                ui.label("Saved debrief").classes("font-bold")
                with ui.row().classes("w-full gap-3"):
                    _bp_metric("Meetings held", str(debrief.get("meetings_held", "—")), [])
                    _bp_metric("Effectiveness", f"{debrief.get('effectiveness', '—')}/100", [])
                    _bp_metric("Best meeting", debrief.get("best_meeting") or "—", [])
                if debrief.get("key_objection"):
                    ui.label(f"Key objection: {debrief['key_objection']}").style(f"color:{COLORS['warning']};font-size:var(--fs-sm);margin-top:6px;")
                if debrief.get("narrative_gap"):
                    ui.label(f"Narrative gap: {debrief['narrative_gap']}").style(f"color:{COLORS['warning']};font-size:var(--fs-sm);")
                if debrief.get("next_targets"):
                    ui.label(f"Next targets: {debrief['next_targets']}").style(f"color:{COLORS['accent_light']};font-size:var(--fs-sm);")

    trip_sel.on_value_change(rebuild_form)
    rebuild_form()


# ─────────────────────────────────────────────────────────────────────────
# Meeting Hub tab
# ─────────────────────────────────────────────────────────────────────────
def _render_meeting_hub_tab():
    scheduled = _load_json("scheduled_meetings.json", [])
    notes = _load_json("post_meeting_notes.json", [])
    today = datetime.now().date()

    # Firm lookup — union of every tracked institution (Buy-Side Intelligence)
    # and every institution with a known contact on file, so the dropdown
    # covers both "funds we're already scoring" and "funds we just have a
    # contact for." with_input + new-value-mode=add-unique makes this a
    # type-to-filter combo box that still accepts a firm that isn't in
    # either list yet (e.g. a brand-new prospect) rather than locking the
    # form to only known names.
    known_contacts = get_institution_contacts()
    firm_options = sorted({i["Fund"] for i in get_seed_buyside_institutions(get_active_client_id())}
                           | set(known_contacts.keys()))

    def _autofill_contact(e, contact_field):
        typed = (e.value or "").strip()
        if not typed:
            return
        info = known_contacts.get(typed)
        if not info:
            # The Firm combo box allows free text (new-value-mode=add-unique)
            # so it's the browser, not us, deciding whether a click on the
            # filtered suggestion or a plain blur/tab-out committed the
            # value — and a blur commits whatever's still typed (e.g.
            # "rutabaga") rather than snapping to the full matched option
            # ("Rutabaga Capital Management"), so an exact-key lookup alone
            # missed this case. Fall back to a case-insensitive, then
            # substring, match against the known contact list.
            typed_lower = typed.lower()
            info = next((v for k, v in known_contacts.items() if k.lower() == typed_lower), None)
            if not info:
                partial = [v for k, v in known_contacts.items() if typed_lower in k.lower()]
                info = partial[0] if len(partial) == 1 else None
        if info and not contact_field.value:
            contact_field.value = info["name"]

    with ui.tabs().classes("w-full") as hub_tabs:
        h1 = ui.tab("Upcoming Meetings")
        h2 = ui.tab("Schedule Meeting")
        h3 = ui.tab("Post-Meeting Notes")
    with ui.tab_panels(hub_tabs, value=h1).classes("w-full"):
        with ui.tab_panel(h1):
            upcoming = sorted([m for m in scheduled if _safe_date(m["Date"]) >= today], key=lambda x: x["Date"])
            # Tiles are clickable filters over the meeting list below (by side);
            # "Notes captured" jumps to the Post-Meeting Notes tab.
            _hub_filter = {"mode": "all"}
            hub_cards_row = ui.row().classes("w-full gap-3")

            def _hub_shown():
                if _hub_filter["mode"] == "sell":
                    return [m for m in upcoming if m["Side"] == "Sell-side"]
                if _hub_filter["mode"] == "buy":
                    return [m for m in upcoming if m["Side"] == "Buy-side"]
                return upcoming

            def set_hub_filter(mode):
                _hub_filter["mode"] = mode
                render_hub_cards()
                render_hub_list()

            def render_hub_cards():
                hub_cards_row.clear()
                with hub_cards_row:
                    _hub_metric("Upcoming", len(upcoming), "All scheduled ahead",
                                _hub_filter["mode"] == "all", lambda: set_hub_filter("all"))
                    _hub_metric("Sell-side", sum(1 for m in upcoming if m["Side"] == "Sell-side"),
                                "Analyst meetings", _hub_filter["mode"] == "sell", lambda: set_hub_filter("sell"))
                    _hub_metric("Buy-side", sum(1 for m in upcoming if m["Side"] == "Buy-side"),
                                "Investor meetings", _hub_filter["mode"] == "buy", lambda: set_hub_filter("buy"))
                    _hub_metric("Notes captured", len(notes), "Open the Notes tab", False,
                                lambda: hub_tabs.set_value(h3))

            render_hub_cards()

            def sync_inbox():
                # "kind" tells mail_gateway who's who — sell-side analysts
                # (CA()) get their models/notes/NDR asks routed for review;
                # buy-side/institutional contacts are classified too (e.g.
                # "speak to management" requests) but never treated as a
                # model source. See core/mail_gateway.py's module docstring.
                contact_lookup = {}
                for firm_name, info in known_contacts.items():
                    if info.get("email"):
                        contact_lookup[info["email"].lower()] = {"name": info["name"], "firm": firm_name, "kind": "institution"}
                for a in CA():
                    if a.get("email"):
                        contact_lookup[a["email"].lower()] = {"name": a["name"], "firm": a.get("firm"), "kind": "analyst"}
                result = mail_gateway.sync_inbox(contact_lookup, client_id=get_active_client_id())
                if not result["ok"]:
                    ui.notify(f"{result['message']}", type="warning")
                else:
                    n = len(result["messages"])
                    routed = sum(1 for m in result["messages"] if m.get("category") != "general")
                    if n:
                        extra = f" · {routed} item(s) flagged for review below" if routed else ""
                        ui.notify(f"Synced — {n} message(s) from known contacts.{extra}")
                    else:
                        ui.notify("Synced — no new messages from known contacts.")
                    _refresh()

            ui.button("Sync IR Inbox", on_click=sync_inbox).props("flat dense").style(
                f"color:{COLORS['accent_light']};font-size:var(--fs-sm);margin-top:4px;")
            if not mail_gateway.is_configured():
                ui.label("Email sync isn't configured yet (no IMAP credentials in .env) — this button will "
                          "explain that rather than pretend to have fetched anything.").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

            # Meeting Hub is about meetings — surface only inbound emails asking for / confirming a
            # meeting. Models, research notes, conference invites, etc. are triaged on the IR Inbox
            # page, not here (that filtering gap is why an analyst model was showing up in the hub).
            _render_pending_inbox_items(
                categories=("speak_to_management", "meeting_confirmation"),
                title="Pending meeting requests")

            hub_list = ui.column().classes("w-full gap-2").style("margin-top:8px;")

            def render_hub_list():
                hub_list.clear()
                shown = _hub_shown()
                with hub_list:
                    label = {"all": "Upcoming", "sell": "Sell-side", "buy": "Buy-side"}[_hub_filter["mode"]]
                    clear = "" if _hub_filter["mode"] == "all" else "  ·  click Upcoming to clear"
                    ui.label(f"Showing {len(shown)} meeting(s) — {label}{clear}").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    if not shown:
                        ui.label("No upcoming meetings in this view.").style(f"color:{COLORS['text_muted']};")
                    for mtg in shown:
                        cap = f"{mtg['Date']} {mtg.get('Time','')} · {mtg['Firm']} · {mtg.get('Side','')} · {mtg.get('Status','—')}"
                        with ui.expansion(f"{mtg['Contact']} ({mtg['Firm']})", caption=cap).classes("w-full").style(
                                f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};border-radius:8px;"):
                            ui.label(f"{mtg['Type']} · Priority: {mtg.get('Priority','—')}").style(
                                f"color:{COLORS['text_secondary']};font-size:var(--fs-base);")
                            if mtg.get("Topic"):
                                ui.label(f"Topic: {mtg['Topic']}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                            _render_linked_documents(mtg["Contact"], mtg["Firm"])

            render_hub_list()

        with ui.tab_panel(h2):
            ui.label("Schedule a New Meeting or Callback").classes("font-bold")
            with ui.row().classes("w-full gap-4"):
                with ui.column().classes("flex-1"):
                    s_contact = ui.input("Contact name *").classes("w-full").props("outlined dense")
                    s_firm = ui.input("Firm *", autocomplete=firm_options).classes("w-full").props("outlined dense")
                    s_firm.on_value_change(lambda e: _autofill_contact(e, s_contact))
                    s_side = ui.select(["Buy-side", "Sell-side"], value="Buy-side").classes("w-full").props("outlined dense")
                    s_priority = ui.select(["High", "Medium", "Low"], value="Medium").classes("w-full").props("outlined dense")
                with ui.column().classes("flex-1"):
                    s_date = ui.input("Date (YYYY-MM-DD)", value=(today + timedelta(days=7)).strftime("%Y-%m-%d")).classes("w-full").props("outlined dense")
                    s_time = ui.input("Time", placeholder="e.g. 2:00 PM ET").classes("w-full").props("outlined dense")
                    s_status = ui.select(["Confirmed", "Tentative", "Pending", "Requested"], value="Pending").classes("w-full").props("outlined dense")
                    s_type = ui.select(["Intro call", "Follow-up call", "1x1 — Investor Conference", "Model update call",
                                         "PT discussion", "NDR meeting", "Callback", "Earnings call Q&A", "Other"], value="Intro call").classes("w-full").props("outlined dense")
                s_topic = ui.textarea("Topic / agenda").classes("w-full").props("outlined autogrow")

            # Optional model/document attach — the file itself is held in
            # memory (pending_upload) until "Add to Meeting Queue" actually
            # creates the meeting record, so the saved document can be
            # linked to that meeting's id rather than floating unattached.
            pending_upload = {"filename": None, "content_type": None, "bytes": None}
            upload_status = ui.label("").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

            async def handle_model_upload(e):
                pending_upload["bytes"] = await e.file.read()
                pending_upload["filename"] = e.file.name
                pending_upload["content_type"] = e.file.type
                upload_status.text = f"Attached: {e.file.name} — will be saved when you add this to the queue."
                upload_status.style(f"color:#15803D;font-size:var(--fs-sm);")

            ui.label("Attach the analyst's model or agenda doc (optional):").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);margin-top:6px;")
            ui.upload(on_upload=handle_model_upload, auto_upload=True).props("flat").classes("w-full")

            def add_meeting():
                if not (s_contact.value and s_firm.value):
                    ui.notify("Contact and firm are required.", type="warning")
                    return
                meeting_id = str(uuid.uuid4())
                sched = _load_json("scheduled_meetings.json", [])
                sched.append({"id": meeting_id, "Contact": s_contact.value, "Firm": s_firm.value, "Side": s_side.value,
                              "Date": s_date.value, "Time": s_time.value, "Type": s_type.value,
                              "Topic": s_topic.value, "Status": s_status.value, "Priority": s_priority.value})
                _save_json("scheduled_meetings.json", sched)
                if pending_upload["bytes"]:
                    documents.save_document(
                        contact=s_contact.value, firm=s_firm.value, doc_type="model",
                        filename=pending_upload["filename"], file_bytes=pending_upload["bytes"],
                        content_type=pending_upload["content_type"], source="manual_upload",
                        uploaded_by=CI().get("name"), linked_meeting_id=meeting_id,
                    )
                ui.notify(f"{s_contact.value} at {s_firm.value} scheduled for {s_date.value}"
                          + (" — model attached" if pending_upload["bytes"] else ""))
                _refresh()

            ui.button("Add to Meeting Queue", on_click=add_meeting).props("color=primary")

        with ui.tab_panel(h3):
            ui.label("Post-Meeting Note Capture — AI Structures Your Raw Notes").classes("font-bold")
            ui.label("Type your raw notes immediately after the call — an AI model organizes them into questions, concerns, signals, and actions.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            with ui.row().classes("w-full gap-4"):
                with ui.column().classes("flex-1"):
                    n_contact = ui.input("Contact name").classes("w-full").props("outlined dense")
                    n_firm = ui.input("Firm", autocomplete=firm_options).classes("w-full").props("outlined dense")
                    n_firm.on_value_change(lambda e: _autofill_contact(e, n_contact))
                    n_side = ui.select(["Buy-side", "Sell-side"], label="Side", value="Buy-side").classes("w-full").props("outlined dense")
                    n_type = ui.select(["Intro call", "Follow-up call", "1x1 — Investor Conference",
                                         "Model update call", "Callback", "Other", "Remove"],
                                        label="Meeting Type", value="Intro call").classes("w-full").props("outlined dense")
                with ui.column().classes("flex-1"):
                    n_raw = ui.textarea("Raw notes", placeholder="Type exactly what you heard and said, no need to format").classes("w-full").props("rows=8 outlined")

                    # Topic chips — a blank textarea right after a call tends
                    # to get skipped or filled with two rushed sentences.
                    # These don't force structure (still one freeform box,
                    # since real calls don't happen in a fixed topic order)
                    # but nudge coverage of the financial topics that matter
                    # most, and the inserted headers also give the AI a much
                    # stronger signal to pull from than unlabeled prose.
                    ui.label("Click a topic to add it to your notes:").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);margin-top:2px;")
                    with ui.row().classes("gap-1.5").style("flex-wrap:wrap;"):
                        for topic in NOTE_TOPICS:
                            def _insert_topic(t=topic):
                                prefix = "\n" if n_raw.value and not n_raw.value.endswith("\n") else ""
                                n_raw.value = f"{n_raw.value}{prefix}{t}: "
                            ui.button(topic, on_click=_insert_topic).props("flat dense size=sm").style(
                                f"background:{COLORS['surface_hover_bg']};font-size:var(--fs-xs);padding:2px 8px;")

            n_pending_upload = {"filename": None, "content_type": None, "bytes": None}
            n_upload_status = ui.label("").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

            async def handle_note_upload(e):
                n_pending_upload["bytes"] = await e.file.read()
                n_pending_upload["filename"] = e.file.name
                n_pending_upload["content_type"] = e.file.type
                n_upload_status.text = f"Attached: {e.file.name}"
                n_upload_status.style("color:#15803D;font-size:var(--fs-sm);")

            ui.label("Attach anything that came with this note (updated model, PDF, etc.) — optional:").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            ui.upload(on_upload=handle_note_upload, auto_upload=True).props("flat").classes("w-full")

            result_area = ui.column().classes("w-full")

            def structure_notes():
                if not (n_raw.value and n_contact.value):
                    ui.notify("Contact name and raw notes are required.", type="warning")
                    return
                ui.notify("Structuring notes...")
                structured = _structure_notes_with_ai(n_raw.value, n_contact.value, n_firm.value, n_type.value)
                all_notes = _load_json("post_meeting_notes.json", [])
                all_notes.append({"Contact": n_contact.value, "Firm": n_firm.value, "Side": n_side.value,
                                  "Date": today.strftime("%Y-%m-%d"), "Type": n_type.value, "Raw": n_raw.value,
                                  "Structured": structured})
                _save_json("post_meeting_notes.json", all_notes)
                if n_pending_upload["bytes"]:
                    documents.save_document(
                        contact=n_contact.value, firm=n_firm.value, doc_type="note_attachment",
                        filename=n_pending_upload["filename"], file_bytes=n_pending_upload["bytes"],
                        content_type=n_pending_upload["content_type"], source="manual_upload",
                        uploaded_by=CI().get("name"),
                    )
                result_area.clear()
                with result_area:
                    _render_structured_note(structured)
                ui.notify(f"Submitted — notes saved for {n_contact.value} at {n_firm.value}"
                          + (" (with attachment)" if n_pending_upload["bytes"] else ""))

            ui.button("Structure Notes with AI", on_click=structure_notes).props("color=primary")
            ui.markdown("---")

            if notes:
                ui.label("Previous Meeting Notes").classes("font-bold")
                for note in sorted(notes, key=lambda x: x.get("Date", ""), reverse=True)[:10]:
                    structured = note.get("Structured", {})
                    sent = structured.get("sentiment", "—") if isinstance(structured, dict) else "—"
                    with ui.expansion(f"{note['Date']} · {note['Contact']} — {note['Firm']} · Sentiment: {sent}").classes("w-full"):
                        _render_structured_note(structured if isinstance(structured, dict) else {})


def _render_structured_note(structured):
    sent = structured.get("sentiment", "—")
    sent_clr = "#15803D" if sent == "Positive" else "#B91C1C" if sent == "Negative" else "#B45309"
    ui.label(f"Sentiment: {sent}").style(f"color:{sent_clr};font-weight:bold;")
    if structured.get("summary"):
        ui.label(structured["summary"]).style(f"color:{COLORS['text_body']};font-size:var(--fs-base);")
    for key, title in [("financial_kpi_takeaways", "Financial/KPI takeaways"),
                        ("key_questions", "Key questions"), ("positive_signals", "Positive signals"),
                        ("concerns_raised", "Concerns raised"), ("follow_up_actions", "Follow-up actions"),
                        ("commitments_made", "Commitments made")]:
        items = structured.get(key)
        if items:
            ui.label(title).style("font-weight:bold;font-size:var(--fs-base);margin-top:4px;")
            for it in items:
                ui.label(f"- {it}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")


_CATEGORY_LABELS = {
    "model": "Model", "research_note": "Research Note", "ndr_request": "NDR Request",
    "conference_invite": "Conference Invite", "speak_to_management": "Speak to Management",
    "meeting_confirmation": "Meeting Confirmation", "shareholder_inquiry": "Shareholder Inquiry",
}


def _fmt_xlsx_cell(v, fmt):
    """Render one spreadsheet cell for the HTML preview, honouring the common Excel number
    formats we author (percent, currency, plain decimal). Text passes through unchanged."""
    if v is None:
        return ""
    if not isinstance(v, (int, float)) or isinstance(v, bool):
        return str(v)
    fmt = fmt or ""
    if "%" in fmt:
        dec = 1 if "0.0" in fmt else 0
        return f"{v * 100:.{dec}f}%"
    if "$" in fmt:
        dec = 2 if "0.00" in fmt else 1
        return f"${v:,.{dec}f}"
    dec = 1 if "0.0" in fmt else 0
    return f"{v:,.{dec}f}"


def _open_attachment_preview(doc_id, filename):
    """Pull the stored attachment up IN THE APP as a legible preview, instead of pushing a raw
    binary download the viewer's OS might open in Notepad (a spreadsheet dumped as text reads as
    mojibake). Spreadsheets render as a styled table mirroring the sheet; PDFs embed inline; text/
    calendar files show as text. A 'Download original' button is always offered as the fallback."""
    import html as _html
    result = documents.get_document_bytes(doc_id)
    dialog = ui.dialog()
    with dialog, ui.card().classes("w-full").style("max-width:900px;background:#FFFFFF;"):
        if not result:
            ui.label("This attachment could not be loaded.").style(f"color:{COLORS['text_muted']};")
            ui.button("Close", on_click=dialog.close).props("flat")
            dialog.open()
            return
        fname, ctype, raw = result
        ext = (fname or "").lower().rsplit(".", 1)[-1]
        with ui.row().classes("w-full items-center justify-between no-wrap"):
            ui.label(fname).classes("font-bold").style(f"color:{COLORS['text_heading']};font-size:var(--fs-md);")
            ui.label("Parsed from the analyst's emailed model" if ext in ("xlsx", "xlsm", "xls")
                     else "Attachment preview").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

        body = ui.column().classes("w-full").style("max-height:70vh;overflow:auto;margin-top:6px;")
        rendered = False
        try:
            if ext in ("xlsx", "xlsm", "xls"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
                MAXCOL = 7

                def _cell_style(c, numeric, spanning):
                    bold = "font-weight:700;" if (c.font and c.font.bold) else ""
                    align = (c.alignment.horizontal if c.alignment and c.alignment.horizontal
                             else ("right" if numeric else "left"))
                    col = ""
                    try:
                        rgb = c.font.color.rgb if c.font and c.font.color else None
                        if isinstance(rgb, str) and len(rgb) == 8 and rgb.upper() != "FF000000":
                            col = f"color:#{rgb[2:]};"
                    except Exception:
                        col = ""
                    # Header/paragraph rows (a single spanning cell) wrap; data cells stay on one line
                    # so the number columns line up instead of one long label stretching column A.
                    wrap = "white-space:normal;" if spanning else "white-space:nowrap;"
                    return f"text-align:{align};{bold}{col}{wrap}vertical-align:top;"

                with body:
                    for ws in wb.worksheets:
                        rows_html = []
                        for row_i in range(1, ws.max_row + 1):
                            cells = [ws.cell(row=row_i, column=ci) for ci in range(1, MAXCOL + 1)]
                            nonempty = [i for i, c in enumerate(cells) if c.value not in (None, "")]
                            if not nonempty:
                                continue
                            if nonempty == [0]:
                                # title / section header / thesis — span the full width and wrap
                                c = cells[0]
                                txt = _html.escape(_fmt_xlsx_cell(c.value, c.number_format))
                                rows_html.append(
                                    f"<tr><td colspan={MAXCOL} style='padding:4px 9px;"
                                    f"{_cell_style(c, False, True)}'>{txt}</td></tr>")
                            else:
                                tds = []
                                for c in cells:
                                    numeric = isinstance(c.value, (int, float)) and not isinstance(c.value, bool)
                                    txt = _html.escape(_fmt_xlsx_cell(c.value, c.number_format))
                                    tds.append(f"<td style='padding:3px 9px;"
                                               f"{_cell_style(c, numeric, False)}'>{txt}</td>")
                                rows_html.append("<tr>" + "".join(tds) + "</tr>")
                        ui.html(
                            "<table style='border-collapse:collapse;font-size:var(--fs-sm);width:100%;"
                            "border:1px solid #E2E8F0;'>" + "".join(rows_html) + "</table>")
                        rendered = True
            elif ext == "csv":
                text = raw.decode("utf-8", errors="replace")
                rows = list(csv.reader(io.StringIO(text)))
                with body:
                    cells = "".join(
                        "<tr>" + "".join(f"<td style='padding:3px 8px;border-bottom:1px solid #EEF;'>"
                                         f"{_html.escape(str(x))}</td>" for x in row) + "</tr>"
                        for row in rows)
                    ui.html(f"<table style='border-collapse:collapse;font-size:var(--fs-sm);'>{cells}</table>")
                    rendered = True
            elif ext == "pdf":
                import base64
                b64 = base64.b64encode(raw).decode("ascii")
                with body:
                    ui.html(f"<embed src='data:application/pdf;base64,{b64}' type='application/pdf' "
                            f"style='width:100%;height:60vh;border:1px solid #E2E8F0;'/>")
                    rendered = True
            elif ext in ("ics", "txt", "eml", "md", "json"):
                text = raw.decode("utf-8", errors="replace")
                with body:
                    ui.html(f"<pre style='white-space:pre-wrap;font-size:var(--fs-sm);color:#334155;'>"
                            f"{_html.escape(text)}</pre>")
                    rendered = True
        except Exception:
            rendered = False

        if not rendered:
            with body:
                ui.label("No inline preview for this file type — use Download to open it in its native app.").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        def _download(fname=fname, raw=raw):
            ui.download(raw, filename=fname)
        with ui.row().classes("w-full justify-end gap-2").style("margin-top:8px;"):
            ui.button("Download original", icon="download", on_click=_download).props("flat dense")
            ui.button("Close", on_click=dialog.close).props("flat dense color=primary")
    dialog.open()


def _render_pending_inbox_items(categories=None, title="Pending Inbox Items"):
    """The human half of the email-routing pipeline: core/mail_gateway.py
    classifies an inbound email (model / research note / NDR request /
    conference invite / speak-to-management) and queues it here rather than
    acting on an AI guess unattended (see core/email_classifier.py's
    docstring for why). One card per pending item, prefilled from whatever
    was extracted, with a category-appropriate confirm action:
      - model              -> core/consensus.py (Markets → Consensus Matrix)
      - research_note      -> CFA-lens breakdown (rating/PT, valuation method,
                              variant view, catalysts/risks, sentiment) so the
                              reviewer can decide fast whether it's worth
                              circulating further internally; optionally also
                              updates consensus if a rating/PT was extracted.
                              "Internal Use Only" is just a flag recorded on
                              the queue entry — actual forwarding is a manual
                              human decision, not something this app sends
                              (sell-side research is licensed content anyway).
      - ndr_request        -> this page's own Inbound NDR/Meeting Requests list
      - conference_invite  -> the Calendar's conference list
      - speak_to_management -> this page's Meeting Hub (Scheduled Meetings)
      - meeting_confirmation -> this page's Meeting Hub (Scheduled Meetings),
                              pre-filled Status "Confirmed" — replaces app.py's
                              old ad hoc IMAP-scan-with-password-prompt button
                              (see core/email_classifier.py's docstring)
    Every card also has Dismiss, for anything that turns out mis-tagged.

    `categories` optionally restricts which classified categories are shown, so a page can surface
    only the items that belong to it — the Meeting Hub passes the meeting-scheduling categories only,
    so a model / research note (which belong in the IR Inbox) never appears there. None = show all."""
    pending = inbox_queue.list_pending_items()
    if categories is not None:
        _cats = set(categories)
        pending = [p for p in pending if p.get("category", "general") in _cats]
    if not pending:
        return

    current = consensus.get_consensus()
    firms_with_data = {a["firm"] for a in CA()}

    with ui.card().classes("w-full").style(f"background:{COLORS['surface_bg']};border:1px solid {COLORS['accent']};margin-top:8px;"):
        ui.label(f"{title} ({len(pending)})").classes("font-bold").style(f"color:{COLORS['text_heading']};")
        ui.label("Classified and pre-filled from the email by AI where possible — review, correct if needed, "
                  "and confirm to send it where it belongs, or dismiss if it was mis-tagged.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        for item in pending:
            extracted = item.get("extracted") or {}
            category = item.get("category", "general")
            with ui.card().classes("w-full").style(f"background:{COLORS['canvas_bg']};border:1px solid {COLORS['border']};margin-top:6px;"):
                with ui.row().classes("w-full items-center justify-between"):
                    _hdr = f"{_CATEGORY_LABELS.get(category, category)} — {item['contact'] or 'unknown sender'}"
                    _hdr += f" ({item['firm']})" if item.get("firm") else (
                        " (individual investor)" if category == "shareholder_inquiry" else "")
                    ui.label(_hdr).classes("font-bold").style(f"color:{COLORS['accent_light']};")
                    ui.label(f"received {item['received_at']}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                ui.label(f"Subject: {item.get('subject') or '(no subject)'}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

                if item.get("doc_id") is not None:
                    def _preview(doc_id=item["doc_id"], fname=item.get("filename")):
                        _open_attachment_preview(doc_id, fname)
                    ui.button(f"{item['filename']}", icon="visibility", on_click=_preview).props(
                        "flat dense size=sm").style(f"color:{COLORS['accent_light']};font-size:var(--fs-xs);").tooltip(
                        "Open a legible preview in the app (Download original inside)")

                def dismiss(item_id=item["id"]):
                    inbox_queue.dismiss_item(item_id)
                    ui.notify("Dismissed.")
                    _refresh()

                if category == "model":
                    firm = item["firm"] if item["firm"] in firms_with_data else (list(firms_with_data)[0] if firms_with_data else item["firm"])
                    default_period = extracted.get("period") if extracted.get("period") in ALL_PERIODS else ALL_PERIODS[0]
                    existing = current["period_estimates"].get(default_period, {}).get(firm, {})

                    with ui.row().classes("w-full gap-3").style("margin-top:6px;"):
                        r_period = ui.select(ALL_PERIODS, value=default_period, label="Period").classes("flex-1").props("outlined dense")
                        r_rating = ui.select(["Buy", "Hold", "Sell", "Not Rated"],
                                              value=extracted.get("rating") or existing.get("Rating") or "Buy", label="Rating").classes("flex-1").props("outlined dense")
                    with ui.row().classes("w-full gap-3"):
                        r_pt = ui.number("Price Target ($)", value=extracted.get("price_target") or existing.get("Price Target") or 0.0, step=0.25).classes("flex-1").props("outlined dense")
                        r_eps = ui.number("EPS Est ($)", value=extracted.get("eps_est") or existing.get("EPS Est") or 0.0, step=0.01).classes("flex-1").props("outlined dense")
                        r_rev = ui.number("Revenue Est ($M)", value=extracted.get("revenue_est") or existing.get("Revenue Est ($M)") or 0.0, step=0.5).classes("flex-1").props("outlined dense")
                        r_ebd = ui.number("EBITDA Est ($M)", value=extracted.get("ebitda_est") or existing.get("EBITDA Est ($M)") or 0.0, step=0.1).classes("flex-1").props("outlined dense")
                    if extracted:
                        ui.label("Numbers pre-filled by AI from the attached file — check them before confirming.").style(
                            f"color:{COLORS['warning']};font-size:var(--fs-xs);")

                    def confirm(item_id=item["id"], firm=item["firm"], r_period=r_period, r_rating=r_rating,
                                r_pt=r_pt, r_eps=r_eps, r_rev=r_rev, r_ebd=r_ebd):
                        ok = consensus.confirm_model_review(
                            item_id, period=r_period.value, firm=firm, rating=r_rating.value,
                            price_target=r_pt.value or None, eps_est=r_eps.value if r_eps.value else None,
                            revenue_est=r_rev.value or None, ebitda_est=r_ebd.value or None,
                        )
                        if ok:
                            ui.notify(f"{firm} consensus updated for {r_period.value}.")
                        else:
                            ui.notify("That item was already actioned.", type="warning")
                        _refresh()
                    confirm_label = "Confirm & Update Consensus"

                elif category == "research_note":
                    # CFA-lens breakdown, not just a summary — the reviewer
                    # (Head of IR) should be able to decide in seconds
                    # whether this is worth circulating further internally.
                    # Forwarding itself stays a manual human decision (see
                    # this function's docstring) — this just gets them a
                    # sharper read than a generic AI summary would.
                    with ui.column().classes("w-full gap-1").style("margin-top:6px;"):
                        if extracted.get("thesis_summary"):
                            ui.label(extracted["thesis_summary"]).style(f"color:{COLORS['text_body']};font-size:var(--fs-base);")
                        meta_bits = []
                        if extracted.get("sentiment"):
                            meta_bits.append(f"Sentiment: {extracted['sentiment']}")
                        if extracted.get("variant_view"):
                            meta_bits.append(f"View: {extracted['variant_view']}")
                        if extracted.get("valuation_method"):
                            meta_bits.append(f"Method: {extracted['valuation_method']}")
                        if meta_bits:
                            ui.label(" · ".join(meta_bits)).style(f"color:{COLORS['accent_light']};font-size:var(--fs-sm);font-weight:bold;")
                        if extracted.get("key_assumptions"):
                            ui.label(f"Key assumptions: {extracted['key_assumptions']}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                        if extracted.get("catalysts_risks"):
                            ui.label(f"Catalysts/risks to watch: {extracted['catalysts_risks']}").style(f"color:{COLORS['warning']};font-size:var(--fs-sm);")
                        if extracted.get("prior_price_target") and extracted.get("price_target"):
                            ui.label(f"PT change: ${extracted['prior_price_target']} → ${extracted['price_target']}").style(
                                f"color:{COLORS['text_body']};font-size:var(--fs-sm);")

                    has_rating_or_pt = bool(extracted.get("rating") or extracted.get("price_target"))
                    if has_rating_or_pt:
                        firm = item["firm"] if item["firm"] in firms_with_data else (list(firms_with_data)[0] if firms_with_data else item["firm"])
                        default_period = extracted.get("period") if extracted.get("period") in ALL_PERIODS else ALL_PERIODS[0]
                        existing = current["period_estimates"].get(default_period, {}).get(firm, {})
                        with ui.row().classes("w-full gap-3").style("margin-top:6px;"):
                            rn_period = ui.select(ALL_PERIODS, value=default_period, label="Period").classes("flex-1").props("outlined dense")
                            rn_rating = ui.select(["Buy", "Hold", "Sell", "Not Rated"],
                                                   value=extracted.get("rating") or existing.get("Rating") or "Buy", label="Rating").classes("flex-1").props("outlined dense")
                            rn_pt = ui.number("Price Target ($)", value=extracted.get("price_target") or existing.get("Price Target") or 0.0, step=0.25).classes("flex-1").props("outlined dense")
                        rn_update_consensus = ui.checkbox("Also update consensus with this rating/PT", value=True)
                        rn_internal_only = ui.checkbox("Internal Use Only (flag before any further internal circulation)", value=True)

                        def confirm(item_id=item["id"], firm=firm, rn_period=rn_period, rn_rating=rn_rating,
                                    rn_pt=rn_pt, rn_update_consensus=rn_update_consensus, rn_internal_only=rn_internal_only):
                            notes = []
                            if rn_update_consensus.value:
                                consensus.update_estimate(rn_period.value, firm, rating=rn_rating.value,
                                                           price_target=rn_pt.value or None, source="email_research_note")
                                notes.append(f"consensus updated for {firm}, {rn_period.value}")
                            notes.append("marked Internal Use Only" if rn_internal_only.value else "reviewed")
                            inbox_queue.mark_confirmed(item_id, outcome="; ".join(notes).capitalize())
                            ui.notify(f"Research note from {firm} — {', '.join(notes)}.")
                            _refresh()
                    else:
                        rn_internal_only = ui.checkbox("Internal Use Only (flag before any further internal circulation)", value=True)

                        def confirm(item_id=item["id"], firm=item["firm"], rn_internal_only=rn_internal_only):
                            outcome = "Marked Internal Use Only" if rn_internal_only.value else "Reviewed"
                            inbox_queue.mark_confirmed(item_id, outcome=outcome)
                            ui.notify(f"Research note from {firm} — {outcome.lower()}.")
                            _refresh()
                    confirm_label = "Confirm Review"

                elif category == "ndr_request":
                    n_city = ui.input("City *", value=extracted.get("city") or "").classes("w-full").style("margin-top:6px;").props("outlined dense")
                    n_metro = ui.input("Metro region", value=extracted.get("metro") or "").classes("w-full").props("outlined dense")
                    n_reason = ui.textarea("Reason / context", value=extracted.get("reason") or "").classes("w-full").props("outlined autogrow")

                    def confirm(item_id=item["id"], contact=item["contact"], firm=item["firm"],
                                n_city=n_city, n_metro=n_metro, n_reason=n_reason):
                        if not n_city.value:
                            ui.notify("City is required.", type="warning")
                            return
                        reqs = _load_ndr_requests()
                        reqs.append({
                            "id": datetime.now().strftime("%Y%m%d%H%M%S"), "analyst": contact, "firm": firm,
                            "city": n_city.value, "metro": n_metro.value or n_city.value, "reason": n_reason.value or "—",
                            "received": datetime.now().strftime("%b %d, %Y"), "resolved": False, "seeded": False,
                        })
                        _save_ndr_requests(reqs)
                        inbox_queue.mark_confirmed(item_id, outcome=f"Logged NDR request for {n_city.value}")
                        ui.notify(f"NDR request from {contact} logged — start a roadshow from it in Outbound → NDR → Plan New NDR.")
                        _refresh()
                    confirm_label = "Log NDR Request"

                elif category == "conference_invite":
                    c_event = ui.input("Event name *", value=extracted.get("event_name") or "").classes("w-full").style("margin-top:6px;").props("outlined dense")
                    c_date = ui.input("Date (YYYY-MM-DD)", value=extracted.get("date") or "").classes("w-full").props("outlined dense")
                    c_loc = ui.input("Location", value=extracted.get("location") or "").classes("w-full").props("outlined dense")
                    # Organizer is parsed from the invite (the host running the conference — Baird,
                    # Sidoti…), which is NOT always the sender's firm. Prefill the parsed value and
                    # fall back to the sender firm; it used to be silently overwritten with the firm.
                    c_org = ui.input("Organizer", value=extracted.get("organizer") or item["firm"] or "").classes("w-full").props("outlined dense")
                    # RSVP / confirm-attendance-by date, parsed from the invite → the calendar's
                    # registration-deadline field, so the "reply by" date isn't lost.
                    c_deadline = ui.input("RSVP / confirm by (YYYY-MM-DD)",
                                          value=extracted.get("rsvp_deadline") or "").classes("w-full").props("outlined dense")

                    # Who from management is available? The first call an IR person makes on an invite —
                    # who can actually go. CEO and CFO first (the prepared-remarks speakers, same lineup
                    # as the earnings transcript), then IR and other execs. Anyone already booked on the
                    # event date is flagged, and the checked names become the calendar event's Attending.
                    from core import conferences
                    from config.client_config import role_roster as _role_roster
                    import re as _re
                    _roster = sorted(_role_roster() or [],
                                     key=lambda r: {"CEO": 0, "CFO": 1, "IR": 2}.get(r.get("role_key"), 3))
                    ui.label("Who from management is available?").style(
                        f"color:{COLORS['text_body']};font-size:var(--fs-sm);font-weight:600;margin-top:8px;")
                    _attend_cbs = {}
                    _avail_box = ui.column().classes("w-full gap-0")

                    def _render_avail(c_date=c_date, _roster=_roster, _attend_cbs=_attend_cbs, _avail_box=_avail_box):
                        # Recompute conflicts against the CURRENT date field so editing the date
                        # live-updates who's booked. Preserve any checkbox choices already made
                        # (clear-and-repopulate the SAME dict so confirm()'s reference stays valid).
                        _prev = {nm: cb.value for nm, cb in _attend_cbs.items()}
                        _attend_cbs.clear()
                        _avail_box.clear()
                        _d = (c_date.value or "").strip()
                        _busy = conferences.busy_attendees_on(_d) if _re.match(r"^\d{4}-\d{2}-\d{2}$", _d) else {}
                        with _avail_box:
                            if not _roster:
                                ui.label("No management roster on file for this client.").style(
                                    f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                                return
                            for _r in _roster:
                                _nm, _rk = _r.get("name"), _r.get("role_key")
                                _conflict = _busy.get((_nm or "").lower())
                                with ui.row().classes("items-center gap-2").style("margin-top:2px;"):
                                    _attend_cbs[_nm] = ui.checkbox(
                                        value=_prev.get(_nm, _rk in ("CEO", "CFO"))).props("dense")
                                    ui.label(f"{_r.get('label')} — {_nm}").style(
                                        f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
                                    if _conflict:
                                        ui.label(f"⚠ already booked: {_conflict}").style("color:#B45309;font-size:var(--fs-xs);")
                                    else:
                                        ui.label("available").style("color:#15803D;font-size:var(--fs-xs);")

                    _render_avail()
                    # Re-check availability when the date changes: immediately on a complete/empty
                    # value (paste or final digit), and on blur as a catch-all — without flickering
                    # the roster on every partial keystroke mid-typing.
                    def _on_date_change(_render_avail=_render_avail, c_date=c_date):
                        _d = (c_date.value or "").strip()
                        if _d == "" or _re.match(r"^\d{4}-\d{2}-\d{2}$", _d):
                            _render_avail()
                    c_date.on_value_change(lambda _e: _on_date_change())
                    c_date.on("blur", lambda _e: _render_avail())

                    def confirm(item_id=item["id"], firm=item["firm"], contact=item["contact"],
                                c_event=c_event, c_date=c_date, c_loc=c_loc, c_org=c_org, c_deadline=c_deadline,
                                _attend_cbs=_attend_cbs):
                        if not c_event.value:
                            ui.notify("Event name is required.", type="warning")
                            return
                        from core import conferences
                        _attending = ", ".join(nm for nm, cb in _attend_cbs.items() if cb.value) or "TBD"
                        _row, added = conferences.add_event(
                            event=c_event.value, date=c_date.value or None, location=c_loc.value or None,
                            organizer=c_org.value or firm, deadline=c_deadline.value or None,
                            attending=_attending, source="Email invite",
                            notes=f"Invitation received by email from {contact} ({firm}).")
                        inbox_queue.mark_confirmed(item_id, outcome=f"Added '{c_event.value}' to Calendar")
                        ui.notify(f"'{c_event.value}' added to Calendar — confirm details there." if added
                                  else f"'{c_event.value}' is already on the Calendar.")
                        _refresh()
                    confirm_label = "Add to Calendar"

                elif category == "speak_to_management":
                    m_contact_role = ui.input("Requested contact", value=extracted.get("requested_contact") or "CFO / IR").classes("w-full").style("margin-top:6px;").props("outlined dense")
                    m_topic = ui.textarea("Topic", value=extracted.get("topic") or "").classes("w-full").props("outlined autogrow")
                    m_date = ui.input("Proposed date (YYYY-MM-DD)", value=(datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")).classes("w-full").props("outlined dense")

                    def confirm(item_id=item["id"], contact=item["contact"], firm=item["firm"],
                                m_topic=m_topic, m_date=m_date):
                        meetings = db.load_json("scheduled_meetings.json", None) or []
                        meetings.append({
                            "id": str(uuid.uuid4()), "Contact": contact, "Firm": firm, "Side": "Buy-side",
                            "Date": m_date.value, "Time": "", "Type": "Callback",
                            "Topic": m_topic.value or "Speak-to-management request from email — confirm date/time.",
                            "Status": "Requested", "Priority": "Medium",
                        })
                        db.save_json("scheduled_meetings.json", meetings)
                        inbox_queue.mark_confirmed(item_id, outcome=f"Scheduled meeting request for {contact}")
                        ui.notify(f"Meeting request from {contact} added to Meeting Hub — confirm date/time there.")
                        _refresh()
                    confirm_label = "Schedule Meeting"

                elif category == "meeting_confirmation":
                    mc_type = ui.select(["1x1 call", "Conference call", "Video call", "In-person", "Other"],
                                         value=extracted.get("meeting_type") or "1x1 call", label="Meeting type").classes("w-full").style("margin-top:6px;").props("outlined dense")
                    mc_date = ui.input("Date (YYYY-MM-DD)", value=extracted.get("date") or (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")).classes("w-full").props("outlined dense")
                    mc_time = ui.input("Time", value=extracted.get("time") or "").classes("w-full").props("outlined dense")
                    mc_notes = ui.textarea("Notes", value=extracted.get("notes") or "").classes("w-full").props("outlined autogrow")

                    def confirm(item_id=item["id"], contact=item["contact"], firm=item["firm"],
                                mc_type=mc_type, mc_date=mc_date, mc_time=mc_time, mc_notes=mc_notes):
                        meetings = db.load_json("scheduled_meetings.json", None) or []
                        meetings.append({
                            "id": str(uuid.uuid4()), "Contact": contact, "Firm": firm, "Side": "Buy-side",
                            "Date": mc_date.value, "Time": mc_time.value, "Type": mc_type.value,
                            "Topic": mc_notes.value or "Confirmed via email — see original message for details.",
                            "Status": "Confirmed", "Priority": "Medium",
                        })
                        db.save_json("scheduled_meetings.json", meetings)
                        inbox_queue.mark_confirmed(item_id, outcome=f"Logged confirmed meeting with {contact} on {mc_date.value}")
                        ui.notify(f"Confirmed meeting with {contact} added to Meeting Hub.")
                        _refresh()
                    confirm_label = "Log Confirmed Meeting"

                elif category == "shareholder_inquiry":
                    from core import shareholder_reply
                    ui.label(f"Topic: {extracted.get('topic') or item.get('subject') or 'shareholder question'}").style(
                        f"color:{COLORS['text_body']};font-size:var(--fs-sm);margin-top:6px;")
                    _bits = " · ".join(b for b in (extracted.get("sub_type"), extracted.get("sentiment")) if b)
                    if _bits:
                        ui.label(_bits).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

                    # Regulation FD compliance flag — the point of this view. A shareholder fishing
                    # for undisclosed/forward info, or a live quiet period, is surfaced up front.
                    _note = shareholder_reply.compliance_note(extracted)
                    if _note:
                        with ui.card().classes("w-full").style(
                                "background:#FEF3C7;border:1px solid #B45309;border-left:4px solid #B45309;"
                                "padding:6px 10px;margin-top:6px;"):
                            ui.label("⚠ Regulation FD — review before replying").style(
                                "color:#8A5A0B;font-size:var(--fs-sm);font-weight:700;")
                            ui.label(_note).style("color:#7A5310;font-size:var(--fs-xs);line-height:1.4;")

                    _draft_box = ui.column().classes("w-full gap-1").style("margin-top:6px;")

                    def _generate_draft(subject=item.get("subject"), body=item.get("body") or "",
                                        extracted=extracted, _box=_draft_box, _reply_to=item.get("sender_email")):
                        d = shareholder_reply.draft_reply(subject, body, extracted)
                        _box.clear()
                        with _box:
                            _ta = ui.textarea("Draft reply — public info only; edit before sending", value=d["draft"]).props(
                                'outlined autogrow input-style="min-height:150px"').classes("w-full").style("font-size:var(--fs-sm);")
                            with ui.row().classes("gap-2"):
                                if _reply_to:
                                    def _reply(_ta=_ta, _to=_reply_to, _subj=subject or ""):
                                        _mt = (f"mailto:{_to}?subject={quote('Re: ' + _subj)}"
                                               f"&body={quote(_ta.value)}")
                                        ui.run_javascript(f"window.location.href = {json.dumps(_mt)}")
                                        ui.notify(f"Opening a reply to {_to} — review and send from your mail client.")
                                    ui.button("Reply", icon="reply", on_click=_reply).props("color=primary dense")

                                def _copy(_ta=_ta):
                                    ui.run_javascript(f"navigator.clipboard.writeText({json.dumps(_ta.value)})")
                                    ui.notify("Draft copied — paste it into your reply and send.", type="positive")
                                ui.button("Copy draft", icon="content_copy", on_click=_copy).props(
                                    ("flat dense" if _reply_to else "flat dense color=primary"))
                                ui.button("Re-draft", icon="refresh", on_click=_generate_draft).props("flat dense")
                            ui.label("IRconnect never sends on your behalf — you review, edit, and send this yourself.").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-2xs);")

                    with _draft_box:
                        ui.button("Draft a compliant reply", icon="smart_toy",
                                  on_click=_generate_draft).props("color=primary dense outline")
                        ui.label("Uses only public facts (next earnings date, filing locations, quiet-period "
                                 "status) — never guidance or unreleased numbers.").style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

                    def confirm(item_id=item["id"], contact=item["contact"]):
                        inbox_queue.mark_confirmed(item_id, outcome=f"Handled shareholder inquiry from {contact}")
                        ui.notify("Marked handled.")
                        _refresh()
                    confirm_label = "Mark handled"

                else:  # any future/unrecognized category — never leave confirm undefined
                    def confirm(item_id=item["id"]):
                        inbox_queue.mark_confirmed(item_id, outcome="Reviewed")
                        ui.notify("Marked reviewed.")
                        _refresh()
                    confirm_label = "Mark reviewed"

                with ui.row().classes("w-full gap-2").style("margin-top:6px;"):
                    ui.button(confirm_label, on_click=confirm).props("color=primary dense")
                    ui.button("Dismiss", on_click=dismiss).props("flat dense")


def _render_linked_documents(contact, firm):
    """Shown on an Upcoming Meetings card — "pull up the model the analyst
    sent in and the most recent note" in one place, instead of hunting
    through email or a shared drive. Looks up by contact+firm (the same
    join key used by post_meeting_notes.json and the Meeting Log), so this
    works whether the document was attached from Schedule Meeting, from
    Post-Meeting Notes, or (once wired) pulled in from the IR inbox."""
    model_doc = documents.get_latest_document(contact=contact, firm=firm, doc_type="model")
    research_doc = documents.get_latest_document(contact=contact, firm=firm, doc_type="research_note")
    note_doc = documents.get_latest_document(contact=contact, firm=firm, doc_type="note_attachment")
    if not (model_doc or research_doc or note_doc):
        return

    def _download(doc_id):
        result = documents.get_document_bytes(doc_id)
        if result:
            fname, _ctype, raw = result
            ui.download(raw, filename=fname)

    with ui.row().classes("w-full gap-2").style("margin-top:4px;flex-wrap:wrap;"):
        for _label, _doc in (("Latest model", model_doc),
                             ("Latest research note", research_doc),
                             ("Latest note attachment", note_doc)):
            if _doc:
                ui.button(f"{_label} — {_doc['filename']}",
                          on_click=lambda doc_id=_doc["id"]: _download(doc_id)) \
                    .props("flat dense size=sm").style(f"color:{COLORS['accent_light']};font-size:var(--fs-xs);")


def _hub_metric(label, value, hint="", active=False, on_click=None):
    """Summary tile shared across Meeting Hub / Target Database / SEC
    Intelligence. `hint` adds a one-line explanation so it isn't just a bare
    number; passing `on_click` makes it a clickable filter that highlights
    (accent top edge) when `active`."""
    border = COLORS["accent"] if active else COLORS["border"]
    bg = COLORS["surface_hover_bg"] if active else COLORS["surface_bg"]
    edge = COLORS["accent"] if active else "transparent"
    classes = "flex-1 text-center" + (" cursor-pointer" if on_click else "")
    card = ui.card().classes(classes).style(
        f"background:{bg};border:1px solid {border};border-top:3px solid {edge};")
    with card:
        ui.label(str(value)).classes("text-lg font-bold").style(f"color:{COLORS['text_heading']};")
        ui.label(label).style(f"color:{COLORS['text_secondary']};font-size:var(--fs-sm);font-weight:600;")
        if hint:
            ui.label(hint).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
    if on_click:
        card.on("click", on_click)
    return card


def _safe_date(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return datetime.now().date()


# ─────────────────────────────────────────────────────────────────────────
# Target Database tab
# ─────────────────────────────────────────────────────────────────────────
def _render_target_db_tab(institutions, client_id, mode_controls=None):
    prospects = _load_json("prospects.json", [])

    ui.label("Target Database").classes("text-lg font-bold")
    ui.label("Search and filter the tracked institution universe, plus manually-added prospects. The automated "
             "prospecting pipeline below (Analyst Coverage Network, live-13F auto-generation, NOBO cross-reference, "
             "bulk paste) is ported and rebuilt on the live SEC 13F fetcher — see each section for details.").style(
        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    _db_filter = {"mode": "all"}
    # Canonical Tier-1-ready set, SHARED with the Ownership summary card via _tier1_ready_names — so the
    # "Tier-1 ready" count here is exactly the "Target Investors" number the user clicked to get here.
    _t1set = _tier1_ready_names(institutions, client_id)
    # Arriving from an Ownership Big-Picture card (e.g. "Current Investors" → holders): open pre-filtered
    # to that bucket so the click lands on exactly the list the card summarized.
    _filter_pref = nav.pop_highlight("db_filter", None)
    if _filter_pref in ("all", "holders", "nonholders", "prospects", "tier1"):
        _db_filter["mode"] = _filter_pref
    db_cards_row = ui.row().classes("w-full gap-3").style("margin-top:6px;")
    # Geographic "where to go" block — the WHERE TO GO read + the "Institutions & Peer Ownership By
    # City/Metro Area" table — sits directly under the summary cards. Returns the metro-table expansion
    # so a summary-card click can collapse it (see set_db_filter) to focus on the filtered results.
    _geo_exp = _render_big_picture(institutions, "geo")
    # Search sits directly UNDER the metro table (the primary action); the mode toggle and the
    # automated-prospecting sections now follow it. Search + results live in a collapsed expansion —
    # the cards above are the summary; it auto-opens when a card filter is clicked, a search runs, or a
    # global-search prefill lands, so a filtered view is never hidden behind a closed panel.
    _search_exp = ui.expansion("Search by fund name", icon="search", value=False).classes("w-full").style(
        f"border:1px solid {COLORS['border']};border-radius:8px;margin-top:8px;")
    with _search_exp:
        search_in = ui.input("Search by fund name").classes("w-full").props("outlined dense")
        _search_btn_row = ui.row().classes("gap-2 items-center")
        results = ui.column().classes("w-full gap-2")
    # Pre/Post-earnings mode toggle + scoring-weights bar, now BELOW Search (moved down per request).
    if mode_controls:
        ui.markdown("---")
        mode_controls()
        ui.markdown("---")

    _OUTCOME_OPTIONS = {"— unresolved —": None, "Won (meeting set / met / owns)": "positive", "Passed": "negative"}
    _OUTCOME_REVERSE = {"positive": "Won (meeting set / met / owns)", "negative": "Passed", None: "— unresolved —"}

    def _db_combined():
        combined = [{"Fund": i["Fund"], "Metro": i["Metro"], "Type": i["Type"], "USIO_Holder": i["USIO_Holder"],
                    "Score": i["Engagement_Score"], "Source": "Tracked", "_pidx": None} for i in institutions]
        _seen = {(x["Fund"] or "").strip().upper() for x in combined}
        # Fold in the peer-owner prospect universe — the SAME set the Ownership summary, the Buyside tab,
        # and the geography table use. Without this the Target DB showed only the ~14 scored tracked
        # non-holders, so "Target Investors (N)" linked here to a much smaller list — the 13-vs-14 mismatch.
        try:
            from core import peer_prospects
            for c in (peer_prospects.build_candidates(client_id, limit=None) or []):
                nm = c.get("filer") or ""
                if not nm or nm.strip().upper() in _seen:
                    continue
                _seen.add(nm.strip().upper())
                combined.append({"Fund": nm, "Metro": (c.get("city") or "—").title(), "Type": "Peer-owner",
                                 "USIO_Holder": False, "Score": round(c.get("conviction") or 0),
                                 "Source": "Peer-owner", "_pidx": None})
        except Exception:
            pass
        for idx, p in enumerate(_load_json("prospects.json", [])):
            if (p.get("fund") or "").strip().upper() in _seen:
                continue
            combined.append({"Fund": p["fund"], "Metro": p.get("metro", "—"), "Type": p.get("style", "—"),
                             "USIO_Holder": False, "Score": p.get("score", 0), "Source": "Prospect", "_pidx": idx,
                             "_outcome": p.get("outcome"), "_has_score": bool(p.get("score_breakdown"))})
        return combined

    def do_search():
        results.clear()
        q = (search_in.value or "").lower().strip()
        combined = _db_combined()
        if q:
            combined = [c for c in combined if q in c["Fund"].lower()]
        mode = _db_filter["mode"]
        if mode == "holders":
            combined = [c for c in combined if c["USIO_Holder"]]
        elif mode == "nonholders":
            combined = [c for c in combined if not c["USIO_Holder"]]
        elif mode == "prospects":
            combined = [c for c in combined if c["Source"] == "Prospect"]
        elif mode == "tier1":
            combined = [c for c in combined if not c["USIO_Holder"] and (c["Fund"] or "").strip().upper() in _t1set]
        combined.sort(key=lambda x: -x["Score"])
        with results:
            _labels = {"all": "All targets", "holders": "Current holders", "nonholders": "Non-holders",
                       "prospects": "Prospects", "tier1": "Tier-1 ready (high-conviction non-holders)"}
            clear = "" if mode == "all" else "  ·  click All targets to clear"
            ui.label(f"{len(combined)} result(s) — {_labels[mode]}{clear}").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            if not combined:
                ui.label("No matches.").style(f"color:{COLORS['text_muted']};")
            # Cap the results window to ~5 rows tall (scroll for the rest) so a 100+ set doesn't dominate.
            _cards_box = ui.column().classes("w-full").style("max-height:360px;overflow-y:auto;gap:0;")
        with _cards_box:
            for c in combined:
                with _list_item_card():
                  with ui.row().classes("w-full items-center justify-between no-wrap"):
                    with ui.column().classes("gap-0").style("flex:1;min-width:0;"):
                        ui.label(f"{pretty_name(c['Fund'])} ({c['Source']})").classes("name-link").on(
                            "click", lambda c=c: _open_account_profile(c)).style(
                            f"color:{COLORS['text_heading']};font-size:var(--fs-lg);font-weight:700;")
                        # Drop any empty / None / placeholder segment (real SEC-sourced holders
                        # often carry no Type) so the meta line never renders a literal "· None ·".
                        _meta = "  ·  ".join(str(s) for s in (
                            c["Metro"], c["Type"], "Holder" if c["USIO_Holder"] else "Non-holder",
                            f"Score {c['Score']}")
                            if s is not None and str(s).strip().lower() not in ("none", "—", ""))
                        ui.label(_meta).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    # Outcome marker — only for prospects that were added WITH a
                    # stored Fit Score breakdown (see run_fit_score_ranking's
                    # add_scored_prospect), since that's what
                    # core.fit_score.suggest_reweight() needs to learn from.
                    # Every "Won" or "Passed" logged here is a labeled data
                    # point the re-weight suggestion picks up next time it runs.
                    if c["Source"] == "Prospect" and c.get("_has_score"):
                        def set_outcome(e, pidx=c["_pidx"]):
                            plist = _load_json("prospects.json", [])
                            plist[pidx]["outcome"] = _OUTCOME_OPTIONS.get(e.value)
                            _save_json("prospects.json", plist)
                            ui.notify("Outcome saved.")

                        ui.select(list(_OUTCOME_OPTIONS.keys()), value=_OUTCOME_REVERSE.get(c.get("_outcome")),
                                  on_change=set_outcome).props("dense outlined").classes("min-w-[220px]")

    def set_db_filter(mode):
        _db_filter["mode"] = mode
        render_db_cards()
        do_search()
        _search_exp.set_value(True)  # a filtered list must not stay hidden
        if _geo_exp is not None:
            _geo_exp.set_value(False)  # collapse the metro table so the filtered results take the screen

    def render_db_cards():
        c = _db_combined()
        db_cards_row.clear()
        with db_cards_row:
            _hub_metric("All targets", len(c), "Tracked + prospects",
                        _db_filter["mode"] == "all", lambda: set_db_filter("all"))
            _hub_metric("Current holders", sum(1 for x in c if x["USIO_Holder"]), f"Own {CT('ticker')} now",
                        _db_filter["mode"] == "holders", lambda: set_db_filter("holders"))
            _hub_metric("Non-holders", sum(1 for x in c if not x["USIO_Holder"]), "Conversion targets",
                        _db_filter["mode"] == "nonholders", lambda: set_db_filter("nonholders"))
            _hub_metric("Tier-1 ready",
                        sum(1 for x in c if not x["USIO_Holder"] and (x["Fund"] or "").strip().upper() in _t1set),
                        "High-conviction · matches the summary", _db_filter["mode"] == "tier1",
                        lambda: set_db_filter("tier1"))
            _hub_metric("Prospects", sum(1 for x in c if x["Source"] == "Prospect"), "Added to database",
                        _db_filter["mode"] == "prospects", lambda: set_db_filter("prospects"))

    # A global-search result for a fund lands here pre-searched (nav.go_to passes
    # the fund name as search_prefill), so the user sees exactly the record they
    # clicked instead of the full list.
    _prefill = nav.pop_highlight("search_prefill", None)
    if _prefill:
        search_in.value = _prefill
        _search_exp.set_value(True)  # arrived from global search — show the hit
    if _filter_pref:
        _search_exp.set_value(True)  # arrived from an Ownership card — open the filtered list, don't hide it
    search_in.on("keydown.enter", do_search)
    with _search_btn_row:
        ui.button("Search", on_click=do_search).props("dense")
    render_db_cards()
    do_search()

    _add_div = ui.markdown("---")
    _add_exp = ui.expansion("Add a prospect manually").classes("w-full")
    with _add_exp:
        p_fund = ui.input("Fund name *").classes("w-full").props("outlined dense")
        p_metro = ui.input("Metro / region").classes("w-full").props("outlined dense")
        p_style = ui.input("Investment style").classes("w-full").props("outlined dense")
        p_score = ui.number("Fit score (0-100)", value=50, min=0, max=100).props("outlined dense")
        p_notes = ui.textarea("Notes").classes("w-full").props("outlined autogrow")

        def add_prospect():
            if not p_fund.value:
                ui.notify("Fund name is required.", type="warning")
                return
            plist = _load_json("prospects.json", [])
            plist.append({"fund": p_fund.value, "metro": p_metro.value, "style": p_style.value,
                          "score": p_score.value, "notes": p_notes.value, "added": datetime.now().strftime("%Y-%m-%d")})
            _save_json("prospects.json", plist)
            ui.notify(f"Added {p_fund.value} to the prospect list.")
            _refresh()

        ui.button("Add Prospect", on_click=add_prospect).props("color=primary")

    _bulk_div = ui.markdown("---")
    _bulk_exp = ui.expansion("Bulk paste-from-website", value=False).classes("w-full")
    with _bulk_exp:
        ui.label("Some sites (WhaleWisdom in particular) render holder tables as styled grids, not plain HTML "
                  "tables — copy/paste from those won't preserve columns cleanly through a file upload. Paste raw "
                  "text here instead; it's parsed by tab or by runs of 2+ spaces.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        bp_raw = ui.textarea("Paste table here", placeholder="Paste tab-separated or space-aligned rows here...").classes("w-full").props("outlined autogrow")
        bp_preview = ui.column().classes("w-full")
        bp_state = {"columns": [], "rows": [], "fund_col_idx": 0}

        def bp_parse():
            bp_preview.clear()
            columns, rows = prospecting.parse_pasted_table(bp_raw.value or "")
            bp_state["columns"], bp_state["rows"] = columns, rows
            with bp_preview:
                if not rows:
                    ui.label("Nothing parsed yet — paste some rows above.").style(f"color:{COLORS['text_muted']};")
                    return
                ui.label(f"Parsed {len(rows)} row(s) — check this looks right before adding.").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                fund_col_sel = ui.select(columns, value=columns[0], label="Which column is the fund/investor name?").classes("w-full").props("outlined dense")

                def set_fund_col():
                    bp_state["fund_col_idx"] = columns.index(fund_col_sel.value)
                fund_col_sel.on_value_change(set_fund_col)

                with ui.column().classes("w-full").style(f"max-height:220px;overflow-y:auto;border:1px solid {COLORS['border']};border-radius:6px;padding:6px;"):
                    for r in rows[:30]:
                        ui.label(" | ".join(r)).style(f"color:{COLORS['text_body']};font-size:var(--fs-xs);")
                    if len(rows) > 30:
                        ui.label(f"... + {len(rows) - 30} more rows").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

                def add_all():
                    plist = _load_json("prospects.json", [])
                    existing_names = {p["fund"] for p in plist}
                    added = 0
                    fi = bp_state["fund_col_idx"]
                    for r in bp_state["rows"]:
                        name = r[fi].strip() if fi < len(r) else ""
                        if not name or name.lower() in ("history", "nan", ""):
                            continue
                        if name in existing_names:
                            continue
                        other = {columns[i]: r[i] for i in range(len(r)) if i != fi and r[i]}
                        plist.append({
                            "fund": name, "metro": "—", "style": "Unknown", "score": 0,
                            "notes": " · ".join(f"{k}: {v}" for k, v in other.items()),
                            "added": datetime.now().strftime("%Y-%m-%d"), "source": "Bulk paste",
                        })
                        existing_names.add(name)
                        added += 1
                    _save_json("prospects.json", plist)
                    ui.notify(f"Added {added} funds to the prospect queue.")
                    _refresh()

                ui.button(f"Add all parsed rows to Prospect Queue", on_click=add_all).props("dense").style("margin-top:6px;")

        bp_raw.on_value_change(bp_parse)
        ui.button("Parse", on_click=bp_parse).props("dense flat")

    _analyst_div = ui.markdown("---")
    _analyst_exp = ui.expansion("Analyst Coverage Network Targeting", value=False).classes("w-full")
    with _analyst_exp:
        ui.label("Institutions already owning other stocks a covering analyst rates Buy have demonstrated they "
                  "trust that analyst's research and understand the surrounding sector thesis — the "
                  "highest-probability prospects once that analyst raises this client's target.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        coverage = analyst_coverage.get_coverage(client_id)
        if not coverage:
            ui.label("No covering analysts on file yet.").style(f"color:{COLORS['text_muted']};")
        else:
            analyst_keys = list(coverage.keys())
            analyst_sel = ui.select(analyst_keys, value=analyst_keys[0], label="Select analyst coverage network to mine").classes("w-full").props("outlined dense")
            coverage_list = ui.column().classes("w-full")

            def render_coverage_list():
                coverage_list.clear()
                cov = analyst_coverage.get_coverage(client_id)
                data = cov.get(analyst_sel.value)
                if not data:
                    return
                with coverage_list:
                    ui.label(f"{data['analyst']} — {data['firm']} · {data['email']} · "
                              f"{len(data['coverage'])} stocks in coverage universe").classes("font-bold").style(
                        f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
                    for stock in sorted(data["coverage"], key=lambda s: -s.get("relevance", 0)):
                        rel = stock.get("relevance", 0)
                        rel_clr = "#15803D" if rel >= 80 else "#B45309" if rel >= 50 else "#94A3B8"
                        with ui.card().classes("w-full").style(f"background:{COLORS['surface_hover_bg']};margin-top:4px;"):
                            with ui.row().classes("w-full justify-between items-center"):
                                ui.label(f"{stock['ticker']} — {stock['name']} · {stock.get('sector', '—')}").classes("font-bold").style(
                                    f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
                                ui.label(f"{stock.get('rating', '—')} · PT ${stock.get('pt', 0):.2f} · Relevance {rel}/100").style(
                                    f"color:{rel_clr};font-size:var(--fs-sm);font-weight:bold;")
                            if stock.get("bridge"):
                                ui.label(stock["bridge"]).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                            if stock.get("shared_dna"):
                                ui.label(" · ".join(t.strip() for t in stock["shared_dna"].split("·") if t.strip())).style(
                                    f"color:{COLORS['accent_light']};font-size:var(--fs-xs);")

            analyst_sel.on_value_change(render_coverage_list)
            render_coverage_list()

            ui.markdown("---")
            with ui.expansion("Add stock to this analyst's coverage").classes("w-full"):
                ac_ticker = ui.input("Ticker").classes("w-full").props("outlined dense")
                ac_name = ui.input("Company name").classes("w-full").props("outlined dense")
                ac_pt = ui.number("Price target ($)", value=0.0, step=0.25).props("outlined dense")
                ac_rating = ui.select(["Buy", "Neutral", "Sell"], value="Buy").props("outlined dense")
                ac_sector = ui.input("Sector").classes("w-full").props("outlined dense")
                ac_relevance = ui.number(f"{CT('ticker')} relevance (0-100)", value=50, min=0, max=100).props("outlined dense")
                ac_bridge = ui.textarea("Why holders of this ticker are prospects").classes("w-full").props("outlined autogrow")

                def add_coverage_entry():
                    if not (ac_ticker.value and ac_name.value):
                        ui.notify("Ticker and company name are required.", type="warning")
                        return
                    ok = analyst_coverage.add_coverage_stock(
                        analyst_sel.value, ac_ticker.value, ac_name.value, ac_pt.value, ac_rating.value,
                        ac_sector.value, relevance=int(ac_relevance.value), bridge=ac_bridge.value, client_id=client_id,
                    )
                    if ok:
                        ui.notify(f"{ac_ticker.value.upper()} added to {analyst_sel.value}'s coverage.")
                        render_coverage_list()
                    else:
                        ui.notify(f"{ac_ticker.value.upper()} is already in this analyst's coverage.", type="warning")

                ui.button("Add", on_click=add_coverage_entry).props("dense")

    ui.markdown("---")
    with ui.expansion("Automated Prospecting Pipeline — Live 13F Holders of Coverage-Network Stocks", value=False).classes("w-full"):
        ui.label("For the selected analyst's other Buy-rated stocks scoring at or above the threshold, pulls real "
                  "SEC 13F institutional holders (already fetched via the 13F refresh in Settings → Data Sources "
                  "— this does not trigger a new download itself), excludes anyone "
                  "already tracked or already in the prospect queue, and ranks the rest.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        coverage2 = analyst_coverage.get_coverage(client_id)
        if not coverage2:
            ui.label("No covering analysts on file — add one above first.").style(f"color:{COLORS['text_muted']};")
        else:
            analyst_keys2 = list(coverage2.keys())
            auto_analyst = ui.select(analyst_keys2, value=analyst_keys2[0], label="Analyst coverage network").classes("w-full").props("outlined dense")
            ui.label(f"Minimum {CT('ticker')} relevance score to include:").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
            auto_threshold = ui.slider(min=0, max=100, value=50, step=5).props("label-always")
            auto_results = ui.column().classes("w-full")

            def run_auto_prospecting():
                auto_results.clear()
                known_names = {i["Fund"] for i in institutions} | {p["fund"] for p in _load_json("prospects.json", [])}
                result = prospecting.generate_coverage_prospects(
                    auto_analyst.value, min_relevance=int(auto_threshold.value),
                    known_fund_names=known_names, client_id=client_id,
                )
                with auto_results:
                    if result["not_fetched_tickers"]:
                        ui.label(f"No live 13F data cached yet for: {', '.join(result['not_fetched_tickers'])} — "
                                  f"add them to the Peer Universe below and refresh 13F from Settings → Data Sources "
                                  f"to include their holders here.").style(f"color:{COLORS['warning']};font-size:var(--fs-xs);")
                    # Show what the quality gates removed, so a short (or empty) list
                    # reads as "the filter worked", not "the pipeline is broken".
                    _filt = result.get("filtered_passive", 0) + result.get("filtered_breadth", 0)
                    if not result["prospects"]:
                        msg = ("No new prospects found at this relevance threshold."
                               if not _filt else
                               f"No qualifying institutional prospects — {_filt} holder(s) were found but "
                               "filtered out as index/passive books, market makers, banks, or ETF issuers. "
                               "Lower the relevance threshold or add more coverage tickers to widen the net.")
                        ui.label(msg).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                        # NB: no early return — the RIA bucket below must still render;
                        # "no institutional targets, but 6 RIAs own it" is a real answer.
                    # Deliberately NOT called "prospects found" — that collided with the
                    # "Prospects" card above and read like the two numbers should match.
                    # They're disjoint by construction: this list EXCLUDES everything
                    # already tracked or already in the prospect queue (known_fund_names),
                    # so these are candidates that are not in the database until ADDed.
                    if result["prospects"]:
                        ui.label(f"{len(result['prospects'])} new suggestion(s) — not in your database yet").classes("font-bold")
                        _fnote = (f" · {_filt} filtered out as passive / market-maker / index noise"
                                  if _filt else "")
                        ui.label("Already-tracked funds and prospects you've added are filtered out of this list. "
                                 f"Click ADD to move one into the Prospects count on the cards above.{_fnote}").style(
                            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                    for p in result["prospects"][:25]:
                        with _list_item_card():
                          with ui.row().classes("w-full items-start justify-between no-wrap"):
                            with ui.column().classes("gap-0 flex-1"):
                                ui.label(f"{pretty_name(p['fund'])} — via {p['source_ticker']} ({p['source_name']})").classes("font-bold").style(
                                    f"color:{COLORS['text_heading']};font-size:var(--fs-lg);")
                                ui.label(p["talking_point"]).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                                # size_known=False means this holder came from the EDGAR full-text-search
                                # path (sec_filings._search_holders_for_ticker) — confirmed to hold a
                                # position, but exact share/dollar size wasn't fetched (see that
                                # function's docstring). Showing "1 shares · $0 reported" in that case
                                # would be misleadingly precise-looking for a value we don't actually have.
                                size_line = (f"{p['shares']:,} shares · ${p['value']:,} reported · relevance {p['relevance']}/100"
                                             if p.get("size_known", True)
                                             else f"confirmed holder (position size not available) · relevance {p['relevance']}/100")
                                ui.label(size_line).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

                            def add_this(p=p):
                                plist = _load_json("prospects.json", [])
                                if any(x["fund"] == p["fund"] for x in plist):
                                    ui.notify(f"{p['fund']} is already in the prospect queue.", type="warning")
                                    return
                                plist.append({
                                    "fund": p["fund"], "metro": "—", "style": "Unknown", "score": p["relevance"],
                                    "notes": p["talking_point"], "added": datetime.now().strftime("%Y-%m-%d"),
                                    "source": f"Coverage Network — {p['source_ticker']}",
                                })
                                _save_json("prospects.json", plist)
                                ui.notify(f"{p['fund']} added to prospect queue.")
                                _refresh()

                            ui.button("Add", on_click=add_this).props("flat dense")

                    # ── RIA / wealth bucket ──────────────────────────────────
                    # These own the stock but aren't institutional NDR targets —
                    # kept out of the list above, yet visible here (only ones with
                    # a real position) since who owns adjacent names through the
                    # advisory channel is still worth knowing.
                    _rias = result.get("ria_holders") or []
                    if _rias:
                        with ui.expansion(f"RIA / wealth managers holding these names ({len(_rias)})",
                                          icon="account_balance_wallet", value=False).classes("w-full").style(
                                f"border:1px solid {COLORS['border']};border-radius:8px;margin-top:10px;"):
                            ui.label("Real positions, but held through client accounts — there's no portfolio "
                                     "manager to pitch, so these aren't NDR targets. Video-call tier: an "
                                     "assistant can set these up, no management travel needed. Add one and the "
                                     "NDR scheduler defaults it to a video call. Ranked by position size.").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                            for r in _rias[:25]:
                                with _list_item_card():
                                  with ui.row().classes("w-full items-center justify-between no-wrap"):
                                    with ui.column().classes("gap-0 flex-1"):
                                        ui.label(r["fund"]).style(
                                            f"color:{COLORS['text_heading']};font-size:var(--fs-lg);font-weight:700;")
                                        ui.label(f"{r['shares']:,} shares of {r['source_ticker']} "
                                                 f"({r['source_name']})").style(
                                            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

                                    def add_ria(r=r):
                                        plist = _load_json("prospects.json", [])
                                        if any(x["fund"] == r["fund"] for x in plist):
                                            ui.notify(f"{r['fund']} is already in the prospect queue.", type="warning")
                                            return
                                        plist.append({
                                            "fund": r["fund"], "metro": "—", "style": "RIA / wealth manager",
                                            "score": r["relevance"],
                                            # Low-touch tier: video call an assistant schedules, not
                                            # management travel. The NDR scheduler reads is_ria() and
                                            # defaults these to a video format for the same reason.
                                            "touch": "video-call",
                                            "outreach": "Video call — assistant can schedule",
                                            "notes": f"RIA/wealth manager — holds {r['shares']:,} shares of "
                                                     f"{r['source_ticker']} ({r['source_name']}). Video-call "
                                                     f"tier: no PM to pitch, so an assistant can set up a call.",
                                            "added": datetime.now().strftime("%Y-%m-%d"),
                                            "source": f"Coverage Network (RIA) — {r['source_ticker']}",
                                        })
                                        _save_json("prospects.json", plist)
                                        ui.notify(f"{r['fund']} added to prospect queue.")
                                        _refresh()

                                    ui.button("Add", on_click=add_ria).props("flat dense")

            ui.button("Auto-Generate Prospect List", on_click=run_auto_prospecting).props("color=primary dense")
            run_auto_prospecting()

    ui.markdown("---")
    with ui.expansion("NOBO Cross-Reference", value=False).classes("w-full"):
        ui.label("Request from the transfer agent (e.g. Computershare) quarterly — Excel or CSV with holder name "
                  "and shares. Compares a list of names (the prospect queue, tracked institutions, the call "
                  "listener log, the IR website visitor log, or a pasted list) against this client's actual NOBO "
                  "shareholder file — matches are already-identified shareholders; everyone else is a real, "
                  "unconfirmed prospect.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        nobo_status = ui.label("").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        def refresh_nobo_status():
            cur = prospecting.get_nobo_list(client_id)
            if cur.get("rows"):
                nobo_status.text = f"{cur['source_label']} — {len(cur['rows'])} holders on file (loaded {cur['_fetched_at'][:10]})"
            else:
                nobo_status.text = "No NOBO file on hand yet — request one from your transfer agent, or upload one below."
        refresh_nobo_status()

        async def handle_nobo_upload(e):
            try:
                raw = await e.file.read()
                fname = e.file.name
                if fname.lower().endswith(".csv"):
                    text = raw.decode("utf-8", errors="replace")
                    rows_in = list(csv.DictReader(io.StringIO(text)))
                else:
                    df = pd.read_excel(io.BytesIO(raw))
                    rows_in = df.to_dict("records")
                if not rows_in:
                    ui.notify("That file has no rows.", type="warning")
                    return

                def find_col(row, *keywords):
                    for k in row.keys():
                        if any(kw in str(k).lower() for kw in keywords):
                            return k
                    return None

                name_col = find_col(rows_in[0], "holder_name", "holder name", "name")
                shares_col = find_col(rows_in[0], "shares")
                if not (name_col and shares_col):
                    ui.notify("Couldn't find a holder-name and shares column — check this file's headers.", type="negative")
                    return
                parsed = []
                for r in rows_in:
                    try:
                        shares = int(float(r.get(shares_col, 0) or 0))
                    except (TypeError, ValueError):
                        shares = 0
                    name = str(r.get(name_col, "")).strip()
                    if name and name.lower() != "nan":
                        parsed.append({"holder_name": name, "shares": shares})
                prospecting.save_nobo_list(parsed, source_label=f"Uploaded: {fname}", client_id=client_id)
                ui.notify(f"NOBO list updated — {len(parsed)} holders.")
                refresh_nobo_status()
            except Exception as ex:
                ui.notify(f"Couldn't read that file: {ex}", type="negative")

        ui.upload(on_upload=handle_nobo_upload, auto_upload=True).props("accept=.csv,.xlsx flat").classes("w-full")

        ui.markdown("---")
        nobo_source_options = ["Prospect Queue", "Tracked Institutions"]
        if os.path.exists(client_data_path("q1_2026_call_listener_log.csv", client_id)):
            nobo_source_options.append("Call Listener Log")
        if os.path.exists(client_data_path("ir_website_visitor_log.csv", client_id)):
            nobo_source_options.append("IR Website Visitor Log")
        nobo_source_options.append("Paste names")

        nobo_compare_src = ui.select(nobo_source_options, value="Prospect Queue", label="Compare against:").classes("w-full").props("outlined dense")
        nobo_paste = ui.textarea("Paste one name per line").classes("w-full").props("outlined autogrow")
        nobo_paste.set_visibility(False)
        nobo_compare_src.on_value_change(lambda: nobo_paste.set_visibility(nobo_compare_src.value == "Paste names"))
        nobo_results = ui.column().classes("w-full")

        def run_nobo_compare():
            nobo_results.clear()
            src = nobo_compare_src.value
            if src == "Prospect Queue":
                names = [p["fund"] for p in _load_json("prospects.json", [])]
            elif src == "Tracked Institutions":
                names = [i["Fund"] for i in institutions]
            elif src == "Call Listener Log":
                cl = pd.read_csv(client_data_path("q1_2026_call_listener_log.csv", client_id))
                names = cl["Firm"].dropna().unique().tolist() if "Firm" in cl.columns else []
            elif src == "IR Website Visitor Log":
                vl = pd.read_csv(client_data_path("ir_website_visitor_log.csv", client_id))
                names = vl["Visitor_Organization"].dropna().unique().tolist() if "Visitor_Organization" in vl.columns else []
            else:
                names = [n.strip() for n in (nobo_paste.value or "").split("\n") if n.strip()]

            with nobo_results:
                if not prospecting.get_nobo_list(client_id).get("rows"):
                    ui.label("Upload a NOBO file above first.").style(f"color:{COLORS['text_muted']};")
                    return
                if not names:
                    ui.label("Nothing to compare — that source is empty.").style(f"color:{COLORS['text_muted']};")
                    return
                matched, unmatched = prospecting.match_against_nobo(names, client_id=client_id)
                ui.label(f"{len(matched)} identified on the NOBO file · {len(unmatched)} not found — real "
                          f"prospects, not already-known shareholders").classes("font-bold")
                with ui.row().classes("w-full gap-4"):
                    with ui.column().classes("flex-1").style(f"background:{COLORS['surface_hover_bg']};border-radius:8px;padding:8px;"):
                        ui.label("Identified shareholders").classes("font-bold").style("color:#15803D;font-size:var(--fs-sm);")
                        if matched:
                            for nm, mn, sh in matched:
                                ui.label(f"{nm} → {mn} — {sh:,} shares").style(f"color:{COLORS['text_body']};font-size:var(--fs-xs);")
                        else:
                            ui.label("None matched.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                    with ui.column().classes("flex-1").style(f"background:{COLORS['surface_hover_bg']};border-radius:8px;padding:8px;"):
                        ui.label("Not on the NOBO file (genuine prospects)").classes("font-bold").style(
                            f"color:{COLORS['text_heading']};font-size:var(--fs-sm);")
                        if unmatched:
                            for nm in unmatched:
                                ui.label(nm).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
                        else:
                            ui.label("Everything compared matched the NOBO file.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

        ui.button("Compare", on_click=run_nobo_compare).props("dense")

    ui.markdown("---")
    _render_peer_universe_manager(institutions)

    # Requested layout: the Search box and the manual-entry tools (Add a prospect · Bulk paste ·
    # Analyst Coverage) drop to the page BOTTOM, so the first section after the metro table is the
    # Automated Prospecting Pipeline. They're created earlier (so their search/init wiring runs in
    # place), then repositioned here into a bottom slot — creation order and handlers unchanged.
    _tools_slot = ui.column().classes("w-full").style("gap:0;")
    with _tools_slot:
        ui.markdown("---")
    for _el in (_add_div, _add_exp, _bulk_div, _bulk_exp, _analyst_div, _analyst_exp):
        _el.move(_tools_slot)


def _render_peer_universe_manager(institutions):
    ui.label("Peer Cross-Targeting — Find Institutions Owning Peers But Not This Client").classes("font-bold")
    peers = _load_peer_universe()

    with ui.expansion("Manage Peer Universe — Add or Remove Tickers", value=False).classes("w-full"):
        ui.label(f"Custom peer group saved to peer_universe.csv · Persists across restarts · {len(peers)} peers currently tracked").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

        manage_list = ui.column().classes("w-full")

        def render_manage_list():
            manage_list.clear()
            current = _load_peer_universe()
            with manage_list:
                for idx, p in enumerate(current):
                    with ui.row().classes("w-full items-center justify-between").style(f"border-bottom:1px solid {COLORS['border']};padding:4px 0;"):
                        ev = f"{p['ev_rev']}x" if p.get("ev_rev") else "no EV/Rev on file"
                        ui.label(f"{p['ticker']} — {p['name']} · {p.get('sector','—')} · {ev}").style(f"color:{COLORS['text_body']};font-size:var(--fs-base);")

                        def do_delete(idx=idx):
                            cur = _load_peer_universe()
                            removed = cur.pop(idx)
                            _save_peer_universe(cur)
                            ui.notify(f"{removed['ticker']} removed.")
                            render_manage_list()

                        # Fit Score inputs — tier drives Comparability fit +
                        # the P1-P4 tier's "core" bonus; weight drives Peer
                        # conviction (see core/fit_score.py). Saved
                        # immediately on change, no separate Save button —
                        # matches the delete button's immediate-effect pattern
                        # right next to it.
                        def set_tier(e, idx=idx):
                            cur = _load_peer_universe()
                            cur[idx]["tier"] = e.value
                            _save_peer_universe(cur)

                        def set_weight(e, idx=idx):
                            try:
                                w = float(e.value)
                            except (TypeError, ValueError):
                                return
                            cur = _load_peer_universe()
                            cur[idx]["weight"] = w
                            _save_peer_universe(cur)

                        # Options are the Fit Score tiers, but a peer may carry a benchmarking
                        # tier ("reference"/"primary") from peer discovery — include the current
                        # value so the select never rejects it (crashed the Target DB tab) and the
                        # stored tier isn't silently rewritten.
                        _tier_opts = ["core", "close", "large"]
                        _tier_val = p.get("tier", "close")
                        if _tier_val not in _tier_opts:
                            _tier_opts = _tier_opts + [_tier_val]
                        ui.select(_tier_opts, value=_tier_val,
                                  on_change=set_tier).props("dense outlined").classes("min-w-[90px]")
                        ui.number(value=p.get("weight", 1.0), min=0.1, max=5.0, step=0.5,
                                  on_change=set_weight).props("dense outlined").classes("min-w-[70px]")
                        ui.button("", on_click=do_delete).props("flat dense")

        render_manage_list()
        ui.label("Tier: \"core\" = closest-size direct comp (full Fit Score points, +5 conviction bonus) · "
                 "\"close\" = real but less-comparable peer (default) · \"large\" = mega-cap/weak-signal peer. "
                 "Weight scales Peer Conviction — higher for peers closest to this client's own size.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-style:italic;")
        ui.markdown("---")

        ui.label("Add a new peer ticker:").classes("font-bold").style("font-size:var(--fs-base);")
        with ui.row().classes("w-full gap-2"):
            new_ticker = ui.input("Ticker", placeholder="e.g. PAYO").classes("flex-1").props("outlined dense")
            new_name = ui.input("Company name", placeholder="e.g. Payoneer Global").classes("flex-1").props("outlined dense")
            new_sector = ui.input("Sector / focus", placeholder="e.g. Cross-Border Payments").classes("flex-1").props("outlined dense")
            new_ev = ui.input("EV/Rev (optional)", placeholder="EV/Rev").classes("min-w-[100px]").props("outlined dense")
        with ui.row().classes("w-full gap-2"):
            new_tier = ui.select(["core", "close", "large"], value="close", label="Fit Score tier").classes("min-w-[120px]").props("outlined dense")
            new_weight = ui.number("Fit Score weight", value=1.0, min=0.1, max=5.0, step=0.5).classes("min-w-[120px]").props("outlined dense")

        def add_peer():
            tkr = (new_ticker.value or "").upper().strip()
            nm = (new_name.value or "").strip()
            if not (tkr and nm):
                ui.notify("Enter at least a ticker and company name.", type="warning")
                return
            cur = _load_peer_universe()
            if tkr in [p["ticker"] for p in cur]:
                ui.notify(f"{tkr} is already in the peer universe.", type="warning")
                return
            try:
                ev_val = float(new_ev.value) if new_ev.value and new_ev.value.strip() else None
            except ValueError:
                ev_val = None
            cur.append({"ticker": tkr, "name": nm, "sector": new_sector.value or "Payments / Fintech", "ev_rev": ev_val,
                        "tier": new_tier.value or "close", "weight": float(new_weight.value or 1.0)})
            _save_peer_universe(cur)
            ui.notify(f"{tkr} — {nm} added and saved" + (f" with EV/Rev {ev_val}x." if ev_val else " — no EV/Rev yet, flagged as needing research."))
            new_ticker.value, new_name.value, new_sector.value, new_ev.value = "", "", "", ""
            new_tier.value, new_weight.value = "close", 1.0
            render_manage_list()
            _refresh()

        ui.button("Add", on_click=add_peer).props("dense")

        missing_ev = [p["ticker"] for p in peers if not p.get("ev_rev")]
        if missing_ev:
            ui.label(f"No EV/Revenue multiple on file for: {', '.join(missing_ev)} — these won't get a specific "
                     f"valuation comparison in outreach drafts until a multiple is added here.").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")

        ui.markdown("---")
        ui.label("Quick-add common fintech peers:").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
        quick_peers = [("PAYO", "Payoneer", "Cross-Border"), ("FLYW", "Flywire", "Ed/Healthcare Payments"),
                       ("RELY", "Remitly", "Digital Remittance"), ("EVTC", "EVERTEC", "LatAm Payments"),
                       ("I", "Inpay", "B2B Cross-Border")]
        with ui.row().classes("w-full gap-2"):
            for qt, qn, qs in quick_peers:
                existing_tickers = [p["ticker"] for p in peers]

                def add_quick(qt=qt, qn=qn, qs=qs):
                    cur = _load_peer_universe()
                    cur.append({"ticker": qt, "name": qn, "sector": qs, "ev_rev": None})
                    _save_peer_universe(cur)
                    ui.notify(f"{qt} added.")
                    render_manage_list()
                    _refresh()

                if qt in existing_tickers:
                    ui.label(f"{qt}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                else:
                    ui.button(f"+ {qt}", on_click=add_quick).props("flat dense")

    ui.markdown("---")
    all_tickers = [p["ticker"] for p in peers]

    # Data-freshness summary — shown BEFORE the selector so it's clear
    # up front which tickers this analysis can actually back with a real
    # SEC 13F filing right now vs. which are still on the original seed
    # guess (see _enrich_peer_holdings_with_live_13f). A ticker only
    # becomes "live" after someone clicks "Refresh 13F Institutional
    # Holders" on the SEC Intelligence tab for it — this list doesn't
    # trigger that fetch itself (13F refresh is heavy, see
    # core/sec_filings.py's module docstring).
    _live_tickers = [t for t in all_tickers if sec_filings.get_cached_13f_holders(t).get("quarter")]
    _seed_tickers = [t for t in all_tickers if t not in _live_tickers]
    _freshness_bits = []
    if _live_tickers:
        _freshness_bits.append(f"Live SEC 13F data: {', '.join(_live_tickers)}")
    if _seed_tickers:
        _freshness_bits.append(f"Seed data only (not yet 13F-refreshed): {', '.join(_seed_tickers)}")
    ui.label(" · ".join(_freshness_bits)).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
    if _seed_tickers:
        ui.label("Refresh a ticker from Settings → Data Sources "
                  "to replace its seed guess with a real filing.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-style:italic;")

    with ui.row().classes("w-full gap-4"):
        peer_select = ui.select(all_tickers, multiple=True, value=all_tickers,
                                 label="Select peers to cross-reference").classes("flex-1").props("outlined dense")
        aum_select = ui.select(["Any", "$100M+", "$500M+", "$1B+"], value="$100M+", label="Min fund AUM").classes("min-w-[140px]").props("outlined dense")

    results_area = ui.column().classes("w-full")

    def run_cross_targeting(from_click=False):
        # This already runs once automatically on page load (see the
        # unconditional call at the bottom of this function), so the results
        # below are visible before anyone touches the button. Clicking
        # "Run cross-targeting" without first changing the peer selection or
        # AUM filter re-renders the identical list — which reads as "the
        # button did nothing" even though it worked. The ui.notify() calls
        # below give every click a visible confirmation regardless of
        # whether the underlying result list actually changed.
        results_area.clear()
        sel = peer_select.value or []
        aum_min = {"Any": 0.0, "$100M+": 100.0, "$500M+": 500.0, "$1B+": 1000.0}[aum_select.value]
        with results_area:
            if not sel:
                ui.label("Select at least one peer ticker to run cross-targeting analysis.").style(f"color:{COLORS['text_muted']};")
                if from_click:
                    ui.notify("Select at least one peer ticker first.", type="warning")
                return
            for t in sel:
                info = next((p for p in peers if p["ticker"] == t), None)
                if info:
                    ui.label(f"{t} — {info['name']} · {info.get('sector','—')}").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

            cross_targets = [i for i in institutions
                              if not i["USIO_Holder"]
                              and any(p in i["Peer_Holdings"] for p in sel)
                              and _parse_aum_millions(i["AUM"]) >= aum_min]
            if not cross_targets:
                ui.label("No cross-target matches at this AUM threshold. This checks your tracked watchlist "
                         "institutions by exact legal-entity name against the live SEC 13F filer roster — a peer "
                         "can have real institutional holders (see Settings → SEC Intelligence) without any of them "
                         "being one of the specific institutions on your watchlist. Add more institutions to the "
                         "watchlist above, or lower the AUM threshold.").style(f"color:{COLORS['text_muted']};margin-top:6px;")
                if from_click:
                    ui.notify("Cross-targeting run — no matches at this AUM threshold.", type="warning")
                return
            ui.label(f"{len(cross_targets)} institutions own {', '.join(sel)} but NOT this client (AUM ≥ {aum_select.value}):").classes("font-bold").style("margin-top:6px;")
            for ct in sorted(cross_targets, key=lambda x: x["Engagement_Score"], reverse=True):
                overlap = [p for p in sel if p in ct["Peer_Holdings"]]
                # Same = confirmed-by-real-13F-filing marker as the
                # institution card's "Peers held" line — kept consistent so
                # the same overlap doesn't look "more real" in one place
                # than the other.
                ct_src = ct.get("Peer_Holdings_Source", {})
                overlap_labels = [p + (" " if ct_src.get(p) == "live" else "") for p in overlap]
                score = ct["Engagement_Score"]
                score_clr = "#15803D" if score >= 80 else "#B45309" if score >= 50 else "#94A3B8"
                with ui.card().classes("w-full").style(f"background:{COLORS['surface_hover_bg']};"):
                    with ui.row().classes("w-full justify-between items-center"):
                        ui.label(f"{ct['Fund']} — {ct['Type']} · AUM {ct['AUM']}").classes("font-bold").style(f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
                        ui.label(f"Score: {score}").style(f"color:{score_clr};font-weight:bold;")
                    ui.label(f"Owns: {', '.join(overlap_labels)} · IR visits (30d): {ct['IR_Visits_30d']} · "
                             f"Q1 call: {'Yes' if ct['Call_Listener'] else 'No'} · Does NOT own this client").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
            if from_click:
                ui.notify(f"Cross-targeting run — {len(cross_targets)} institution(s) found.", type="positive")

    ui.button("Run cross-targeting", on_click=lambda: run_cross_targeting(from_click=True)).props("color=primary dense")
    run_cross_targeting()

    ui.markdown("---")
    ui.label("Fit Score Ranking — All Confirmed Peer Holders (Live 13F)").classes("font-bold")
    ui.label(
        "Different scope from the watchlist-only Peer Cross-Targeting above: this scores EVERY institution with "
        "a confirmed live 13F position in your selected peers — not just names already on your tracked buy-side "
        "roster — using the full six-component Fit Score (Peer conviction /30, New-buyer /20, Comparability fit "
        "/15, Turnover /15, Purchasing power /10, Contactability /10 = /100, P1-P4 tier). Peer conviction/Fit are "
        "computed directly from the tier + weight set per peer above. Purchasing power is a real number for the "
        "top-ranked prospects shown below — each fund's whole 13F book, fetched on demand from its own SEC filing "
        "the first time it appears in a top result (falls back to a neutral placeholder until then). Turnover and "
        "Contactability are rough automated heuristics (name-pattern + existing-relationship matching) until "
        "real outcomes accumulate — see \"Which score signals are converting?\" below."
    ).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    fit_results_area = ui.column().classes("w-full")

    async def run_fit_score_ranking(from_click=False):
        # async + asyncio.to_thread: score_all_holders() now does a bounded
        # lazy backfill of real purchasing-power book totals for its top ~40
        # results (see fit_score.py's backfill_top_n, sec_filings.py's
        # ensure_book_totals) — up to ~40 sequential SEC network requests.
        # Same class of bug as the old blocking 13F-refresh buttons (see
        # refresh_13d13g/refresh_13f above): a plain synchronous handler here
        # would freeze the whole app's event loop for every connected user
        # long enough to trip "Connection lost." Running the scoring call in
        # a background thread keeps the UI responsive while it works.
        fit_results_area.clear()
        sel = peer_select.value or []
        if not sel:
            with fit_results_area:
                ui.label("Select at least one peer ticker above to run Fit Score ranking.").style(f"color:{COLORS['text_muted']};")
            if from_click:
                ui.notify("Select at least one peer ticker first.", type="warning")
            return
        if from_click:
            ui.notify("Scoring confirmed holders — fetching real purchasing-power figures for top results, this may take a few seconds.")
        try:
            scored = await asyncio.to_thread(fit_score.score_all_holders, sel, get_active_client_id())
        except Exception as e:
            with fit_results_area:
                ui.label(f"Fit Score ranking failed: {e}").style(f"color:{COLORS['warning']};")
            if from_click:
                ui.notify(f"Fit Score ranking failed: {e}", type="negative")
            return
        with fit_results_area:
            if not scored:
                ui.label("No confirmed holders yet for the selected peers — refresh 13F Institutional Holders "
                          "from Settings → Data Sources first.").style(f"color:{COLORS['text_muted']};")
                if from_click:
                    ui.notify("No confirmed 13F holders yet for these peers.", type="warning")
                return
            ui.label(f"{len(scored)} institution(s) scored, ranked by Fit Score:").classes("font-bold").style("margin-top:6px;")
            tier_clrs = {"P1": "#15803D", "P2": "#3B8A5C", "P3": "#B45309", "P4": "#94A3B8"}
            for r in scored[:40]:
                tier_clr = tier_clrs.get(r["tier"], "#94A3B8")
                def add_scored_prospect(r=r):
                    plist = _load_json("prospects.json", [])
                    if any(p.get("fund") == r["fund"] for p in plist):
                        ui.notify(f"{r['fund']} is already in the prospect queue.", type="warning")
                        return
                    plist.append({
                        "fund": r["fund"], "metro": "—", "style": r["tier_label"], "score": r["composite"],
                        "score_breakdown": r, "outcome": None,
                        "notes": f"Fit Score {r['composite']}/100 · {r['tier_label']} · holds {', '.join(r['peers_held'])} "
                                 f"· via Peer Cross-Targeting Fit Score ranking",
                        "added": datetime.now().strftime("%Y-%m-%d"), "source": "Fit Score ranking",
                    })
                    _save_json("prospects.json", plist)
                    ui.notify(f"{r['fund']} added to prospect queue.")

                with ui.card().classes("w-full").style(f"background:{COLORS['surface_hover_bg']};"):
                    with ui.row().classes("w-full justify-between items-center"):
                        ui.label(r["fund"]).classes("font-bold").style(f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
                        with ui.row().classes("items-center gap-2"):
                            ui.label(f"{r['tier_label']} · {r['composite']}/100").style(f"color:{tier_clr};font-weight:bold;font-size:var(--fs-base);")
                            ui.button("Add", on_click=add_scored_prospect).props("flat dense")
                    ui.label(f"Holds: {', '.join(r['peers_held'])}" +
                             (" · NEW buyer this quarter" if r["newbuyer_pts"] else "")).style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    with ui.row().classes("w-full gap-2").style("margin-top:2px;flex-wrap:wrap;"):
                        for label, val, maxv in [
                            ("Conviction", r["conviction"], 30), ("New-buyer", r["newbuyer_pts"], 20),
                            ("Fit", r["fit"], 15), ("Turnover", r["turnover_pts"], 15),
                            ("Purch. power", r["pp_pts"], 10), ("Contact", r["contact_pts"], 10),
                        ]:
                            ui.label(f"{label} {val}/{maxv}").style(
                                f"color:{COLORS['text_muted']};font-size:var(--fs-xs);background:{COLORS['surface_bg']};"
                                f"padding:1px 6px;border-radius:6px;")
                    ui.label(f"Turnover: {r['turnover_class']} ({r['turnover_why']}) · "
                              f"Purchasing power: {r['pp_class']} ({r['pp_why']}) · "
                              f"Contactability: {r['contact_class']} ({r['contact_why']})").style(
                        f"color:{COLORS['text_muted']};font-size:var(--fs-xs);font-style:italic;margin-top:2px;")
            if from_click:
                ui.notify(f"Fit Score ranking run — {len(scored)} institution(s) scored.", type="positive")

    ui.button("Rank by Fit Score", on_click=lambda: run_fit_score_ranking(from_click=True)).props("color=primary dense")

    with ui.expansion("Fit Score Weights — Re-weight the Components", value=False).classes("w-full"):
        ui.label("Current weights (must sum to 100). Same 'transparent, re-weightable' design as the source — "
                  "every component always shows in its own column, the composite is never a black box.").style(
            f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
        current_weights = fit_score.get_weights(get_active_client_id())
        weight_inputs = {}
        with ui.row().classes("w-full gap-2"):
            for key, label in [("conviction", "Conviction"), ("newbuyer", "New-buyer"), ("fit", "Fit"),
                                ("turnover", "Turnover"), ("pp", "Purch. power"), ("contact", "Contact")]:
                weight_inputs[key] = ui.number(label, value=current_weights.get(key, fit_score.DEFAULT_WEIGHTS[key]),
                                                min=0, max=100, step=1).classes("min-w-[100px]").props("outlined dense")

        def save_weight_changes():
            new_weights = {k: (inp.value or 0) for k, inp in weight_inputs.items()}
            total = sum(new_weights.values())
            if total != 100:
                ui.notify(f"Weights sum to {total}, not 100 — adjust before saving.", type="warning")
                return
            fit_score.save_weights(new_weights, get_active_client_id())
            ui.notify("Weights saved — re-run Fit Score ranking to see the effect.", type="positive")

        ui.button("Save weights", on_click=save_weight_changes).props("dense")

        ui.markdown("---")
        ui.label("Which score signals are converting? Re-weight the score:").classes("font-bold").style("font-size:var(--fs-base);")
        reweight_area = ui.column().classes("w-full")

        def run_reweight_suggestion():
            reweight_area.clear()
            result = fit_score.suggest_reweight(get_active_client_id())
            with reweight_area:
                if result["mode"] == "insufficient":
                    ui.label(result["message"]).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    return
                ui.label(f"From {result['n_pos']} positive (met/owns) and {result['n_neg']} pass outcomes:").style(
                    f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                for c in result["components"]:
                    move_str = f"+{c['move']}" if c["move"] > 0 else str(c["move"])
                    move_clr = "#15803D" if c["move"] > 0 else "#B91C1C" if c["move"] < 0 else COLORS["text_muted"]
                    ui.label(f"{c['label']}: current {c['cur']} → suggested {c['suggested']} ({move_str}) · "
                              f"lift {c['lift']} (avg {c['mean_pos']} converters vs {c['mean_neg']} passes)").style(
                        f"color:{move_clr};font-size:var(--fs-sm);")

        ui.button("Check for a re-weight suggestion", on_click=run_reweight_suggestion).props("dense")


# ─────────────────────────────────────────────────────────────────────────
# SEC Intelligence tab — 13D/13G ownership-stake alerts + 13F institutional
# holders for this client and its full peer universe (config.client_config
# .CP(), the same peer list Peer Cross-Targeting above manages). Data comes
# from core/sec_filings.py, which is cache-first: this tab reads whatever
# is cached (kept warm by app_nicegui.py's startup hook for 13D/13G) rather
# than blocking page render on a live SEC call — the two Refresh buttons
# below are the only things that hit the network from inside a page render,
# and both are explicit, user-initiated actions.
# ─────────────────────────────────────────────────────────────────────────
def _render_sec_intelligence_tab():
    ui.label("SEC Intelligence").classes("text-lg font-bold")
    ui.label(
        "Ownership-stake filings (13D/13G) and institutional holders (13F) for this client and its full peer "
        "universe — tracked tickers come straight from the Peer Cross-Targeting list in Target Database, so "
        "adding a competitor there adds it here too."
    ).style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    log = sec_filings.get_last_refresh_log()
    with ui.row().classes("w-full items-center gap-3").style("margin-top:6px;"):
        if log:
            ui.label(f"Last 13D/13G refresh: {log.get('finished_at', '—')[:19].replace('T', ' ')}").style(
                f"color:{COLORS['text_muted']};font-size:var(--fs-sm);"
            )
        else:
            ui.label("No refresh has run yet — the app's startup hook kicks one off automatically each launch, "
                      "or trigger one now:").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")

    # Surface a top-level 13F failure explicitly. Real bug found 2026-07-10:
    # if the SEC bulk download/unzip/parse itself fails (network error, SEC
    # blocking the request, an unexpected file format), refresh_all catches
    # it and stamps "13f_error" onto every entry in the run log — but NO
    # per-ticker 13F cache key ever gets written, since the failure happens
    # before refresh_13f_for_tracked_tickers's per-ticker loop even starts.
    # get_cached_13f_holders then returns its generic "not yet fetched"
    # default for every single ticker, indistinguishable from "nobody has
    # clicked Refresh 13F yet" — the real error was silently dropped, only
    # ever visible in this log dict which nothing rendered. Surfacing it
    # here means a top-level failure (e.g. this being the very first live
    # run against sec.gov from this machine) is visible instead of looking
    # like "no institutional data exists for any tracked ticker."
    if log:
        _13f_errors = {
            r["ticker"]: r["13f_error"]
            for r in log.get("results", [])
            if r.get("13f_error")
        }
        if _13f_errors:
            _sample_ticker, _sample_err = next(iter(_13f_errors.items()))
            with ui.row().classes("w-full").style(
                f"background:{COLORS['surface_bg']};border-left:4px solid {COLORS['danger']};"
                f"border-radius:6px;padding:8px 12px;margin-top:4px;"
            ):
                ui.label(
                    f"13F refresh failed for all {len(_13f_errors)} tracked ticker(s) — the SEC bulk "
                    f"download/parse itself errored before any institutional-holder data could be saved. "
                    f"This is not \"no data exists\"; it's a failed fetch. Error: {_sample_err}"
                ).style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);")

    # At-a-glance summary of what's cached across the tracked universe.
    _sec_tickers = sec_filings.tracked_tickers()
    _sec_13d = sum(len(sec_filings.get_cached_13d_13g(t, refresh_if_stale=False).get("filings", []))
                   for t, _n in _sec_tickers)
    _sec_hold = sum(len(sec_filings.get_cached_13f_holders(t).get("holders", []))
                    for t, _n in _sec_tickers)
    with ui.row().classes("w-full gap-3").style("margin-top:8px;"):
        _hub_metric("Tickers tracked", len(_sec_tickers), "Client + peer universe")
        _hub_metric("13D/13G on file", _sec_13d, "Ownership-stake filings")
        _hub_metric("Institutional holders", _sec_hold, "From latest 13F")

    # Data-provenance rollup — how big the tracked universe is and where every
    # name came from. This is the honest, concrete answer to "why isn't the
    # database bigger?": it grows from authoritative sources, each tagged.
    _cid = get_active_client_id()
    _merged = _merge_sec_universe([dict(i) for i in get_seed_buyside_institutions(_cid)], _cid)
    from collections import Counter as _Counter
    _by_src = _Counter(i.get("Source", "Seed (demo)") for i in _merged)
    try:
        _nobo_inst = sum(1 for h in nobo_engine.get_active_pulls(_cid)["current"]["holders"]
                         if h.get("type") == "Institutional")
    except Exception:
        _nobo_inst = 0
    ui.label("Universe by data source").classes("font-bold").style("margin-top:12px;")
    ui.label(f"{len(_merged)} distinct names in the tracked universe · plus {_nobo_inst} institutional NOBO "
             "holders (Ownership → NOBO Ownership). Hand-typed seed today; refresh below to grow the real "
             "SEC-sourced names live from EDGAR.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
    with ui.row().classes("w-full gap-2 flex-wrap").style("margin-top:4px;"):
        for _src, _n in _by_src.most_common():
            ui.label(f"{_src}: {_n}").style(
                f"background:{_source_color(_src)};color:#fff;border-radius:6px;padding:2px 10px;"
                "font-size:var(--fs-sm);font-weight:600;")

    sec_container = ui.column().classes("w-full gap-2").style("margin-top:8px;")

    def render_sec_list():
        sec_container.clear()
        with sec_container:
            for ticker, name in sec_filings.tracked_tickers():
                d13 = sec_filings.get_cached_13d_13g(ticker, refresh_if_stale=False)
                f13 = sec_filings.get_cached_13f_holders(ticker)
                filings = d13.get("filings", [])
                holders = f13.get("holders", [])
                label = f"{ticker} — {name} · {len(filings)} 13D/13G on file"
                if f13.get("quarter"):
                    label += f" · {len(holders)} institutional holders ({f13['quarter']})"
                with ui.expansion(label).classes("w-full"):
                    if d13.get("_error"):
                        ui.label(f"13D/13G fetch error: {d13['_error']}").style(f"color:{COLORS['warning']};font-size:var(--fs-sm);")
                    if not filings:
                        ui.label("No 13D/13G filings cached yet for this ticker.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    else:
                        for f in filings[:10]:
                            ui.label(f"{f['date']} · {f['form']} · {f['title']}").style(f"color:{COLORS['text_body']};font-size:var(--fs-sm);")
                            if f.get("link"):
                                ui.link("View on EDGAR ↗", f["link"], new_tab=True).style("font-size:var(--fs-sm);")

                    ui.markdown("---")
                    if f13.get("_error") and f13["_error"] != "not yet fetched":
                        ui.label(f"13F fetch error: {f13['_error']}").style(f"color:{COLORS['warning']};font-size:var(--fs-sm);")
                    if not holders:
                        ui.label("No 13F institutional-holder data cached yet — use 'Refresh 13F Institutional "
                                 "Holders' below.").style(f"color:{COLORS['text_muted']};font-size:var(--fs-sm);")
                    else:
                        for h in holders[:10]:
                            # size_known=False: confirmed via EDGAR full-text search, but this holder's
                            # exact share/dollar size wasn't fetched (see
                            # sec_filings._search_holders_for_ticker's docstring) — don't show the
                            # sentinel shares=1/value=0 as if it were a real reported figure.
                            size_text = f"{h['shares']:,} shares (${h['value']:,} reported)" if h.get("size_known", True) \
                                else "position size not available"
                            ui.label(f"{h['filer']} — {size_text}").style(
                                f"color:{COLORS['text_body']};font-size:var(--fs-sm);"
                            )

    render_sec_list()

    # Was: plain `def` handlers calling sec_filings.refresh_all(...)
    # directly — a real, blocking network call (13F alone downloads a
    # multi-ticker, tens-of-MB SEC bulk dataset). NiceGUI runs a sync
    # click handler straight on the server's single asyncio event loop, so
    # this froze the ENTIRE app for every connected user until the whole
    # download finished (or hung/timed out) — not just this button, and
    # often not even flushing the "Refreshing..." toast first, since a
    # non-yielding sync call blocks outgoing websocket messages too. That's
    # exactly what "I clicked it and nothing happened" looks like from the
    # browser: no toast, no error, no visible progress, just silence.
    # app_nicegui.py's startup hooks already established the correct fix
    # for this exact class of call (asyncio.to_thread, see
    # _kick_off_sec_refresh) — applying the same pattern here so a manual
    # refresh behaves like the automatic startup one: the UI stays
    # responsive and the toast/notify actually show up immediately.
    async def refresh_13d13g():
        # force_13d13g=True: a manual "Refresh Now" click must always hit
        # SEC live, not silently return an already-cached snapshot because
        # it's under the 24h TTL (see get_cached_13d_13g's force param —
        # without this, repeated clicks reported "refresh complete" while
        # actually doing nothing, which is exactly what "the data is stuck
        # on old filings" looks like from the browser).
        ui.notify("Refreshing 13D/13G filings from SEC EDGAR — this hits the network now and may take a few seconds per ticker.")
        try:
            await asyncio.to_thread(sec_filings.refresh_all, False, True)
            ui.notify("13D/13G refresh complete.", type="positive")
        except Exception as e:
            ui.notify(f"Refresh failed: {e}", type="negative")
        render_sec_list()

    async def refresh_13f():
        # As of 2026-07-12 this runs one lightweight EDGAR full-text search per
        # tracked ticker (see sec_filings.refresh_13f_for_tracked_tickers) instead
        # of downloading SEC's entire quarterly bulk dataset — much faster, but
        # still a handful of sequential network requests, so it stays behind
        # asyncio.to_thread and an explicit button rather than a page render.
        ui.notify(
            "Refreshing 13F institutional holders from SEC EDGAR (targeted per-ticker search) — "
            "this may take a few seconds per tracked ticker.",
            type="warning",
        )
        try:
            await asyncio.to_thread(sec_filings.refresh_all, True)
            ui.notify("13F refresh complete.", type="positive")
        except Exception as e:
            ui.notify(f"Refresh failed: {e}", type="negative")
        render_sec_list()

    async def refresh_universe():
        # One-click: pull COMPLETE 13F holders (bulk structured dataset,
        # filtered by CUSIP — exact shares, every holder) for USIO and every
        # peer, plus 13D/13G, then report how many real names populate the
        # universe by source. This is the "make it bigger, accurately" action.
        ui.notify("Refreshing from SEC EDGAR — downloading the ~100MB quarterly 13F bulk dataset and pulling "
                  "13D/13G. This is the complete, by-CUSIP holder list; it takes 1–3 minutes.", type="warning")
        try:
            await asyncio.to_thread(sec_filings.refresh_13f_bulk_all, sec_filings.tracked_tickers())
            await asyncio.to_thread(sec_filings.refresh_all, False, True)
            recs = _sec_universe_records(get_active_client_id())
            bs = _Counter(r["Source"] for r in recs)
            summary = ", ".join(f"{n} {s}" for s, n in bs.most_common()) or "no new names this quarter"
            ui.notify(f"SEC EDGAR refresh complete — {len(recs)} real names now in your universe ({summary}). "
                      "Reopen Buy-Side Intelligence to see them tagged by source.", type="positive")
        except Exception as e:
            ui.notify(f"Refresh failed: {e}", type="negative")
        render_sec_list()

    # Data pulls (13F / 13D-G / full universe) are triggered only from
    # Settings → Data Sources now — one deliberate place for an expensive SEC
    # refresh, so opening a tab never kicks one off. This tab shows the cached
    # result; reopen it after a pull to see fresh holders. (The refresh_* handlers
    # above are unused here — the live controls live in
    # settings_page._render_data_sources.)
    ui.markdown("---")
    with ui.card().classes("w-full").style(
            f"background:{COLORS['surface_bg']};border:1px solid {COLORS['border']};"
            f"border-left:4px solid {COLORS['accent']};"):
        ui.label("Data pulls live in Settings").classes("font-bold").style(
            f"color:{COLORS['text_heading']};font-size:var(--fs-base);")
        ui.label("Refreshing SEC 13F / 13D-G data hits the network and can take minutes, so it's kicked off only "
                 "from Settings → Data Sources — never by navigating here.").style(
            f"color:{COLORS['text_secondary']};font-size:var(--fs-sm);")
        ui.button("Open Settings → Data Sources", icon="settings",
                  on_click=lambda: nav.go_to("Settings", "Data Sources")).props(
            "flat dense color=primary").style("margin-top:4px;")

    ui.markdown("---")
    ui.label(
        "If a ticker fails to resolve to a SEC CIK (rare — recent IPOs, ADRs, or a ticker change), add a manual "
        "override in core/sec_filings.py's MANUAL_CIK_OVERRIDES dict."
    ).style(f"color:{COLORS['text_muted']};font-size:var(--fs-xs);")
