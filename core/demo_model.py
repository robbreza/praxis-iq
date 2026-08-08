"""core/demo_model.py — generator for the illustrative demo tenant's sell-side analyst model.

A single, reproducible source for the clean NLKP model .xlsx used by the demo tenant's IR-Inbox
"model" ingestion demo. Both the seeder (scripts/seed_illustrative_tenant.py) and the standalone
rebuild script (scripts/build_demo_model.py) call build_model_xlsx(), so the file is generated the
same way everywhere — properly Unicode-encoded, no mojibake.

The model is illustrative: a fictional covering analyst (Cascade Securities / J. Meridian, CFA) on
Northlake Payments (NLKP). Nothing here is real data.
"""
import io

# Analyst / file identity — shared so the document, the inbox item, and the download all agree.
FIRM = "Cascade Securities"
ANALYST = "James Meridian"
FILENAME = "NLKP_Q2_2026_Model.xlsx"
CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SENDER_EMAIL = "jmeridian@cascade-research.com"

# The numbers the classifier would have "extracted" from the model — prefilled into the inbox
# review card. Kept in lock-step with the Q2'26E column of build_model_xlsx().
EXTRACTED = {"period": "Q2 2026E", "rating": "Buy", "price_target": 42.70,
             "eps_est": 0.13, "revenue_est": 25.9, "ebitda_est": None}

PERIODS = ["Q1'26A", "Q2'26E", "Q3'26E", "Q4'26E", "FY2026E", "FY2027E"]


def build_model_xlsx() -> bytes:
    """Construct the clean workbook and return its bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NLKP Model"

    ink = "1E293B"
    accent = "1E40AF"
    grey = "64748B"
    thin = Side(style="thin", color="D3DBE4")
    box = Border(bottom=thin)

    def put(cell, value, *, bold=False, size=11, color=ink, align="left", fill=None,
            italic=False, border=False, num=None):
        c = ws[cell]
        c.value = value
        c.font = Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if border:
            c.border = box
        if num:
            c.number_format = num

    # ── Header ──
    put("A1", "Northlake Payments, Inc.  (NASDAQ: NLKP)", bold=True, size=15, color=accent)
    put("A2", "Sell-Side Earnings Model — Q2 2026 Preview", bold=True, size=11, color=ink)
    put("A3", "Cascade Securities · Equity Research   |   J. Meridian, CFA   |   August 3, 2026",
        size=10, color=grey, italic=True)

    put("A5", "Rating", bold=True); put("B5", "BUY", bold=True, color="15803D")
    put("A6", "Price target"); put("B6", 42.70, num='$#,##0.00', align="right")
    put("A7", "Last price"); put("B7", 32.84, num='$#,##0.00', align="right")
    put("A8", "Implied upside"); put("B8", 0.30, num='0%', align="right", color="15803D", bold=True)

    # ── P&L table ──
    hdr = 10
    put(f"A{hdr}", "($ in millions, except per-share)", bold=True, size=10, color=grey, border=True)
    for j, p in enumerate(PERIODS):
        col = chr(ord("B") + j)
        put(f"{col}{hdr}", p, bold=True, size=10, color=ink, align="right", fill="EEF2F7", border=True)

    def row(label, vals, r, *, indent=False, num='#,##0.0', bold=False, color=ink):
        put(f"A{r}", ("   " + label) if indent else label, bold=bold, color=color)
        for j, v in enumerate(vals):
            col = chr(ord("B") + j)
            put(f"{col}{r}", v, align="right", num=(num if isinstance(v, (int, float)) else None),
                bold=bold, color=color)

    row("Revenue", [24.6, 25.9, 26.7, 27.4, 104.6, 117.2], 11, bold=True)
    row("Payment Processing", [18.7, 19.5, 19.9, 20.3, 78.4, 85.9], 12, indent=True, color=grey)
    row("Software & Platform", [5.9, 6.4, 6.8, 7.1, 26.2, 31.3], 13, indent=True, color=grey)
    row("Revenue growth (YoY)", [0.081, 0.089, 0.093, 0.101, 0.092, 0.120], 14, num='0.0%', color=grey)
    row("Gross profit", [6.5, 6.9, 7.2, 7.5, 28.1, 32.4], 15)
    row("Gross margin", [0.264, 0.266, 0.270, 0.274, 0.269, 0.276], 16, num='0.0%', color=grey)
    row("Operating income", [2.1, 2.3, 2.5, 2.7, 9.6, 12.1], 17)
    row("Operating margin", [0.085, 0.089, 0.094, 0.099, 0.092, 0.103], 18, num='0.0%', color=grey)
    row("Net income", [1.6, 1.7, 1.9, 2.0, 7.2, 9.1], 19)
    row("Diluted EPS", [0.11, 0.13, 0.13, 0.16, 0.53, 0.63], 20, num='$0.00', bold=True)
    row("Diluted shares (M)", [14.3, 14.3, 14.4, 14.4, 14.4, 14.5], 21, num='#,##0.0', color=grey)

    # ── Valuation ──
    put("A23", "Valuation", bold=True, size=11, color=accent)
    val = [
        ("Market capitalization", "$473M", "$32.84 × 14.4M shares"),
        ("Less: net cash", "($18M)", "cash less debt, 6/30 est."),
        ("Enterprise value", "$455M", ""),
        ("EV / Gross Profit (FY26E)", "16.2x", "peer median ~13x — premium on mix shift + growth"),
        ("EV / Gross Profit (FY27E)", "14.0x", ""),
        ("Price target basis", "$42.70", "~18x FY27E gross profit, ~52x FY27E EPS"),
    ]
    for i, (k, v, note) in enumerate(val):
        r = 24 + i
        put(f"A{r}", k)
        put(f"B{r}", v, align="right", bold=True)
        if note:
            put(f"C{r}", note, color=grey, italic=True, size=10)

    # ── Thesis ──
    tr = 24 + len(val) + 1
    put(f"A{tr}", "Thesis", bold=True, size=11, color=accent)
    thesis = (
        "NLKP is compounding Software & Platform revenue (25% of mix, +19% YoY) on top of a steady "
        "payment-processing base, lifting blended gross margin ~70bps/yr. We model Q2 revenue of "
        "$25.9M (Street $25.7M) and EPS of $0.12. A guide reiteration plus a software-attach update "
        "should support the multiple; our $42.70 target is ~18x FY27E gross profit."
    )
    put(f"A{tr + 1}", thesis, color=ink, size=10)
    ws[f"A{tr + 1}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{tr + 1}:G{tr + 4}")

    ws.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
